# Standard library imports
import sys
import logging
import os
import json
import threading
import re
import traceback
import asyncio
from http.server import BaseHTTPRequestHandler, HTTPServer
from typing import Dict, List, Optional, Any, Union
import io
from datetime import datetime

# Load environment variables from config.env
try:
    from dotenv import load_dotenv
    load_dotenv('config.env')
except ImportError:
    pass  # dotenv not available, use system environment variables

# Configure matplotlib for headless environments without filesystem dependencies
# Note: No filesystem configuration needed for in-memory operations

# Third-party imports
import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import okama as ok

# Optional Excel support
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl library not available. Excel export will use CSV format.")

# Configure matplotlib backend for headless environments (CI/CD)
if os.getenv('DISPLAY') is None and os.getenv('MPLBACKEND') is None:
    matplotlib.use('Agg')

# Suppress matplotlib warnings for missing CJK glyphs
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='matplotlib')

# Optional imports
try:
    import tabulate
    TABULATE_AVAILABLE = True
except ImportError:
    TABULATE_AVAILABLE = False
    print("Warning: tabulate library not available. Using simple text formatting.")

# Telegram imports
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove, KeyboardButton
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler

# Check Python version compatibility
if sys.version_info < (3, 7):
    print("ERROR: Python 3.7+ required. Current version:", sys.version)
    raise RuntimeError("Python 3.7+ required")

# Local imports
from config import Config
from services.yandexgpt_service import YandexGPTService
from services.tushare_service import TushareService
from services.gemini_service import GeminiService
from services.examples_service import ExamplesService

from services.chart_styles import chart_styles
from services.context_store import JSONUserContextStore

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Health check for deployment
def health_check():
    """Simple health check for deployment"""
    logger.info(f"✅ Environment: {'PRODUCTION' if os.getenv('PRODUCTION') else 'LOCAL'}")
    logger.info(f"✅ Python version: {sys.version}")
    logger.info(f"✅ Bot token configured: {'Yes' if Config.TELEGRAM_BOT_TOKEN else 'No'}")
    return True

class ShansAi:
    """Telegram bot for financial analysis with Okama library"""
    
    def __init__(self):
        """Initialize the bot with required services"""
        Config.validate()
        
        # Initialize logger
        self.logger = logging.getLogger(__name__)
        
        # Initialize services
        self.yandexgpt_service = YandexGPTService()
        self.chart_styles = chart_styles
        self.examples_service = ExamplesService()
        
        # Initialize Tushare service if API key is available
        try:
            self.tushare_service = TushareService()
        except ValueError:
            self.tushare_service = None
            self.logger.warning("Tushare service not initialized - API key not provided")
        
        # Initialize Gemini service for data analysis
        try:
            self.gemini_service = GeminiService()
            if self.gemini_service.is_available():
                self.logger.info("Gemini service initialized successfully")
            else:
                self.logger.warning("Gemini service not available - check credentials")
                # Log detailed status for debugging
                status = self.gemini_service.get_service_status()
                self.logger.info(f"Gemini status: {status}")
        except Exception as e:
            self.gemini_service = None
            self.logger.warning(f"Gemini service not initialized: {e}")
        
        # Initialize simple chart analysis as fallback
        
        # Known working asset symbols for suggestions (legacy)
        self.known_assets = {
            'US': ['VOO.US', 'SPY.US', 'QQQ.US', 'AGG.US', 'AAPL.US', 'TSLA.US', 'MSFT.US'],
            'INDX': ['RGBITR.INDX', 'MCFTR.INDX', 'SPX.INDX', 'IXIC.INDX'],
            'COMM': ['GC.COMM', 'SI.COMM', 'CL.COMM', 'BRENT.COMM'],
            'FX': ['EURUSD.FX', 'GBPUSD.FX', 'USDJPY.FX'],
            'MOEX': ['SBER.MOEX', 'GAZP.MOEX', 'LKOH.MOEX'],
            'LSE': ['VOD.LSE', 'HSBA.LSE', 'BP.LSE'],
            'SSE': ['600000.SH', '000001.SH'],
            'SZSE': ['000001.SZ', '399005.SZ'],
            'BSE': ['900001.BJ', '800001.BJ'],
            'HKEX': ['00001.HK', '00700.HK']
        }
        
        # Risk-free rate mapping for different currencies and regions
        self.risk_free_rate_mapping = {
            'USD': 'US_EFFR.RATE',  # US Federal Reserve Effective Federal Funds Rate
            'EUR': 'EU_DFR.RATE',   # European Central Bank key rate
            'GBP': 'UK_BR.RATE',    # Bank of England Bank Rate
            'RUB': 'RUS_CBR.RATE',  # Bank of Russia key rate
            'CNY': 'CHN_LPR1.RATE', # China one-year loan prime rate
            'JPY': 'US_EFFR.RATE',  # Use US rate as proxy for JPY (no specific JPY rate available)
            'CHF': 'EU_DFR.RATE',   # Use EU rate as proxy for CHF
            'CAD': 'US_EFFR.RATE',  # Use US rate as proxy for CAD
            'AUD': 'US_EFFR.RATE',  # Use US rate as proxy for AUD
            'ILS': 'ISR_IR.RATE',   # Bank of Israel interest rate
            'HKD': None,            # Hong Kong Dollar - not supported by okama, use fixed rate
        }
        
        # User session storage (in-memory for fast access)
        self.user_sessions = {}
        # Persistent context store
        self.context_store = JSONUserContextStore()

    def get_risk_free_rate(self, currency: str, period_years: float = None) -> float:
        """
        Get appropriate risk-free rate for given currency using okama rates
        
        Args:
            currency: Currency code (USD, EUR, RUB, etc.)
            period_years: Investment period in years (for selecting appropriate rate)
            
        Returns:
            Risk-free rate as decimal (e.g., 0.05 for 5%)
        """
        try:
            # Special handling for HKD - fixed rate (as requested)
            if currency.upper() == 'HKD':
                fixed_rate = 0.0285  # 2.85% fixed rate for Hong Kong Dollar
                self.logger.info(f"Using fixed risk-free rate for HKD: {fixed_rate:.4f}")
                return fixed_rate
            
            # Special handling for RUB - use more realistic rates based on current OFZ yields
            # Skip okama API for RUB as it returns unrealistic values (33%+)
            if currency.upper() == 'RUB':
                if period_years is not None:
                    if period_years <= 0.25:  # 3 months or less
                        fixed_rate = 0.08  # 8% - OFZ 3M approximation
                    elif period_years <= 0.5:  # 6 months or less
                        fixed_rate = 0.085  # 8.5% - OFZ 6M approximation
                    elif period_years <= 1.0:  # 1 year or less
                        fixed_rate = 0.09  # 9% - OFZ 1Y approximation
                    elif period_years <= 3.0:  # 3 years or less
                        fixed_rate = 0.095  # 9.5% - OFZ 3Y approximation
                    elif period_years <= 5.0:  # 5 years or less
                        fixed_rate = 0.10  # 10% - OFZ 5Y approximation
                    else:  # More than 5 years
                        fixed_rate = 0.105  # 10.5% - OFZ 10Y approximation
                else:
                    fixed_rate = 0.09  # 9% - default OFZ 1Y rate
                
                self.logger.info(f"Using OFZ-based risk-free rate for RUB ({period_years} years): {fixed_rate:.4f}")
                return fixed_rate
            
            # Try to get rate from okama.Rate for other currencies
            try:
                rate_data = self._get_okama_rate_data(currency, period_years)
                if rate_data is not None:
                    self.logger.info(f"Using okama rate for {currency}: {rate_data:.4f}")
                    return rate_data
            except Exception as e:
                self.logger.warning(f"Could not get okama rate for {currency}: {e}")
            
            # Special handling for CNY - period-dependent rates
            if currency.upper() == 'CNY':
                if period_years is not None and period_years > 5:
                    fixed_rate = 0.04  # 4% - 5-year rate approximation
                else:
                    fixed_rate = 0.035  # 3.5% - 1-year rate approximation
                
                self.logger.info(f"Using period-based risk-free rate for CNY ({period_years} years): {fixed_rate:.4f}")
                return fixed_rate
            
            # For all other currencies, use fallback rates directly
            # This avoids okama API calls that fail due to RATE namespace not being supported
            self.logger.info(f"Using fallback risk-free rate for {currency} (okama RATE namespace not supported)")
            
        except Exception as e:
            self.logger.warning(f"Error in risk-free rate calculation for {currency}: {e}")
        
        # Fallback rates if okama data is not available
        fallback_rates = {
            'USD': 0.05,  # 5% - current Fed funds rate
            'EUR': 0.04,  # 4% - current ECB rate
            'GBP': 0.05,  # 5% - current BoE rate
            'RUB': 0.10,  # 10% - OFZ 5Y rate (updated to realistic current rate)
            'CNY': 0.035, # 3.5% - current LPR rate
            'JPY': 0.05,  # 5% - use US rate as proxy
            'CHF': 0.04,  # 4% - use EU rate as proxy
            'CAD': 0.05,  # 5% - use US rate as proxy
            'AUD': 0.05,  # 5% - use US rate as proxy
            'ILS': 0.045, # 4.5% - current BoI rate
            'HKD': 0.0285, # 2.85% - Hong Kong Dollar fixed rate (not supported by okama)
        }
        
        fallback_rate = fallback_rates.get(currency.upper(), 0.05)  # Default to 5%
        self.logger.info(f"Using fallback risk-free rate for {currency}: {fallback_rate:.4f}")
        return fallback_rate

    def _get_okama_rate_data(self, currency: str, period_years: float = None) -> float:
        """
        Get risk-free rate from okama.Rate for given currency and period
        
        Args:
            currency: Currency code
            period_years: Investment period in years
            
        Returns:
            Risk-free rate as decimal
        """
        try:
            # Map currencies to appropriate okama rate symbols
            rate_symbols = self._get_rate_symbol_for_currency(currency, period_years)
            
            if not rate_symbols:
                return None
            
            # Try to get rate data from okama
            for rate_symbol in rate_symbols:
                try:
                    rate_obj = ok.Rate(rate_symbol)
                    rate_values = rate_obj.values_monthly
                    
                    if rate_values is not None and len(rate_values) > 0:
                        # Calculate average rate over the period
                        avg_rate = rate_values.mean()
                        return float(avg_rate)
                        
                except Exception as e:
                    self.logger.debug(f"Could not get rate for {rate_symbol}: {e}")
                    continue
            
            return None
            
        except Exception as e:
            self.logger.warning(f"Error getting okama rate data: {e}")
            return None

    def _get_rate_symbol_for_currency(self, currency: str, period_years: float = None) -> list:
        """
        Get appropriate rate symbols for currency and period
        
        Args:
            currency: Currency code
            period_years: Investment period in years
            
        Returns:
            List of rate symbols to try
        """
        currency_upper = currency.upper()
        
        if currency_upper == 'USD':
            return ['US_EFFR.RATE']
        elif currency_upper == 'EUR':
            return ['EU_DFR.RATE', 'EU_MLR.RATE', 'EU_MRO.RATE']
        elif currency_upper == 'RUB':
            if period_years is not None:
                if period_years <= 0.25:  # 3 months or less
                    return ['RUONIA.RATE', 'RUONIA_AVG_1M.RATE']
                elif period_years <= 0.5:  # 6 months or less
                    return ['RUONIA_AVG_3M.RATE', 'RUONIA_AVG_1M.RATE']
                elif period_years <= 1.0:  # 1 year or less
                    return ['RUONIA_AVG_6M.RATE', 'RUONIA_AVG_3M.RATE']
                else:  # More than 1 year
                    return ['RUS_CBR.RATE', 'RUONIA_AVG_6M.RATE']
            else:
                return ['RUS_CBR.RATE', 'RUONIA_AVG_6M.RATE']
        elif currency_upper == 'CNY':
            if period_years is not None and period_years > 5:
                return ['CHN_LPR5.RATE']
            else:
                return ['CHN_LPR1.RATE']
        elif currency_upper == 'GBP':
            return ['UK_BR.RATE']
        elif currency_upper == 'ILS':
            return ['ISR_IR.RATE']
        else:
            return []

    def calculate_sharpe_ratio(self, returns: Union[float, pd.Series], volatility: float, 
                              currency: str = 'USD', period_years: float = None, 
                              asset_data: Any = None) -> float:
        """
        Calculate Sharpe ratio using appropriate risk-free rate for currency
        
        Args:
            returns: Annual return (float) or returns series (pd.Series)
            volatility: Annual volatility
            currency: Currency code for risk-free rate selection
            period_years: Investment period in years
            asset_data: Okama asset/portfolio object (optional, for direct calculation)
            
        Returns:
            Sharpe ratio
        """
        try:
            # If we have asset_data with get_sharpe_ratio method, use it with appropriate risk-free rate
            if asset_data is not None and hasattr(asset_data, 'get_sharpe_ratio'):
                risk_free_rate = self.get_risk_free_rate(currency, period_years)
                sharpe_ratio = asset_data.get_sharpe_ratio(rf_return=risk_free_rate)
                
                # Handle different return types
                if hasattr(sharpe_ratio, 'iloc'):
                    return float(sharpe_ratio.iloc[0])
                elif hasattr(sharpe_ratio, '__getitem__'):
                    return float(sharpe_ratio[0])
                else:
                    return float(sharpe_ratio)
            
            # Manual calculation
            if isinstance(returns, pd.Series):
                # Calculate annual return from series
                if len(returns) > 0:
                    annual_return = (1 + returns).prod() ** (252 / len(returns)) - 1  # Annualized
                else:
                    annual_return = 0.0
            else:
                annual_return = float(returns)
            
            # Get appropriate risk-free rate
            risk_free_rate = self.get_risk_free_rate(currency, period_years)
            
            # Calculate Sharpe ratio: (return - risk_free_rate) / volatility
            if volatility > 0:
                sharpe_ratio = (annual_return - risk_free_rate) / volatility
                self.logger.info(f"Calculated Sharpe ratio: ({annual_return:.4f} - {risk_free_rate:.4f}) / {volatility:.4f} = {sharpe_ratio:.4f}")
                return sharpe_ratio
            else:
                self.logger.warning("Volatility is zero or negative, cannot calculate Sharpe ratio")
                return 0.0
                
        except Exception as e:
            self.logger.warning(f"Failed to calculate Sharpe ratio: {e}")
            return 0.0
        
        # User history management
        self.user_history: Dict[int, List[dict]] = {}       # chat_id -> list[{"role": "...", "parts": [str]}]
        self.context_enabled: Dict[int, bool] = {}          # chat_id -> bool
        self.MAX_HISTORY_MESSAGES = 20
        self.MAX_TELEGRAM_CHUNK = 4000

    def clean_symbol(self, symbol: str) -> str:
        """Очищает символ от случайных символов и нормализует его"""
        if not symbol:
            return symbol
            
        # Удаляем обратные слеши и другие проблемные символы
        cleaned = symbol.replace('\\', '').replace('/', '').replace('"', '').replace("'", '')
        
        # Удаляем лишние пробелы
        cleaned = cleaned.strip()
        
        # Удаляем множественные пробелы
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        # Нормализуем namespace (конвертируем lowercase в uppercase)
        cleaned = self._normalize_symbol_namespace(cleaned)
        
        return cleaned
    
    def _is_chinese_or_hongkong_symbol(self, symbol: str) -> bool:
        """Проверяет, является ли символ китайским или гонконгским"""
        if not symbol:
            return False
            
        # Проверяем namespace
        if '.' in symbol:
            namespace = symbol.split('.')[1].upper()
            # Китайские биржи
            if namespace in ['SSE', 'SZSE', 'SH', 'SZ']:
                return True
            # Гонконгская биржа
            if namespace in ['HK', 'HKEX']:
                return True
        
        # Проверяем сам тикер на китайские символы
        # Китайские символы в диапазоне Unicode
        for char in symbol:
            if '\u4e00' <= char <= '\u9fff':  # Основной диапазон китайских иероглифов
                return True
            if '\u3400' <= char <= '\u4dbf':  # Расширенный диапазон A
                return True
            if '\u20000' <= ord(char) <= '\u2a6df':  # Расширенный диапазон B
                return True
        
        return False
    
    def _clean_symbol_for_parsing(self, symbol: str) -> str:
        """Очищает символ для парсинга портфеля, сохраняя двоеточие"""
        if not symbol:
            return symbol
            
        # Удаляем обратные слеши и другие проблемные символы, но НЕ двоеточие
        cleaned = symbol.replace('\\', '').replace('/', '').replace('"', '').replace("'", '')
        
        # Удаляем лишние пробелы
        cleaned = cleaned.strip()
        
        # Удаляем множественные пробелы
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        # Нормализуем namespace (конвертируем lowercase в uppercase)
        cleaned = self._normalize_symbol_namespace_for_parsing(cleaned)
        
        return cleaned
    
    def _normalize_symbol_namespace_for_parsing(self, symbol: str) -> str:
        """
        Нормализовать регистр namespace в символе для парсинга портфеля
        
        Args:
            symbol: Символ в формате TICKER.NAMESPACE
            
        Returns:
            str: Символ с нормализованным namespace (uppercase)
        """
        if '.' not in symbol:
            return symbol
        
        ticker, namespace = symbol.split('.', 1)
        
        # Known namespace mappings (lowercase -> uppercase)
        namespace_mappings = {
            'moex': 'MOEX',
            'us': 'US',
            'lse': 'LSE',
            'xetr': 'XETR',
            'xfra': 'XFRA',
            'xstu': 'XSTU',
            'xams': 'XAMS',
            'xtae': 'XTAE',
            'pif': 'PIF',
            'fx': 'FX',
            'cc': 'CC',
            'indx': 'INDX',
            'comm': 'COMM',
            'hk': 'HK',
            'sh': 'SH',
            'sz': 'SZ'
        }
        
        # Convert namespace to uppercase if it's in our mappings
        normalized_namespace = namespace_mappings.get(namespace.lower(), namespace.upper())
        
        return f"{ticker}.{normalized_namespace}"
    
    def smart_parse_portfolio_input(self, text: str) -> Dict[str, Any]:
        """
        Умный парсинг ввода портфеля с прощением мелких ошибок
        
        Поддерживает различные форматы:
        - Стандартный: "SBER.MOEX:0.3, GAZP.MOEX:0.7"
        - Список символов: "SBER.MOEX, GAZP.MOEX, LKOH.MOEX" (равные доли)
        - Смешанный: "SBER.MOEX:0.3, GAZP.MOEX, LKOH.MOEX:0.2" (остаток распределяется поровну)
        
        Args:
            text: Входной текст для парсинга
            
        Returns:
            Dict с ключами:
            - 'success': bool - успешность парсинга
            - 'portfolio_data': List[Tuple[str, float]] - список (символ, доля)
            - 'message': str - сообщение пользователю
            - 'suggestions': List[str] - предложения по исправлению
        """
        if not text or not text.strip():
            return {
                'success': False,
                'portfolio_data': [],
                'message': "❌ Пустой ввод. Укажите активы для портфеля.",
                'suggestions': ["Пример: SBER.MOEX:0.3, GAZP.MOEX:0.7"]
            }
        
        # Очищаем и нормализуем ввод
        cleaned_text = text.strip()
        
        # Сначала заменяем запятые в числах на точки, чтобы они не мешали разбору
        # Ищем паттерн "число,число" и заменяем на "число.число"
        import re
        cleaned_text = re.sub(r'(\d+),(\d+)', r'\1.\2', cleaned_text)
        
        # Удаляем лишние пробелы вокруг запятых и двоеточий
        cleaned_text = re.sub(r'\s*,\s*', ', ', cleaned_text)
        cleaned_text = re.sub(r'\s*:\s*', ':', cleaned_text)
        
        # Определяем разделитель: если есть запятые, используем их, иначе пробелы
        if ',' in cleaned_text:
            # Разбиваем по запятым
            parts = [part.strip() for part in cleaned_text.split(',') if part.strip()]
        else:
            # Разбиваем по пробелам, но только между парами символ:вес
            # Используем регулярное выражение для поиска паттернов "символ:вес"
            import re
            # Ищем все паттерны "символ:число" и разделяем по ним
            pattern = r'([A-Za-z0-9._]+:\d+(?:\.\d+)?)'
            matches = re.findall(pattern, cleaned_text)
            
            if matches:
                # Если найдены паттерны с весами, используем их
                parts = []
                remaining_text = cleaned_text
                
                for match in matches:
                    # Находим позицию паттерна в тексте
                    pos = remaining_text.find(match)
                    if pos > 0:
                        # Добавляем текст до паттерна как отдельные символы
                        before = remaining_text[:pos].strip()
                        if before:
                            # Разделяем по пробелам и добавляем как отдельные символы
                            for symbol in before.split():
                                if symbol.strip():
                                    parts.append(symbol.strip())
                    
                    # Добавляем найденный паттерн
                    parts.append(match)
                    
                    # Обновляем оставшийся текст
                    remaining_text = remaining_text[pos + len(match):].strip()
                
                # Добавляем оставшиеся символы
                if remaining_text:
                    for symbol in remaining_text.split():
                        if symbol.strip():
                            parts.append(symbol.strip())
            else:
                # Если паттерны не найдены, просто разделяем по пробелам
                parts = [part.strip() for part in cleaned_text.split() if part.strip()]
        
        if not parts:
            return {
                'success': False,
                'portfolio_data': [],
                'message': "❌ Не найдены активы в вводе.",
                'suggestions': ["Пример: SBER.MOEX:0.3, GAZP.MOEX:0.7"]
            }
        
        portfolio_data = []
        symbols_without_weights = []
        total_explicit_weight = 0.0
        suggestions = []
        
        for part in parts:
            if ':' in part:
                # Формат "символ:доля"
                try:
                    symbol_part, weight_part = part.split(':', 1)
                    # Очищаем символ без удаления двоеточия
                    symbol = self._clean_symbol_for_parsing(symbol_part.strip()).upper()
                    
                    # Проверяем, что символ не пустой
                    if not symbol:
                        suggestions.append(f"Пустой символ в части: '{part}'")
                        continue
                    
                    # Парсим долю (запятые уже заменены на точки выше)
                    weight_str = weight_part.strip()
                    
                    try:
                        weight = float(weight_str)
                    except ValueError:
                        suggestions.append(f"Некорректная доля '{weight_part.strip()}' для {symbol}. Используйте числа от 0 до 1.")
                        continue
                    
                    if weight <= 0 or weight > 1:
                        suggestions.append(f"Доля для {symbol} ({weight}) должна быть от 0 до 1.")
                        continue
                    
                    portfolio_data.append((symbol, weight))
                    total_explicit_weight += weight
                    
                except Exception as e:
                    suggestions.append(f"Ошибка парсинга '{part}': {str(e)}")
                    continue
                    
            else:
                # Только символ без доли
                symbol = self.clean_symbol(part.strip()).upper()
                
                if not symbol:
                    suggestions.append(f"Пустой символ в части: '{part}'")
                    continue
                
                symbols_without_weights.append(symbol)
        
        # Если есть символы без весов, распределяем оставшийся вес поровну
        if symbols_without_weights:
            remaining_weight = 1.0 - total_explicit_weight
            
            if remaining_weight <= 0:
                suggestions.append("Сумма явно указанных долей уже равна или превышает 1.0")
                remaining_weight = 0.1  # Минимальный вес для символов без весов
            
            weight_per_symbol = remaining_weight / len(symbols_without_weights)
            
            for symbol in symbols_without_weights:
                portfolio_data.append((symbol, weight_per_symbol))
        
        # Проверяем результат
        if not portfolio_data:
            return {
                'success': False,
                'portfolio_data': [],
                'message': "❌ Не удалось распознать ни одного актива.",
                'suggestions': suggestions or ["Пример: SBER.MOEX:0.3, GAZP.MOEX:0.7"]
            }
        
        # Проверяем сумму долей
        total_weight = sum(weight for _, weight in portfolio_data)
        
        if abs(total_weight - 1.0) > 0.01:
            if abs(total_weight - 1.0) <= 0.11:  # Увеличиваем порог для учета погрешности вычислений
                # Нормализуем веса
                normalized_data = []
                for symbol, weight in portfolio_data:
                    normalized_weight = weight / total_weight
                    normalized_data.append((symbol, normalized_weight))
                
                portfolio_data = normalized_data
                suggestions.append(f"Веса нормализованы (сумма была {total_weight:.3f})")
            else:
                suggestions.append(f"Сумма долей ({total_weight:.3f}) должна быть близка к 1.0")
        
        # Формируем сообщение
        if suggestions:
            message = f"✅ Портфель создан с автоматическими исправлениями:\n"
            for symbol, weight in portfolio_data:
                message += f"• {symbol}: {weight:.3f}\n"
            message += f"\nИсправления:\n" + "\n".join(f"• {s}" for s in suggestions)
        else:
            message = f"✅ Портфель успешно создан:\n"
            for symbol, weight in portfolio_data:
                message += f"• {symbol}: {weight:.3f}\n"
        
        return {
            'success': True,
            'portfolio_data': portfolio_data,
            'message': message,
            'suggestions': suggestions
        }


    # --- Asset Service Methods ---
    
    def search_assets_with_selection(self, identifier: str) -> Dict[str, Union[str, Any]]:
        """
        Search for assets with possibility to select from multiple results.
        
        Returns:
        - Single result: {'symbol': str, 'type': str, 'source': str, 'name': str}
        - Multiple results: {'results': list, 'type': str, 'query': str}
        - Error: {'error': str}
        """
        try:
            raw = (identifier or '').strip()
            if not raw:
                return {'error': 'Пустой идентификатор актива'}

            upper = raw.upper()

            # If already okama-style ticker like XXX.SUFFIX
            if '.' in upper and len(upper.split('.')) == 2 and all(part for part in upper.split('.')):
                # Check if it's a Chinese exchange symbol
                if self.tushare_service and self.tushare_service.is_tushare_symbol(upper):
                    return {'symbol': upper, 'type': 'ticker', 'source': 'tushare'}
                else:
                    return {'symbol': upper, 'type': 'ticker', 'source': 'input'}

            # Search in okama database
            okama_results = []
            try:
                import okama as ok
                search_result = ok.search(raw)
                if len(search_result) > 0:
                    # Convert to list of results
                    for _, row in search_result.iterrows():
                        okama_results.append({
                            'symbol': row['symbol'],
                            'name': row.get('name', ''),
                            'source': 'okama'
                        })
            except Exception as e:
                self.logger.warning(f"Okama search failed for '{raw}': {e}")

            # Search in tushare database
            tushare_results = []
            if self.tushare_service:
                try:
                    tushare_search = self.tushare_service.search_symbols(raw)
                    if tushare_search:
                        for result in tushare_search:
                            tushare_results.append({
                                'symbol': result['symbol'],
                                'name': result['name'],
                                'source': 'tushare'
                            })
                except Exception as e:
                    self.logger.warning(f"Tushare search failed for '{raw}': {e}")

            # Combine and deduplicate results
            all_results = okama_results + tushare_results
            unique_results = []
            seen_symbols = set()
            
            for result in all_results:
                if result['symbol'] not in seen_symbols:
                    unique_results.append(result)
                    seen_symbols.add(result['symbol'])

            # If no results found
            if not unique_results:
                if self._looks_like_ticker(raw):
                    return {'symbol': upper, 'type': 'ticker', 'source': 'plain'}
                else:
                    return {'error': f'"{raw}" не найден в базе данных okama и tushare'}

            # If only one result, return it directly
            if len(unique_results) == 1:
                result = unique_results[0]
                return {
                    'symbol': result['symbol'],
                    'type': 'ticker' if self._looks_like_ticker(raw) else 'company_name',
                    'source': result['source'],
                    'name': result['name']
                }

            # Multiple results - return for selection
            return {
                'results': unique_results[:20],  # Limit to 20 results
                'type': 'ticker' if self._looks_like_ticker(raw) else 'company_name',
                'query': raw
            }

        except Exception as e:
            return {'error': f"Ошибка при поиске активов: {str(e)}"}

    def resolve_symbol_or_isin(self, identifier: str) -> Dict[str, Union[str, Any]]:
        """
        Resolve user-provided identifier to an okama-compatible ticker.

        Supports:
        - Ticker in okama format (e.g., 'AAPL.US', 'SBER.MOEX')
        - Plain ticker (e.g., 'AAPL') – automatically adds appropriate namespace
        - ISIN (e.g., 'US0378331005') – tries to resolve via okama search
        - Company names (e.g., 'Apple', 'Tesla') – searches via okama search

        Returns dict: { 'symbol': str, 'type': 'ticker'|'isin'|'company_name', 'source': str }
        or { 'error': str } on failure.
        """
        try:
            raw = (identifier or '').strip()
            if not raw:
                return {'error': 'Пустой идентификатор актива'}

            upper = raw.upper()

            # If already okama-style ticker like XXX.SUFFIX
            if '.' in upper and len(upper.split('.')) == 2 and all(part for part in upper.split('.')):
                # Check if it's a Chinese exchange symbol
                if self.tushare_service and self.tushare_service.is_tushare_symbol(upper):
                    return {'symbol': upper, 'type': 'ticker', 'source': 'tushare'}
                else:
                    return {'symbol': upper, 'type': 'ticker', 'source': 'input'}

            if self._looks_like_isin(upper):
                # For ISIN, search for the corresponding symbol
                try:
                    import okama as ok
                    search_result = ok.search(upper)
                    if len(search_result) > 0:
                        # Found the asset, use the first result
                        symbol = search_result.iloc[0]['symbol']
                        return {'symbol': symbol, 'type': 'isin', 'source': 'okama_search'}
                    else:
                        # ISIN not found, return error
                        return {'error': f'ISIN {upper} не найден в базе данных okama'}
                except Exception as e:
                    # Search failed, return error
                    return {'error': f'Ошибка поиска ISIN {upper}: {str(e)}'}

            # Try to search for company name or plain ticker
            try:
                import okama as ok
                search_result = ok.search(raw)  # Use original case for better search
                if len(search_result) > 0:
                    # Found the asset, select the best result
                    symbol = self._select_best_search_result(search_result, raw)
                    name = search_result.iloc[0].get('name', '')
                    
                    # Determine type based on input
                    if self._looks_like_ticker(raw):
                        input_type = 'ticker'
                    else:
                        input_type = 'company_name'
                    
                    return {
                        'symbol': symbol, 
                        'type': input_type, 
                        'source': 'okama_search',
                        'name': name
                    }
                else:
                    # Not found in Okama, try Tushare if available
                    if self.tushare_service:
                        tushare_results = self.tushare_service.search_symbols(raw)
                        if tushare_results:
                            # Return the first result
                            result = tushare_results[0]
                            return {
                                'symbol': result['symbol'],
                                'type': 'ticker',
                                'source': 'tushare_search',
                                'name': result['name']
                            }
                    
                    # Not found, try as plain ticker
                    if self._looks_like_ticker(raw):
                        return {'symbol': upper, 'type': 'ticker', 'source': 'plain'}
                    else:
                        return {'error': f'"{raw}" не найден в базе данных okama и tushare'}
            except Exception as e:
                # Search failed, try as plain ticker
                if self._looks_like_ticker(raw):
                    return {'symbol': upper, 'type': 'ticker', 'source': 'plain'}
                else:
                    return {'error': f'Ошибка поиска "{raw}": {str(e)}'}

        except Exception as e:
            return {'error': f"Ошибка при разборе идентификатора: {str(e)}"}

    def determine_data_source(self, symbol: str) -> str:
        """
        Determine which data source to use (okama or tushare) based on symbol
        
        Args:
            symbol: Symbol in format like 'AAPL.US' or '600000.SH'
            
        Returns:
            'tushare' for Chinese exchanges, 'okama' for others
        """
        if not self.tushare_service:
            return 'okama'
        
        return 'tushare' if self.tushare_service.is_tushare_symbol(symbol) else 'okama'

    def _create_asset_selection_keyboard(self, results: List[Dict], query: str) -> InlineKeyboardMarkup:
        """Создать клавиатуру для выбора актива из множественных результатов"""
        keyboard = []
        
        # Показываем все найденные результаты
        for i in range(len(results)):
            result = results[i]
            symbol = result['symbol']
            name = result['name']
            source = result['source']
            
            # Создаем иконку для источника
            source_icon = "🌍" if source == "okama" else "🇨🇳"
            
            # Ограничиваем длину названия для читаемости
            display_name = name[:30] + "..." if len(name) > 30 else name
            
            button_text = f"{source_icon} {symbol} - {display_name}"
            
            # Создаем callback_data с символом и запросом
            callback_data = f"select_asset_{symbol}_{query}"
            
            keyboard.append([InlineKeyboardButton(button_text, callback_data=callback_data)])
        
        # Добавить кнопку отмены
        keyboard.append([InlineKeyboardButton("❌ Отмена", callback_data=f"cancel_selection_{query}")])
        
        return InlineKeyboardMarkup(keyboard)

    def _looks_like_isin(self, val: str) -> bool:
        """
        Check if string looks like an ISIN code
        
        Args:
            val: String to check
            
        Returns:
            True if string matches ISIN format
        """
        if len(val) != 12:
            return False
        if not (val[0:2].isalpha() and val[0:2].isupper()):
            return False
        if not val[-1].isdigit():
            return False
        mid = val[2:11]
        return mid.isalnum()

    def _looks_like_ticker(self, val: str) -> bool:
        """
        Check if string looks like a ticker symbol
        
        Args:
            val: String to check
            
        Returns:
            True if string looks like a ticker
        """
        if not val or len(val) < 1:
            return False
        
        # Ticker should be mostly alphanumeric, 1-5 characters
        if len(val) > 5:
            return False
        
        # Should be mostly uppercase letters and numbers
        if not val.isalnum():
            return False
        
        # Should start with a letter
        if not val[0].isalpha():
            return False
        
        return True

    def _select_best_search_result(self, search_result, original_input: str) -> str:
        """
        Select the best symbol from search results based on priority and relevance
        
        Args:
            search_result: DataFrame with search results
            original_input: Original user input
            
        Returns:
            Best symbol string
        """
        # Allowed exchanges from okama configuration (all official namespaces)
        allowed_exchanges = ['US', 'LSE', 'XETR', 'XFRA', 'XSTU', 'XAMS', 'MOEX', 'XTAE', 'PIF', 'FX', 'CC', 'INDX', 'COMM', 'RE', 'CBR', 'PF', 'INFL', 'RATE', 'RATIO', 'SSE', 'SZSE', 'HKEX', 'BSE']
        
        # Priority order for exchanges (subset of allowed exchanges)
        priority_exchanges = ['US', 'MOEX', 'LSE', 'XETR', 'XFRA', 'XAMS', 'SSE', 'SZSE', 'HKEX']
        
        # First, try to find exact match with priority exchanges
        for _, row in search_result.iterrows():
            symbol = row['symbol']
            name = row.get('name', '').lower()
            
            # Check if symbol matches original input exactly (case-insensitive)
            if symbol.split('.')[0].upper() == original_input.upper():
                if '.' in symbol and symbol.split('.')[-1] in priority_exchanges:
                    return symbol
        
        # Second, try to find name match with priority exchanges
        original_lower = original_input.lower()
        for _, row in search_result.iterrows():
            symbol = row['symbol']
            name = row.get('name', '').lower()
            
            # Check if name contains original input or vice versa
            if (original_lower in name or name in original_lower) and len(name) > 3:
                if '.' in symbol and symbol.split('.')[-1] in priority_exchanges:
                    return symbol
        
        # Third, try to find any result with priority exchanges
        for exchange in priority_exchanges:
            for _, row in search_result.iterrows():
                symbol = row['symbol']
                if '.' in symbol and symbol.split('.')[-1] == exchange:
                    return symbol
        
        # Fourth, try to find any result with allowed exchanges
        for exchange in allowed_exchanges:
            for _, row in search_result.iterrows():
                symbol = row['symbol']
                if '.' in symbol and symbol.split('.')[-1] == exchange:
                    # Verify the symbol is actually supported by okama
                    try:
                        ok.Asset(symbol)
                        return symbol
                    except Exception:
                        continue  # Skip this symbol if it's not supported
        
        # If no allowed exchange found, try the first result
        first_symbol = search_result.iloc[0]['symbol']
        try:
            ok.Asset(first_symbol)
            return first_symbol
        except Exception:
            # If even the first result fails, return it anyway (will be handled by caller)
            return first_symbol



    def get_random_examples(self, count: int = 3) -> list:
        """Get random examples from known assets, including Chinese and Hong Kong assets"""
        import random
        all_assets = []
        # Include all assets including Chinese and Hong Kong assets
        for category, assets in self.known_assets.items():
            all_assets.extend(assets)
        # Get random sample and format with backticks
        selected_assets = random.sample(all_assets, min(count, len(all_assets)))
        return [f"`{asset}`" for asset in selected_assets]




    async def _handle_error(self, update: Update, error: Exception, context: str = "Unknown operation") -> None:
        """Общая функция для обработки ошибок"""
        error_msg = f"❌ Ошибка в {context}: {str(error)}"
        self.logger.error(f"{error_msg} - {traceback.format_exc()}")
        
        try:
            await self._send_message_safe(update, error_msg)
        except Exception as send_error:
            self.logger.error(f"Failed to send error message: {send_error}")
            # Try to send a simple error message
            try:
                if hasattr(update, 'message') and update.message is not None:
                    await update.message.reply_text("Произошла ошибка при обработке запроса")
            except Exception as final_error:
                self.logger.error(f"Final error message sending failed: {final_error}")
        
    def _find_portfolio_by_symbol(self, portfolio_symbol: str, saved_portfolios: Dict, user_id: int = None) -> Optional[str]:
        """
        Найти портфель по символу с использованием различных стратегий поиска.
        
        Args:
            portfolio_symbol: Символ портфеля для поиска
            saved_portfolios: Словарь сохраненных портфелей
            user_id: ID пользователя для логирования (опционально)
            
        Returns:
            Найденный ключ портфеля или None, если не найден
        """
        log_prefix = f"User {user_id}: " if user_id else ""
        
        # 1. Точное совпадение
        if portfolio_symbol in saved_portfolios:
            if user_id:
                self.logger.info(f"{log_prefix}Found exact match for portfolio: '{portfolio_symbol}'")
            return portfolio_symbol
        
        # 2. Поиск по активам (в случае разного порядка или формата символа)
        try:
            portfolio_assets = set(portfolio_symbol.split(','))
            for key, portfolio_info in saved_portfolios.items():
                saved_assets = set(portfolio_info.get('symbols', []))
                if portfolio_assets == saved_assets:
                    if user_id:
                        self.logger.info(f"{log_prefix}Found portfolio by assets match: '{key}' for requested '{portfolio_symbol}'")
                    return key
        except Exception as e:
            if user_id:
                self.logger.warning(f"{log_prefix}Error during assets matching: {e}")
        
        # 3. Поиск без учета регистра
        for key in saved_portfolios.keys():
            if key.lower() == portfolio_symbol.lower():
                if user_id:
                    self.logger.info(f"{log_prefix}Found case-insensitive match: '{key}' for requested '{portfolio_symbol}'")
                return key
        
        # 4. Поиск без пробелов
        portfolio_symbol_no_spaces = portfolio_symbol.replace(' ', '')
        for key in saved_portfolios.keys():
            if key.replace(' ', '') == portfolio_symbol_no_spaces:
                if user_id:
                    self.logger.info(f"{log_prefix}Found no-spaces match: '{key}' for requested '{portfolio_symbol}'")
                return key
        
        # Не найдено
        if user_id:
            self.logger.error(f"{log_prefix}Portfolio not found: '{portfolio_symbol}'")
            self.logger.error(f"{log_prefix}Available portfolios: {list(saved_portfolios.keys())}")
        
        return None

    def _check_existing_portfolio(self, symbols: List[str], weights: List[float], saved_portfolios: Dict) -> Optional[str]:
        """
        Проверяет, существует ли портфель с такими же активами и пропорциями.
        
        Args:
            symbols: Список символов активов
            weights: Список весов активов
            saved_portfolios: Словарь сохраненных портфелей
            
        Returns:
            Символ существующего портфеля или None, если не найден
        """
        # Нормализуем веса для сравнения (сумма = 1.0)
        total_weight = sum(weights)
        normalized_weights = [w / total_weight for w in weights]
        
        for portfolio_symbol, portfolio_info in saved_portfolios.items():
            existing_symbols = portfolio_info.get('symbols', [])
            existing_weights = portfolio_info.get('weights', [])
            
            # Проверяем количество активов
            if len(symbols) != len(existing_symbols):
                continue
                
            # Проверяем, что все символы совпадают (с учетом регистра)
            if set(symbol.upper() for symbol in symbols) != set(symbol.upper() for symbol in existing_symbols):
                continue
                
            # Нормализуем существующие веса
            existing_total = sum(existing_weights)
            if existing_total == 0:
                continue
            normalized_existing_weights = [w / existing_total for w in existing_weights]
            
            # Проверяем, что веса совпадают с точностью до 0.001
            weight_matches = True
            for i, (new_weight, existing_weight) in enumerate(zip(normalized_weights, normalized_existing_weights)):
                if abs(new_weight - existing_weight) > 0.001:
                    weight_matches = False
                    break
                    
            if weight_matches:
                return portfolio_symbol
                
        return None

    def _parse_portfolio_data(self, portfolio_data_str: str) -> tuple[list, list]:
        """Parse portfolio data string with weights (symbol:weight,symbol:weight)"""
        try:
            if not portfolio_data_str or not portfolio_data_str.strip():
                return [], []
                
            symbols = []
            weights = []
            
            for item in portfolio_data_str.split(','):
                item = item.strip()
                if not item:  # Skip empty items
                    continue
                    
                if ':' in item:
                    symbol, weight_str = item.split(':', 1)
                    symbol = symbol.strip()
                    if symbol:  # Only add non-empty symbols
                        symbols.append(symbol)
                        weights.append(float(weight_str.strip()))
                else:
                    # Fallback: treat as symbol without weight
                    symbols.append(item)
                    weights.append(1.0 / len([x for x in portfolio_data_str.split(',') if x.strip()]))
            
            return symbols, weights
        except Exception as e:
            self.logger.error(f"Error parsing portfolio data: {e}")
            return [], []

    def _normalize_or_equalize_weights(self, symbols: list, weights: list) -> list:
        """Return valid weights for okama: normalize if close to 1, else equal weights.

        - If weights list is empty or length mismatch with symbols, fallback to equal weights.
        - If any weight is non-positive, fallback to equal weights.
        - If sum differs from 1.0 beyond small epsilon, normalize to sum=1.0.
        """
        try:
            if not symbols:
                return []

            num_assets = len(symbols)

            # Fallback to equal weights when missing or invalid length
            if not weights or len(weights) != num_assets:
                return [1.0 / num_assets] * num_assets

            # Validate positivity
            if any((w is None) or (w <= 0) for w in weights):
                return [1.0 / num_assets] * num_assets

            total = sum(weights)
            if total <= 0:
                return [1.0 / num_assets] * num_assets

            # If already ~1 within tolerance, keep; else normalize precisely
            if abs(total - 1.0) <= 1e-6:
                return [float(w) for w in weights]

            normalized = [float(w) / total for w in weights]

            # Guard against numerical drift
            norm_total = sum(normalized)
            if abs(norm_total - 1.0) > 1e-9:
                # Final correction by scaling the first element
                diff = 1.0 - norm_total
                normalized[0] += diff
            return normalized
        except Exception:
            # Safe fallback
            return [1.0 / len(symbols)] * len(symbols) if symbols else []

    def _create_portfolio_with_period(self, symbols: list, weights: list, currency: str, user_context: dict) -> object:
        """Create portfolio with period from user context"""
        import okama as ok
        from datetime import datetime, timedelta
        
        current_period = user_context.get('current_period')
        if current_period:
            years = int(current_period[:-1])
            end_date = datetime.now()
            start_date = end_date - timedelta(days=years * 365)
            portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency,
                                   first_date=start_date.strftime('%Y-%m-%d'), 
                                   last_date=end_date.strftime('%Y-%m-%d'))
            self.logger.info(f"Created portfolio with period {current_period}")
        else:
            portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency)
            self.logger.info("Created portfolio without period (MAX)")
        
        return portfolio

    def _get_user_context(self, user_id: int) -> Dict[str, Any]:
        """Получить контекст пользователя (с поддержкой персистентности)."""
        # Load from persistent store, and mirror into memory for hot path
        ctx = self.context_store.get_user_context(user_id)
        self.user_sessions[user_id] = ctx
        return ctx
    
    def _update_user_context(self, user_id: int, **kwargs):
        """Обновить контекст пользователя (и сохранить)."""
        # Update persistent store; keep in-memory mirror in sync
        updated = self.context_store.update_user_context(user_id, **kwargs)
        self.user_sessions[user_id] = updated

    def _add_to_analyzed_tickers(self, user_id: int, symbol: str):
        """Добавить тикер в историю анализируемых активов пользователя"""
        user_context = self._get_user_context(user_id)
        analyzed_tickers = user_context.get('analyzed_tickers', [])
        
        # Удаляем тикер если он уже есть (чтобы переместить в начало)
        if symbol in analyzed_tickers:
            analyzed_tickers.remove(symbol)
        
        # Добавляем в начало списка
        analyzed_tickers.insert(0, symbol)
        
        # Ограничиваем историю до 20 тикеров
        analyzed_tickers = analyzed_tickers[:20]
        
        # Обновляем контекст
        self._update_user_context(user_id, analyzed_tickers=analyzed_tickers)
    
    def _add_to_conversation_history(self, user_id: int, message: str, response: str):
        """Добавить сообщение в историю разговора (с сохранением)."""
        self.context_store.add_conversation_entry(user_id, message, response)
        # Refresh in-memory cache
        self.user_sessions[user_id] = self.context_store.get_user_context(user_id)
    
    def _get_context_summary(self, user_id: int) -> str:
        """Получить краткое резюме контекста пользователя"""
        context = self._get_user_context(user_id)
        summary = []
        
        if context['last_assets']:
            summary.append(f"Последние активы: {', '.join(context['last_assets'][-3:])}")
        
        if context['last_period']:
            summary.append(f"Период: {context['last_period']}")
        
        return "; ".join(summary) if summary else "Новый пользователь"
    
    def _get_currency_by_symbol(self, symbol: str) -> tuple[str, str]:
        """
        Определить валюту по символу с учетом китайских бирж
        
        Returns:
            tuple: (currency, currency_info)
        """
        try:
            if '.' in symbol:
                namespace = symbol.split('.')[1]
                
                # Китайские биржи
                if namespace == 'HK':
                    return "HKD", f"автоматически определена по бирже HKEX ({symbol})"
                elif namespace == 'SH':
                    return "CNY", f"автоматически определена по бирже SSE ({symbol})"
                elif namespace == 'SZ':
                    return "CNY", f"автоматически определена по бирже SZSE ({symbol})"
                elif namespace == 'BJ':
                    return "CNY", f"автоматически определена по бирже BSE ({symbol})"
                
                # Другие биржи
                elif namespace == 'MOEX':
                    return "RUB", f"автоматически определена по бирже MOEX ({symbol})"
                elif namespace == 'US':
                    return "USD", f"автоматически определена по бирже US ({symbol})"
                elif namespace == 'LSE':
                    return "GBP", f"автоматически определена по бирже LSE ({symbol})"
                elif namespace == 'FX':
                    return "USD", f"автоматически определена по бирже FX ({symbol})"
                elif namespace == 'COMM':
                    return "USD", f"автоматически определена по бирже COMM ({symbol})"
                elif namespace == 'INDX':
                    return "USD", f"автоматически определена по бирже INDX ({symbol})"
                else:
                    return "USD", f"автоматически определена по умолчанию ({symbol})"
            else:
                return "USD", f"автоматически определена по умолчанию ({symbol})"
        except Exception as e:
            self.logger.warning(f"Could not determine currency for {symbol}: {e}")
            return "USD", f"автоматически определена по умолчанию ({symbol})"
    
    def _get_inflation_ticker_by_currency(self, currency: str) -> str:
        """
        Получить тикер инфляции по валюте
        
        Returns:
            str: тикер инфляции (например, 'CNY.INFL' для CNY)
        """
        inflation_mapping = {
            'USD': 'US.INFL',
            'RUB': 'RUS.INFL', 
            'EUR': 'EU.INFL',
            'GBP': 'GB.INFL',
            'CNY': 'CNY.INFL',  # Китайская инфляция
            'HKD': 'HK.INFL'    # Гонконгская инфляция
        }
        return inflation_mapping.get(currency, 'US.INFL')
    
    def _parse_currency_and_period(self, args: List[str], preserve_weights: bool = False) -> tuple[List[str], Optional[str], Optional[str]]:
        """
        Parse currency and period parameters from command arguments.
        
        Args:
            args: List of command arguments
            preserve_weights: If True, preserve symbol:weight format; if False, strip weights (for compare command)
            
        Returns:
            Tuple of (symbols, currency, period) where:
            - symbols: List of symbols (with or without weights depending on preserve_weights)
            - currency: Currency code (e.g., 'USD', 'RUB') or None
            - period: Period string (e.g., '5Y', '10Y') or None
        """
        if not args:
            return [], None, None
            
        # Valid currency codes
        valid_currencies = {'USD', 'RUB', 'EUR', 'GBP', 'CNY', 'HKD', 'JPY'}
        
        # Valid period patterns (e.g., '5Y', '10Y', '1Y', '2Y', etc.)
        import re
        period_pattern = re.compile(r'^(\d+)Y$', re.IGNORECASE)
        
        symbols = []
        currency = None
        period = None
        
        for arg in args:
            arg_upper = arg.upper()
            
            # Check if it's a currency code
            if arg_upper in valid_currencies:
                if currency is None:
                    currency = arg_upper
                else:
                    self.logger.warning(f"Multiple currencies specified, using first: {currency}")
                continue
            
            # Check if it's a period (e.g., '5Y', '10Y')
            period_match = period_pattern.match(arg)
            if period_match:
                if period is None:
                    period = arg_upper
                else:
                    self.logger.warning(f"Multiple periods specified, using first: {period}")
                continue
            
            # Check if this is a symbol:weight format
            if ':' in arg:
                if preserve_weights:
                    # For portfolio command, preserve the full symbol:weight format
                    symbols.append(arg)
                else:
                    # For compare command, ignore weights and take only the symbol part
                    symbol_part = arg.split(':', 1)[0].strip().rstrip(',')
                    if symbol_part:  # Only add non-empty symbols
                        symbols.append(symbol_part)
                continue
            
            # If it's neither currency nor period, it's a symbol
            # Strip trailing commas that might be left from command parsing
            symbol = arg.rstrip(',')
            if symbol:  # Only add non-empty symbols
                symbols.append(symbol)
        
        return symbols, currency, period
    
    def _is_chinese_symbol(self, symbol: str) -> bool:
        """
        Проверить, является ли символ китайским
        
        Returns:
            bool: True если символ китайский
        """
        if not self.tushare_service:
            return False
        return self.tushare_service.is_tushare_symbol(symbol)
    
    def _normalize_symbol_namespace(self, symbol: str) -> str:
        """
        Нормализовать регистр namespace в символе
        
        Args:
            symbol: Символ в формате TICKER.NAMESPACE
            
        Returns:
            str: Символ с нормализованным namespace (uppercase)
        """
        if '.' not in symbol:
            return symbol
        
        ticker, namespace = symbol.split('.', 1)
        
        # Known namespace mappings (lowercase -> uppercase)
        namespace_mappings = {
            'moex': 'MOEX',
            'us': 'US',
            'lse': 'LSE',
            'xetr': 'XETR',
            'xfra': 'XFRA',
            'xstu': 'XSTU',
            'xams': 'XAMS',
            'xtae': 'XTAE',
            'pif': 'PIF',
            'fx': 'FX',
            'cc': 'CC',
            'indx': 'INDX',
            'comm': 'COMM',
            're': 'RE',
            'cbr': 'CBR',
            'pf': 'PF'
        }
        
        # Convert namespace to uppercase if it's in our mappings
        normalized_namespace = namespace_mappings.get(namespace.lower(), namespace.upper())
        
        return f"{ticker}.{normalized_namespace}"
    
    def _get_chinese_symbol_data(self, symbol: str) -> Optional[Dict[str, Any]]:
        """
        Получить данные китайского символа через Tushare
        
        Returns:
            Dict с данными символа или None
        """
        if not self.tushare_service or not self._is_chinese_symbol(symbol):
            return None
        
        try:
            # Получаем базовую информацию о символе
            symbol_info = self.tushare_service.get_symbol_info(symbol)
            if symbol_info:
                return symbol_info
        except Exception as e:
            self.logger.warning(f"Could not get Chinese symbol data for {symbol}: {e}")
        
        return None
    
    async def _create_hybrid_chinese_comparison(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """
        Создать гибридное сравнение китайских символов
        - Максимальный период по датам
        - Данные по CNY.INFL из okama
        - Скрытые xlabel и ylabel
        - Заголовок: Сравнение название тикеров, биржа, валюта
        """
        try:
            self.logger.info(f"Creating hybrid comparison for Chinese symbols: {symbols}")
            
            # Определяем валюту по первому символу
            currency, currency_info = self._get_currency_by_symbol(symbols[0])
            inflation_ticker = self._get_inflation_ticker_by_currency(currency)
            
            # Получаем данные китайских символов через Tushare (максимальный период)
            chinese_data = {}
            all_dates = set()
            
            for symbol in symbols:
                if self._is_chinese_symbol(symbol):
                    try:
                        symbol_info = self.tushare_service.get_symbol_info(symbol)
                        # Получаем месячные данные для лучшего отображения
                        historical_data = self.tushare_service.get_monthly_data(symbol, start_date='19900101')
                        
                        if not historical_data.empty:
                            # Устанавливаем дату как индекс
                            historical_data = historical_data.set_index('trade_date')
                            chinese_data[symbol] = {
                                'info': symbol_info,
                                'data': historical_data
                            }
                            all_dates.update(historical_data.index)
                            self.logger.info(f"Got monthly data for Chinese symbol {symbol}: {len(historical_data)} records")
                    except Exception as e:
                        self.logger.warning(f"Could not get data for Chinese symbol {symbol}: {e}")
            
            if not chinese_data:
                await self._send_message_safe(update, 
                    f"❌ Не удалось получить данные для китайских символов: {', '.join(symbols)}")
                return
            
            # Получаем данные по инфляции из okama для CNY и HKD активов
            inflation_data = None
            self.logger.info(f"Currency: {currency}, inflation_ticker: {inflation_ticker}")
            self.logger.info(f"Condition check: currency in ['CNY', 'HKD'] = {currency in ['CNY', 'HKD']}, inflation_ticker in ['CNY.INFL', 'HK.INFL'] = {inflation_ticker in ['CNY.INFL', 'HK.INFL']}")
            
            if currency in ['CNY', 'HKD'] and inflation_ticker in ['CNY.INFL', 'HK.INFL']:
                try:
                    import okama as ok
                    self.logger.info(f"Creating okama Asset for {inflation_ticker}")
                    inflation_asset = ok.Asset(inflation_ticker)
                    self.logger.info(f"Got inflation asset, wealth_index type: {type(inflation_asset.wealth_index)}")
                    # Получаем месячные данные инфляции для соответствия основным данным
                    inflation_data = inflation_asset.wealth_index.resample('M').last()
                    self.logger.info(f"Got monthly inflation data for {inflation_ticker}: {len(inflation_data)} records")
                    self.logger.info(f"Inflation data sample: {inflation_data.head()}")
                except Exception as e:
                    self.logger.warning(f"Could not get inflation data for {inflation_ticker}: {e}")
                    import traceback
                    self.logger.warning(f"Inflation error traceback: {traceback.format_exc()}")
            else:
                self.logger.info(f"Skipping inflation data: currency={currency}, inflation_ticker={inflation_ticker}")
            
            # Подготавливаем данные для стандартного метода сравнения
            comparison_data = {}
            symbols_list = []
            
            for symbol, data_dict in chinese_data.items():
                historical_data = data_dict['data']
                symbol_info = data_dict['info']
                
                if not historical_data.empty:
                    # Нормализуем данные к базовому значению (1000) как в okama
                    normalized_data = historical_data['close'] / historical_data['close'].iloc[0] * 1000
                    
                    # Получаем английское название символа для легенды
                    symbol_name = symbol_info.get('name', symbol)
                    # Приоритет английскому названию если доступно
                    if 'enname' in symbol_info and symbol_info['enname'] and symbol_info['enname'].strip():
                        symbol_name = symbol_info['enname']
                    
                    self.logger.info(f"Symbol {symbol}: name='{symbol_info.get('name', 'N/A')}', enname='{symbol_info.get('enname', 'N/A')}', final='{symbol_name}'")
                    
                    symbols_list.append(symbol)
                    
                    if len(symbol_name) > 30:
                        symbol_name = symbol_name[:27] + "..."
                    
                    # Добавляем в данные для сравнения
                    comparison_data[f"{symbol} - {symbol_name}"] = normalized_data
            
            # Добавляем данные по инфляции если доступны
            if inflation_data is not None and not inflation_data.empty:
                # Нормализуем инфляцию к базовому значению (1000)
                normalized_inflation = inflation_data / inflation_data.iloc[0] * 1000
                self.logger.info(f"Adding inflation line: {len(normalized_inflation)} points, range: {normalized_inflation.min():.2f} - {normalized_inflation.max():.2f}")
                comparison_data[f"{inflation_ticker} - Inflation"] = normalized_inflation
            else:
                self.logger.warning(f"Inflation data is None or empty: {inflation_data is None}, {inflation_data.empty if inflation_data is not None else 'N/A'}")
            
            # Создаем DataFrame для стандартного метода сравнения
            import pandas as pd
            comparison_df = pd.DataFrame(comparison_data)
            
            # Формируем заголовок
            title_parts = ["Comparison"]
            if symbols_list:
                symbols_str = ", ".join(symbols_list)
                title_parts.append(symbols_str)
            title_parts.append(f"Currency: {currency}")
            title = ", ".join(title_parts)
            
            # Используем стандартный метод создания графика сравнения
            fig, ax = self.chart_styles.create_comparison_chart(
                data=comparison_df,
                symbols=list(comparison_data.keys()),
                currency=currency,
                data_source='tushare',
                title=title,
                xlabel='',  # Скрываем подпись оси X
                ylabel=''   # Скрываем подпись оси Y
            )
            
            # Сохраняем график в bytes
            img_buffer = io.BytesIO()
            fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Очищаем matplotlib
            import matplotlib.pyplot as plt
            plt.close(fig)
            
            # Создаем caption
            caption = f"📈 Сравнение: {', '.join(symbols)}\n\n"
            caption += f"💱 Валюта: {currency} ({currency_info})\n"
            caption += f"📊 Инфляция: {inflation_ticker}\n"

            
            # Отправляем график
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_bytes,
                caption=caption
            )
            
            self.logger.info(f"Successfully created hybrid comparison for {len(symbols)} Chinese symbols")
            
        except Exception as e:
            self.logger.error(f"Error creating hybrid Chinese comparison: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при создании гибридного сравнения: {str(e)}")
    
    # =======================
    # Вспомогательные функции для истории
    # =======================






    def _split_text(self, text: str):
        for i in range(0, len(text), self.MAX_TELEGRAM_CHUNK):
            yield text[i:i + self.MAX_TELEGRAM_CHUNK]

    async def send_long_message(self, update: Update, text: str):
        if not text:
            text = "Пустой ответ."
        if hasattr(update, 'message') and update.message is not None:
            for chunk in self._split_text(text):
                await update.message.reply_text(chunk)
    
    def _escape_markdown(self, text: str) -> str:
        """Escape special Markdown characters"""
        if not text:
            return text
        
        # Escape special Markdown characters
        escape_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
        for char in escape_chars:
            text = text.replace(char, f'\\{char}')
        return text

    def _format_describe_table(self, asset_list) -> str:
        """Format ok.AssetList.describe() data with adaptive formatting for Telegram"""
        try:
            if not TABULATE_AVAILABLE:
                # Fallback to simple text formatting if tabulate is not available
                return self._format_describe_table_simple(asset_list)
            
            # Get describe data
            describe_data = asset_list.describe()
            
            if describe_data is None or describe_data.empty:
                return "⚖️ Данные для сравнения недоступны"
            
            # Count columns (assets) to choose best format
            num_assets = len(describe_data.columns)
            
            if num_assets <= 2:
                # For 1-2 assets, use pipe format (most readable)
                markdown_table = tabulate.tabulate(
                    describe_data, 
                    headers='keys', 
                    tablefmt='pipe',
                    floatfmt='.2f'
                )
                return f"🔍 **Статистика активов:**\n```\n{markdown_table}\n```"
            
            elif num_assets <= 4:
                # For 3-4 assets, use simple format (compact but readable)
                markdown_table = tabulate.tabulate(
                    describe_data, 
                    headers='keys', 
                    tablefmt='simple',
                    floatfmt='.2f'
                )
                return f"🔍 **Статистика активов:**\n```\n{markdown_table}\n```"
            
            else:
                # For 5+ assets, use vertical format (most mobile-friendly)
                return self._format_describe_table_vertical(describe_data)
            
        except Exception as e:
            self.logger.error(f"Error formatting describe table: {e}")
            return "📊 Ошибка при формировании таблицы статистики"
    
    def _format_describe_table_simple(self, asset_list) -> str:
        """Simple text formatting fallback for describe table with adaptive formatting"""
        try:
            describe_data = asset_list.describe()
            
            if describe_data is None or describe_data.empty:
                return "⚖️ Данные для сравнения недоступны"
            
            # Count columns (assets) to choose best format
            num_assets = len(describe_data.columns)
            
            if num_assets <= 2:
                # For 1-2 assets, use simple table format
                return self._format_simple_table(describe_data)
            elif num_assets <= 4:
                # For 3-4 assets, use compact table format
                return self._format_compact_table(describe_data)
            else:
                # For 5+ assets, use vertical format
                return self._format_describe_table_vertical(describe_data)
            
        except Exception as e:
            self.logger.error(f"Error in simple describe table formatting: {e}")
            return "📊 Ошибка при формировании таблицы статистики"
    
    def _format_simple_table(self, describe_data) -> str:
        """Format as simple markdown table"""
        try:
            result = ["🔍 **Статистика активов:**\n"]
            
            # Get column names (asset symbols)
            columns = describe_data.columns.tolist()
            
            # Get row names (metrics)
            rows = describe_data.index.tolist()
            
            # Create simple table
            header = "| Метрика | " + " | ".join([f"`{col}`" for col in columns]) + " |"
            separator = "|" + "|".join([" --- " for _ in range(len(columns) + 1)]) + "|"
            
            result.append(f"```\n{header}\n{separator}")
            
            # Data rows
            for row in rows:
                row_data = [str(row)]
                for col in columns:
                    value = describe_data.loc[row, col]
                    if isinstance(value, (int, float)):
                        if pd.isna(value):
                            row_data.append("N/A")
                        else:
                            row_data.append(f"{value:.2f}")
                    else:
                        row_data.append(str(value))
                
                result.append("| " + " | ".join(row_data) + " |")
            
            result.append("```")
            return "\n".join(result)
            
        except Exception as e:
            self.logger.error(f"Error in simple table formatting: {e}")
            return "📊 Ошибка при формировании таблицы"
    
    def _format_compact_table(self, describe_data) -> str:
        """Format as compact table"""
        try:
            result = ["🔍 **Статистика активов:**\n"]
            
            # Get column names (asset symbols)
            columns = describe_data.columns.tolist()
            
            # Get row names (metrics)
            rows = describe_data.index.tolist()
            
            # Create compact table
            header = "Metric".ljust(20) + " | " + " | ".join([col.ljust(8) for col in columns])
            separator = "-" * len(header)
            
            result.append(f"```\n{header}\n{separator}")
            
            # Data rows
            for row in rows:
                row_str = str(row).ljust(20) + " | "
                values = []
                for col in columns:
                    value = describe_data.loc[row, col]
                    if pd.isna(value):
                        values.append("N/A".ljust(8))
                    elif isinstance(value, (int, float)):
                        values.append(f"{value:.2f}".ljust(8))
                    else:
                        values.append(str(value).ljust(8))
                row_str += " | ".join(values)
                result.append(row_str)
            
            result.append("```")
            return "\n".join(result)
            
        except Exception as e:
            self.logger.error(f"Error in compact table formatting: {e}")
            return "📊 Ошибка при формировании таблицы"
    
    def _format_describe_table_vertical(self, describe_data) -> str:
        """Format describe data in vertical format for mobile-friendly display"""
        try:
            result = ["🔍 **Статистика активов:**\n"]
            
            # Get column names (asset symbols)
            columns = describe_data.columns.tolist()
            
            # Get row names (metrics)
            rows = describe_data.index.tolist()
            
            # Create vertical format - one metric per line
            for row in rows:
                result.append(f"📊 **{row}:**")
                for col in columns:
                    value = describe_data.loc[row, col]
                    if pd.isna(value):
                        result.append(f"  • `{col}`: N/A")
                    elif isinstance(value, (int, float)):
                        result.append(f"  • `{col}`: {value:.2f}")
                    else:
                        result.append(f"  • `{col}`: {value}")
                result.append("")  # Empty line between metrics
            
            return "\n".join(result)
            
        except Exception as e:
            self.logger.error(f"Error in vertical describe table formatting: {e}")
            return "📊 Ошибка при формировании таблицы статистики"

    async def _send_photo_safe(self, update: Update, photo_bytes: bytes, caption: str = None, reply_markup=None, context: ContextTypes.DEFAULT_TYPE = None, parse_mode: str = 'Markdown'):
        """Безопасная отправка фотографии с обработкой ошибок"""
        try:
            import io
            
            # Проверяем, что update не None
            if not update or not update.effective_chat:
                self.logger.error("Update or effective_chat is None in _send_photo_safe")
                return
            
            # Получаем bot из context или из update
            bot = None
            if context and hasattr(context, 'bot'):
                bot = context.bot
            elif hasattr(update, 'bot'):
                bot = update.bot
            else:
                self.logger.error("Cannot find bot instance for sending photo")
                return
            
            # Отправляем фотографию с parse_mode по умолчанию Markdown
            await bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=io.BytesIO(photo_bytes),
                caption=caption,
                parse_mode=parse_mode,
                reply_markup=reply_markup
            )
            
        except Exception as e:
            self.logger.error(f"Error sending photo: {e}")
            # Fallback: отправляем только текст с тем же parse_mode
            if caption:
                await self._send_message_safe(update, caption, reply_markup=reply_markup)

    async def _send_message_safe(self, update: Update, text: str, reply_markup=None, parse_mode='Markdown'):
        """Безопасная отправка сообщения с автоматическим разбиением на части - исправлено для обработки None"""
        try:
            # Проверяем, что update не None
            if update is None:
                self.logger.error("Cannot send message: update is None")
                return
            
            # Проверяем, что text не пустой
            if not text or text.strip() == "":
                self.logger.error("Cannot send message: text is empty")
                return
            
            # Дополнительная проверка для callback query
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                # Проверяем еще раз для callback query
                if not text or text.strip() == "":
                    self.logger.error("Cannot send callback message: text is empty")
                    return
            
            # Если это callback query, используем специальную функцию
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                self.logger.info("_send_message_safe: Redirecting to _send_callback_message for callback query")
                # Для callback query нужно использовать context.bot.send_message напрямую
                try:
                    await update.callback_query.message.reply_text(text, parse_mode=parse_mode, reply_markup=reply_markup)
                except Exception as e:
                    self.logger.error(f"Error sending callback message: {e}")
                    # Fallback: попробуем через context.bot
                    try:
                        await update.callback_query.bot.send_message(
                            chat_id=update.callback_query.message.chat_id,
                            text=text,
                            parse_mode=parse_mode,
                            reply_markup=reply_markup
                        )
                    except Exception as fallback_error:
                        self.logger.error(f"Fallback callback message sending also failed: {fallback_error}")
                return
            
            # Проверяем, что message не None для обычных сообщений
            if not hasattr(update, 'message') or update.message is None:
                self.logger.error("Cannot send message: update.message is None")
                return
            
            # Проверяем, что text действительно является строкой
            if not isinstance(text, str):
                self.logger.warning(f"_send_message_safe received non-string data: {type(text)}")
                text = str(text)
            
            # Проверяем длину строки
            if len(text) <= 4000:
                self.logger.info(f"Sending message with reply_markup: {reply_markup is not None}")
                if reply_markup:
                    self.logger.info(f"Reply markup type: {type(reply_markup)}")
                    self.logger.info(f"Reply markup content: {reply_markup.to_dict() if hasattr(reply_markup, 'to_dict') else 'No to_dict method'}")
                
                # Попробуем отправить с parse_mode, если не получится - без него
                try:
                    await update.message.reply_text(text, parse_mode=parse_mode, reply_markup=reply_markup)
                except Exception as parse_error:
                    self.logger.warning(f"Failed to send with parse_mode '{parse_mode}': {parse_error}")
                    # Попробуем без parse_mode, но с кнопками
                    try:
                        await update.message.reply_text(text, reply_markup=reply_markup)
                    except Exception as no_parse_error:
                        self.logger.warning(f"Failed to send with reply_markup: {no_parse_error}")
                        # Последняя попытка - только текст
                        await update.message.reply_text(text)
            else:
                # Для длинных сообщений с кнопками отправляем первую часть с кнопками
                if reply_markup:
                    first_part = text[:4000]
                    try:
                        await update.message.reply_text(first_part, parse_mode=parse_mode, reply_markup=reply_markup)
                    except Exception as long_parse_error:
                        self.logger.warning(f"Failed to send long message with parse_mode '{parse_mode}': {long_parse_error}")
                        # Попробуем без parse_mode, но с кнопками
                        try:
                            await update.message.reply_text(first_part, reply_markup=reply_markup)
                        except Exception as long_no_parse_error:
                            self.logger.warning(f"Failed to send long message with reply_markup: {long_no_parse_error}")
                            # Последняя попытка - только текст
                            await update.message.reply_text(first_part)
                    
                    # Остальную часть отправляем без кнопок
                    remaining_text = text[4000:]
                    if remaining_text:
                        await self.send_long_message(update, remaining_text)
                else:
                    await self.send_long_message(update, text)
        except Exception as e:
            self.logger.error(f"Error in _send_message_safe: {e}")
            # Fallback: попробуем отправить как обычный текст, но сохраним кнопки
            try:
                if hasattr(update, 'message') and update.message is not None:
                    # Попробуем отправить без parse_mode, но с кнопками
                    try:
                        await update.message.reply_text(f"Ошибка форматирования: {str(text)[:1000]}...", reply_markup=reply_markup)
                    except Exception as markdown_error:
                        self.logger.warning(f"Failed to send with reply_markup, trying without: {markdown_error}")
                        # Если не получилось с кнопками, попробуем без них
                        await update.message.reply_text(f"Ошибка форматирования: {str(text)[:1000]}...")
            except Exception as fallback_error:
                self.logger.error(f"Fallback message sending failed: {fallback_error}")
                try:
                    if hasattr(update, 'message') and update.message is not None:
                        await update.message.reply_text("Произошла ошибка при отправке сообщения")
                except Exception as final_error:
                    self.logger.error(f"Final fallback message sending failed: {final_error}")
    
    def _truncate_caption(self, text: Any) -> str:
        """Обрезать подпись до допустимой длины Telegram.
        Гарантирует, что длина не превышает Config.MAX_CAPTION_LENGTH.
        """
        try:
            if text is None:
                return ""
            if not isinstance(text, str):
                text = str(text)
            max_len = getattr(Config, 'MAX_CAPTION_LENGTH', 1024)
            if len(text) <= max_len:
                return text
            ellipsis = "…"
            if max_len > len(ellipsis):
                return text[: max_len - len(ellipsis)] + ellipsis
            return text[:max_len]
        except Exception as e:
            self.logger.error(f"Error truncating caption: {e}")
            safe_text = "" if text is None else str(text)
            return safe_text[:1024]
    
    def _format_correlation_values(self, correlation_matrix: pd.DataFrame) -> str:
        """Форматировать численные значения корреляции для отображения под матрицей"""
        try:
            if correlation_matrix is None or correlation_matrix.empty:
                return ""
            
            values_text = ""
            
            # Get upper triangle values only (avoid duplicates)
            symbols = correlation_matrix.columns.tolist()
            
            for i, symbol1 in enumerate(symbols):
                for j, symbol2 in enumerate(symbols):
                    if i < j:  # Only upper triangle
                        corr_value = correlation_matrix.iloc[i, j]
                        values_text += f"• {symbol1} ↔ {symbol2}: {corr_value:.3f}\n"
            
            return values_text
            
        except Exception as e:
            self.logger.error(f"Error formatting correlation values: {e}")
            return ""
    
    async def _send_additional_charts(self, update: Update, context: ContextTypes.DEFAULT_TYPE, asset_list, symbols: list, currency: str):
        """Отправить дополнительные графики анализа (drawdowns, dividend yield)"""
        try:
            # Send typing indicator
            await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
            
            # Create drawdowns chart
            await self._create_drawdowns_chart(update, context, asset_list, symbols, currency)
            
            # Create dividend yield chart if available
            await self._create_dividend_yield_chart(update, context, asset_list, symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error creating additional charts: {e}")
            await self._send_message_safe(update, f"⚠️ Не удалось создать дополнительные графики: {str(e)}")
    

    
    async def _create_drawdowns_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, asset_list, symbols: list, currency: str):
        """Создать график drawdowns"""
        try:
            # Check if drawdowns data is available
            if not hasattr(asset_list, 'drawdowns') or asset_list.drawdowns.empty:
                await self._send_message_safe(update, "ℹ️ Данные о drawdowns недоступны для выбранных активов")
                return
            
            # Create drawdowns chart using chart_styles
            fig, ax = chart_styles.create_drawdowns_chart(
                asset_list.drawdowns, symbols, currency, data_source='okama'
            )
            
            # Save chart to bytes with memory optimization
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(fig)
            
            # Send drawdowns chart without keyboard
            await context.bot.send_photo(
                chat_id=update.effective_chat.id, 
                photo=io.BytesIO(img_bytes),
                caption=self._truncate_caption(f"📉 Периоды падения и восстановления")
            )
            
        except Exception as e:
            self.logger.error(f"Error creating drawdowns chart: {e}")
            await self._send_message_safe(update, f"⚠️ Не удалось создать график drawdowns: {str(e)}")
    
    async def _create_dividend_yield_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, asset_list, symbols: list, currency: str):
        """Создать график dividend yield"""
        try:
            # Check if dividend yield data is available
            if not hasattr(asset_list, 'dividend_yield') or asset_list.dividend_yield.empty:
                await self._send_message_safe(update, "ℹ️ Данные о дивидендной доходности недоступны для выбранных активов")
                return
            
            # Create dividend yield chart using chart_styles
            fig, ax = chart_styles.create_dividend_yield_chart(
                asset_list.dividend_yield, symbols, data_source='okama'
            )
            
            # Save chart to bytes with memory optimization
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(fig)
            
            # Send dividend yield chart without keyboard
            await context.bot.send_photo(
                chat_id=update.effective_chat.id, 
                photo=io.BytesIO(img_bytes),
                caption=self._truncate_caption(f"Сравнение дивидендной доходности {len(symbols)} активов\n\nИстория дивидендных выплат и доходность")
            )
            
        except Exception as e:
            self.logger.error(f"Error creating dividend yield chart: {e}")
            await self._send_message_safe(update, f"⚠️ Не удалось создать график дивидендной доходности: {str(e)}")
    
    async def _create_correlation_matrix(self, update: Update, context: ContextTypes.DEFAULT_TYPE, asset_list, symbols: list, currency: str):
        """Создать корреляционную матрицу активов"""
        try:
            self.logger.info(f"Starting correlation matrix creation for symbols: {symbols}")
            
            # Check if assets_ror data is available
            if not hasattr(asset_list, 'assets_ror'):
                self.logger.warning("assets_ror attribute does not exist")
                await self._send_message_safe(update, "ℹ️ Данные о доходности активов недоступны для создания корреляционной матрицы")
                return
                
            if asset_list.assets_ror is None:
                self.logger.warning("assets_ror is None")
                await self._send_message_safe(update, "ℹ️ Данные о доходности активов недоступны для создания корреляционной матрицы")
                return
                
            if asset_list.assets_ror.empty:
                self.logger.warning("assets_ror is empty")
                await self._send_message_safe(update, "ℹ️ Данные о доходности активов недоступны для создания корреляционной матрицы")
                return
            
            # Get correlation matrix
            correlation_matrix = asset_list.assets_ror.corr()
            
            self.logger.info(f"Correlation matrix created successfully, shape: {correlation_matrix.shape}")
            
            if correlation_matrix.empty:
                self.logger.warning("Correlation matrix is empty")
                await self._send_message_safe(update, "ℹ️ Не удалось вычислить корреляционную матрицу")
                return
            
            # Create correlation matrix visualization using chart_styles
            fig, ax = chart_styles.create_correlation_matrix_chart(
                correlation_matrix, data_source='okama'
            )
            
            # Save chart to bytes with memory optimization
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(fig)
            
            # Prepare correlation values text for caption
            correlation_values_text = self._format_correlation_values(correlation_matrix)
            
            # Send correlation matrix without keyboard
            self.logger.info("Sending correlation matrix image...")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id, 
                photo=io.BytesIO(img_bytes),
                caption=self._truncate_caption(f"🔗 Корреляция {len(symbols)} активов\n\n• +1: полная положительная корреляция\n• 0: отсутствие корреляции\n• -1: полная отрицательная корреляция\n\n{correlation_values_text}")
            )
            self.logger.info("Correlation matrix image sent successfully")
            
            plt.close(fig)
            
        except Exception as e:
            self.logger.error(f"Error creating correlation matrix: {e}")
            # Check if this is an FX-related error
            if "FX" in str(e) and "not found" in str(e):
                await self._send_message_safe(update, f"⚠️ Не удалось создать корреляционную матрицу: некоторые валютные пары недоступны в базе данных okama.\n\nПопробуйте использовать другие активы или проверьте доступность символов.")
            else:
                await self._send_message_safe(update, f"⚠️ Не удалось создать корреляционную матрицу: {str(e)}")
    
    async def start_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /start command with welcome message and interactive buttons"""
        # Ensure no reply keyboard is shown
        await self._ensure_no_reply_keyboard(update, context)
        
        user = update.effective_user
        user_name = user.first_name or "User"
        # Remove any special characters that could break Markdown
        user_name = user_name.replace("*", "").replace("_", "").replace("`", "").replace("[", "").replace("]", "")
        
        welcome_message = f"""👋 Здравствуйте! Я помогаю принимать взвешенные инвестиционные решения на основе данных, а не эмоций. Анализирую акции, ETF, валюты и товары 12 бирж, всего более 120 000 инструментов.

Попробуйте одну из ключевых функций прямо сейчас:

🔍 Анализ: полная сводка по любой бумаге, валюте или товару /info

⚖️ Сравнение: объективная оценка нескольких активов по десяткам метрик /compare

💼 Портфель: создание, анализ и прогнозирование доходности ваших портфелей /portfolio

📚 Просмотр всех доступных данных и символов /list

© Okama, tushare, YandexGPT, Google Gemini.
"""

        # Create reply keyboard with interactive buttons
        reply_markup = self._create_start_reply_keyboard()
        
        await self._send_message_safe(update, welcome_message, reply_markup=reply_markup)

    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /help command with full help"""
        # Ensure no reply keyboard is shown
        await self._ensure_no_reply_keyboard(update, context)
        
        welcome_message = f"""📘 *Справка по командам*

🔹 *Основные команды*

`/info <тикер>` — информация об активе (графики цен, ключевые метрики)
Пример: `/info GAZP.MOEX`

`/compare <тикер1> <тикер2>` ... — сравнение нескольких активов (график накопленной доходности, инфляция)
Пример: `/compare SPY.US SBER.MOEX`

`/portfolio <тикер1:вес> <тикер2:вес>` ... — создание и анализ портфеля (состав, риски, доходность, прогнозы)
Пример: `/portfolio SBER.MOEX:0.4 GAZP.MOEX:0.3 LKOH.MOEX:0.3`

`/list` — список доступных данных и тм
`/search <название или ISIN>` — поиск актива по базе okama и tushare
Пример: `/search Apple`



🔹 *Примеры команд*

*Сравнение активов*
`/compare SBER.MOEX,GAZP.MOEX` — сравнить Сбербанк и Газпром
`/compare SPY.US, QQQ.US, VOO.US` — сравнить ETF на США
`/compare GC.COMM CL.COMM` — сравнить золото и нефть
`/compare VOO.US,BND.US,GC.COMM` — акции, облигации и золото

*Создание портфеля*
`/portfolio SPY.US:0.5 QQQ.US:0.3 BND.US:0.2` — портфель США (50% S&P 500, 30% NASDAQ, 20% облигации)
`/portfolio SBER.MOEX:0.4 GAZP.MOEX:0.3 LKOH.MOEX:0.3` — портфель из российских акций

⚠️ *Важно*: Вся информация предоставляется исключительно в информационных целях и не является инвестиционной рекомендацией."""

        await self._send_message_safe(update, welcome_message)
    
    async def show_info_help(self, update: Update):
        """Показать справку по команде /info"""
        help_text = """🔍 Команда /info - Информация об активе

Используйте команду `/info [тикер] [период]` для получения полной информации об активе.

Примеры команд:
• `/info AAPL.US` - информация об Apple
• `/info SBER.MOEX` - информация о Сбербанке
• `/info GC.COMM 5Y` - золото за 5 лет
• `/info SPX.INDX 10Y` - S&P 500 за 10 лет

Поддерживаемые периоды:
• 1Y, 2Y, 5Y, 10Y, MAX
• По умолчанию: 10Y для акций, 5Y для макро
"""
        
        await self._send_message_safe(update, help_text)
    
    async def show_namespace_help(self, update: Update):
        """Показать справку по команде /list"""
        help_text = """📚 Команда /list - Пространства имен

Используйте команду `/list` для просмотра всех доступных пространств имен.

• `/list US` - американские акции
• `/list MOEX` - российские акции
• `/list INDX` - мировые индексы
• `/list FX` - валютные пары
• `/list COMM` - товарные активы

"""
        
        await self._send_message_safe(update, help_text)
    
    async def _show_tushare_namespace_symbols(self, update: Update, context: ContextTypes.DEFAULT_TYPE, namespace: str, is_callback: bool = False, page: int = 0):
        """Show symbols for Chinese exchanges using Tushare"""
        try:
            if not self.tushare_service:
                error_msg = "❌ Сервис Tushare недоступен"
                if is_callback:
                    await context.bot.send_message(
                        chat_id=update.callback_query.message.chat_id,
                        text=error_msg
                    )
                else:
                    await self._send_message_safe(update, error_msg)
                return
            
            # Get symbols from Tushare
            try:
                symbols_data = self.tushare_service.get_exchange_symbols(namespace)
                total_count = self.tushare_service.get_exchange_symbols_count(namespace)
                
                if not symbols_data:
                    error_msg = f"❌ Символы для биржи '{namespace}' не найдены"
                    if is_callback:
                        await context.bot.send_message(
                            chat_id=update.callback_query.message.chat_id,
                            text=error_msg
                        )
                    else:
                        await self._send_message_safe(update, error_msg)
                    return
            except Exception as e:
                error_msg = f"❌ Ошибка при получении символов для '{namespace}': {str(e)}"
                if is_callback:
                    await context.bot.send_message(
                        chat_id=update.callback_query.message.chat_id,
                        text=error_msg
                    )
                else:
                    await self._send_message_safe(update, error_msg)
                return
            
            # Format response
            exchange_names = {
                'SSE': 'Shanghai Stock Exchange',
                'SZSE': 'Shenzhen Stock Exchange', 
                'BSE': 'Beijing Stock Exchange',
                'HKEX': 'Hong Kong Stock Exchange'
            }
            
            # Pagination logic
            symbols_per_page = 20
            total_pages = (total_count + symbols_per_page - 1) // symbols_per_page
            current_page = min(page, total_pages - 1) if total_pages > 0 else 0
            
            # Calculate start and end indices for current page
            start_idx = current_page * symbols_per_page
            end_idx = min(start_idx + symbols_per_page, len(symbols_data))
            
            response = f"📊 **{exchange_names.get(namespace, namespace)}**\n\n"
            response += f"📈 Всего: {total_count:,}\n"
            response += f"📄 Страница {current_page + 1} из {total_pages}\n\n"
            
            # Get symbols for current page
            page_symbols = symbols_data[start_idx:end_idx]
            
            # Create bullet list format
            symbol_list = []
            
            for symbol_info in page_symbols:
                symbol = symbol_info['symbol']
                name = symbol_info['name']
                
                # Simple escaping for list display - only escape characters that interfere with bold formatting
                escaped_name = name.replace('*', '\\*').replace('_', '\\_')
                
                # Truncate name to maximum 40 characters
                if len(escaped_name) > 40:
                    escaped_name = escaped_name[:37] + "..."
                
                # Create bullet list item with bold ticker
                symbol_list.append(f"• **`{symbol}`** - {escaped_name}")
            
            # Add symbol list to response
            if symbol_list:
                response += "\n".join(symbol_list) + "\n"
            
            response += f"\n💡 Скопируйте тикер и вставьте в команды `/info`, `/compare`, `/portfolio`"
            
            # Create appropriate keyboard based on context
            if is_callback:
                # For callback messages, use inline keyboard (original behavior)
                keyboard = []
                
                # Navigation buttons (only if more than one page)
                if total_pages > 1:
                    nav_buttons = []
                    
                    # Previous button
                    if current_page > 0:
                        nav_buttons.append(InlineKeyboardButton(
                            "⬅️ Назад", 
                            callback_data=f"nav_namespace_{namespace}_{current_page - 1}"
                        ))
                    
                    # Page indicator
                    nav_buttons.append(InlineKeyboardButton(
                        f"{current_page + 1}/{total_pages}", 
                        callback_data="noop"
                    ))
                    
                    # Next button
                    if current_page < total_pages - 1:
                        nav_buttons.append(InlineKeyboardButton(
                            "➡️ Вперед", 
                            callback_data=f"nav_namespace_{namespace}_{current_page + 1}"
                        ))
                    
                    keyboard.append(nav_buttons)
                
                # Excel export button
                keyboard.append([
                    InlineKeyboardButton(
                        f"📊 Полный список в Excel ({total_count:,})", 
                        callback_data=f"excel_namespace_{namespace}"
                    )
                ])
                
                # Home button after Excel
                keyboard.append([
                    InlineKeyboardButton("🏠 Домой", callback_data="namespace_home")
                ])
                
                # Analysis, Compare, Portfolio buttons
                keyboard.append([
                    InlineKeyboardButton("🔍 Анализ", callback_data="namespace_analysis"),
                    InlineKeyboardButton("⚖️ Сравнить", callback_data="namespace_compare"),
                    InlineKeyboardButton("💼 В портфель", callback_data="namespace_portfolio")
                ])
                
                reply_markup = InlineKeyboardMarkup(keyboard)
            else:
                # For direct command calls, use reply keyboard
                reply_markup = self._create_list_namespace_reply_keyboard(namespace, current_page, total_pages, total_count)
                
                # Save current namespace context for reply keyboard handling
                user_id = update.effective_user.id
                self._update_user_context(user_id, 
                    current_namespace=namespace,
                    current_namespace_page=current_page
                )
            
            if is_callback:
                await context.bot.edit_message_text(
                    chat_id=update.callback_query.message.chat_id,
                    message_id=update.callback_query.message.message_id,
                    text=response,
                    parse_mode='Markdown',
                    reply_markup=reply_markup
                )
            else:
                await self._send_message_safe(update, response, reply_markup=reply_markup)
                
        except Exception as e:
            error_msg = f"❌ Ошибка при получении данных для '{namespace}': {str(e)}"
            if is_callback:
                await context.bot.send_message(
                    chat_id=update.callback_query.message.chat_id,
                    text=error_msg
                )
            else:
                await self._send_message_safe(update, error_msg)
    
    async def _show_namespace_symbols(self, update: Update, context: ContextTypes.DEFAULT_TYPE, namespace: str, is_callback: bool = False, page: int = 0):
        """Единый метод для показа символов в пространстве имен с навигацией"""
        try:
            # Check if it's a Chinese exchange
            chinese_exchanges = ['SSE', 'SZSE', 'BSE', 'HKEX']
            if namespace in chinese_exchanges:
                await self._show_tushare_namespace_symbols(update, context, namespace, is_callback, page)
                return
            
            symbols_df = ok.symbols_in_namespace(namespace)
            
            if symbols_df.empty:
                error_msg = f"❌ Пространство имен '{namespace}' не найдено или пусто"
                if is_callback:
                    # Для callback сообщений отправляем через context.bot
                    await context.bot.send_message(
                        chat_id=update.callback_query.message.chat_id,
                        text=error_msg
                    )
                else:
                    await self._send_message_safe(update, error_msg)
                return
            
            # Show statistics first
            total_symbols = len(symbols_df)
            symbols_per_page = 20  # Показываем по 20 символов на страницу
            
            # Calculate pagination
            total_pages = (total_symbols + symbols_per_page - 1) // symbols_per_page
            current_page = min(page, total_pages - 1) if total_pages > 0 else 0
            
            # Calculate start and end indices
            start_idx = current_page * symbols_per_page
            end_idx = min(start_idx + symbols_per_page, total_symbols)
            
            # Navigation info instead of first symbols
            response = f"📊 **{namespace}** - Всего тикеров: {total_symbols:,}\n\n"
            response += f"📋 Показаны символы {start_idx + 1}-{end_idx} из {total_symbols}\n"
            response += f"📄 Страница {current_page + 1} из {total_pages}\n\n"
            
            # Get symbols for current page
            page_symbols = symbols_df.iloc[start_idx:end_idx]
            
            # Create bullet list format
            symbol_list = []
            
            for _, row in page_symbols.iterrows():
                symbol = row['symbol'] if pd.notna(row['symbol']) else 'N/A'
                name = row['name'] if pd.notna(row['name']) else 'N/A'
                
                # Simple escaping for list display - only escape characters that interfere with bold formatting
                escaped_name = name.replace('*', '\\*').replace('_', '\\_')
                
                # Truncate name to maximum 40 characters
                if len(escaped_name) > 40:
                    escaped_name = escaped_name[:37] + "..."
                
                # Create bullet list item with bold ticker
                symbol_list.append(f"• **`{symbol}`** - {escaped_name}")
            
            # Add symbol list to response
            if symbol_list:
                response += "\n".join(symbol_list) + "\n"
            else:
                # If no symbols found, add a message
                response += "❌ Символы не найдены для данной страницы\n"

            response += "\n💡 Используйте /search для поиска активов"

            # Create reply keyboard for navigation
            reply_markup = self._create_list_namespace_reply_keyboard(namespace, current_page, total_pages, total_symbols)
            
            # Save current namespace context for reply keyboard handling
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                current_namespace=namespace,
                current_namespace_page=current_page
            )
            
            # Отправляем сообщение с таблицей и кнопками
            self.logger.info(f"About to send message. Response length: {len(response)}, Is callback: {is_callback}")
            if is_callback:
                # Для callback сообщений отправляем новое сообщение с reply keyboard
                await context.bot.send_message(
                    chat_id=update.callback_query.message.chat_id,
                    text=response,
                    parse_mode='Markdown',
                    reply_markup=reply_markup
                )
            else:
                await self._send_message_safe(update, response, reply_markup=reply_markup)
            
        except Exception as e:
            error_msg = f"❌ Ошибка при получении данных для '{namespace}': {str(e)}"
            if is_callback:
                # Для callback сообщений отправляем через context.bot
                await context.bot.send_message(
                    chat_id=update.callback_query.message.chat_id,
                    text=error_msg
                )
            else:
                await self._send_message_safe(update, error_msg)
    
    async def _show_namespace_symbols_with_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE, namespace: str, page: int = 0):
        """Show namespace symbols with reply keyboard - for namespace button clicks"""
        try:
            symbols_df = ok.symbols_in_namespace(namespace)
            
            if symbols_df.empty:
                error_msg = f"❌ Пространство имен '{namespace}' не найдено или пусто"
                await context.bot.send_message(
                    chat_id=update.callback_query.message.chat_id,
                    text=error_msg
                )
                return
            
            # Show statistics first
            total_symbols = len(symbols_df)
            symbols_per_page = 20  # Показываем по 20 символов на страницу
            
            # Calculate pagination
            total_pages = (total_symbols + symbols_per_page - 1) // symbols_per_page
            current_page = min(page, total_pages - 1) if total_pages > 0 else 0
            
            # Calculate start and end indices
            start_idx = current_page * symbols_per_page
            end_idx = min(start_idx + symbols_per_page, total_symbols)
            
            # Navigation info instead of first symbols
            response = f"📊 **{namespace}** - Всего символов: {total_symbols:,}\n\n"
            response += f"📋 **Навигация:** Показаны символы {start_idx + 1}-{end_idx} из {total_symbols}\n"
            response += f"📄 Страница {current_page + 1} из {total_pages}\n\n"
            
            # Get symbols for current page
            page_symbols = symbols_df.iloc[start_idx:end_idx]
            
            # Create bullet list format
            symbol_list = []
            
            for _, row in page_symbols.iterrows():
                symbol = row['symbol'] if pd.notna(row['symbol']) else 'N/A'
                name = row['name'] if pd.notna(row['name']) else 'N/A'
                
                # Escape special characters for Markdown
                escaped_name = name.replace('*', '\\*').replace('_', '\\_').replace('[', '\\[').replace(']', '\\]')
                
                # Create bullet list item with bold ticker
                symbol_list.append(f"• **`{symbol}`** - {escaped_name}")
            
            # Add symbol list to response
            if symbol_list:
                response += "\n".join(symbol_list) + "\n"
            
            # Create reply keyboard
            reply_markup = self._create_list_namespace_reply_keyboard(namespace, current_page, total_pages, total_symbols)
            
            # Save current namespace context for reply keyboard handling
            user_id = update.callback_query.from_user.id
            self._update_user_context(user_id, 
                current_namespace=namespace,
                current_namespace_page=current_page
            )
            
            # Send new message with reply keyboard
            await context.bot.send_message(
                chat_id=update.callback_query.message.chat_id,
                text=response,
                parse_mode='Markdown',
                reply_markup=reply_markup
            )
            
        except Exception as e:
            error_msg = f"❌ Ошибка при получении данных для '{namespace}': {str(e)}"
            await context.bot.send_message(
                chat_id=update.callback_query.message.chat_id,
                text=error_msg
            )

    async def _show_tushare_namespace_symbols_with_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE, namespace: str, page: int = 0):
        """Show Tushare namespace symbols with reply keyboard - for namespace button clicks"""
        try:
            if not self.tushare_service:
                await context.bot.send_message(
                    chat_id=update.callback_query.message.chat_id,
                    text="❌ Сервис Tushare не настроен"
                )
                return
            
            # Get symbols from Tushare
            symbols_data = self.tushare_service.get_exchange_symbols(namespace)
            
            if not symbols_data:
                await context.bot.send_message(
                    chat_id=update.callback_query.message.chat_id,
                    text=f"❌ Пространство имен '{namespace}' не найдено или пусто"
                )
                return
            
            # Show statistics first
            total_count = len(symbols_data)
            symbols_per_page = 20  # Показываем по 20 символов на страницу
            
            # Calculate pagination
            total_pages = (total_count + symbols_per_page - 1) // symbols_per_page
            current_page = min(page, total_pages - 1) if total_pages > 0 else 0
            
            # Calculate start and end indices
            start_idx = current_page * symbols_per_page
            end_idx = min(start_idx + symbols_per_page, total_count)
            
            # Navigation info instead of first symbols
            response = f"📊 **{namespace}** - Всего символов: {total_count:,}\n\n"
            response += f"📋 **Навигация:** Показаны символы {start_idx + 1}-{end_idx} из {total_count}\n"
            response += f"📄 Страница {current_page + 1} из {total_pages}\n\n"
            
            # Get symbols for current page
            page_symbols = symbols_data[start_idx:end_idx]
            
            # Create bullet list format
            symbol_list = []
            
            for symbol_data in page_symbols:
                symbol = symbol_data.get('symbol', 'N/A')
                name = symbol_data.get('name', 'N/A')
                
                # Escape special characters for Markdown
                escaped_name = name.replace('*', '\\*').replace('_', '\\_').replace('[', '\\[').replace(']', '\\]')
                
                # Create bullet list item with bold ticker
                symbol_list.append(f"• **`{symbol}`** - {escaped_name}")
            
            # Add symbol list to response
            if symbol_list:
                response += "\n".join(symbol_list) + "\n"
            
            # Create reply keyboard
            reply_markup = self._create_list_namespace_reply_keyboard(namespace, current_page, total_pages, total_count)
            
            # Save current namespace context for reply keyboard handling
            user_id = update.callback_query.from_user.id
            self._update_user_context(user_id, 
                current_namespace=namespace,
                current_namespace_page=current_page
            )
            
            # Send new message with reply keyboard
            await context.bot.send_message(
                chat_id=update.callback_query.message.chat_id,
                text=response,
                parse_mode='Markdown',
                reply_markup=reply_markup
            )
            
        except Exception as e:
            error_msg = f"❌ Ошибка при получении данных для '{namespace}': {str(e)}"
            await context.bot.send_message(
                chat_id=update.callback_query.message.chat_id,
                text=error_msg
            )

    async def info_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /info command - показывает ежедневный график с базовой информацией и AI анализом"""
        # Ensure no reply keyboard is shown
        await self._ensure_no_reply_keyboard(update, context)
        
        if not context.args:
            # Get random examples for user
            examples = self.examples_service.get_info_examples(5)
            examples_text = "\n".join([f"• {example}" for example in examples])
            
            # Set flag that user is waiting for info input and clear portfolio flags
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                waiting_for_info=True,
                waiting_for_portfolio=False,
                waiting_for_portfolio_weights=False,
                waiting_for_compare=False
            )
            
            await self._send_message_safe(update, 
                f"🔍 *Анализ*\n\n"
                f"*Примеры:*\n{examples_text}\n\n"
                f"💬 Просто напишите название, тикер или ISIN инструмента")
            return
        
        symbol = context.args[0]
        
        # Update user context - clear all waiting flags
        user_id = update.effective_user.id
        self._update_user_context(user_id, 
            waiting_for_info=False,
            waiting_for_portfolio=False,
            waiting_for_portfolio_weights=False,
            waiting_for_compare=False
        )
        
        await self._send_ephemeral_message(update, context, f"📊 Поиск {symbol}...", delete_after=3)
        
        try:
            # Search for assets with selection possibility
            search_result = self.search_assets_with_selection(symbol)
            
            if 'error' in search_result:
                await self._send_message_safe(update, f"❌ Ошибка: {search_result['error']}")
                return
            
            # If multiple results, show selection menu
            if 'results' in search_result:
                results = search_result['results']
                query = search_result['query']
                
                message = f"🔍 *Найдено {len(results)} активов по запросу '{query}':*\n\n"
                message += "Выберите нужный актив:"
                
                keyboard = self._create_asset_selection_keyboard(results, query)
                
                await self._send_message_safe(update, message, reply_markup=keyboard)
                return
            
            # Single result - proceed with info
            resolved_symbol = search_result['symbol']
            
            # Update user context with the selected asset
            self._update_user_context(user_id, 
                                    last_assets=[resolved_symbol] + self._get_user_context(user_id).get('last_assets', []))
            
            # Add to analyzed tickers history
            self._add_to_analyzed_tickers(user_id, resolved_symbol)
            
            # Determine data source
            data_source = self.determine_data_source(resolved_symbol)
            
            if data_source == 'tushare':
                # Use Tushare service for Chinese exchanges
                await self._handle_tushare_info(update, resolved_symbol, context)
            else:
                # Use Okama for other exchanges
                await self._handle_okama_info(update, resolved_symbol, context)
                
        except Exception as e:
            self.logger.error(f"Error in info command for {symbol}: {e}")
            await self._send_message_safe(update, f"❌ Ошибка: {str(e)}")

    async def handle_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle text messages - treat as asset symbol for /info or portfolio for /portfolio"""
        if not update.message or not update.message.text:
            return
        
        original_text = update.message.text.strip()
        
        # Check if this is a Reply Keyboard button BEFORE cleaning
        if self._is_reply_keyboard_button(original_text):
            await self._handle_reply_keyboard_button(update, context, original_text)
            return
        
        text = self.clean_symbol(original_text)
        if not text:
            return
        
        user_id = update.effective_user.id
        user_context = self._get_user_context(user_id)
        
        # Debug: Log user context state
        self.logger.info(f"User {user_id} context: {user_context}")
        self.logger.info(f"Waiting for portfolio: {user_context.get('waiting_for_portfolio', False)}")
        self.logger.info(f"Input text: {text}")
        self.logger.info(f"Original text: {original_text}")
        
        # Check if user is waiting for portfolio input
        if user_context.get('waiting_for_portfolio', False):
            self.logger.info(f"Processing as portfolio input: {text}")
            # Process as portfolio input
            await self._handle_portfolio_input(update, context, text)
            return
        
        # Check if user is waiting for portfolio weights input (from compare command)
        if user_context.get('waiting_for_portfolio_weights', False):
            self.logger.info(f"Processing as portfolio weights input: {text}")
            
            # Check if this is from portfolio command (tickers only) or compare command
            if user_context.get('portfolio_tickers'):
                # This is from portfolio command with tickers only
                await self._handle_portfolio_tickers_weights_input(update, context, text)
            else:
                # This is from compare command - handle as compare input instead of portfolio weights
                await self._handle_compare_input(update, context, text)
            return
        
        # Check if user is waiting for info input
        if user_context.get('waiting_for_info', False):
            self.logger.info(f"Processing as info input: {text}")
            # Clear all waiting flags
            self._update_user_context(user_id, 
                waiting_for_info=False,
                waiting_for_portfolio=False,
                waiting_for_portfolio_weights=False,
                waiting_for_compare=False
            )
            # Process as info command with the symbol
            context.args = [text]
            await self.info_command(update, context)
            return
        
        # Check if user is waiting for compare input or has stored compare symbol
        if user_context.get('waiting_for_compare', False) or user_context.get('compare_first_symbol') or user_context.get('compare_base_symbol'):
            self.logger.info(f"Processing as compare input: {text}")
            # Process as compare input
            await self._handle_compare_input(update, context, text)
            return
        
        # Check if text contains multiple symbols (space or comma separated)
        # This allows users to send "SPY.US QQQ.US" directly as a comparison request
        symbols = []
        if ',' in text:
            # Handle comma-separated symbols
            for symbol_part in text.split(','):
                symbol_part = self.clean_symbol(symbol_part.strip())
                if symbol_part:
                    if any(portfolio_indicator in symbol_part.upper() for portfolio_indicator in ['PORTFOLIO_', 'PF_', 'PORTFOLIO_', '.PF', '.pf']):
                        symbols.append(symbol_part)
                    else:
                        symbols.append(symbol_part.upper())
        elif ' ' in text:
            # Handle space-separated symbols
            for symbol in text.split():
                symbol = self.clean_symbol(symbol.strip())
                if symbol:
                    if any(portfolio_indicator in symbol.upper() for portfolio_indicator in ['PORTFOLIO_', 'PF_', 'PORTFOLIO_', '.PF', '.pf']):
                        symbols.append(symbol)
                    else:
                        symbols.append(symbol.upper())
        
        # If we detected multiple symbols, treat as comparison request
        if len(symbols) >= 2:
            self.logger.info(f"Detected multiple symbols in message: {symbols}")
            # Process as compare input
            await self._handle_compare_input(update, context, text)
            return
        
        # Treat text as single asset symbol and process with /info logic
        symbol = text
        
        await self._send_ephemeral_message(update, context, f"🔍 Ищу актив '{symbol}'...", delete_after=3)
        
        try:
            # Search for assets with selection possibility
            search_result = self.search_assets_with_selection(symbol)
            
            if 'error' in search_result:
                await self._send_message_safe(update, f"❌ Ошибка: {search_result['error']}")
                return
            
            # If multiple results, show selection menu
            if 'results' in search_result:
                results = search_result['results']
                query = search_result['query']
                
                message = f"🔍 *Найдено {len(results)} активов по запросу '{query}':*\n\n"
                message += "Выберите нужный актив:"
                
                keyboard = self._create_asset_selection_keyboard(results, query)
                
                await self._send_message_safe(update, message, reply_markup=keyboard)
                return
            
            # Single result - proceed with info
            resolved_symbol = search_result['symbol']
            
            # Update user context with the selected asset
            self._update_user_context(user_id, 
                                    last_assets=[resolved_symbol] + user_context.get('last_assets', []))
            
            # Add to analyzed tickers history
            self._add_to_analyzed_tickers(user_id, resolved_symbol)
            
            # Determine data source
            data_source = self.determine_data_source(resolved_symbol)
            
            if data_source == 'tushare':
                # Use Tushare service for Chinese exchanges
                await self._handle_tushare_info(update, resolved_symbol, context)
            else:
                # Use Okama for other exchanges
                await self._handle_okama_info(update, resolved_symbol, context)
                
        except Exception as e:
            self.logger.error(f"Error in handle_message for {symbol}: {e}")
            await self._send_message_safe(update, f"❌ Ошибка: {str(e)}")

    async def _handle_okama_info(self, update: Update, symbol: str, context: ContextTypes.DEFAULT_TYPE = None):
        """Handle info display for Okama assets with new interactive structure"""
        try:
            # Получаем объект актива
            try:
                asset = ok.Asset(symbol)
                
                # Получаем ключевые метрики за 1 год
                key_metrics = await self._get_asset_key_metrics(asset, symbol, period='1Y')
                
                # Формируем структурированный ответ
                info_text = self._format_asset_info_response(asset, symbol, key_metrics)
                
                # Создаем reply keyboard для управления
                reply_markup = self._create_info_reply_keyboard()
                
                # Save current symbol context for reply keyboard handling
                user_id = update.effective_user.id
                self._update_user_context(user_id, 
                    current_info_symbol=symbol
                )
                
                # Получаем график доходности за 1 год
                self.logger.info(f"Getting daily chart for {symbol}")
                chart_data = await self._get_daily_chart(symbol)
                self.logger.info(f"Chart data result: {chart_data is not None}")
                
                if chart_data:
                    # Отправляем график с информацией в caption
                    caption = f"📈 График доходности за 1 год\n\n{info_text}"
                    self.logger.info(f"Sending chart with caption length: {len(caption)}")
                    await self._send_photo_safe(update, chart_data, caption=caption, reply_markup=reply_markup, context=context)
                else:
                    # Если график не удалось получить, отправляем только текст
                    self.logger.warning(f"Could not get chart for {symbol}, sending text only")
                    await self._send_message_safe(update, info_text, reply_markup=reply_markup)
                
            except Exception as e:
                # При ошибке получения данных актива отправляем только сообщение об ошибке без кнопок
                error_text = f"❌ Ошибка при получении информации об активе: {str(e)}"
                await self._send_message_safe(update, error_text)
            
        except Exception as e:
            self.logger.error(f"Error in _handle_okama_info for {symbol}: {e}")
            await self._send_message_safe(update, f"❌ Ошибка: {str(e)}")

    async def _handle_tushare_info(self, update: Update, symbol: str, context: ContextTypes.DEFAULT_TYPE = None):
        """Handle info display for Tushare assets with new interactive structure"""
        try:
            if not self.tushare_service:
                await self._send_message_safe(update, "*❌ Сервис Tushare недоступен*")
                return
            
            # Get symbol information from Tushare
            symbol_info = self.tushare_service.get_symbol_info(symbol)
            
            if 'error' in symbol_info:
                # При ошибке отправляем сообщение об ошибке с кнопками для консистентности
                error_text = f"❌ Ошибка: {symbol_info['error']}"
                
                # Создаем кнопки даже при ошибке для консистентности с Okama
                keyboard = self._create_info_interactive_keyboard(symbol)
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                await self._send_message_safe(update, error_text, reply_markup=reply_markup)
                return
            
            # Format information according to new structure
            info_text = self._format_tushare_info_response(symbol_info, symbol)
            
            # Create reply keyboard for management
            reply_markup = self._create_info_reply_keyboard()
            
            # Save current symbol context for reply keyboard handling
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                current_info_symbol=symbol
            )
            
            # Try to get chart data
            chart_data = await self._get_tushare_chart(symbol)
            
            if chart_data:
                # Create enhanced caption with English information
                chart_caption = self._format_tushare_chart_caption(symbol_info, symbol, "1 год")
                caption = f"{chart_caption}\n\n{info_text}"
                await self._send_photo_safe(update, chart_data, caption=caption, reply_markup=reply_markup, context=context)
            else:
                await self._send_message_safe(update, info_text, reply_markup=reply_markup)
            
        except Exception as e:
            self.logger.error(f"Error in _handle_tushare_info for {symbol}: {e}")
            await self._send_message_safe(update, f"❌ Ошибка: {str(e)}")

    def _format_tushare_chart_caption(self, symbol_info: Dict[str, Any], symbol: str, period_text: str) -> str:
        """Format chart caption with English information for Chinese/Hong Kong assets"""
        try:
            # Get English name
            english_name = symbol_info.get('english_name', symbol_info.get('name', symbol.split('.')[0]))
            
            # Get exchange and location information
            exchange = symbol_info.get('exchange', 'N/A')
            area = symbol_info.get('area', 'N/A')
            
            # Map exchanges to English names
            exchange_map = {
                'SSE': 'Shanghai Stock Exchange',
                'SZSE': 'Shenzhen Stock Exchange', 
                'HKEX': 'Hong Kong Exchange',
                'SH': 'Shanghai Stock Exchange',
                'SZ': 'Shenzhen Stock Exchange',
                'HK': 'Hong Kong Exchange'
            }
            
            # Map areas to English names
            area_map = {
                'China': 'China',
                'Hong Kong': 'Hong Kong',
                'Shanghai': 'Shanghai',
                'Shenzhen': 'Shenzhen'
            }
            
            english_exchange = exchange_map.get(exchange, exchange)
            english_location = area_map.get(area, area)
            
            # Determine currency
            currency = 'HKD' if symbol.endswith('.HK') else 'CNY'
            
            # Format listing date if available
            listing_date = symbol_info.get('list_date', '')
            formatted_date = ''
            if listing_date and listing_date != 'N/A' and len(str(listing_date)) == 8:
                try:
                    # Convert YYYYMMDD to YYYY-MM-DD
                    date_str = str(listing_date)
                    formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                except:
                    formatted_date = str(listing_date)
            
            # Create caption
            caption = f"📈 {english_name}\n"
            caption += f"🏢 {english_exchange}\n"
            caption += f"📍 {english_location}\n"
            caption += f"💰 {currency}"
            
            if formatted_date:
                caption += f"\n📅 Listing Date: {formatted_date}"
            
            return caption
            
        except Exception as e:
            self.logger.error(f"Error formatting chart caption for {symbol}: {e}")
            return f"📈 {symbol}"

    def _get_max_period_years_tushare(self, symbol_info: Dict[str, Any]) -> str:
        """Get maximum available period in years for Tushare asset"""
        try:
            list_date = symbol_info.get('list_date', '')
            if list_date and len(str(list_date)) == 8:
                # Calculate years since listing
                from datetime import datetime
                try:
                    listing_date = datetime.strptime(str(list_date), '%Y%m%d')
                    years_since_listing = (datetime.now() - listing_date).days / 365.25
                    
                    if years_since_listing < 1:
                        return "максимальный период данных (лет)"
                    elif years_since_listing < 2:
                        return f"максимальный период данных ({years_since_listing:.1f} лет)"
                    else:
                        return f"максимальный период данных ({years_since_listing:.0f} лет)"
                except:
                    return "максимальный период данных (лет)"
            else:
                return "максимальный период данных (лет)"
                
        except Exception as e:
            self.logger.error(f"Error calculating max period years for Tushare: {e}")
            return "максимальный период данных (лет)"

    def _calculate_tushare_period_metrics(self, symbol_info: Dict[str, Any], symbol: str, period: str) -> Dict[str, Any]:
        """Calculate metrics for specific period for Tushare assets"""
        try:
            if not self.tushare_service:
                return {}
            
            # Get ts_code
            ts_code = symbol_info.get('ts_code')
            if not ts_code:
                return {}
            
            # Determine days based on period
            if period == '1Y':
                days = 252
            elif period == '5Y':
                days = 1260
            elif period == 'MAX':
                days = 2520  # 10 years max
            else:
                days = 252
            
            # Get daily data for the period
            daily_data = self.tushare_service.get_daily_data_by_days(ts_code, days)
            if daily_data is None or daily_data.empty:
                return {}
            
            # Calculate returns
            daily_data = daily_data.sort_values('trade_date')
            daily_data['returns'] = daily_data['close'].pct_change()
            returns = daily_data['returns'].dropna()
            
            if returns.empty:
                return {}
            
            # Calculate annual return (CAGR)
            if len(returns) > 1:
                total_return = (daily_data['close'].iloc[-1] / daily_data['close'].iloc[0]) - 1
                years = len(returns) / 252.0  # Approximate trading days per year
                annual_return = (1 + total_return) ** (1 / years) - 1 if years > 0 else 0
            else:
                annual_return = 0
            
            # Calculate volatility (annualized)
            volatility = returns.std() * (252 ** 0.5) if len(returns) > 1 else 0
            
            return {
                'annual_return': annual_return,
                'volatility': volatility
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating Tushare period metrics for {symbol} period {period}: {e}")
            return {}

    def _format_tushare_info_response(self, symbol_info: Dict[str, Any], symbol: str, period: str = '1Y') -> str:
        """Format Tushare info response according to new structure"""
        try:
            # Block 1: Header - who is this?
            asset_name = symbol_info.get('name', symbol)
            english_name = symbol_info.get('english_name', '')
            exchange = symbol_info.get('exchange', 'N/A')
            industry = symbol_info.get('industry', 'N/A')
            area = symbol_info.get('area', 'N/A')
            
            # Format header with Chinese and English names
            if english_name and english_name != asset_name:
                header = f"📊 {asset_name} ({symbol})\n"
                header += f"🌐 {english_name}\n"
            else:
                header = f"📊 {asset_name} ({symbol})\n"
            header += f"📍 {area} | {industry} | {exchange}"
            
            # Block 2: Key metrics showcase
            period_text = {
                '1Y': '1 год',
                '5Y': '5 лет',
                'MAX': self._get_max_period_years_tushare(symbol_info)
            }.get(period, period)
            metrics_text = f"\n\nКлючевые показатели (за {period_text}):\n"
            
            # Current price
            if 'current_price' in symbol_info:
                price = symbol_info['current_price']
                price_text = f"• Цена: {price:.2f} CNY"
                
                # Show only percentage change to avoid duplicate brackets
                if 'pct_chg' in symbol_info:
                    pct_chg = symbol_info['pct_chg']
                    change_sign = "+" if pct_chg >= 0 else ""
                    price_text += f" ({change_sign}{pct_chg:.2f}%)"
                
                metrics_text += f"{price_text}\n"
            
            # Volume
            if 'volume' in symbol_info:
                volume = symbol_info['volume']
                metrics_text += f"• Объем торгов: {volume:,.0f}\n"
            
            # Add calculated metrics for the specific period
            period_metrics = self._calculate_tushare_period_metrics(symbol_info, symbol, period)
            if period_metrics:
                if 'annual_return' in period_metrics:
                    annual_return = period_metrics['annual_return']
                    return_sign = "+" if annual_return >= 0 else ""
                    metrics_text += f"• Доходность (годовая): {return_sign}{annual_return:.1%}\n"
                
                if 'volatility' in period_metrics:
                    volatility = period_metrics['volatility']
                    metrics_text += f"• Волатильность: {volatility:.1%}\n"
            else:
                # Fallback to original metrics if period calculation fails
                if 'annual_return' in symbol_info:
                    annual_return = symbol_info['annual_return']
                    return_sign = "+" if annual_return >= 0 else ""
                    metrics_text += f"• Доходность (годовая): {return_sign}{annual_return:.1%}\n"
                
                if 'volatility' in symbol_info:
                    volatility = symbol_info['volatility']
                    metrics_text += f"Волатильность: {volatility:.1%}\n"
            
            # List date
            if 'list_date' in symbol_info:
                list_date = symbol_info['list_date']
                metrics_text += f"• Дата листинга: {list_date}\n"
            
            return header + metrics_text
            
        except Exception as e:
            self.logger.error(f"Error formatting Tushare info response for {symbol}: {e}")
            return f"🔍 {symbol}\n\n❌ Ошибка при форматировании информации"

    async def _get_tushare_chart(self, symbol: str) -> Optional[bytes]:
        """Get chart for Tushare asset"""
        try:
            if not self.tushare_service:
                return None
            
            # Get symbol info to get ts_code
            symbol_info = self.tushare_service.get_symbol_info(symbol)
            if 'error' in symbol_info or 'ts_code' not in symbol_info:
                self.logger.warning(f"No ts_code found for {symbol}")
                return None
            
            ts_code = symbol_info['ts_code']
            
            # Get 1 year of daily data
            daily_data = self.tushare_service.get_daily_data_by_days(ts_code, 252)
            if daily_data is None or daily_data.empty:
                self.logger.warning(f"No daily data available for {symbol}")
                return None
            
            # Create chart using ChartStyles
            chart_data = self._create_tushare_price_chart(symbol, daily_data, symbol_info, '1Y')
            return chart_data
            
        except Exception as e:
            self.logger.error(f"Error getting Tushare chart for {symbol}: {e}")
            return None

    async def _get_tushare_chart_for_period(self, symbol: str, period: str) -> Optional[bytes]:
        """Get Tushare chart for specific period"""
        try:
            if not self.tushare_service:
                return None
            
            # Get symbol info to get ts_code
            symbol_info = self.tushare_service.get_symbol_info(symbol)
            if 'error' in symbol_info or 'ts_code' not in symbol_info:
                self.logger.warning(f"No ts_code found for {symbol}")
                return None
            
            ts_code = symbol_info['ts_code']
            
            # Determine days based on period
            if period == '1Y':
                days = 252
            elif period == '5Y':
                days = 1260
            elif period == 'MAX':
                days = 2520  # 10 years max
            else:
                days = 252
            
            # Get daily data for the period
            daily_data = self.tushare_service.get_daily_data_by_days(ts_code, days)
            if daily_data is None or daily_data.empty:
                self.logger.warning(f"No daily data available for {symbol} period {period}")
                return None
            
            # Create chart using ChartStyles
            chart_data = self._create_tushare_price_chart(symbol, daily_data, symbol_info, period)
            return chart_data
            
        except Exception as e:
            self.logger.error(f"Error getting Tushare chart for {symbol} period {period}: {e}")
            return None

    def _create_tushare_price_chart(self, symbol: str, daily_data, symbol_info: Dict[str, Any], period: str = '1Y') -> Optional[bytes]:
        """Create price chart for Tushare asset using ChartStyles"""
        try:
            import io
            import pandas as pd
            from services.chart_styles import ChartStyles
            
            # Prepare data for chart
            daily_data = daily_data.sort_values('trade_date')
            daily_data['date'] = pd.to_datetime(daily_data['trade_date'], format='%Y%m%d')
            
            # Create price series
            price_series = daily_data.set_index('date')['close']
            
            # Get asset information
            asset_name = symbol_info.get('name', symbol)
            english_name = symbol_info.get('english_name', '')
            
            # Determine currency based on exchange
            currency = 'HKD' if symbol.endswith('.HK') else 'CNY'
            
            # Create chart using ChartStyles
            chart_styles = ChartStyles()
            fig, ax = chart_styles.create_price_chart(
                data=price_series,
                symbol=symbol,
                currency=currency,
                period=period,  # Use the actual period
                data_source='tushare'
            )
            
            # Set title with proper format: Тикер | Английское название | Валюта | Срок
            # Use English name if available, otherwise fall back to asset name
            display_name = english_name if english_name and english_name != asset_name else asset_name
            title = f"{symbol} | {display_name} | {currency} | {period}"
            ax.set_title(title, **chart_styles.title)
            
            # Убираем подписи осей
            ax.set_xlabel('')
            ax.set_ylabel('')
            
            # Convert to bytes
            buffer = io.BytesIO()
            fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            chart_bytes = buffer.getvalue()
            buffer.close()
            
            return chart_bytes
            
        except Exception as e:
            self.logger.error(f"Error creating Tushare price chart for {symbol}: {e}")
            return None

    def _get_data_for_period(self, asset, period: str):
        """Get filtered data for the specified period"""
        try:
            if not hasattr(asset, 'close_daily'):
                return None, None
            
            data = asset.close_daily
            
            if period == '1Y':
                # Last 252 trading days (approximately 1 year)
                filtered_data = data.tail(252)
            elif period == '5Y':
                # Last 1260 trading days (approximately 5 years)
                filtered_data = data.tail(1260)
            elif period == 'MAX':
                # All available data
                filtered_data = data
            else:
                # Default to 1Y
                filtered_data = data.tail(252)
            
            # Calculate returns for the filtered period
            returns = filtered_data.pct_change().dropna()
            
            return filtered_data, returns
            
        except Exception as e:
            self.logger.error(f"Error filtering data for period {period}: {e}")
            return None, None

    async def _get_asset_key_metrics(self, asset, symbol: str, period: str = '1Y') -> Dict[str, Any]:
        """Get key metrics for an asset for the specified period"""
        try:
            metrics = {}
            
            # Get filtered data for the period
            filtered_data, returns = self._get_data_for_period(asset, period)
            
            if filtered_data is None or len(filtered_data) == 0:
                self.logger.warning(f"No data available for {symbol} for period {period}")
                return {}
            
            # Get current price (always the latest price)
            try:
                if hasattr(asset, 'close_daily'):
                    current_price = asset.close_daily.iloc[-1]
                    metrics['current_price'] = current_price
                    
                    # Calculate price change for the selected period
                    if filtered_data is not None and len(filtered_data) > 1:
                        # For period-specific price change, use start and end of filtered data
                        start_price = filtered_data.iloc[0]
                        end_price = filtered_data.iloc[-1]
                        price_change = ((end_price - start_price) / start_price) * 100
                        metrics['price_change_pct'] = price_change
                    elif len(asset.close_daily) > 1:
                        # Fallback: use last two prices for daily change
                        prev_price = asset.close_daily.iloc[-2]
                        price_change = ((current_price - prev_price) / prev_price) * 100
                        metrics['price_change_pct'] = price_change
                else:
                    metrics['current_price'] = None
                    metrics['price_change_pct'] = None
            except Exception as e:
                self.logger.warning(f"Could not get current price for {symbol}: {e}")
                metrics['current_price'] = None
                metrics['price_change_pct'] = None
            
            # Calculate CAGR for the specific period
            try:
                if returns is not None and len(returns) > 0:
                    # Calculate total return for the period
                    total_return = (1 + returns).prod() - 1
                    
                    # Calculate years based on the period
                    if period == '1Y':
                        years = 1.0
                    elif period == '5Y':
                        years = 5.0
                    elif period == 'MAX':
                        # Calculate actual years from data
                        years = len(filtered_data) / 252.0  # Approximate trading days per year
                    else:
                        years = 1.0
                    
                    if years > 0:
                        cagr = (1 + total_return) ** (1 / years) - 1
                        metrics['cagr'] = cagr
                    else:
                        metrics['cagr'] = None
                else:
                    metrics['cagr'] = None
            except Exception as e:
                self.logger.warning(f"Could not calculate CAGR for {symbol} period {period}: {e}")
                metrics['cagr'] = None
            
            # Calculate volatility for the specific period
            try:
                if returns is not None and len(returns) > 0:
                    # Calculate annualized volatility for the period
                    volatility = returns.std() * (252 ** 0.5)  # Annualized for daily data
                    metrics['volatility'] = volatility
                else:
                    metrics['volatility'] = None
            except Exception as e:
                self.logger.warning(f"Could not calculate volatility for {symbol} period {period}: {e}")
                metrics['volatility'] = None
            
            # Get dividend yield (this is typically a current metric, not period-specific)
            try:
                if hasattr(asset, 'dividend_yield'):
                    dividend_yield = asset.dividend_yield
                    if hasattr(dividend_yield, 'iloc'):
                        metrics['dividend_yield'] = dividend_yield.iloc[-1]
                    else:
                        metrics['dividend_yield'] = dividend_yield
                else:
                    metrics['dividend_yield'] = None
            except Exception as e:
                self.logger.warning(f"Could not get dividend yield for {symbol}: {e}")
                metrics['dividend_yield'] = None
            
            # Add period information to metrics
            metrics['period'] = period
            metrics['data_points'] = len(filtered_data)
            
            return metrics
            
        except Exception as e:
            self.logger.error(f"Error getting key metrics for {symbol} period {period}: {e}")
            return {}

    def _format_asset_info_response(self, asset, symbol: str, key_metrics: Dict[str, Any]) -> str:
        """Format the asset info response according to the new structure"""
        try:
            # Block 1: Header - who is this?
            asset_name = getattr(asset, 'name', symbol)
            country = getattr(asset, 'country', 'N/A')
            asset_type = getattr(asset, 'asset_type', 'N/A')
            exchange = getattr(asset, 'exchange', 'N/A')
            isin = getattr(asset, 'isin', 'N/A')
            
            header = f"📊 {asset_name} ({symbol})\n"
            header += f"📍 {country} | {asset_type} | {exchange}"
            if isin and isin != 'N/A':
                header += f" | ISIN: {isin}"
            
            # Block 2: Key metrics showcase
            period = key_metrics.get('period', '1Y')
            period_text = {
                '1Y': '1 год',
                '5Y': '5 лет', 
                'MAX': 'MAX'
            }.get(period, '1 год')
            
            metrics_text = f"\n\nКлючевые показатели (за {period_text}):\n"
            
            # Current price
            if key_metrics.get('current_price') is not None:
                price = key_metrics['current_price']
                currency = getattr(asset, 'currency', 'USD')
                price_text = f"Цена: {price:.2f} {currency}"
                
                if key_metrics.get('price_change_pct') is not None:
                    change_pct = key_metrics['price_change_pct']
                    change_sign = "+" if change_pct >= 0 else ""
                    price_text += f" ({change_sign}{change_pct:.2f}%)"
                
                metrics_text += f"{price_text}\n"
            
            # CAGR
            if key_metrics.get('cagr') is not None:
                cagr = key_metrics['cagr']
                cagr_sign = "+" if cagr >= 0 else ""
                metrics_text += f"Доходность (CAGR): {cagr_sign}{cagr:.1%}\n"
            
            # Volatility
            if key_metrics.get('volatility') is not None:
                volatility = key_metrics['volatility']
                metrics_text += f"Волатильность: {volatility:.1%}\n"
            
            # Dividend yield
            if key_metrics.get('dividend_yield') is not None:
                dividend_yield = key_metrics['dividend_yield']
                metrics_text += f"Див. доходность (LTM): {dividend_yield:.2%}\n"
            
            return header + metrics_text
            
        except Exception as e:
            self.logger.error(f"Error formatting asset info response for {symbol}: {e}")
            return f"🔍 {symbol}\n\n❌ Ошибка при форматировании информации"

    def _create_info_interactive_keyboard(self, symbol: str) -> List[List[InlineKeyboardButton]]:
        """Create interactive keyboard for info command"""
        return self._create_info_interactive_keyboard_with_period(symbol, "1Y")

    def _create_info_interactive_keyboard_with_period(self, symbol: str, active_period: str) -> List[List[InlineKeyboardButton]]:
        """Create interactive keyboard for info command with active period highlighted"""
        # Create period buttons with active period highlighted
        period_buttons = []
        periods = [
            ("1Y", "1 год"),
            ("5Y", "5 лет"),
            ("MAX", "MAX")
        ]
        
        for period_code, period_text in periods:
            if period_code == active_period:
                button_text = f"✅ {period_text}"
            else:
                button_text = period_text
            period_buttons.append(
                InlineKeyboardButton(button_text, callback_data=f"info_period_{symbol}_{period_code}")
            )
        
        keyboard = [
            # Row 1: Period switching
            period_buttons,
            # Row 2: Actions
            [
                InlineKeyboardButton("💵 Дивиденды", callback_data=f"info_dividends_{symbol}"),
                InlineKeyboardButton("⚖️ Сравнить", callback_data=f"info_compare_{symbol}"),
                InlineKeyboardButton("💼 В портфель", callback_data=f"info_portfolio_{symbol}")
            ]
        ]
        return keyboard

    async def _get_daily_chart(self, symbol: str) -> Optional[bytes]:
        """Получить ежедневный график за последний год используя ChartStyles"""
        try:
            import io
            
            def create_daily_chart():
                self.logger.info(f"Creating daily chart for {symbol}")
                # Устанавливаем backend для headless режима
                import matplotlib
                matplotlib.use('Agg')
                
                asset = ok.Asset(symbol)
                self.logger.info(f"Asset created for {symbol}")
                
                # Получаем данные за последний год
                daily_data = asset.close_daily
                self.logger.info(f"Daily data shape: {daily_data.shape if hasattr(daily_data, 'shape') else 'No shape'}")
                
                # Берем последние 252 торговых дня (примерно год)
                filtered_data = daily_data.tail(252)
                self.logger.info(f"Filtered data shape: {filtered_data.shape if hasattr(filtered_data, 'shape') else 'No shape'}")
                
                # Получаем информацию об активе для заголовка
                asset_name = getattr(asset, 'name', symbol)
                currency = getattr(asset, 'currency', '')
                self.logger.info(f"Asset name: {asset_name}, currency: {currency}")
                
                # Используем ChartStyles для создания графика
                self.logger.info("Creating chart with ChartStyles")
                fig, ax = chart_styles.create_price_chart(
                    data=filtered_data,
                    symbol=symbol,
                    currency=currency,
                    period='1Y',
                    data_source='okama'
                )
                self.logger.info("Chart created successfully")
                
                # Обновляем заголовок с нужным форматом
                title = f"{symbol} | {asset_name} | {currency} | 1Y"
                ax.set_title(title, **chart_styles.title)
                
                # Убираем подписи осей
                ax.set_xlabel('')
                ax.set_ylabel('')
                
                # Сохраняем в bytes
                output = io.BytesIO()
                chart_styles.save_figure(fig, output)
                output.seek(0)
                
                # Очистка
                chart_styles.cleanup_figure(fig)
                
                result = output.getvalue()
                self.logger.info(f"Chart bytes length: {len(result)}")
                return result
            
            # Выполняем с таймаутом
            self.logger.info(f"Starting chart creation for {symbol}")
            chart_data = await asyncio.wait_for(
                asyncio.to_thread(create_daily_chart),
                timeout=30.0
            )
            
            self.logger.info(f"Chart creation completed for {symbol}")
            return chart_data
            
        except Exception as e:
            self.logger.error(f"Error getting daily chart for {symbol}: {e}")
            self.logger.error(f"Error type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None








    async def namespace_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /list command"""
        # Ensure no reply keyboard is shown
        await self._ensure_no_reply_keyboard(update, context)
        
        try:
            
            if not context.args:
                # Show available namespaces
                namespaces = ok.namespaces
                
                # Prepare data for list
                namespace_data = []
                
                # Create data with only code and description
                for namespace, description in namespaces.items():
                    namespace_data.append([namespace, description])
                
                # Add Chinese exchanges manually (not in ok.namespaces)
                chinese_exchanges = {
                    'SSE': 'Shanghai Stock Exchange',
                    'SZSE': 'Shenzhen Stock Exchange', 
                    'BSE': 'Beijing Stock Exchange',
                    'HKEX': 'Hong Kong Stock Exchange'
                }
                
                for exchange_code, exchange_name in chinese_exchanges.items():
                    namespace_data.append([exchange_code, exchange_name])
                
                # Sort by namespace
                namespace_data.sort(key=lambda x: x[0])
                response = "📚 База данных\n\n"
                
                # Create bulleted list format
                for row in namespace_data:
                    response += f"• {row[0]} - {row[1]}\n"
                response += "\n"
                
                response += "💡 Используйте кнопки ниже для выбора биржи\n\nИспользуйте команду /search для поиска активов"
                
                # Создаем reply keyboard для пространств имен
                reply_markup = self._create_namespace_reply_keyboard()
                
                await self._send_message_safe(update, response, reply_markup=reply_markup)
                
            else:
                # Show symbols in specific namespace
                namespace = self.clean_symbol(context.args[0]).upper()
                
                # Use the unified method that handles both okama and tushare
                await self._show_namespace_symbols(update, context, namespace, is_callback=False, page=0)
                    
        except ImportError:
            await self._send_message_safe(update, "*❌ Библиотека okama не установлена*")
        except Exception as e:
            self.logger.error(f"Error in namespace command: {e}")
            await self._send_message_safe(update, f"❌ Ошибка: {str(e)}")

    async def search_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /search command for searching assets by name or ISIN"""
        # Ensure no reply keyboard is shown
        await self._ensure_no_reply_keyboard(update, context)
        
        try:
            if not context.args:
                await self._send_message_safe(update, 
                    "🔍 **Поиск активов**\n\n"
                    "Используйте команду `/search <запрос>` для поиска тикеров по названию или ISIN.\n\n"
                    "**Примеры:**\n"
                    "• `/search Apple` - найти акции Apple\n"
                    "• `/search SBER` - найти Сбербанк\n"
                    "• `/search US0378331005` - найти по ISIN\n"
                    "• `/search золото` - найти золото\n"
                    "• `/search SP500` - найти индекс S&P 500\n\n"
                    "Поиск работает в базе данных okama и tushare на английском языке."
                )
                return
            
            # Clean query from special characters and validate length
            query = ' '.join(context.args)
            # Remove special characters except letters, numbers, spaces, and basic punctuation
            import re
            cleaned_query = re.sub(r'[^\w\s\-\.]', '', query)
            cleaned_query = cleaned_query.strip()
            
            if len(cleaned_query) < 3:
                await self._send_message_safe(update, "❌ Запрос должен содержать минимум 3 символа")
                return
            
            await self._send_message_safe(update, f"🔍 Ищу активы по запросу: `{cleaned_query}`...")
            
            # Search in okama
            okama_results = []
            tushare_results = []
            
            try:
                import okama as ok
                search_result = ok.search(cleaned_query)
                if not search_result.empty:
                    for _, row in search_result.head(30).iterrows():  # Increased limit to 30 results
                        symbol = row.get('symbol', '')
                        name = row.get('name', '')
                        if symbol and name:
                            okama_results.append({
                                'symbol': symbol,
                                'name': name
                            })
            except Exception as e:
                self.logger.warning(f"Okama search error: {e}")
            
            # Search in tushare for Chinese exchanges with English names
            try:
                if self.tushare_service:
                    # Search in all Chinese exchanges
                    for exchange in ['SSE', 'SZSE', 'BSE', 'HKEX']:
                        try:
                            exchange_results = self.tushare_service.search_symbols_english(cleaned_query, exchange)
                            for result in exchange_results[:10]:  # Increased limit to 10 per exchange
                                tushare_results.append({
                                    'symbol': result['symbol'],
                                    'name': result['name']
                                })
                        except Exception as e:
                            self.logger.warning(f"Tushare search error for {exchange}: {e}")
            except Exception as e:
                self.logger.warning(f"Tushare search error: {e}")
            
            # Combine and format results
            all_results = okama_results + tushare_results
            
            if not all_results:
                await self._send_message_safe(update, 
                    f"❌ По запросу `{cleaned_query}` ничего не найдено.\n\n"
                    "**Попробуйте:**\n"
                    "• Использовать английское название\n"
                    "• Использовать тикер (например, AAPL вместо Apple)\n"
                    "• Проверить правильность написания\n"
                    "• Использовать `/list` для просмотра доступных активов"
                )
                return
            
            # Format results in markdown table format similar to /list namespace
            response = f"🔍 **Результаты поиска по запросу:** `{cleaned_query}`\n\n"
            response += f"Найдено активов: **{len(all_results)}**\n\n"
            
            # Create table using tabulate or fallback to simple format
            if TABULATE_AVAILABLE and len(all_results) > 0:
                headers = ["Тикер", "Название"]
                table_data = []
                
                for result in all_results[:50]:  # Increased limit to 50 results for display
                    table_data.append([
                        f"`{result['symbol']}`",
                        result['name'][:60] + "..." if len(result['name']) > 60 else result['name']
                    ])
                
                # Use plain format for best Telegram display
                table = tabulate.tabulate(table_data, headers=headers, tablefmt="plain")
                response += f"```\n{table}\n```\n\n"
            else:
                # Fallback to simple text format
                response += "Тикер | Название\n"
                response += "--- | ---\n"
                for result in all_results[:50]:  # Increased limit to 50 results for display
                    name = result['name'][:60] + "..." if len(result['name']) > 60 else result['name']
                    response += f"`{result['symbol']}` | {name}\n"
                response += "\n"
            
            if len(all_results) > 50:
                response += f"*Показаны первые 50 из {len(all_results)} результатов*\n\n"
            
            response += "💡 Используйте нужный тикер в командах `/info`, `/compare` или `/portfolio`"
            
            await self._send_message_safe(update, response)
            
        except Exception as e:
            self.logger.error(f"Error in search command: {e}")
            await self._send_message_safe(update, f"❌ Ошибка поиска: {str(e)}")

    async def compare_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /compare command for comparing multiple assets"""
        # Ensure no reply keyboard is shown initially
        await self._ensure_no_reply_keyboard(update, context)
        
        try:
            if not context.args:
                # Get user's saved portfolios for help message
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                saved_portfolios = user_context.get('saved_portfolios', {})
                
                # Clear any existing compare context when starting fresh
                self._update_user_context(user_id, compare_first_symbol=None, waiting_for_compare=False)
                
                # Get analyzed tickers from context for examples
                analyzed_tickers = user_context.get('analyzed_tickers', [])
                
                # Get random examples for user using context tickers if available
                examples = self.examples_service.get_compare_examples(3, analyzed_tickers, saved_portfolios)
                examples_text = "\n".join([f"{example}" for example in examples])
                
                help_text = "⚖️ *Сравнение*\n\n"
                help_text += f"*Примеры:*\n{examples_text}\n\n"

                # Add saved portfolios information
                if saved_portfolios:
                    help_text += "💼 Ваши портфели:\n"
                    for portfolio_symbol, portfolio_info in saved_portfolios.items():
                        symbols = portfolio_info.get('symbols', [])
                        weights = portfolio_info.get('weights', [])
                        
                        # Create formatted portfolio string with symbols and weights
                        if symbols and weights and len(symbols) == len(weights):
                            portfolio_parts = []
                            for i, (symbol, weight) in enumerate(zip(symbols, weights)):
                                # Escape underscores in symbol names for markdown
                                escaped_symbol = symbol.replace('_', '\\_')
                                portfolio_parts.append(f"{escaped_symbol}:{weight:.1%}")
                            portfolio_str = ' '.join(portfolio_parts)
                        else:
                            # Escape underscores in symbol names for markdown
                            escaped_symbols = [symbol.replace('_', '\\_') for symbol in symbols]
                            portfolio_str = ', '.join(escaped_symbols)
                        
                        # Escape underscores in portfolio symbol for markdown
                        escaped_symbol = portfolio_symbol.replace('_', '\\_')
                        help_text += f"• `{escaped_symbol}` ({portfolio_str})\n\n"
                    
                
                # Add usage tips
                help_text += "💡 Можно сравнивать портфели и обычные активы (для этого нужно сначала создать портфель)\n"
                help_text += "💡 Первый актив в списке определяет базовую валюту для инфляции\n\n"
                help_text += "💬 Введите тикеры для сравнения через пробел:"
                
                await self._send_message_safe(update, help_text)
                
                # Set waiting flag for compare input
                self._update_user_context(user_id, waiting_for_compare=True)
                return

            # Parse currency and period parameters from command arguments
            # Check if they were already parsed by _handle_compare_input
            if hasattr(context, 'specified_currency') and hasattr(context, 'specified_period'):
                symbols = context.args
                specified_currency = context.specified_currency
                specified_period = context.specified_period
                self.logger.info(f"Using pre-parsed parameters from _handle_compare_input: currency={specified_currency}, period={specified_period}")
            else:
                symbols, specified_currency, specified_period = self._parse_currency_and_period(context.args)
            
            # Clean up symbols (remove empty strings and whitespace)
            symbols = [symbol for symbol in symbols if symbol.strip()]
            
            # Normalize namespace case (convert lowercase namespaces to uppercase)
            symbols = [self._normalize_symbol_namespace(symbol) for symbol in symbols]
            
            # Log the parsed parameters for debugging
            self.logger.info(f"Parsed symbols: {symbols}")
            self.logger.info(f"Parsed currency: {specified_currency}")
            self.logger.info(f"Parsed period: {specified_period}")
            
            # Get user_id for context operations
            user_id = update.effective_user.id
            
            # Check if we have a stored first symbol and only 1 symbol in current input
            user_context = self._get_user_context(user_id)
            stored_first_symbol = user_context.get('compare_first_symbol')
            
            if len(symbols) == 1 and stored_first_symbol is not None:
                # Combine stored symbol with current input
                symbols = [stored_first_symbol] + symbols
                # Clear the stored symbol
                self._update_user_context(user_id, compare_first_symbol=None)
                self.logger.info(f"Combined with stored symbol: {symbols}")
            
            if len(symbols) < 2:
                await self._send_message_safe(update, "❌ Необходимо указать минимум 2 символа для сравнения")
                return
            
            if len(symbols) > 5:
                await self._send_message_safe(update, "❌ Максимум 5 активов для сравнения. Пожалуйста, введите список для сравнения заново (не более 5 активов)")
                # Clear any stored symbols and reset waiting state
                self._update_user_context(user_id, compare_first_symbol=None, compare_base_symbol=None, waiting_for_compare=False)
                return

            # Check for duplicate symbols
            unique_symbols = list(dict.fromkeys(symbols))  # Preserve order while removing duplicates
            duplicates_found = len(symbols) != len(unique_symbols)
            
            if duplicates_found:
                if len(symbols) == 2:
                    # Exactly 2 symbols and they are the same
                    await self._send_message_safe(update, "❌ Можно сравнивать только разные активы. Указаны одинаковые символы.")
                    return
                else:
                    # 3+ symbols with duplicates - remove duplicates and continue
                    symbols = unique_symbols
                    duplicate_count = len(symbols) - len(unique_symbols)
                    await self._send_message_safe(update, f"⚠️ Обнаружены повторяющиеся активы. Исключены дубликаты из сравнения.")
                    self.logger.info(f"Removed {duplicate_count} duplicate symbols from comparison")

            # Process portfolio symbols and expand them
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Log saved portfolios for debugging
            self.logger.info(f"User {user_id} has {len(saved_portfolios)} saved portfolios: {list(saved_portfolios.keys())}")
            
            expanded_symbols = []
            portfolio_descriptions = []
            portfolio_contexts = []  # Store portfolio context for buttons
            
            for symbol in symbols:
                # Log symbol being processed for debugging
                self.logger.info(f"Processing symbol: '{symbol}'")
                
                # Check if this is a saved portfolio symbol (various formats)
                # First check exact match, then check case-insensitive match
                is_portfolio = symbol in saved_portfolios
                
                # Additional check: avoid treating asset symbols as portfolios
                # If the symbol looks like an asset symbol (contains .), 
                # and there's no explicit portfolio indicator, treat it as asset
                if is_portfolio and ('.' in symbol and 
                    not any(indicator in symbol.upper() for indicator in ['PORTFOLIO_', 'PF_', '.PF', '.pf'])):
                    # This looks like an asset symbol, not a portfolio
                    is_portfolio = False
                
                if not is_portfolio:
                    # Check case-insensitive match for portfolio symbols
                    # But only for symbols that look like portfolio names (not asset symbols)
                    for portfolio_key in saved_portfolios.keys():
                        if (symbol.lower() == portfolio_key.lower() or
                            symbol.upper() == portfolio_key.upper()):
                            # Additional check: avoid treating asset symbols as portfolios
                            if ('.' in symbol and 
                                not any(indicator in symbol.upper() for indicator in ['PORTFOLIO_', 'PF_', '.PF', '.pf'])):
                                # This looks like an asset symbol, not a portfolio
                                is_portfolio = False
                                break
                            else:
                                # Use the exact key from saved_portfolios
                                symbol = portfolio_key
                                is_portfolio = True
                                break
                
                self.logger.info(f"Symbol '{symbol}' is_portfolio: {is_portfolio}, in saved_portfolios: {symbol in saved_portfolios}")
                
                if is_portfolio:
                    # This is a saved portfolio, expand it
                    portfolio_info = saved_portfolios[symbol]
                    
                    # Ensure we have the correct portfolio symbol from context
                    portfolio_symbol = portfolio_info.get('portfolio_symbol', symbol)
                    portfolio_symbols = portfolio_info['symbols']
                    portfolio_weights = portfolio_info['weights']
                    portfolio_currency = portfolio_info['currency']
                    
                    # Validate portfolio data
                    if not portfolio_symbols or not portfolio_weights:
                        self.logger.error(f"Invalid portfolio data for {symbol}: missing symbols or weights")
                        await self._send_message_safe(update, f"❌ Ошибка: неполные данные портфеля {symbol}")
                        return
                    
                    if len(portfolio_symbols) != len(portfolio_weights):
                        self.logger.error(f"Portfolio {symbol} has mismatched symbols and weights")
                        await self._send_message_safe(update, f"❌ Ошибка: несоответствие символов и весов в портфеле {symbol}")
                        return
                    
                    # Create portfolio using okama
                    try:
                        portfolio = ok.Portfolio(portfolio_symbols, weights=portfolio_weights, ccy=portfolio_currency)
                        
                        # Add portfolio wealth index to expanded symbols
                        expanded_symbols.append(portfolio.wealth_index)
                        
                        # Use the original symbol for description to maintain consistency
                        portfolio_descriptions.append(f"{symbol} ({', '.join(portfolio_symbols)})")
                        
                        # Store portfolio context for buttons - use descriptive name for display
                        portfolio_contexts.append({
                            'symbol': f"{symbol} ({', '.join(portfolio_symbols)})",  # Descriptive name with asset composition
                            'portfolio_symbols': portfolio_symbols,
                            'portfolio_weights': portfolio_weights,
                            'portfolio_currency': portfolio_currency,
                            'portfolio_object': portfolio,
                            'original_portfolio_symbol': portfolio_symbol  # Store original portfolio symbol
                        })
                        
                        self.logger.info(f"Expanded portfolio {symbol} with {len(portfolio_symbols)} assets")
                        self.logger.info(f"Portfolio currency: {portfolio_currency}, weights: {portfolio_weights}")
                        self.logger.info(f"DEBUG: Added portfolio description: '{symbol} ({', '.join(portfolio_symbols)})'")
                        
                    except Exception as e:
                        self.logger.error(f"Error expanding portfolio {symbol}: {e}")
                        await self._send_message_safe(update, f"❌ Ошибка при обработке портфеля {symbol}: {str(e)}")
                        return
                else:
                    # Regular asset symbol
                    expanded_symbols.append(symbol)
                    portfolio_descriptions.append(symbol)
                    portfolio_contexts.append({
                        'symbol': symbol,
                        'portfolio_symbols': [symbol],
                        'portfolio_weights': [1.0],
                        'portfolio_currency': None,  # Will be determined later
                        'portfolio_object': None,
                        'original_portfolio_symbol': None  # Not a portfolio
                    })
            
            # Update symbols list with expanded portfolio descriptions
            symbols = portfolio_descriptions
            
            # Debug logging for symbols and expanded_symbols
            self.logger.info(f"DEBUG: portfolio_descriptions: {portfolio_descriptions}")
            self.logger.info(f"DEBUG: symbols after update: {symbols}")
            self.logger.info(f"DEBUG: expanded_symbols types: {[type(s) for s in expanded_symbols]}")
            self.logger.info(f"DEBUG: expanded_symbols values: {expanded_symbols}")
            
            # Additional debug: check for any problematic symbols
            for i, desc in enumerate(portfolio_descriptions):
                self.logger.info(f"DEBUG: portfolio_descriptions[{i}]: '{desc}'")
            for i, exp_sym in enumerate(expanded_symbols):
                self.logger.info(f"DEBUG: expanded_symbols[{i}]: '{exp_sym}' (type: {type(exp_sym)})")
            
            loading_message = await self._send_message_safe(update, f"⚖️ Сравниваю {', '.join(symbols)}...")

            # Create comparison using okama
            
            # Determine base currency - use specified currency if provided, otherwise auto-detect
            if specified_currency:
                currency = specified_currency
                currency_info = f"указана пользователем ({specified_currency})"
                self.logger.info(f"Using user-specified currency: {currency}")
            else:
                # Auto-detect currency from the first asset
                first_symbol = symbols[0]
                try:
                    # Extract the original symbol from the description (remove the asset list part)
                    original_first_symbol = first_symbol.split(' (')[0] if ' (' in first_symbol else first_symbol
                    
                    # Check if first symbol is a portfolio symbol
                    is_first_portfolio = original_first_symbol in saved_portfolios
                    
                    if not is_first_portfolio:
                        # Check case-insensitive match for portfolio symbols
                        for portfolio_key in saved_portfolios.keys():
                            if (original_first_symbol.lower() == portfolio_key.lower() or
                                original_first_symbol.upper() == portfolio_key.upper()):
                                original_first_symbol = portfolio_key
                                is_first_portfolio = True
                                break
                    
                    if is_first_portfolio:
                        # First symbol is a portfolio, use its currency
                        portfolio_info = saved_portfolios[original_first_symbol]
                        currency = portfolio_info.get('currency', 'USD')
                        currency_info = f"автоматически определена по портфелю ({original_first_symbol})"
                        self.logger.info(f"Using portfolio currency for {original_first_symbol}: {currency}")
                    else:
                        # Use our new currency detection function
                        currency, currency_info = self._get_currency_by_symbol(first_symbol)
                        
                        self.logger.info(f"Auto-detected currency for {first_symbol}: {currency}")
                    
                except Exception as e:
                    self.logger.warning(f"Could not auto-detect currency, using USD: {e}")
                    currency = "USD"
                    currency_info = "по умолчанию (USD)"
            
            try:
                # Check if we have portfolios in the comparison
                has_portfolios = any(isinstance(symbol, (pd.Series, pd.DataFrame)) for symbol in expanded_symbols)
                
                if has_portfolios:
                    # We have portfolios, need to create proper comparison using ok.AssetList
                    self.logger.info("Creating comparison with portfolios using ok.AssetList")
                    
                    # Prepare assets list for ok.AssetList
                    assets_for_comparison = []
                    
                    for i, symbol in enumerate(expanded_symbols):
                        if isinstance(symbol, (pd.Series, pd.DataFrame)):
                            # This is a portfolio wealth index - we need to create portfolio object
                            portfolio_context = portfolio_contexts[i] if i < len(portfolio_contexts) else None
                            
                            if portfolio_context:
                                try:
                                    # Use existing portfolio object if available, otherwise create new one
                                    if 'portfolio_object' in portfolio_context and portfolio_context['portfolio_object'] is not None:
                                        portfolio = portfolio_context['portfolio_object']
                                        self.logger.info(f"Using existing portfolio object for {portfolio_context['symbol']}")
                                    else:
                                        # Create portfolio object using okama
                                        portfolio = ok.Portfolio(
                                            portfolio_context['portfolio_symbols'], 
                                            weights=portfolio_context['portfolio_weights'], 
                                            ccy=portfolio_context['portfolio_currency']
                                        )
                                        self.logger.info(f"Created new portfolio object for {portfolio_context['symbol']}")
                                    
                                    assets_for_comparison.append(portfolio)
                                    self.logger.info(f"Added portfolio {portfolio_context['symbol']} to comparison")
                                except Exception as portfolio_error:
                                    self.logger.error(f"Error creating portfolio {portfolio_context['symbol']}: {portfolio_error}")
                                    # Delete loading message before showing error
                                    if loading_message:
                                        try:
                                            await loading_message.delete()
                                        except Exception as delete_error:
                                            self.logger.warning(f"Could not delete loading message: {delete_error}")
                                    await self._send_message_safe(update, f"❌ Ошибка при создании портфеля {portfolio_context['symbol']}: {str(portfolio_error)}")
                                    return
                            else:
                                self.logger.warning(f"No portfolio context found for index {i}, using generic portfolio")
                                # Create a generic portfolio if context is missing
                                try:
                                    # Extract symbols from the description
                                    desc = symbols[i]
                                    if ' (' in desc:
                                        portfolio_symbols = desc.split(' (')[1].rstrip(')').split(', ')
                                        portfolio_weights = [1.0/len(portfolio_symbols)] * len(portfolio_symbols)
                                        portfolio = ok.Portfolio(portfolio_symbols, weights=portfolio_weights, ccy=currency)
                                        assets_for_comparison.append(portfolio)
                                        self.logger.info(f"Added generic portfolio to comparison")
                                    else:
                                        self.logger.error(f"Could not extract portfolio symbols from description: {desc}")
                                        # Delete loading message before showing error
                                        if loading_message:
                                            try:
                                                await loading_message.delete()
                                            except Exception as delete_error:
                                                self.logger.warning(f"Could not delete loading message: {delete_error}")
                                        await self._send_message_safe(update, f"❌ Ошибка при обработке портфеля: {desc}")
                                        return
                                except Exception as e:
                                    self.logger.error(f"Error creating generic portfolio: {e}")
                                    # Delete loading message before showing error
                                    if loading_message:
                                        try:
                                            await loading_message.delete()
                                        except Exception as delete_error:
                                            self.logger.warning(f"Could not delete loading message: {delete_error}")
                                    await self._send_message_safe(update, f"❌ Ошибка при создании портфеля: {str(e)}")
                                    return
                        else:
                            # Regular asset symbol
                            assets_for_comparison.append(symbol)
                            self.logger.info(f"Added asset {symbol} to comparison")
                    
                    # Determine currency from first asset or portfolio
                    if assets_for_comparison:
                        first_asset = assets_for_comparison[0]
                        if hasattr(first_asset, 'currency'):
                            currency = first_asset.currency
                            currency_info = f"автоматически определена по первому активу/портфелю"
                        else:
                            # Try to determine from symbol
                            if '.' in str(first_asset):
                                namespace = str(first_asset).split('.')[1]
                                if namespace == 'MOEX':
                                    currency = "RUB"
                                    currency_info = f"автоматически определена по первому активу ({first_asset})"
                                elif namespace == 'US':
                                    currency = "USD"
                                    currency_info = f"автоматически определена по первому активу ({first_asset})"
                                elif namespace == 'LSE':
                                    currency = "GBP"
                                    currency_info = f"автоматически определена по первому активу ({first_asset})"
                                else:
                                    currency = "USD"
                                    currency_info = "по умолчанию (USD)"
                            else:
                                currency = "USD"
                                currency_info = "по умолчанию (USD)"
                    
                    # Check if we have Chinese symbols that need special handling
                    chinese_symbols = []
                    okama_symbols = []
                    
                    for symbol in symbols:
                        if self._is_chinese_symbol(symbol):
                            chinese_symbols.append(symbol)
                        else:
                            okama_symbols.append(symbol)
                    
                    # If we have Chinese symbols, use hybrid approach
                    if chinese_symbols:
                        self.logger.info(f"Found Chinese symbols: {chinese_symbols}")
                        
                        # Check if we have only Chinese symbols (pure Chinese comparison)
                        if len(chinese_symbols) == len(symbols):
                            # Delete loading message before hybrid comparison
                            if loading_message:
                                try:
                                    await loading_message.delete()
                                except Exception as e:
                                    self.logger.warning(f"Could not delete loading message: {e}")
                            # Pure Chinese comparison - use hybrid approach
                            await self._create_hybrid_chinese_comparison(update, context, chinese_symbols)
                            return
                        else:
                            # Delete loading message before showing mixed comparison message
                            if loading_message:
                                try:
                                    await loading_message.delete()
                                except Exception as e:
                                    self.logger.warning(f"Could not delete loading message: {e}")
                            # Mixed comparison - show message
                            await self._send_message_safe(update, 
                                f"⚠️ Смешанное сравнение (китайские + прочие символы) пока не поддерживается.\n"
                                f"Для сравнения только китайских символов используйте: /compare {' '.join(chinese_symbols)}\n\n"
                                f"Поддерживаемые символы для сравнения: {', '.join(okama_symbols) if okama_symbols else 'нет'}")
                            return
                    
                    # No Chinese symbols - use standard okama approach
                    self.logger.info(f"No Chinese symbols found, using standard okama comparison for: {symbols}")
                    
                    # Create comparison using ok.AssetList (proper way to compare portfolios with assets)
                    try:
                        self.logger.info(f"Creating AssetList with {len(assets_for_comparison)} assets/portfolios")
                        # Add inflation support for Chinese symbols
                        inflation_ticker = self._get_inflation_ticker_by_currency(currency)
                        
                        # Apply period filter if specified
                        self.logger.info(f"DEBUG: Portfolio comparison - specified_period = {specified_period}, currency = {currency}")
                        if specified_period:
                            years = int(specified_period[:-1])  # Extract number from '5Y'
                            from datetime import timedelta
                            end_date = datetime.now()
                            start_date = end_date - timedelta(days=years * 365)
                            self.logger.info(f"DEBUG: Creating AssetList with portfolios and period {specified_period}, start_date={start_date.strftime('%Y-%m-%d')}, end_date={end_date.strftime('%Y-%m-%d')}")
                            comparison = ok.AssetList(assets_for_comparison, ccy=currency, inflation=True, 
                                                    first_date=start_date.strftime('%Y-%m-%d'), 
                                                    last_date=end_date.strftime('%Y-%m-%d'))
                            self.logger.info(f"Successfully created AssetList comparison with period {specified_period} and inflation ({inflation_ticker}) using first_date/last_date parameters")
                        else:
                            self.logger.info(f"DEBUG: No period specified for portfolio comparison, creating AssetList without period filter")
                            comparison = ok.AssetList(assets_for_comparison, ccy=currency, inflation=True)
                            self.logger.info(f"Successfully created AssetList comparison with inflation ({inflation_ticker})")
                    except Exception as asset_list_error:
                        self.logger.error(f"Error creating AssetList: {asset_list_error}")
                        # Delete loading message before showing error
                        if loading_message:
                            try:
                                await loading_message.delete()
                            except Exception as delete_error:
                                self.logger.warning(f"Could not delete loading message: {delete_error}")
                        await self._send_message_safe(update, f"❌ Ошибка при создании сравнения: {str(asset_list_error)}")
                        return
                    
                else:
                    # Regular comparison without portfolios
                    self.logger.info("Creating regular comparison without portfolios")
                    
                    # Check if we have Chinese symbols that need special handling
                    chinese_symbols = []
                    okama_symbols = []
                    
                    for symbol in symbols:
                        if self._is_chinese_symbol(symbol):
                            chinese_symbols.append(symbol)
                        else:
                            okama_symbols.append(symbol)
                    
                    # If we have Chinese symbols, use hybrid approach
                    if chinese_symbols:
                        self.logger.info(f"Found Chinese symbols: {chinese_symbols}")
                        
                        # Check if we have only Chinese symbols (pure Chinese comparison)
                        if len(chinese_symbols) == len(symbols):
                            # Delete loading message before hybrid comparison
                            if loading_message:
                                try:
                                    await loading_message.delete()
                                except Exception as e:
                                    self.logger.warning(f"Could not delete loading message: {e}")
                            # Pure Chinese comparison - use hybrid approach
                            await self._create_hybrid_chinese_comparison(update, context, chinese_symbols)
                            return
                        else:
                            # Delete loading message before showing mixed comparison message
                            if loading_message:
                                try:
                                    await loading_message.delete()
                                except Exception as e:
                                    self.logger.warning(f"Could not delete loading message: {e}")
                            # Mixed comparison - show message
                            await self._send_message_safe(update, 
                                f"⚠️ Смешанное сравнение (китайских с прочими символами) будет реализовано в новых версиях.\n"
                                f"Китайские символы можно сравнить с китайскими, для этого используйте: /compare {' '.join(chinese_symbols)}\n\n"
                            )
                            return
                    
                    # No Chinese symbols - use standard okama approach
                    self.logger.info(f"No Chinese symbols found, using standard okama comparison for: {symbols}")
                    
                    # Add inflation support for non-Chinese symbols
                    inflation_ticker = self._get_inflation_ticker_by_currency(currency)
                    
                    # Apply period filter if specified
                    self.logger.info(f"DEBUG: specified_period = {specified_period}, currency = {currency}")
                    if specified_period:
                        years = int(specified_period[:-1])  # Extract number from '5Y'
                        from datetime import timedelta
                        end_date = datetime.now()
                        start_date = end_date - timedelta(days=years * 365)
                        self.logger.info(f"DEBUG: Creating AssetList with period {specified_period}, start_date={start_date.strftime('%Y-%m-%d')}, end_date={end_date.strftime('%Y-%m-%d')}")
                        comparison = ok.AssetList(symbols, ccy=currency, inflation=True,
                                                first_date=start_date.strftime('%Y-%m-%d'), 
                                                last_date=end_date.strftime('%Y-%m-%d'))
                        self.logger.info(f"Successfully created regular comparison with period {specified_period} and inflation ({inflation_ticker}) using first_date/last_date parameters")
                    else:
                        self.logger.info(f"DEBUG: No period specified, creating AssetList without period filter")
                        comparison = ok.AssetList(symbols, ccy=currency, inflation=True)
                        self.logger.info(f"Successfully created regular comparison with inflation ({inflation_ticker})")
                
                # Store context for buttons - use clean portfolio symbols for current_symbols
                clean_symbols = []
                display_symbols = []
                for i, symbol in enumerate(symbols):
                    if isinstance(expanded_symbols[i], (pd.Series, pd.DataFrame)):
                        # This is a portfolio - use clean symbol from context
                        if i < len(portfolio_contexts):
                            clean_symbols.append(portfolio_contexts[i]['symbol'].split(' (')[0])  # Extract clean symbol
                            display_symbols.append(portfolio_contexts[i]['symbol'])  # Keep descriptive name
                        else:
                            clean_symbols.append(symbol)
                            display_symbols.append(symbol)
                    else:
                        # This is a regular asset - use original symbol from portfolio_descriptions
                        clean_symbols.append(symbol)
                        display_symbols.append(symbol)
                
                user_context['current_symbols'] = clean_symbols
                user_context['display_symbols'] = display_symbols  # Store descriptive names for display
                user_context['current_currency'] = currency
                user_context['current_period'] = specified_period  # Store period for buttons
                user_context['last_analysis_type'] = 'comparison'
                user_context['portfolio_contexts'] = portfolio_contexts  # Store portfolio contexts
                user_context['expanded_symbols'] = expanded_symbols  # Store expanded symbols
                user_context['last_assets'] = clean_symbols  # Store symbols for Reply Keyboard buttons
                user_context['last_currency'] = currency  # Store currency for Reply Keyboard buttons
                user_context['last_period'] = specified_period  # Store period for Reply Keyboard buttons
                
                # Store describe table for AI analysis
                try:
                    describe_table = self._format_describe_table(comparison)
                    user_context['describe_table'] = describe_table
                except Exception as e:
                    self.logger.error(f"Error storing describe table: {e}")
                    user_context['describe_table'] = "📊 Данные для анализа недоступны"
                
                # Create comparison chart with updated title format
                chart_title = f"Сравнение {', '.join(symbols)} | {currency}"
                if specified_period:
                    chart_title += f" | {specified_period}"
                
                fig, ax = chart_styles.create_comparison_chart(
                    comparison.wealth_indexes, symbols, currency, title=chart_title
                )
                
                # Save chart to bytes with memory optimization
                img_buffer = io.BytesIO()
                chart_styles.save_figure(fig, img_buffer)
                img_buffer.seek(0)
                img_bytes = img_buffer.getvalue()
                
                # Clear matplotlib cache to free memory
                chart_styles.cleanup_figure(fig)
                
                # Chart analysis is now only available via buttons
                
                # Create summary metrics table for separate message
                summary_table = self._create_summary_metrics_table(
                    symbols, currency, expanded_symbols, portfolio_contexts, specified_period
                )
                
                # Create enhanced caption without summary table
                caption = self._create_enhanced_chart_caption(
                    symbols, currency, specified_period
                )
                
                # Describe table will be sent in separate message
                
                # Chart analysis is only available via buttons
                
                # No period selection keyboard needed
                reply_markup = None
                
                # Delete loading message before sending results
                if loading_message:
                    try:
                        await loading_message.delete()
                    except Exception as e:
                        self.logger.warning(f"Could not delete loading message: {e}")
                
                # Create compare reply keyboard
                compare_reply_keyboard = self._create_compare_reply_keyboard()
                
                # Send comparison chart with buttons and reply keyboard
                await self._send_photo_safe(
                    update=update,
                    photo_bytes=img_bytes,
                    caption=self._truncate_caption(caption),
                    reply_markup=compare_reply_keyboard,
                    context=context,
                    parse_mode='HTML'  # Try HTML instead of Markdown for better compatibility
                )
                
                # Update user context to track active keyboard
                user_id = update.effective_user.id
                self._update_user_context(user_id, active_reply_keyboard="compare")
                self.logger.info("Compare reply keyboard set with chart")
                
                # Table statistics now available via Metrics button
                
                # AI analysis is now only available via buttons
                
            except Exception as e:
                self.logger.error(f"Error creating comparison: {e}")
                # Delete loading message before showing error
                if loading_message:
                    try:
                        await loading_message.delete()
                    except Exception as delete_error:
                        self.logger.warning(f"Could not delete loading message: {delete_error}")
                await self._send_message_safe(update, f"❌ Ошибка при создании сравнения: {str(e)}")
                
        except Exception as e:
            self.logger.error(f"Error in compare command: {e}")
            # Delete loading message before showing error
            if loading_message:
                try:
                    await loading_message.delete()
                except Exception as delete_error:
                    self.logger.warning(f"Could not delete loading message: {delete_error}")
            await self._send_message_safe(update, f"❌ Ошибка в команде сравнения: {str(e)}")


    async def portfolio_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /portfolio command for creating portfolio with weights"""
        # Ensure no reply keyboard is shown initially
        await self._ensure_no_reply_keyboard(update, context)
        
        try:
            if not context.args:
                # Get user context for recently analyzed tickers
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                
                help_text = "💼 *Создание портфеля*\n\n"
                
                # Get analyzed tickers from context for examples
                analyzed_tickers = user_context.get('analyzed_tickers', [])
                
                # Get random examples for user using context tickers if available
                examples = self.examples_service.get_portfolio_examples(3, analyzed_tickers)
                examples_text = "\n".join([f"• {example}" for example in examples])
                
                help_text += "*Примеры команд:*\n"
                help_text += f"{examples_text}\n\n"
                help_text += "💡 Доли должны суммироваться в 1.0 (100%), максимум 10 активов в портфеле\n\n"

                help_text += "💬 Введите тикеры для создания портфеля:"
                
                await self._send_message_safe(update, help_text)
                
                # Set flag to wait for portfolio input
                self.logger.info(f"Setting waiting_for_portfolio=True for user {user_id}")
                self._update_user_context(user_id, waiting_for_portfolio=True)
                
                # Verify the flag was set
                updated_context = self._get_user_context(user_id)
                self.logger.info(f"Updated context waiting_for_portfolio: {updated_context.get('waiting_for_portfolio', False)}")
                return

            # Parse currency and period parameters from command arguments
            # For portfolio command, preserve the symbol:weight format
            symbols, specified_currency, specified_period = self._parse_currency_and_period(context.args, preserve_weights=True)
            
            # Extract symbols and weights from command arguments
            raw_args = ' '.join(symbols)
            portfolio_data = []
            tickers_only = []
            
            for arg in raw_args.split():
                if ':' in arg:
                    symbol_part, weight_part = arg.split(':', 1)
                    original_symbol = self.clean_symbol(symbol_part.strip())
                    # Преобразуем символ в верхний регистр
                    symbol = original_symbol.upper()
                    
                    try:
                        weight_str = weight_part.strip()
                        self.logger.info(f"DEBUG: Converting weight '{weight_str}' to float for symbol '{symbol}'")
                        
                        # Поддержка процентов
                        if weight_str.endswith('%'):
                            weight = float(weight_str[:-1]) / 100.0
                        else:
                            weight = float(weight_str)
                            
                    except Exception as e:
                        self.logger.error(f"Error converting weight '{weight_part.strip()}' to float: {e}")
                        await self._send_message_safe(update, f"❌ Некорректная доля для {symbol}: '{weight_part.strip()}'. Доля должна быть числом от 0 до 1 или процентом (например, 50%)")
                        return
                    
                    if weight <= 0 or weight > 1:
                        await self._send_message_safe(update, f"❌ Некорректная доля для {symbol}: {weight}. Доля должна быть от 0 до 1")
                        return
                    
                    portfolio_data.append((symbol, weight))
                    
                else:
                    # Это тикер без веса
                    original_symbol = self.clean_symbol(arg.strip())
                    symbol = original_symbol.upper()
                    tickers_only.append(symbol)
            
            # Если есть только тикеры без весов
            if tickers_only and not portfolio_data:
                # Если только один тикер, запрашиваем веса (вместо автоматического создания с весом 100%)
                if len(tickers_only) == 1:
                    single_ticker = tickers_only[0]
                    
                    # Проверяем на китайские и гонконгские символы
                    if self._is_chinese_or_hongkong_symbol(single_ticker):
                        await self._send_message_safe(update, 
                            "🚧 **Поддержка китайских и гонконгских символов в разработке**\n\n"
                            "К сожалению, создание портфелей с китайскими и гонконгскими активами "
                            "пока не поддерживается. Эта функциональность находится в разработке.\n\n"
                            "💡 Попробуйте использовать активы с других бирж:\n"
                            "• `SPY.US` - американские ETF\n"
                            "• `SBER.MOEX` - российские акции\n"
                            "• `VTI.US` - глобальные ETF"
                        )
                        return
                    
                    # Запрашиваем веса для одного актива
                    await self._request_portfolio_weights(update, tickers_only, specified_currency, specified_period)
                    return
                else:
                    # Несколько тикеров без весов - запрашиваем веса
                    await self._request_portfolio_weights(update, tickers_only, specified_currency, specified_period)
                    return
            
            # Если есть смешанный ввод (тикеры с весами и без), это ошибка
            if tickers_only and portfolio_data:
                await self._send_message_safe(update, 
                    "❌ Некорректный формат ввода.\n\n"
                    "Укажите веса для всех активов или только тикеры для равномерного распределения.\n\n"
                    "Примеры:\n"
                    "• `SPY.US QQQ.US BND.US` - только тикеры\n"
                    "• `SPY.US:0.5 QQQ.US:0.3 BND.US:0.2` - тикеры с весами"
                )
                return
            
            if not portfolio_data:
                await self._send_message_safe(update, "❌ Не указаны активы для портфеля")
                return
            
            # Check if weights sum to approximately 1.0
            total_weight = sum(weight for _, weight in portfolio_data)
            if abs(total_weight - 1.0) > 0.01:
                # Предлагаем исправление, если сумма близка к 1
                if abs(total_weight - 1.0) <= 0.1:
                    corrected_weights = []
                    for symbol, weight in portfolio_data:
                        corrected_weight = weight / total_weight
                        corrected_weights.append((symbol, corrected_weight))
                    
                    await self._send_message_safe(update, 
                        f"⚠️ Сумма долей ({total_weight:.3f}) не равна 1.0\n\n"
                        f"Исправленные доли:\n"
                        f"{chr(10).join([f'• {symbol}: {weight:.3f}' for symbol, weight in corrected_weights])}\n\n"
                        f"Попробуйте команду:\n"
                        f"`/portfolio {' '.join([f'{symbol}:{weight:.3f}' for symbol, weight in corrected_weights])}`"
                    )
                else:
                    await self._send_message_safe(update, 
                        f"❌ Сумма долей должна быть равна 1.0, текущая сумма: {total_weight:.3f}\n\n"
                        f"Пример правильной команды:\n"
                        f"`/portfolio LQDT.MOEX:0.78 OBLG.MOEX:0.16 GOLD.MOEX:0.06`"
                    )
                return
            
            if len(portfolio_data) > 10:
                await self._send_message_safe(update, "❌ Максимум 10 активов в портфеле")
                return
            
            symbols = [symbol for symbol, _ in portfolio_data]
            weights = [weight for _, weight in portfolio_data]
            
            # Проверяем на китайские и гонконгские символы
            chinese_hk_symbols = [symbol for symbol in symbols if self._is_chinese_or_hongkong_symbol(symbol)]
            if chinese_hk_symbols:
                await self._send_message_safe(update, 
                    "🚧 **Портфельный анализ для китайских и гонконгских активов находится в разработке**\n\n"
                    f"Обнаружены неподдерживаемые активы: {', '.join(chinese_hk_symbols)}\n\n"
                    "К сожалению, создание портфелей с китайскими и гонконгскими активами "
                    "пока не поддерживается. Эта функциональность находится в разработке.\n\n"
                    "💡 Попробуйте использовать активы с других бирж:\n"
                    "• `SPY.US` - американские ETF\n"
                    "• `SBER.MOEX` - российские акции\n"
                    "• `VTI.US` - глобальные ETF\n\n"
                    "🔄 Для создания нового портфеля используйте команду `/portfolio`"
                )
                return
            
            await self._send_ephemeral_message(update, context, f"Создаю портфель: {', '.join(symbols)}...", delete_after=3)
            
            # Create portfolio using okama
            
            self.logger.info(f"DEBUG: About to create portfolio with symbols: {symbols}, weights: {weights}")
            self.logger.info(f"DEBUG: Symbols types: {[type(s) for s in symbols]}")
            self.logger.info(f"DEBUG: Weights types: {[type(w) for w in weights]}")
            
            # Determine base currency - use specified currency if provided, otherwise auto-detect
            if specified_currency:
                currency = specified_currency
                currency_info = f"указана пользователем ({specified_currency})"
                self.logger.info(f"Using user-specified currency for portfolio: {currency}")
            else:
                # Auto-detect currency from the first asset
                first_symbol = symbols[0]
                try:
                    if '.' in first_symbol:
                        namespace = first_symbol.split('.')[1]
                        if namespace == 'MOEX':
                            currency = "RUB"
                            currency_info = f"автоматически определена по первому активу ({first_symbol})"
                        elif namespace == 'US':
                            currency = "USD"
                            currency_info = f"автоматически определена по первому активу ({first_symbol})"
                        elif namespace == 'LSE':
                            currency = "GBP"
                            currency_info = f"автоматически определена по первому активу ({first_symbol})"
                        elif namespace == 'FX':
                            currency = "USD"
                            currency_info = f"автоматически определена по первому активу ({first_symbol})"
                        elif namespace == 'COMM':
                            currency = "USD"
                            currency_info = f"автоматически определена по первому активу ({first_symbol})"
                        elif namespace == 'INDX':
                            currency = "USD"
                            currency_info = f"автоматически определена по первому активу ({first_symbol})"
                        else:
                            currency = "USD"
                            currency_info = "по умолчанию (USD)"
                    else:
                        currency = "USD"
                        currency_info = "по умолчанию (USD)"
                    
                    self.logger.info(f"Auto-detected currency for portfolio {first_symbol}: {currency}")
                    
                except Exception as e:
                    self.logger.warning(f"Could not auto-detect currency, using USD: {e}")
                    currency = "USD"
                    currency_info = "по умолчанию (USD)"
            
            try:
                # Final validation of weights before creating portfolio
                final_total = sum(weights)
                if abs(final_total - 1.0) > 0.001:
                    await self._send_message_safe(update, 
                        f"❌ Ошибка валидации весов: сумма {final_total:.6f} не равна 1.0\n"
                        f"Веса: {', '.join([f'{w:.6f}' for w in weights])}"
                    )
                    return
                
                # Create Portfolio with detected currency and period
                try:
                    self.logger.info(f"DEBUG: About to create ok.Portfolio with symbols={symbols}, ccy={currency}, weights={weights}")
                    self.logger.info(f"DEBUG: Symbols types: {[type(s) for s in symbols]}")
                    self.logger.info(f"DEBUG: Weights types: {[type(w) for w in weights]}")
                    
                    # Apply period filter if specified
                    if specified_period:
                        years = int(specified_period[:-1])  # Extract number from '5Y'
                        from datetime import timedelta
                        end_date = datetime.now()
                        start_date = end_date - timedelta(days=years * 365)
                        portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency,
                                               first_date=start_date.strftime('%Y-%m-%d'), 
                                               last_date=end_date.strftime('%Y-%m-%d'))
                        self.logger.info(f"DEBUG: Successfully created portfolio with period {specified_period}")
                    else:
                        portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency)
                        self.logger.info(f"DEBUG: Successfully created portfolio")
                except Exception as e:
                    self.logger.error(f"DEBUG: Error creating portfolio: {e}")
                    self.logger.error(f"DEBUG: Error type: {type(e)}")
                    raise e
                
                self.logger.info(f"Created Portfolio with weights: {weights}, total: {sum(weights):.6f}")
                
                # Create portfolio information text
                portfolio_text = f"💼 **Портфель создан успешно!**\n\n"
                
                # Add basic metrics to portfolio text
                try:
                    metrics_text = self._get_portfolio_basic_metrics(portfolio, symbols, weights, currency)
                    portfolio_text += metrics_text
                except Exception as e:
                    self.logger.warning(f"Could not add metrics to portfolio text: {e}")
                
                # Portfolio information is already set above as raw object
                
                # Generate portfolio symbol using PF namespace and okama's assigned symbol
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                
                # Count existing portfolios for this user
                portfolio_count = user_context.get('portfolio_count', 0) + 1
                
                # Use PF namespace with custom symbol (okama's symbol is composition string, not suitable for bot)
                portfolio_symbol = f"PF{portfolio_count}"
                
                # Create compact portfolio data string for callback (only symbols to avoid Button_data_invalid)
                portfolio_data_str = ','.join(symbols)
                
                # Add portfolio symbol display
                portfolio_text += f"\n\n💼 Сравнить портфель с другими активами: `/compare {portfolio_symbol}`\n"
                
                # No Inline Keyboard needed - only Reply Keyboard will be used
                reply_markup = None
                
                # Send ephemeral message about creating chart
                await self._send_ephemeral_message(update, context, "📈 Создаю график накопленной доходности...", delete_after=3)
                
                # Create and send wealth chart with portfolio info in caption and buttons
                await self._create_portfolio_wealth_chart_with_info(update, context, portfolio, symbols, currency, weights, portfolio_symbol, portfolio_text, reply_markup)
                
                # Store portfolio data in context
                user_id = update.effective_user.id
                self.logger.info(f"Storing portfolio data in context for user {user_id}")
                self.logger.info(f"Symbols: {symbols}")
                self.logger.info(f"Currency: {currency}")
                self.logger.info(f"Weights: {weights}")
                self.logger.info(f"Portfolio symbol: {portfolio_symbol}")
                
                # Store portfolio information in user context
                self._update_user_context(
                    user_id, 
                    last_assets=symbols,
                    last_analysis_type='portfolio',
                    last_period=specified_period or 'MAX',
                    current_symbols=symbols,
                    current_currency=currency,
                    current_currency_info=currency_info,
                    portfolio_weights=weights,
                    portfolio_count=portfolio_count,
                    current_period=specified_period
                )
                
                # Verify what was saved
                saved_context = self._get_user_context(user_id)
                self.logger.info(f"Saved context keys: {list(saved_context.keys())}")
                self.logger.info(f"Saved current_symbols: {saved_context.get('current_symbols')}")
                self.logger.info(f"Saved current_currency: {saved_context.get('current_currency')}")
                self.logger.info(f"Saved portfolio_weights: {saved_context.get('portfolio_weights')}")
                
                # Get current saved portfolios and check for existing portfolio
                saved_portfolios = user_context.get('saved_portfolios', {})
                
                # Check if portfolio with same assets and proportions already exists
                existing_portfolio_symbol = self._check_existing_portfolio(symbols, weights, saved_portfolios)
                
                if existing_portfolio_symbol:
                    # Use existing portfolio symbol and update the message
                    portfolio_symbol = existing_portfolio_symbol
                    portfolio_text += f"\n\n💼 Имя портфеля: `{portfolio_symbol}`\n"
                    portfolio_text += f"✅ Портфель с такими же активами и пропорциями уже существует\n"
                    portfolio_text += f"💼 Используется ранее сохраненный портфель"
                    
                    # Update portfolio count without incrementing
                    portfolio_count = user_context.get('portfolio_count', 0)
                else:
                    # Increment portfolio count for new portfolio
                    portfolio_count = user_context.get('portfolio_count', 0) + 1
                    portfolio_text += f"\n\n💼 Символ портфеля: `{portfolio_symbol}` (namespace PF)\n"
                    portfolio_text += f"💼 Портфель сохранен в контексте для использования в ]/compare"
                
                # Get additional portfolio attributes for comprehensive storage
                portfolio_attributes = {}
                try:
                    # Generate portfolio name
                    portfolio_name = self._generate_portfolio_name(symbols, weights)
                    
                    # Basic portfolio info
                    portfolio_attributes.update({
                        'symbols': symbols,
                        'weights': weights,
                        'currency': currency,
                        'created_at': datetime.now().isoformat(),
                        'description': f"Портфель: {', '.join(symbols)}",
                        'portfolio_symbol': portfolio_symbol,  # Ensure symbol is preserved
                        'portfolio_name': portfolio_name,
                        'total_weight': sum(weights),
                        'asset_count': len(symbols),
                        'period': specified_period
                    })
                    
                    # Portfolio performance metrics
                    if hasattr(portfolio, 'mean_return_annual'):
                        try:
                            value = portfolio.mean_return_annual
                            # Handle Period objects and other non-numeric types
                            if hasattr(value, 'to_timestamp'):
                                try:
                                    value = value.to_timestamp()
                                except Exception:
                                    # If to_timestamp fails, try to convert to string first
                                    if hasattr(value, 'strftime'):
                                        value = str(value)
                                    else:
                                        value = str(value)
                            
                            # Handle Timestamp objects and other datetime-like objects
                            if hasattr(value, 'timestamp'):
                                try:
                                    value = value.timestamp()
                                except Exception:
                                    value = str(value)
                            elif not isinstance(value, (int, float)):
                                # Try to convert non-numeric values
                                value = str(value)
                                import re
                                match = re.search(r'[\d.-]+', value)
                                if match:
                                    value = float(match.group())
                                else:
                                    value = 0.0
                            portfolio_attributes['mean_return_annual'] = float(value)
                        except Exception as e:
                            self.logger.warning(f"Could not convert mean_return_annual to float: {e}")
                            portfolio_attributes['mean_return_annual'] = None
                    
                    if hasattr(portfolio, 'volatility_annual'):
                        try:
                            value = portfolio.volatility_annual
                            # Handle Period objects and other non-numeric types
                            if hasattr(value, 'to_timestamp'):
                                try:
                                    value = value.to_timestamp()
                                except Exception:
                                    # If to_timestamp fails, try to convert to string first
                                    if hasattr(value, 'strftime'):
                                        value = str(value)
                                    else:
                                        value = str(value)
                            
                            # Handle Timestamp objects and other datetime-like objects
                            if hasattr(value, 'timestamp'):
                                try:
                                    value = value.timestamp()
                                except Exception:
                                    value = str(value)
                            elif not isinstance(value, (int, float)):
                                # Try to convert non-numeric values
                                value = str(value)
                                import re
                                match = re.search(r'[\d.-]+', value)
                                if match:
                                    value = float(match.group())
                                else:
                                    value = 0.0
                            portfolio_attributes['volatility_annual'] = float(value)
                        except Exception as e:
                            self.logger.warning(f"Could not convert volatility_annual to float: {e}")
                            portfolio_attributes['volatility_annual'] = None
                    
                    if hasattr(portfolio, 'sharpe_ratio'):
                        try:
                            value = portfolio.sharpe_ratio
                            # Handle Period objects and other non-numeric types
                            if hasattr(value, 'to_timestamp'):
                                try:
                                    value = value.to_timestamp()
                                except Exception:
                                    # If to_timestamp fails, try to convert to string first
                                    if hasattr(value, 'strftime'):
                                        value = str(value)
                                    else:
                                        value = str(value)
                            
                            # Handle Timestamp objects and other datetime-like objects
                            if hasattr(value, 'timestamp'):
                                try:
                                    value = value.timestamp()
                                except Exception:
                                    value = str(value)
                            elif not isinstance(value, (int, float)):
                                # Try to convert non-numeric values
                                value = str(value)
                                import re
                                match = re.search(r'[\d.-]+', value)
                                if match:
                                    value = float(match.group())
                                else:
                                    value = 0.0
                            portfolio_attributes['sharpe_ratio'] = float(value)
                        except Exception as e:
                            self.logger.warning(f"Could not convert sharpe_ratio to float: {e}")
                            portfolio_attributes['sharpe_ratio'] = None
                    
                    # Portfolio dates
                    if hasattr(portfolio, 'first_date'):
                        portfolio_attributes['first_date'] = str(portfolio.first_date)
                    if hasattr(portfolio, 'last_date'):
                        portfolio_attributes['last_date'] = str(portfolio.last_date)
                    if hasattr(portfolio, 'period_length'):
                        portfolio_attributes['period_length'] = str(portfolio.period_length)
                    
                    # Final portfolio value
                    try:
                        final_value = portfolio.wealth_index.iloc[-1]
                        if hasattr(final_value, '__iter__') and not isinstance(final_value, str):
                            if hasattr(final_value, 'iloc'):
                                final_value = final_value.iloc[0]
                            elif hasattr(final_value, '__getitem__'):
                                final_value = final_value[0]
                            else:
                                final_value = list(final_value)[0]
                        
                        # Handle Period objects and other non-numeric types
                        if hasattr(final_value, 'to_timestamp'):
                            try:
                                final_value = final_value.to_timestamp()
                            except Exception:
                                # If to_timestamp fails, try to convert to string first
                                if hasattr(final_value, 'strftime'):
                                    final_value = str(final_value)
                                else:
                                    final_value = str(final_value)
                        
                        # Handle Timestamp objects and other datetime-like objects
                        if hasattr(final_value, 'timestamp'):
                            try:
                                final_value = final_value.timestamp()
                            except Exception:
                                final_value = str(final_value)
                        elif not isinstance(final_value, (int, float)):
                            # Try to convert non-numeric values
                            final_value_str = str(final_value)
                            import re
                            match = re.search(r'[\d.-]+', final_value_str)
                            if match:
                                final_value = float(match.group())
                            else:
                                final_value = 0.0
                        
                        portfolio_attributes['final_value'] = float(final_value)
                    except Exception as e:
                        self.logger.warning(f"Could not get final value for storage: {e}")
                        portfolio_attributes['final_value'] = None
                    
                    # Store as JSON string for better compatibility
                    portfolio_json = json.dumps(portfolio_attributes, ensure_ascii=False, default=str)
                    portfolio_attributes['json_data'] = portfolio_json
                    
                except Exception as e:
                    self.logger.warning(f"Could not get all portfolio attributes: {e}")
                    # Fallback to basic storage
                    portfolio_name = self._generate_portfolio_name(symbols, weights)
                    portfolio_attributes = {
                        'symbols': symbols,
                        'weights': weights,
                        'currency': currency,
                        'created_at': datetime.now().isoformat(),
                        'description': f"Портфель: {', '.join(symbols)}",
                        'portfolio_symbol': portfolio_symbol,
                        'portfolio_name': portfolio_name,
                        'total_weight': sum(weights),
                        'asset_count': len(symbols),
                        'period': specified_period,
                        'json_data': json.dumps({
                            'symbols': symbols,
                            'weights': weights,
                            'currency': currency,
                            'portfolio_symbol': portfolio_symbol
                        }, ensure_ascii=False)
                    }
                
                # Add the new portfolio to saved portfolios (always save, even if similar exists)
                saved_portfolios[portfolio_symbol] = portfolio_attributes
                
                # Update saved portfolios in context (single update)
                self.logger.info(f"Updating user context with portfolio: {portfolio_symbol}")
                self.logger.info(f"Portfolio attributes: {portfolio_attributes}")
                self.logger.info(f"Current saved portfolios: {saved_portfolios}")
                
                self._update_user_context(
                    user_id,
                    saved_portfolios=saved_portfolios,
                    portfolio_count=portfolio_count
                )
                
                # Verify context was saved
                saved_context = self._get_user_context(user_id)
                self.logger.info(f"Saved context keys: {list(saved_context.keys())}")
                self.logger.info(f"Saved current_symbols: {saved_context.get('current_symbols')}")
                self.logger.info(f"Saved last_assets: {saved_context.get('last_assets')}")
                self.logger.info(f"Saved portfolio_weights: {saved_context.get('portfolio_weights')}")
                
            except Exception as e:
                self.logger.error(f"Error creating portfolio: {e}")
                await self._send_message_safe(update, 
                    f"❌ Ошибка при создании портфеля: {str(e)}\n\n"
                    "💡 Возможные причины:\n"
                    "• Один из символов недоступен\n"
                    "• Проблемы с данными\n"
                    "• Неверный формат символа\n\n"
                    "Проверьте:\n"
                    "• Правильность написания символов\n"
                    "• Доступность данных для указанных активов"
                )
                
        except Exception as e:
            self.logger.error(f"Error in portfolio command: {e}")
            # Clear user context to prevent fallback to compare command
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                waiting_for_portfolio=False,
                waiting_for_portfolio_weights=False,
                waiting_for_compare=False,
                portfolio_tickers=None,
                portfolio_base_symbols=None
            )
            await self._send_message_safe(update, f"❌ Ошибка при выполнении команды портфеля: {str(e)}")


    async def _request_portfolio_weights(self, update: Update, tickers: list, currency: str = None, period: str = None):
        """Request portfolio weights from user when only tickers are provided"""
        try:
            user_id = update.effective_user.id
            
            # Set context to wait for portfolio weights
            self._update_user_context(user_id, 
                waiting_for_portfolio_weights=True,
                portfolio_tickers=tickers,
                portfolio_currency=currency,
                portfolio_period=period
            )
            
            # Create message with tickers and request for weights
            tickers_text = ' '.join(tickers)
            message = f"💼 **Укажите доли активов**\n\n"
            message += f"Активы: `{tickers_text}`\n\n"
            message += "**Введите веса в том же порядке:**\n"
            
            # Show suggested equal weights
            equal_weight = 1.0 / len(tickers)
            suggested_weights = []
            for i, ticker in enumerate(tickers):
                suggested_weights.append(f"{ticker}:{equal_weight:.3f}")
            
            message += f"• Равномерное распределение: `{' '.join(suggested_weights)}`\n"
            message += f"• Или укажите свои доли: `{' '.join([f'{ticker}:0.XX' for ticker in tickers])}`\n\n"
            message += "**Форматы весов:**\n"
            message += "• Десятичные: `0.5` (50%)\n"
            message += "• Проценты: `50%`\n\n"
            message += "⚠️ Сумма долей должна быть равна 1.0 (100%)"
            
            await self._send_message_safe(update, message)
            
        except Exception as e:
            self.logger.error(f"Error requesting portfolio weights: {e}")
            await self._send_message_safe(update, "❌ Ошибка при запросе весов портфеля")

    async def _handle_portfolio_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle portfolio input from user message"""
        try:
            user_id = update.effective_user.id
            
            # Clear waiting flag
            self._update_user_context(user_id, waiting_for_portfolio=False)
            
            # Parse currency and period parameters from input text
            text_args = text.split()
            
            # For portfolio command, we need to preserve the full symbol:weight format
            # So we'll parse currency and period manually, keeping the original arguments
            valid_currencies = {'USD', 'RUB', 'EUR', 'GBP', 'CNY', 'HKD', 'JPY'}
            import re
            period_pattern = re.compile(r'^(\d+)Y$', re.IGNORECASE)
            
            portfolio_args = []
            specified_currency = None
            specified_period = None
            
            for arg in text_args:
                arg_upper = arg.upper()
                
                # Check if it's a currency code
                if arg_upper in valid_currencies:
                    if specified_currency is None:
                        specified_currency = arg_upper
                    continue
                
                # Check if it's a period (e.g., '5Y', '10Y')
                period_match = period_pattern.match(arg)
                if period_match:
                    if specified_period is None:
                        specified_period = arg_upper
                    continue
                
                # If it's neither currency nor period, it's a portfolio argument
                portfolio_args.append(arg)
            
            # Extract symbols and weights from portfolio arguments
            portfolio_data = []
            tickers_only = []
            
            for arg in portfolio_args:
                if ':' in arg:
                    symbol_part, weight_part = arg.split(':', 1)
                    original_symbol = self.clean_symbol(symbol_part.strip())
                    # Преобразуем символ в верхний регистр
                    symbol = original_symbol.upper()
                    
                    try:
                        weight_str = weight_part.strip()
                        self.logger.info(f"DEBUG: Converting weight '{weight_str}' to float for symbol '{symbol}'")
                        
                        # Поддержка процентов
                        if weight_str.endswith('%'):
                            weight = float(weight_str[:-1]) / 100.0
                        else:
                            weight = float(weight_str)
                            
                    except Exception as e:
                        self.logger.error(f"Error converting weight '{weight_part.strip()}' to float: {e}")
                        await self._send_message_safe(update, f"❌ Некорректная доля для {symbol}: '{weight_part.strip()}'. Доля должна быть числом от 0 до 1 или процентом (например, 50%)")
                        return
                    
                    if weight <= 0 or weight > 1:
                        await self._send_message_safe(update, f"❌ Некорректная доля для {symbol}: {weight}. Доля должна быть от 0 до 1")
                        return
                    
                    portfolio_data.append((symbol, weight))
                    
                else:
                    # Это тикер без веса
                    original_symbol = self.clean_symbol(arg.strip())
                    symbol = original_symbol.upper()
                    tickers_only.append(symbol)
            
            # Если есть только тикеры без весов
            if tickers_only and not portfolio_data:
                # Если только один тикер, запрашиваем веса (вместо автоматического создания с весом 100%)
                if len(tickers_only) == 1:
                    single_ticker = tickers_only[0]
                    
                    # Проверяем на китайские и гонконгские символы
                    if self._is_chinese_or_hongkong_symbol(single_ticker):
                        await self._send_message_safe(update, 
                            "🚧 **Поддержка китайских и гонконгских символов в разработке**\n\n"
                            "К сожалению, создание портфелей с китайскими и гонконгскими активами "
                            "пока не поддерживается. Эта функциональность находится в разработке.\n\n"
                            "💡 Попробуйте использовать активы с других бирж:\n"
                            "• `SPY.US` - американские ETF\n"
                            "• `SBER.MOEX` - российские акции\n"
                            "• `VTI.US` - глобальные ETF"
                        )
                        return
                    
                    # Запрашиваем веса для одного актива
                    await self._request_portfolio_weights(update, tickers_only, specified_currency, specified_period)
                    return
                else:
                    # Несколько тикеров без весов - запрашиваем веса
                    await self._request_portfolio_weights(update, tickers_only, specified_currency, specified_period)
                    return
            
            # Если есть смешанный ввод (тикеры с весами и без), это ошибка
            if tickers_only and portfolio_data:
                await self._send_message_safe(update, 
                    "❌ Некорректный формат ввода.\n\n"
                    "Укажите веса для всех активов или только тикеры для равномерного распределения.\n\n"
                    "Примеры:\n"
                    "• `SPY.US QQQ.US BND.US` - только тикеры\n"
                    "• `SPY.US:0.5 QQQ.US:0.3 BND.US:0.2` - тикеры с весами"
                )
                return
            
            if not portfolio_data:
                await self._send_message_safe(update, "❌ Не указаны активы для портфеля")
                return
            
            # Check if weights sum to approximately 1.0
            total_weight = sum(weight for _, weight in portfolio_data)
            if abs(total_weight - 1.0) > 0.01:
                # Предлагаем исправление, если сумма близка к 1
                if abs(total_weight - 1.0) <= 0.1:
                    corrected_weights = []
                    for symbol, weight in portfolio_data:
                        corrected_weight = weight / total_weight
                        corrected_weights.append((symbol, corrected_weight))
                    
                    await self._send_message_safe(update, 
                        f"⚠️ Сумма долей ({total_weight:.3f}) не равна 1.0\n\n"
                        f"Исправленные доли:\n"
                        f"{chr(10).join([f'• {symbol}: {weight:.3f}' for symbol, weight in corrected_weights])}\n\n"
                        f"Попробуйте команду:\n"
                        f"`/portfolio {' '.join([f'{symbol}:{weight:.3f}' for symbol, weight in corrected_weights])}`"
                    )
                else:
                    await self._send_message_safe(update, 
                        f"❌ Сумма долей должна быть равна 1.0, текущая сумма: {total_weight:.3f}\n\n"
                        f"Пример правильной команды:\n"
                        f"`/portfolio LQDT.MOEX:0.78 OBLG.MOEX:0.16 GOLD.MOEX:0.06`"
                    )
                return
            
            if len(portfolio_data) > 10:
                await self._send_message_safe(update, "❌ Максимум 10 активов в портфеле")
                return
            
            symbols = [symbol for symbol, _ in portfolio_data]
            weights = [weight for _, weight in portfolio_data]
            
            await self._send_ephemeral_message(update, context, f"Создаю портфель: {', '.join(symbols)}...", delete_after=3)
            
            # Create portfolio using okama
            self.logger.info(f"DEBUG: About to create portfolio with symbols: {symbols}, weights: {weights}")
            self.logger.info(f"DEBUG: Symbols types: {[type(s) for s in symbols]}")
            self.logger.info(f"DEBUG: Weights types: {[type(w) for w in weights]}")
            
            # Determine base currency - use specified currency if provided, otherwise auto-detect
            if specified_currency:
                currency = specified_currency
                currency_info = f"указана пользователем ({specified_currency})"
                self.logger.info(f"Using user-specified currency for portfolio: {currency}")
            else:
                # Auto-detect currency from the first asset
                first_symbol = symbols[0]
                try:
                    # Create asset to get its currency
                    first_asset = ok.Asset(first_symbol)
                    currency = first_asset.currency
                    currency_info = f"автоматически определена по первому активу ({first_symbol})"
                    self.logger.info(f"Currency determined from asset {first_symbol}: {currency}")
                except Exception as e:
                    self.logger.warning(f"Could not determine currency from asset {first_symbol}: {e}")
                    # Fallback to namespace-based detection using our function
                    currency, currency_info = self._get_currency_by_symbol(first_symbol)
            
            # Create portfolio using okama with period support
            try:
                # Apply period filter if specified
                if specified_period:
                    years = int(specified_period[:-1])  # Extract number from '5Y'
                    from datetime import timedelta
                    end_date = datetime.now()
                    start_date = end_date - timedelta(days=years * 365)
                    portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency,
                                           first_date=start_date.strftime('%Y-%m-%d'), 
                                           last_date=end_date.strftime('%Y-%m-%d'))
                    self.logger.info(f"Created portfolio with period {specified_period}")
                else:
                    portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency)
                    self.logger.info(f"Created portfolio with maximum available period")
                
                # Create portfolio information text (without raw object)
                portfolio_text = f"💼 **Портфель создан успешно!**\n\n"
                
                # Add basic metrics to portfolio text
                try:
                    metrics_text = self._get_portfolio_basic_metrics(portfolio, symbols, weights, currency)
                    portfolio_text += metrics_text
                except Exception as e:
                    self.logger.warning(f"Could not add metrics to portfolio text: {e}")
                
                # Generate portfolio symbol using PF namespace and okama's assigned symbol
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                
                # Count existing portfolios for this user
                portfolio_count = user_context.get('portfolio_count', 0) + 1
                
                # Use PF namespace with custom symbol (okama's symbol is composition string, not suitable for bot)
                portfolio_symbol = f"PF{portfolio_count}"
                
                # Create compact portfolio data string for callback (only symbols to avoid Button_data_invalid)
                portfolio_data_str = ','.join(symbols)
                
                # Add portfolio symbol display
                portfolio_text += f"\n\n⚖️ Сравнить портфель: `/compare {portfolio_symbol}`\n"
                
                # No Inline Keyboard needed - only Reply Keyboard will be used
                reply_markup = None
                
                # Send ephemeral message about creating chart
                await self._send_ephemeral_message(update, context, "📈 Создаю график накопленной доходности...", delete_after=3)
                
                # Create and send wealth chart with portfolio info in caption and buttons
                await self._create_portfolio_wealth_chart_with_info(update, context, portfolio, symbols, currency, weights, portfolio_symbol, portfolio_text, reply_markup)
                
                # Store portfolio data in context
                user_id = update.effective_user.id
                self.logger.info(f"Storing portfolio data in context for user {user_id}")
                self.logger.info(f"Symbols: {symbols}")
                self.logger.info(f"Currency: {currency}")
                self.logger.info(f"Weights: {weights}")
                self.logger.info(f"Portfolio symbol: {portfolio_symbol}")
                
                # Store portfolio information in user context
                self._update_user_context(
                    user_id, 
                    last_assets=symbols,
                    last_analysis_type='portfolio',
                    last_period=specified_period or 'MAX',
                    current_symbols=symbols,
                    current_currency=currency,
                    current_currency_info=currency_info,
                    portfolio_weights=weights,
                    portfolio_count=portfolio_count,
                    current_period=specified_period
                )
                
                # Verify what was saved
                saved_context = self._get_user_context(user_id)
                self.logger.info(f"Saved context keys: {list(saved_context.keys())}")
                self.logger.info(f"Saved current_symbols: {saved_context.get('current_symbols')}")
                self.logger.info(f"Saved current_currency: {saved_context.get('current_currency')}")
                self.logger.info(f"Saved portfolio_weights: {saved_context.get('portfolio_weights')}")
                
                # Get current saved portfolios and add the new portfolio
                saved_portfolios = user_context.get('saved_portfolios', {})
                
                # Generate portfolio name
                portfolio_name = self._generate_portfolio_name(symbols, weights)
                
                # Create portfolio attributes for storage
                portfolio_attributes = {
                    'symbols': symbols,
                    'weights': weights,
                    'currency': currency,
                    'created_at': datetime.now().isoformat(),
                    'description': f"Портфель: {', '.join(symbols)}",
                    'portfolio_symbol': portfolio_symbol,
                    'portfolio_name': portfolio_name,
                    'total_weight': sum(weights),
                    'asset_count': len(symbols),
                    'period': specified_period
                }
                
                # Add portfolio to saved portfolios
                saved_portfolios[portfolio_symbol] = portfolio_attributes
                
                # Update saved portfolios in context
                self._update_user_context(
                    user_id,
                    saved_portfolios=saved_portfolios,
                    portfolio_count=portfolio_count
                )
                
                # Verify what was saved
                final_saved_context = self._get_user_context(user_id)
                self.logger.info(f"Final saved portfolios count: {len(final_saved_context.get('saved_portfolios', {}))}")
                self.logger.info(f"Final saved portfolios keys: {list(final_saved_context.get('saved_portfolios', {}).keys())}")
                
            except Exception as e:
                self.logger.error(f"Error creating portfolio: {e}")
                # Restore waiting flag so user can try again
                self._update_user_context(user_id, waiting_for_portfolio=True)
                await self._send_message_safe(update, 
                    f"❌ Ошибка при создании портфеля: {str(e)}\n\n"
                    "💡 Возможные причины:\n"
                    "• Один из символов недоступен\n"
                    "• Проблемы с данными\n"
                    "• Неверный формат символа\n\n"
                    "Проверьте:\n"
                    "• Правильность написания символов\n"
                    "• Доступность данных для указанных активов\n\n"
                    "🔄 Попробуйте ввести портфель снова:"
                )
                
        except Exception as e:
            self.logger.error(f"Error in portfolio input handler: {e}")
            # Clear user context to prevent fallback to compare command
            self._update_user_context(user_id, 
                waiting_for_portfolio=False,
                waiting_for_portfolio_weights=False,
                waiting_for_compare=False,
                portfolio_tickers=None,
                portfolio_base_symbols=None
            )
            await self._send_message_safe(update, f"❌ Ошибка при обработке ввода портфеля: {str(e)}\n\n🔄 Попробуйте ввести портфель снова:")

    async def _handle_portfolio_weights_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle portfolio weights input from compare command"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Get base symbols from context
            base_symbols = user_context.get('portfolio_base_symbols', [])
            if not base_symbols:
                await self._send_message_safe(update, "❌ Ошибка: не найдены базовые символы для портфеля")
                return
            
            # Clear waiting flag
            self._update_user_context(user_id, 
                waiting_for_portfolio_weights=False,
                portfolio_base_symbols=None
            )
            
            # Parse currency and period parameters from input text
            text_args = text.split()
            
            # For portfolio command, we need to preserve the full symbol:weight format
            # So we'll parse currency and period manually, keeping the original arguments
            valid_currencies = {'USD', 'RUB', 'EUR', 'GBP', 'CNY', 'HKD', 'JPY'}
            import re
            period_pattern = re.compile(r'^(\d+)Y$', re.IGNORECASE)
            
            portfolio_args = []
            specified_currency = None
            specified_period = None
            
            for arg in text_args:
                arg_upper = arg.upper()
                
                # Check if it's a currency code
                if arg_upper in valid_currencies:
                    if specified_currency is None:
                        specified_currency = arg_upper
                    continue
                
                # Check if it's a period (e.g., '5Y', '10Y')
                period_match = period_pattern.match(arg)
                if period_match:
                    if specified_period is None:
                        specified_period = arg_upper
                    continue
                
                # If it's neither currency nor period, it's a portfolio argument
                portfolio_args.append(arg)
            
            # Use smart parsing for portfolio arguments
            portfolio_text = ' '.join(portfolio_args)
            parse_result = self.smart_parse_portfolio_input(portfolio_text)
            
            if not parse_result['success']:
                # Show error message with suggestions
                error_msg = parse_result['message']
                if parse_result['suggestions']:
                    error_msg += "\n\n💡 Подсказки:\n" + "\n".join(f"• {s}" for s in parse_result['suggestions'])
                await self._send_message_safe(update, error_msg)
                return
            
            portfolio_data = parse_result['portfolio_data']
            
            # Show success message with any corrections made
            if parse_result['suggestions']:
                await self._send_message_safe(update, parse_result['message'])
            
            # Check if weights sum to approximately 1.0
            total_weight = sum(weight for _, weight in portfolio_data)
            if abs(total_weight - 1.0) > 0.01:
                # Предлагаем исправление, если сумма близка к 1
                if abs(total_weight - 1.0) <= 0.1:
                    corrected_weights = []
                    for symbol, weight in portfolio_data:
                        corrected_weight = weight / total_weight
                        corrected_weights.append((symbol, corrected_weight))
                    
                    await self._send_message_safe(update, 
                        f"⚠️ Сумма долей ({total_weight:.3f}) не равна 1.0\n\n"
                        f"Исправленные доли:\n"
                        f"{chr(10).join([f'• {symbol}: {weight:.3f}' for symbol, weight in corrected_weights])}\n\n"
                        f"Попробуйте команду:\n"
                        f"`/portfolio {' '.join([f'{symbol}:{weight:.3f}' for symbol, weight in corrected_weights])}`"
                    )
                else:
                    await self._send_message_safe(update, 
                        f"❌ Сумма долей должна быть равна 1.0, текущая сумма: {total_weight:.3f}\n\n"
                        f"Пример правильной команды:\n"
                        f"`/portfolio {base_symbols[0]}:0.6 {base_symbols[1] if len(base_symbols) > 1 else 'QQQ.US'}:0.4`"
                    )
                return
            
            if len(portfolio_data) > 10:
                await self._send_message_safe(update, "❌ Максимум 10 активов в портфеле")
                return
            
            symbols = [symbol for symbol, _ in portfolio_data]
            weights = [weight for _, weight in portfolio_data]
            
            await self._send_ephemeral_message(update, context, f"Создаю портфель: {', '.join(symbols)}...", delete_after=3)
            
            # Create portfolio using okama
            self.logger.info(f"DEBUG: About to create portfolio with symbols: {symbols}, weights: {weights}")
            self.logger.info(f"DEBUG: Symbols types: {[type(s) for s in symbols]}")
            self.logger.info(f"DEBUG: Weights types: {[type(w) for w in weights]}")
            
            # Determine base currency - use specified currency if provided, otherwise auto-detect
            if specified_currency:
                currency = specified_currency
                currency_info = f"указана пользователем ({specified_currency})"
                self.logger.info(f"Using user-specified currency for portfolio: {currency}")
            else:
                # Auto-detect currency from the first asset
                first_symbol = symbols[0]
                try:
                    # Create asset to get its currency
                    first_asset = ok.Asset(first_symbol)
                    currency = first_asset.currency
                    currency_info = f"автоматически определена по первому активу ({first_symbol})"
                    self.logger.info(f"Currency determined from asset {first_symbol}: {currency}")
                except Exception as e:
                    self.logger.warning(f"Could not determine currency from asset {first_symbol}: {e}")
                    # Fallback to namespace-based detection using our function
                    currency, currency_info = self._get_currency_by_symbol(first_symbol)
            
            # Create portfolio using okama with period support
            try:
                # Apply period filter if specified
                if specified_period:
                    years = int(specified_period[:-1])  # Extract number from '5Y'
                    from datetime import timedelta
                    end_date = datetime.now()
                    start_date = end_date - timedelta(days=years * 365)
                    portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency,
                                           first_date=start_date.strftime('%Y-%m-%d'), 
                                           last_date=end_date.strftime('%Y-%m-%d'))
                    self.logger.info(f"Created portfolio with period {specified_period}")
                else:
                    portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency)
                    self.logger.info(f"Created portfolio with maximum available period")
                
                # Create portfolio information text (without raw object)
                portfolio_text = f"💼 **Портфель создан успешно!**\n\n"
                
                # Add basic metrics to portfolio text
                try:
                    metrics_text = self._get_portfolio_basic_metrics(portfolio, symbols, weights, currency)
                    portfolio_text += metrics_text
                except Exception as e:
                    self.logger.warning(f"Could not add metrics to portfolio text: {e}")
                
                # Generate portfolio symbol using PF namespace and okama's assigned symbol
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                
                # Count existing portfolios for this user
                portfolio_count = user_context.get('portfolio_count', 0) + 1
                
                # Use PF namespace with okama's assigned symbol
                # Use PF namespace with custom symbol (okama's symbol is composition string, not suitable for bot)
                portfolio_symbol = f"PF{portfolio_count}"
                
                # Create compact portfolio data string for callback (only symbols to avoid Button_data_invalid)
                portfolio_data_str = ','.join(symbols)
                
                # Add portfolio symbol display
                portfolio_text += f"\n\n⚖️ Сравнить портфель: `/compare {portfolio_symbol}`\n"
                
                # Add buttons in 2 columns
                keyboard = [
                    [InlineKeyboardButton("📈 Доходность (накоп.)", callback_data=f"portfolio_wealth_chart_{portfolio_symbol}"),
                     InlineKeyboardButton("💰 Доходность (ГГ)", callback_data=f"portfolio_returns_{portfolio_symbol}")],
                    [InlineKeyboardButton("📉 Просадки", callback_data=f"portfolio_drawdowns_{portfolio_symbol}"),
                     InlineKeyboardButton("📊 Метрики", callback_data=f"portfolio_risk_metrics_{portfolio_symbol}")],
                    [InlineKeyboardButton("🎲 Монте Карло", callback_data=f"portfolio_monte_carlo_{portfolio_symbol}"),
                     InlineKeyboardButton("📈 Процентили 10, 50, 90", callback_data=f"portfolio_forecast_{portfolio_symbol}")],
                    [InlineKeyboardButton("📊 Портфель vs Активы", callback_data=f"portfolio_compare_assets_{portfolio_symbol}"),
                     InlineKeyboardButton("⚖️ Сравнить", callback_data=f"portfolio_compare_{portfolio_symbol}")],
                    [InlineKeyboardButton("💰 Дивиденды", callback_data=f"portfolio_dividends_{portfolio_symbol}")],
                    [InlineKeyboardButton("🤖 AI-анализ", callback_data=f"portfolio_ai_analysis_{portfolio_symbol}")] if self.gemini_service and self.gemini_service.is_available() else []
                ]
                
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                # Send portfolio information with buttons
                await self._send_message_safe(update, portfolio_text, reply_markup=reply_markup)
                
                # Update user context with portfolio information
                self._update_user_context(
                    user_id, 
                    last_assets=symbols,
                    last_analysis_type='portfolio',
                    last_period=specified_period or 'MAX',
                    current_symbols=symbols,
                    current_currency=currency,
                    current_currency_info=currency_info,
                    portfolio_weights=weights,
                    portfolio_count=portfolio_count,
                    current_period=specified_period
                )
                
                # Verify what was saved
                saved_context = self._get_user_context(user_id)
                self.logger.info(f"Saved context keys: {list(saved_context.keys())}")
                self.logger.info(f"Saved current_symbols: {saved_context.get('current_symbols')}")
                self.logger.info(f"Saved current_currency: {saved_context.get('current_currency')}")
                self.logger.info(f"Saved portfolio_weights: {saved_context.get('portfolio_weights')}")
                
                # Get current saved portfolios and add the new portfolio
                saved_portfolios = user_context.get('saved_portfolios', {})
                
                # Generate portfolio name
                portfolio_name = self._generate_portfolio_name(symbols, weights)
                
                # Create portfolio attributes for storage
                portfolio_attributes = {
                    'symbols': symbols,
                    'weights': weights,
                    'currency': currency,
                    'created_at': datetime.now().isoformat(),
                    'description': f"Портфель: {', '.join(symbols)}",
                    'portfolio_symbol': portfolio_symbol,
                    'portfolio_name': portfolio_name,
                    'total_weight': sum(weights),
                    'asset_count': len(symbols),
                    'period': specified_period
                }
                
                # Add portfolio to saved portfolios
                saved_portfolios[portfolio_symbol] = portfolio_attributes
                
                # Update saved portfolios in context
                self._update_user_context(
                    user_id,
                    saved_portfolios=saved_portfolios,
                    portfolio_count=portfolio_count
                )
                
                # Verify what was saved
                final_saved_context = self._get_user_context(user_id)
                self.logger.info(f"Final saved portfolios count: {len(final_saved_context.get('saved_portfolios', {}))}")
                self.logger.info(f"Final saved portfolios keys: {list(final_saved_context.get('saved_portfolios', {}).keys())}")
                
            except Exception as e:
                self.logger.error(f"Error creating portfolio: {e}")
                await self._send_message_safe(update, 
                    f"❌ Ошибка при создании портфеля: {str(e)}\n\n"
                    "💡 Возможные причины:\n"
                    "• Один из символов недоступен\n"
                    "• Проблемы с данными\n"
                    "• Неверный формат символа\n\n"
                    "Проверьте:\n"
                    "• Правильность написания символов\n"
                    "• Доступность данных для указанных активов"
                )
                
        except Exception as e:
            self.logger.error(f"Error in portfolio weights input handler: {e}")
            # Clear user context to prevent fallback to compare command
            self._update_user_context(user_id, 
                waiting_for_portfolio=False,
                waiting_for_portfolio_weights=False,
                waiting_for_compare=False,
                portfolio_tickers=None,
                portfolio_base_symbols=None
            )
            await self._send_message_safe(update, f"❌ Ошибка при обработке ввода портфеля: {str(e)}\n\n🔄 Попробуйте ввести веса снова:")

    async def _handle_portfolio_tickers_weights_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle portfolio weights input when user provided only tickers"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Get tickers from context
            tickers = user_context.get('portfolio_tickers', [])
            if not tickers:
                await self._send_message_safe(update, "❌ Ошибка: не найдены тикеры для портфеля")
                return
            
            # Clear waiting flag and context
            self._update_user_context(user_id, 
                waiting_for_portfolio_weights=False,
                portfolio_tickers=None,
                portfolio_currency=None,
                portfolio_period=None
            )
            
            # Parse currency and period parameters from input text
            text_args = text.split()
            
            valid_currencies = {'USD', 'RUB', 'EUR', 'GBP', 'CNY', 'HKD', 'JPY'}
            import re
            period_pattern = re.compile(r'^(\d+)Y$', re.IGNORECASE)
            
            portfolio_args = []
            specified_currency = None
            specified_period = None
            
            for arg in text_args:
                arg_upper = arg.upper()
                
                # Check if it's a currency code
                if arg_upper in valid_currencies:
                    if specified_currency is None:
                        specified_currency = arg_upper
                    continue
                
                # Check if it's a period (e.g., '5Y', '10Y')
                period_match = period_pattern.match(arg)
                if period_match:
                    if specified_period is None:
                        specified_period = arg_upper
                    continue
                
                # If it's neither currency nor period, it's a portfolio argument
                portfolio_args.append(arg)
            
            # Parse portfolio arguments
            portfolio_data = []
            
            for arg in portfolio_args:
                if ':' in arg:
                    symbol_part, weight_part = arg.split(':', 1)
                    original_symbol = self.clean_symbol(symbol_part.strip())
                    symbol = original_symbol.upper()
                    
                    try:
                        weight_str = weight_part.strip()
                        
                        # Поддержка процентов
                        if weight_str.endswith('%'):
                            weight = float(weight_str[:-1]) / 100.0
                        else:
                            weight = float(weight_str)
                            
                    except Exception as e:
                        self.logger.error(f"Error converting weight '{weight_part.strip()}' to float: {e}")
                        await self._send_message_safe(update, f"❌ Некорректная доля для {symbol}: '{weight_part.strip()}'. Доля должна быть числом от 0 до 1 или процентом (например, 50%)")
                        return
                    
                    if weight <= 0 or weight > 1:
                        await self._send_message_safe(update, f"❌ Некорректная доля для {symbol}: {weight}. Доля должна быть от 0 до 1")
                        return
                    
                    portfolio_data.append((symbol, weight))
                    
                else:
                    # Это только число - попробуем сопоставить с тикерами по порядку
                    try:
                        weight_str = arg.strip()
                        
                        # Поддержка процентов
                        if weight_str.endswith('%'):
                            weight = float(weight_str[:-1]) / 100.0
                        else:
                            weight = float(weight_str)
                            
                        if weight <= 0 or weight > 1:
                            await self._send_message_safe(update, f"❌ Некорректная доля: {weight}. Доля должна быть от 0 до 1")
                            return
                        
                        # Добавляем вес без символа - будет сопоставлен позже
                        portfolio_data.append((None, weight))
                        
                    except Exception as e:
                        await self._send_message_safe(update, f"❌ Некорректный формат: {arg}. Используйте формат символ:доля или только доли")
                        return
            
            # Если пользователь указал только числа, сопоставляем их с тикерами по порядку
            if portfolio_data and all(symbol is None for symbol, _ in portfolio_data):
                if len(portfolio_data) != len(tickers):
                    await self._send_message_safe(update, 
                        f"❌ Количество долей ({len(portfolio_data)}) не совпадает с количеством тикеров ({len(tickers)})\n\n"
                        f"Тикеры: {', '.join(tickers)}\n"
                        f"Укажите доли для каждого тикера в том же порядке"
                    )
                    return
                
                # Сопоставляем доли с тикерами
                portfolio_data = [(tickers[i], weight) for i, (_, weight) in enumerate(portfolio_data)]
            
            # Проверяем, что все тикеры имеют доли
            portfolio_symbols = [symbol for symbol, _ in portfolio_data]
            missing_tickers = [ticker for ticker in tickers if ticker not in portfolio_symbols]
            
            if missing_tickers:
                await self._send_message_safe(update, 
                    f"❌ Не указаны доли для тикеров: {', '.join(missing_tickers)}\n\n"
                    f"Укажите доли для всех тикеров в том же порядке:\n"
                    f"`{' '.join([f'{ticker}:0.XX' for ticker in tickers])}`"
                )
                return
            
            # Проверяем сумму долей
            total_weight = sum(weight for _, weight in portfolio_data)
            if abs(total_weight - 1.0) > 0.01:
                # Предлагаем исправление, если сумма близка к 1
                if abs(total_weight - 1.0) <= 0.1:
                    corrected_weights = []
                    for symbol, weight in portfolio_data:
                        corrected_weight = weight / total_weight
                        corrected_weights.append((symbol, corrected_weight))
                    
                    await self._send_message_safe(update, 
                        f"⚠️ Сумма долей ({total_weight:.3f}) не равна 1.0\n\n"
                        f"Исправленные доли:\n"
                        f"{chr(10).join([f'• {symbol}: {weight:.3f}' for symbol, weight in corrected_weights])}\n\n"
                        f"Попробуйте команду:\n"
                        f"`/portfolio {' '.join([f'{symbol}:{weight:.3f}' for symbol, weight in corrected_weights])}`"
                    )
                else:
                    await self._send_message_safe(update, 
                        f"❌ Сумма долей должна быть равна 1.0, текущая сумма: {total_weight:.3f}\n\n"
                        f"Пример правильной команды:\n"
                        f"`/portfolio {' '.join([f'{ticker}:0.XX' for ticker in tickers])}`"
                    )
                return
            
            if len(portfolio_data) > 10:
                await self._send_message_safe(update, "❌ Максимум 10 активов в портфеле")
                return
            
            symbols = [symbol for symbol, _ in portfolio_data]
            weights = [weight for _, weight in portfolio_data]
            
            await self._send_ephemeral_message(update, context, f"Создаю портфель: {', '.join(symbols)}...", delete_after=3)
            
            # Create portfolio using okama
            self.logger.info(f"DEBUG: About to create portfolio with symbols: {symbols}, weights: {weights}")
            
            # Determine base currency
            if specified_currency:
                currency = specified_currency
                currency_info = f"указана пользователем ({specified_currency})"
            else:
                # Auto-detect currency from the first asset
                first_symbol = symbols[0]
                try:
                    first_asset = ok.Asset(first_symbol)
                    currency = first_asset.currency
                    currency_info = f"автоматически определена по первому активу ({first_symbol})"
                except Exception as e:
                    self.logger.warning(f"Could not determine currency from asset {first_symbol}: {e}")
                    currency, currency_info = self._get_currency_by_symbol(first_symbol)
            
            # Create portfolio using okama with period support
            try:
                if specified_period:
                    years = int(specified_period[:-1])
                    from datetime import timedelta
                    end_date = datetime.now()
                    start_date = end_date - timedelta(days=years * 365)
                    portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency,
                                           first_date=start_date.strftime('%Y-%m-%d'), 
                                           last_date=end_date.strftime('%Y-%m-%d'))
                else:
                    portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency)
                
                # Create portfolio information text
                portfolio_text = f"💼 **Портфель создан успешно!**\n\n"
                
                # Add basic metrics to portfolio text
                try:
                    metrics_text = self._get_portfolio_basic_metrics(portfolio, symbols, weights, currency)
                    portfolio_text += metrics_text
                except Exception as e:
                    self.logger.warning(f"Could not add metrics to portfolio text: {e}")
                
                # Generate portfolio symbol
                user_context = self._get_user_context(user_id)
                portfolio_count = user_context.get('portfolio_count', 0) + 1
                portfolio_symbol = f"PF{portfolio_count}"
                
                # Add portfolio symbol display
                portfolio_text += f"\n\n⚖️ Сравнить портфель: `/compare {portfolio_symbol}`\n"
                
                # Add buttons
                keyboard = [
                    [InlineKeyboardButton("💼 Анализ портфеля", callback_data=f"portfolio_analysis_{portfolio_symbol}")],
                    [InlineKeyboardButton("📈 График доходности", callback_data=f"portfolio_chart_{portfolio_symbol}"),
                     InlineKeyboardButton("📋 Метрики риска", callback_data=f"portfolio_risk_{portfolio_symbol}")],
                    [InlineKeyboardButton("💼 Сохранить портфель", callback_data=f"portfolio_save_{portfolio_symbol}")]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                await self._send_message_safe(update, portfolio_text, reply_markup=reply_markup)
                
                # Save portfolio to user context
                portfolio_attributes = {
                    'symbols': symbols,
                    'weights': weights,
                    'currency': currency,
                    'created_at': datetime.now().isoformat(),
                    'description': f"Портфель: {', '.join(symbols)}",
                    'portfolio_symbol': portfolio_symbol,
                    'total_weight': sum(weights),
                    'asset_count': len(symbols),
                    'period': specified_period
                }
                
                saved_portfolios = user_context.get('saved_portfolios', {})
                saved_portfolios[portfolio_symbol] = portfolio_attributes
                
                self._update_user_context(
                    user_id,
                    saved_portfolios=saved_portfolios,
                    portfolio_count=portfolio_count
                )
                
            except Exception as e:
                self.logger.error(f"Error creating portfolio: {e}")
                await self._send_message_safe(update, 
                    f"❌ Ошибка при создании портфеля: {str(e)}\n\n"
                    "💡 Возможные причины:\n"
                    "• Один из символов недоступен\n"
                    "• Проблемы с данными\n"
                    "• Неверный формат символа\n\n"
                    "Проверьте:\n"
                    "• Правильность написания символов\n"
                    "• Доступность данных для указанных активов"
                )
                
        except Exception as e:
            self.logger.error(f"Error in portfolio tickers weights input handler: {e}")
            # Clear user context to prevent fallback to compare command
            self._update_user_context(user_id, 
                waiting_for_portfolio=False,
                waiting_for_portfolio_weights=False,
                waiting_for_compare=False,
                portfolio_tickers=None,
                portfolio_base_symbols=None
            )
            await self._send_message_safe(update, f"❌ Ошибка при обработке ввода весов портфеля: {str(e)}\n\n🔄 Попробуйте ввести веса снова:")

    async def _handle_compare_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle compare input from user message"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Parse input text using the same logic as compare_command
            # Split the text into arguments and use the same parsing function
            text_args = text.split()
            symbols, specified_currency, specified_period = self._parse_currency_and_period(text_args)
            
            # Clean up symbols (remove empty strings and whitespace)
            symbols = [symbol for symbol in symbols if symbol.strip()]
            
            # Log the parsed parameters for debugging
            self.logger.info(f"Parsed symbols: {symbols}")
            self.logger.info(f"Parsed currency: {specified_currency}")
            self.logger.info(f"Parsed period: {specified_period}")
            
            # Check if we have a stored first symbol from previous input or from compare button
            stored_first_symbol = user_context.get('compare_first_symbol')
            compare_base_symbol = user_context.get('compare_base_symbol')
            
            if len(symbols) == 1:
                if stored_first_symbol is None and compare_base_symbol is None:
                    # First symbol - store it and ask for more
                    self._update_user_context(user_id, compare_first_symbol=symbols[0], waiting_for_compare=True)
                    # Generate random examples for the message
                    random_examples = self.get_random_examples(3)
                    examples_text = ", ".join([f"`{example}`" for example in random_examples])
                    await self._send_message_safe(update, f"Вы указали только 1 символ, а для сравнения нужно 2 и больше, напишите дополнительный символ для сравнения, например {examples_text}")
                    return
                else:
                    # We have a stored symbol or base symbol, combine with new input
                    base_symbol = stored_first_symbol or compare_base_symbol
                    combined_symbols = [base_symbol] + symbols
                    
                    # Check if combined symbols exceed the limit
                    if len(combined_symbols) > 5:
                        await self._send_message_safe(update, "❌ Максимум 5 активов для сравнения. Пожалуйста, введите список для сравнения заново (не более 5 активов)")
                        # Clear both stored symbols and waiting flag
                        self._update_user_context(user_id, compare_first_symbol=None, compare_base_symbol=None, waiting_for_compare=False)
                        return
                    
                    # Clear both stored symbols and waiting flag
                    self._update_user_context(user_id, compare_first_symbol=None, compare_base_symbol=None, waiting_for_compare=False)
                    
                    # Process the comparison with combined symbols
                    context.args = combined_symbols
                    context.specified_currency = specified_currency
                    context.specified_period = specified_period
                    
                    await self.compare_command(update, context)
                    return
            
            elif len(symbols) == 0:
                # Empty input - clear stored symbols and show help
                self._update_user_context(user_id, compare_first_symbol=None, compare_base_symbol=None, waiting_for_compare=False)
                await self._send_message_safe(update, "❌ Необходимо указать минимум 2 символа для сравнения")
                return
            
            elif len(symbols) > 5:
                await self._send_message_safe(update, "❌ Максимум 5 активов для сравнения. Пожалуйста, введите список для сравнения заново (не более 5 активов)")
                # Clear any stored symbols and reset waiting state
                self._update_user_context(user_id, compare_first_symbol=None, compare_base_symbol=None, waiting_for_compare=False)
                return
            
            # We have 2 or more symbols - clear any stored symbols and process normally
            self._update_user_context(user_id, compare_first_symbol=None, compare_base_symbol=None, waiting_for_compare=False)
            
            # Process the comparison using the same logic as compare_command
            # We'll reuse the existing comparison logic by calling compare_command with args
            context.args = symbols
            
            # Store parsed currency and period in context for compare_command to use
            context.specified_currency = specified_currency
            context.specified_period = specified_period
            
            await self.compare_command(update, context)
            
        except Exception as e:
            self.logger.error(f"Error in compare input handler: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке ввода сравнения: {str(e)}")

    def _safe_markdown(self, text: str) -> str:
        """Safe Markdown cleaning to prevent parsing errors - simple version"""
        try:
            if not text or not isinstance(text, str):
                return text or ""
            
            import re
            
            # Simple approach: fix most common issues that cause parsing errors
            # Fix unclosed ** - count and balance them
            bold_count = text.count('**')
            if bold_count % 2 == 1:
                # Remove last **
                last_bold = text.rfind('**')
                if last_bold != -1:
                    text = text[:last_bold] + text[last_bold + 2:]
                    self.logger.warning("Fixed unclosed bold marker")
            
            # Fix unclosed * - count and balance them
            italic_count = text.count('*')
            if italic_count % 2 == 1:
                # Remove last *
                last_italic = text.rfind('*')
                if last_italic != -1:
                    text = text[:last_italic] + text[last_italic + 1:]
                    self.logger.warning("Fixed unclosed italic marker")
            
            # Fix unclosed ` - count and balance them
            code_count = text.count('`')
            if code_count % 2 == 1:
                # Remove last `
                last_code = text.rfind('`')
                if last_code != -1:
                    text = text[:last_code] + text[last_code + 1:]
                    self.logger.warning("Fixed unclosed inline code")
            
            # Fix unclosed code blocks ``` - count and balance them
            block_count = text.count('```')
            if block_count % 2 == 1:
                # Remove last ```
                last_block = text.rfind('```')
                if last_block != -1:
                    text = text[:last_block] + text[last_block + 3:]
                    self.logger.warning("Fixed unclosed code block")
            
            # Escape problematic underscores
            text = re.sub(r'(?<!\*)_(?!\*)', r'\_', text)
            
            return text
            
        except Exception as e:
            self.logger.warning(f"Error in safe markdown cleaning: {e}")
            # Last resort: remove all markdown
            import re
            text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)
            text = re.sub(r'\*(.*?)\*', r'\1', text)
            text = re.sub(r'`(.*?)`', r'\1', text)
            text = re.sub(r'```.*?```', '', text, flags=re.DOTALL)
            return text

    def _clean_markdown(self, text: str) -> str:
        """Clean and fix Markdown formatting to prevent parsing errors"""
        try:
            if not text or not isinstance(text, str):
                return text or ""
            
            # Remove or fix common Markdown issues
            import re
            
            # Fix unclosed bold/italic markers
            # Count ** and * to ensure they're balanced
            bold_count = text.count('**')
            italic_count = text.count('*')
            
            # If odd number of **, remove the last one
            if bold_count % 2 == 1:
                # Find the last ** and remove it
                last_bold = text.rfind('**')
                if last_bold != -1:
                    text = text[:last_bold] + text[last_bold + 2:]
                    self.logger.warning("Fixed unclosed bold marker")
            
            # If odd number of *, remove the last one
            if italic_count % 2 == 1:
                # Find the last * and remove it
                last_italic = text.rfind('*')
                if last_italic != -1:
                    text = text[:last_italic] + text[last_italic + 1:]
                    self.logger.warning("Fixed unclosed italic marker")
            
            # Fix unclosed code blocks - remove the opening ``` and everything after it
            code_block_count = text.count('```')
            if code_block_count % 2 == 1:
                # Find the last ``` and remove it and everything after it
                last_code = text.rfind('```')
                if last_code != -1:
                    # Find the start of the line with the last ```
                    line_start = text.rfind('\n', 0, last_code) + 1
                    text = text[:line_start] + text[last_code + 3:]
                    self.logger.warning("Fixed unclosed code block")
            
            # Fix unclosed inline code
            inline_code_count = text.count('`')
            if inline_code_count % 2 == 1:
                # Remove the last `
                last_inline = text.rfind('`')
                if last_inline != -1:
                    text = text[:last_inline] + text[last_inline + 1:]
                    self.logger.warning("Fixed unclosed inline code")
            
            # Remove or escape problematic characters
            # Escape underscores that are not part of markdown
            text = re.sub(r'(?<!\*)_(?!\*)', r'\_', text)
            
            # Remove or fix problematic sequences
            text = re.sub(r'\*\*\*\*+', '**', text)  # More than 2 asterisks
            text = re.sub(r'\*\*\*', '**', text)     # 3 asterisks
            
            return text
            
        except Exception as e:
            self.logger.warning(f"Error cleaning markdown: {e}")
            # If cleaning fails, remove all markdown
            import re
            text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)  # Remove bold
            text = re.sub(r'\*(.*?)\*', r'\1', text)      # Remove italic
            text = re.sub(r'`(.*?)`', r'\1', text)        # Remove inline code
            text = re.sub(r'```.*?```', '', text, flags=re.DOTALL)  # Remove code blocks
            return text

    async def _remove_keyboard_from_previous_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Удалить клавиатуру с предыдущего сообщения для улучшения UX"""
        try:
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
        except Exception as e:
            self.logger.warning(f"Could not remove keyboard from previous message: {e}")

    async def _remove_keyboard_after_successful_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Удалить клавиатуру с предыдущего сообщения только после успешного создания нового сообщения"""
        try:
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
                self.logger.info("Successfully removed keyboard from previous message")
        except Exception as e:
            self.logger.warning(f"Could not remove keyboard from previous message after successful message creation: {e}")

    async def _remove_keyboard_before_new_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Удалить клавиатуру с предыдущего сообщения перед отправкой нового сообщения"""
        try:
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                self.logger.info(f"Attempting to remove keyboard from message ID: {update.callback_query.message.message_id}")
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
                self.logger.info("Successfully removed keyboard from previous message before sending new message")
            else:
                self.logger.warning("No callback_query found, cannot remove keyboard")
        except Exception as e:
            self.logger.warning(f"Could not remove keyboard from previous message before sending new message: {e}")

    async def _send_callback_message_with_keyboard_removal(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, parse_mode: str = None, reply_markup=None):
        """Отправить сообщение в callback query с удалением клавиатуры с предыдущего сообщения"""
        try:
            self.logger.info("Starting _send_callback_message_with_keyboard_removal")
            
            # Remove keyboard from previous message before sending new message
            await self._remove_keyboard_before_new_message(update, context)
            
            # Send new message with keyboard using context.bot.send_message directly
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                self.logger.info(f"Sending new message to chat_id: {update.callback_query.message.chat_id}")
                await context.bot.send_message(
                    chat_id=update.callback_query.message.chat_id,
                    text=text,
                    parse_mode=parse_mode,
                    reply_markup=reply_markup
                )
                self.logger.info("Successfully sent new message with keyboard")
            else:
                # Fallback to regular callback message if no callback_query
                self.logger.warning("No callback_query found, using fallback")
                await self._send_callback_message(update, context, text, parse_mode=parse_mode)
        except Exception as e:
            self.logger.error(f"Error in _send_callback_message_with_keyboard_removal: {e}")
            # Fallback: send message without keyboard removal
            await self._send_callback_message(update, context, text, parse_mode=parse_mode)

    async def _send_portfolio_message_with_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, parse_mode: str = None):
        """Отправить сообщение портфеля с reply keyboard"""
        try:
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            
            # Send message
            await self._send_message_safe(update, text, parse_mode=parse_mode)
            
        except Exception as e:
            self.logger.error(f"Error in _send_portfolio_message_with_reply_keyboard: {e}")
            # Fallback: send message without keyboard
            await self._send_message_safe(update, text)

    async def _send_portfolio_ai_analysis_with_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, parse_mode: str = None):
        """Отправить длинный AI анализ портфеля с reply keyboard, поддерживая разбиение на части"""
        try:
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            
            # Check if text is longer than Telegram limit
            if len(text) <= 4000:
                # Short message - send normally
                await self._send_message_safe(update, text, parse_mode=parse_mode)
            else:
                # Long message - split into parts
                parts = self._split_text_smart(text)
                
                for i, part in enumerate(parts):
                    # Add part indicator for multi-part messages
                    if len(parts) > 1:
                        part_text = f"📄 **Часть {i+1} из {len(parts)}:**\n\n{part}"
                    else:
                        part_text = part
                    
                    # Clean Markdown for each part
                    if parse_mode == 'Markdown':
                        part_text = self._safe_markdown(part_text)
                    
                    # Send all parts using safe method
                    await self._send_message_safe(update, part_text, parse_mode=parse_mode)
                    
                    # Small delay between parts to avoid rate limiting
                    if i < len(parts) - 1:
                        await asyncio.sleep(0.5)
            
        except Exception as e:
            self.logger.error(f"Error in _send_portfolio_ai_analysis_with_keyboard: {e}")
            # Fallback: send message without keyboard using safe method
            await self._send_message_safe(update, text)

    async def _remove_reply_keyboard_silently(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Тихо скрыть reply keyboard без отправки сообщения пользователю"""
        try:
            # Проверяем, что update и context не None
            if update is None or context is None:
                self.logger.error("Cannot remove reply keyboard: update or context is None")
                return
            
            chat_id = None
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                chat_id = update.callback_query.message.chat_id
            elif hasattr(update, 'message') and update.message is not None:
                chat_id = update.message.chat_id
            else:
                self.logger.error("Cannot remove reply keyboard: no chat_id available")
                return
            
            # Попробуем несколько способов удаления клавиатуры
            
            # Способ 1: Отправка сообщения с ReplyKeyboardRemove и удаление
            try:
                message = await context.bot.send_message(
                    chat_id=chat_id,
                    text="",  # Пустой текст
                    reply_markup=ReplyKeyboardRemove()
                )
                
                # Удаляем сообщение через небольшую задержку
                await asyncio.sleep(0.1)
                await context.bot.delete_message(chat_id=chat_id, message_id=message.message_id)
                self.logger.info("Reply keyboard removed using method 1 (send + delete)")
                return
                
            except Exception as method1_error:
                self.logger.warning(f"Method 1 failed: {method1_error}")
            
            # Способ 2: Отправка сообщения с ReplyKeyboardRemove без удаления
            try:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text="",  # Пустой текст
                    reply_markup=ReplyKeyboardRemove()
                )
                self.logger.info("Reply keyboard removed using method 2 (send only)")
                return
                
            except Exception as method2_error:
                self.logger.warning(f"Method 2 failed: {method2_error}")
            
            # Способ 3: Использование edit_message_reply_markup для callback queries
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                try:
                    await context.bot.edit_message_reply_markup(
                        chat_id=chat_id,
                        message_id=update.callback_query.message.message_id,
                        reply_markup=ReplyKeyboardRemove()
                    )
                    self.logger.info("Reply keyboard removed using method 3 (edit_message_reply_markup)")
                    return
                except Exception as method3_error:
                    self.logger.warning(f"Method 3 failed: {method3_error}")
            
            # Если все способы не сработали
            self.logger.error("All methods to remove reply keyboard failed")
            
        except Exception as e:
            self.logger.error(f"Error removing reply keyboard silently: {e}")

    async def _remove_reply_keyboard_alternative(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Альтернативный способ удаления reply keyboard - отправка сообщения с невидимым символом"""
        try:
            chat_id = None
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                chat_id = update.callback_query.message.chat_id
            elif hasattr(update, 'message') and update.message is not None:
                chat_id = update.message.chat_id
            else:
                return
            
            # Отправляем сообщение с невидимым символом и ReplyKeyboardRemove
            await context.bot.send_message(
                chat_id=chat_id,
                text="\u200B",  # Невидимый символ (Zero Width Space)
                reply_markup=ReplyKeyboardRemove()
            )
            self.logger.info("Reply keyboard removed using alternative method (invisible character)")
            
        except Exception as e:
            self.logger.error(f"Error in alternative keyboard removal: {e}")

    async def _manage_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE, keyboard_type: str = None):
        """
        Универсальное управление reply keyboard на основе контекста
        
        Args:
            update: Telegram update object
            context: Telegram context object
            keyboard_type: Тип клавиатуры для показа ("portfolio", "compare") или None для скрытия
        """
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            current_keyboard = user_context.get('active_reply_keyboard')
            
            # Если нужно скрыть клавиатуру
            if keyboard_type is None:
                if current_keyboard is not None:
                    self.logger.info(f"Removing active reply keyboard: {current_keyboard}")
                    try:
                        await self._remove_reply_keyboard_silently(update, context)
                    except Exception as e:
                        self.logger.warning(f"Primary keyboard removal failed: {e}, trying alternative method")
                        await self._remove_reply_keyboard_alternative(update, context)
                    self._update_user_context(user_id, active_reply_keyboard=None)
                return
            
            # Если нужно показать клавиатуру
            if current_keyboard != keyboard_type:
                # Скрываем текущую клавиатуру если она есть
                if current_keyboard is not None:
                    self.logger.info(f"Switching from {current_keyboard} to {keyboard_type} keyboard")
                    try:
                        await self._remove_reply_keyboard_silently(update, context)
                    except Exception as e:
                        self.logger.warning(f"Primary keyboard removal failed during switch: {e}, trying alternative method")
                        await self._remove_reply_keyboard_alternative(update, context)
                
                # Показываем новую клавиатуру
                if keyboard_type == "portfolio":
                    await self._show_portfolio_reply_keyboard(update, context)
                elif keyboard_type == "compare":
                    await self._show_compare_reply_keyboard(update, context)
                elif keyboard_type == "list":
                    # Для list клавиатуры не нужно показывать отдельное сообщение
                    # Клавиатура уже показывается в _show_namespace_symbols
                    pass
                else:
                    self.logger.warning(f"Unknown keyboard type: {keyboard_type}")
                    return
                
                # Обновляем контекст
                self._update_user_context(user_id, active_reply_keyboard=keyboard_type)
                self.logger.info(f"Active reply keyboard set to: {keyboard_type}")
            else:
                self.logger.info(f"Reply keyboard {keyboard_type} is already active")
                
        except Exception as e:
            self.logger.error(f"Error managing reply keyboard: {e}")

    async def _ensure_no_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Убедиться что reply keyboard скрыта (для команд которые не должны показывать клавиатуру)"""
        try:
            # Отправляем сообщение с ReplyKeyboardRemove для немедленного скрытия
            # Используем обычный пробел вместо невидимого символа
            await self._send_message_safe(
                update, 
                " ",  # Обычный пробел
                reply_markup=ReplyKeyboardRemove(),
                parse_mode=None
            )
            
            # Обновляем контекст пользователя
            user_id = update.effective_user.id
            self._update_user_context(user_id, active_reply_keyboard=None)
            self.logger.info("Reply keyboard removed using ReplyKeyboardRemove")
            
        except Exception as e:
            self.logger.error(f"Error removing reply keyboard: {e}")
            # Fallback к старому методу
            await self._manage_reply_keyboard(update, context, keyboard_type=None)

    async def _send_ephemeral_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, parse_mode: str = None, delete_after: int = 5, reply_markup=None):
        """Отправить исчезающее сообщение, которое удаляется через указанное время"""
        try:
            # Проверяем, что update и context не None
            if update is None or context is None:
                self.logger.error("Cannot send ephemeral message: update or context is None")
                return
            
            # Clean Markdown if parse_mode is Markdown
            if parse_mode == 'Markdown':
                text = self._safe_markdown(text)
            
            chat_id = None
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                chat_id = update.callback_query.message.chat_id
            elif hasattr(update, 'message') and update.message is not None:
                chat_id = update.message.chat_id
            else:
                self.logger.error("Cannot send ephemeral message: no chat_id available")
                return
            
            # Отправляем сообщение
            message = await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                parse_mode=parse_mode,
                reply_markup=reply_markup
            )
            
            # Планируем удаление сообщения через указанное время
            async def delete_message():
                try:
                    await asyncio.sleep(delete_after)
                    await context.bot.delete_message(chat_id=chat_id, message_id=message.message_id)
                except Exception as delete_error:
                    self.logger.warning(f"Could not delete ephemeral message: {delete_error}")
            
            # Запускаем удаление в фоне
            asyncio.create_task(delete_message())
            
        except Exception as e:
            self.logger.error(f"Error sending ephemeral message: {e}")
            # Fallback: отправляем обычное сообщение
            await self._send_callback_message(update, context, text, parse_mode)

    async def _send_callback_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, parse_mode: str = None, reply_markup=None):
        """Отправить сообщение в callback query - исправлено для обработки None и разбивки длинных сообщений"""
        try:
            # Проверяем, что update и context не None
            if update is None or context is None:
                self.logger.error("Cannot send message: update or context is None")
                return
            
            # Clean Markdown if parse_mode is Markdown
            if parse_mode == 'Markdown':
                text = self._safe_markdown(text)
            
            # Разбиваем длинные сообщения на части
            max_length = 4000  # Оставляем запас для безопасности
            if len(text) > max_length:
                self.logger.info(f"Splitting long message ({len(text)} chars) into multiple parts")
                await self._send_long_callback_message(update, context, text, parse_mode)
                return
            
            # Добавляем детальное логирование для отладки
            self.logger.info(f"_send_callback_message: hasattr callback_query: {hasattr(update, 'callback_query')}")
            self.logger.info(f"_send_callback_message: callback_query is not None: {update.callback_query is not None if hasattr(update, 'callback_query') else 'No attr'}")
            self.logger.info(f"_send_callback_message: hasattr message: {hasattr(update, 'message')}")
            self.logger.info(f"_send_callback_message: message is not None: {update.message is not None if hasattr(update, 'message') else 'No attr'}")
            
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                # Для callback query используем context.bot.send_message
                self.logger.info("_send_callback_message: Using callback_query path")
                try:
                    await context.bot.send_message(
                        chat_id=update.callback_query.message.chat_id,
                        text=text,
                        parse_mode=parse_mode,
                        reply_markup=reply_markup
                    )
                except Exception as callback_error:
                    self.logger.error(f"Error sending callback message: {callback_error}")
                    # Fallback: попробуем отправить через context.bot напрямую
                    try:
                        await context.bot.send_message(
                            chat_id=update.callback_query.message.chat_id,
                            text=f"❌ Ошибка отправки сообщения: {text[:500]}...",
                            reply_markup=reply_markup
                        )
                    except Exception as fallback_error:
                        self.logger.error(f"Fallback callback message sending also failed: {fallback_error}")
            elif hasattr(update, 'message') and update.message is not None:
                # Для обычных сообщений используем _send_message_safe
                self.logger.info("_send_callback_message: Using message path")
                await self._send_message_safe(update, text, reply_markup=reply_markup)
            else:
                # Если ни то, ни другое - логируем ошибку
                self.logger.error("_send_callback_message: Cannot send message - neither callback_query nor message available")
                self.logger.error(f"Update type: {type(update)}")
                self.logger.error(f"Update attributes: {dir(update) if update else 'None'}")
        except Exception as e:
            self.logger.error(f"Error sending callback message: {e}")
            # Fallback: попробуем отправить через context.bot
            try:
                if hasattr(update, 'callback_query') and update.callback_query is not None:
                    await context.bot.send_message(
                        chat_id=update.callback_query.message.chat_id,
                        text=f"❌ Ошибка отправки: {text[:500]}..."
                    )
            except Exception as fallback_error:
                self.logger.error(f"Fallback message sending also failed: {fallback_error}")

    async def _send_long_callback_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, parse_mode: str = None):
        """Отправить длинное сообщение по частям через callback query"""
        try:
            # Clean Markdown if parse_mode is Markdown
            if parse_mode == 'Markdown':
                text = self._safe_markdown(text)
            
            # Разбиваем текст на части
            parts = self._split_text_smart(text)
            
            for i, part in enumerate(parts):
                # Добавляем индикатор части для многочастных сообщений
                if len(parts) > 1:
                    part_text = f"📄 **Часть {i+1} из {len(parts)}:**\n\n{part}"
                else:
                    part_text = part
                
                # Clean Markdown for each part
                if parse_mode == 'Markdown':
                    part_text = self._safe_markdown(part_text)
                
                # Отправляем каждую часть
                if hasattr(update, 'callback_query') and update.callback_query is not None:
                    try:
                        await context.bot.send_message(
                            chat_id=update.callback_query.message.chat_id,
                            text=part_text,
                            parse_mode=parse_mode
                        )
                    except Exception as part_error:
                        self.logger.error(f"Error sending message part {i+1}: {part_error}")
                        # Fallback для этой части
                        await self._send_message_safe(update, part_text)
                elif hasattr(update, 'message') and update.message is not None:
                    await self._send_message_safe(update, part_text)
                
                # Небольшая пауза между частями для избежания rate limiting
                if i < len(parts) - 1:  # Не делаем паузу после последней части
                    import asyncio
                    await asyncio.sleep(0.5)
                    
        except Exception as e:
            self.logger.error(f"Error sending long callback message: {e}")
            # Fallback: отправляем обрезанную версию
            try:
                if hasattr(update, 'callback_query') and update.callback_query is not None:
                    await context.bot.send_message(
                        chat_id=update.callback_query.message.chat_id,
                        text=f"❌ Ошибка разбивки сообщения: {text[:1000]}..."
                    )
            except Exception as fallback_error:
                self.logger.error(f"Fallback long message sending also failed: {fallback_error}")

    def _split_text_smart(self, text: str) -> list:
        """Умное разбиение текста на части с учетом структуры"""
        max_length = 4000
        if len(text) <= max_length:
            return [text]
        
        parts = []
        current_part = ""
        
        # Разбиваем по строкам для лучшего сохранения структуры
        lines = text.split('\n')
        
        for line in lines:
            # Если добавление строки не превышает лимит
            if len(current_part) + len(line) + 1 <= max_length:
                if current_part:
                    current_part += '\n' + line
                else:
                    current_part = line
            else:
                # Если текущая часть не пустая, сохраняем её
                if current_part:
                    parts.append(current_part)
                    current_part = ""
                
                # Если одна строка слишком длинная, разбиваем её
                if len(line) > max_length:
                    # Разбиваем длинную строку по словам
                    words = line.split(' ')
                    temp_line = ""
                    for word in words:
                        if len(temp_line) + len(word) + 1 <= max_length:
                            if temp_line:
                                temp_line += ' ' + word
                            else:
                                temp_line = word
                        else:
                            if temp_line:
                                parts.append(temp_line)
                                temp_line = word
                            else:
                                # Если одно слово слишком длинное, обрезаем его
                                parts.append(word[:max_length])
                                temp_line = word[max_length:]
                    if temp_line:
                        current_part = temp_line
                else:
                    current_part = line
        
        # Добавляем последнюю часть
        if current_part:
            parts.append(current_part)
        
        return parts

    async def button_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle button callbacks for additional analysis"""
        query = update.callback_query
        await query.answer()
        
        self.logger.info(f"Button callback received: {query.data}")
        self.logger.info(f"Full callback query: {query}")
        
        try:
            # Parse callback data
            callback_data = query.data
            self.logger.info(f"Processing callback data: {callback_data}")
            
            
            # Handle start command callbacks
            if callback_data.startswith("start_"):
                self.logger.info(f"Processing start command callback: {callback_data}")
                if callback_data == "start_help":
                    # Execute help command
                    self.logger.info("Executing help command from callback")
                    await self.help_command(update, context)
                elif callback_data == "start_info":
                    # Execute info command without parameters
                    self.logger.info("Executing info command from callback")
                    context.args = []
                    await self.info_command(update, context)
                elif callback_data == "start_compare":
                    # Execute compare command without parameters
                    self.logger.info("Executing compare command from callback")
                    context.args = []
                    await self.compare_command(update, context)
                elif callback_data == "start_portfolio":
                    # Execute portfolio command without parameters
                    self.logger.info("Executing portfolio command from callback")
                    context.args = []
                    await self.portfolio_command(update, context)
                elif callback_data == "start_list":
                    # Execute list command without parameters
                    self.logger.info("Executing list command from callback")
                    context.args = []
                    await self.namespace_command(update, context)
                elif callback_data.startswith("start_info_"):
                    # Extract symbol and execute info command
                    symbol = callback_data.replace("start_info_", "")
                    context.args = [symbol]
                    await self.info_command(update, context)
                elif callback_data.startswith("start_compare_"):
                    # Extract symbols and execute compare command
                    symbols_str = callback_data.replace("start_compare_", "")
                    symbols = symbols_str.split("_")
                    context.args = symbols
                    await self.compare_command(update, context)
                elif callback_data.startswith("start_portfolio_"):
                    # Extract portfolio data and execute portfolio command
                    portfolio_str = callback_data.replace("start_portfolio_", "")
                    portfolio_parts = portfolio_str.split("_")
                    # Convert dashes back to colons for portfolio weights
                    portfolio_args = []
                    for part in portfolio_parts:
                        if "-" in part:
                            portfolio_args.append(part.replace("-", ":"))
                        else:
                            portfolio_args.append(part)
                    context.args = portfolio_args
                    await self.portfolio_command(update, context)
                return
            
            # Handle asset selection callbacks
            if callback_data.startswith("select_asset_"):
                # Extract symbol and query from callback data
                parts = callback_data.replace("select_asset_", "").split("_")
                if len(parts) >= 2:
                    symbol = parts[0]
                    query = "_".join(parts[1:])  # Reconstruct query in case it contains underscores
                    
                    # Update user context with selected asset
                    user_id = update.effective_user.id
                    user_context = self._get_user_context(user_id)
                    self._update_user_context(user_id, 
                                            last_assets=[symbol] + user_context.get('last_assets', []))
                    
                    # Add to analyzed tickers history
                    self._add_to_analyzed_tickers(user_id, symbol)
                    
                    # Execute info command with selected symbol
                    context.args = [symbol]
                    await self.info_command(update, context)
                return
            
            # Handle cancel selection callbacks
            if callback_data.startswith("cancel_selection_"):
                query_text = callback_data.replace("cancel_selection_", "")
                await query.edit_message_text(f"❌ Выбор актива отменен для запроса '{query_text}'")
                return
            
            if callback_data == "drawdowns" or callback_data == "drawdowns_compare" or callback_data == "compare_drawdowns":
                self.logger.info("Drawdowns button clicked")
                
                # Get data from user context
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                last_analysis_type = user_context.get('last_analysis_type')
                
                self.logger.info(f"Last analysis type: {last_analysis_type}")
                
                if last_analysis_type == 'portfolio':
                    symbols = user_context.get('current_symbols', [])
                    await self._handle_portfolio_drawdowns_button(update, context, symbols)
                else:
                    symbols = user_context.get('current_symbols', [])
                    await self._handle_drawdowns_button(update, context, symbols)
            elif callback_data == "dividends" or callback_data == "dividends_compare" or callback_data == "compare_dividends":
                self.logger.info("Dividends button clicked")
                
                # Get data from user context
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                symbols = user_context.get('current_symbols', [])
                await self._handle_dividends_button(update, context, symbols)
            elif callback_data == "correlation" or callback_data == "correlation_compare" or callback_data == "compare_correlation":
                self.logger.info("Correlation button clicked")
                
                # Get data from user context
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                symbols = user_context.get('current_symbols', [])
                await self._handle_correlation_button(update, context, symbols)
            elif callback_data == "risk_metrics" or callback_data == "risk_metrics_compare" or callback_data == "compare_risk_metrics":
                self.logger.info("Risk metrics button clicked")
                
                # Get data from user context
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                symbols = user_context.get('current_symbols', [])
                await self._handle_risk_metrics_button(update, context, symbols)
            elif callback_data == "monte_carlo" or callback_data == "monte_carlo_compare" or callback_data == "compare_monte_carlo":
                self.logger.info("Monte Carlo button clicked")
                
                # Get data from user context
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                symbols = user_context.get('current_symbols', [])
                await self._handle_monte_carlo_button(update, context, symbols)
            elif callback_data == "forecast" or callback_data == "forecast_compare" or callback_data == "compare_forecast":
                self.logger.info("Forecast button clicked")
                
                # Get data from user context
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                symbols = user_context.get('current_symbols', [])
                await self._handle_forecast_button(update, context, symbols)
            elif callback_data.startswith('info_daily_chart_'):
                symbol = self.clean_symbol(callback_data.replace('info_daily_chart_', ''))
                self.logger.info(f"Info daily chart button clicked for symbol: {symbol}")
                await self._handle_daily_chart_button(update, context, symbol)
            elif callback_data.startswith('daily_chart_'):
                symbol = self.clean_symbol(callback_data.replace('daily_chart_', ''))
                self.logger.info(f"Daily chart button clicked for symbol: {symbol}")
                await self._handle_daily_chart_button(update, context, symbol)
            elif callback_data.startswith('info_monthly_chart_'):
                symbol = self.clean_symbol(callback_data.replace('info_monthly_chart_', ''))
                self.logger.info(f"Info monthly chart button clicked for symbol: {symbol}")
                await self._handle_monthly_chart_button(update, context, symbol)
            elif callback_data.startswith('monthly_chart_'):
                symbol = self.clean_symbol(callback_data.replace('monthly_chart_', ''))
                self.logger.info(f"Monthly chart button clicked for symbol: {symbol}")
                await self._handle_monthly_chart_button(update, context, symbol)
            elif callback_data.startswith('all_chart_'):
                symbol = self.clean_symbol(callback_data.replace('all_chart_', ''))
                self.logger.info(f"All chart button clicked for symbol: {symbol}")
                await self._handle_all_chart_button(update, context, symbol)
            elif callback_data.startswith('info_period_'):
                # Handle period switching for info command
                parts = callback_data.replace('info_period_', '').split('_')
                if len(parts) >= 2:
                    symbol = self.clean_symbol(parts[0])
                    period = '_'.join(parts[1:])
                    self.logger.info(f"Info period button clicked for symbol: {symbol}, period: {period}")
                    await self._handle_info_period_button(update, context, symbol, period)
            elif callback_data.startswith('info_risks_'):
                symbol = self.clean_symbol(callback_data.replace('info_risks_', ''))
                self.logger.info(f"Info risks button clicked for symbol: {symbol}")
                await self._handle_info_risks_button(update, context, symbol)
            elif callback_data.startswith('info_metrics_'):
                symbol = self.clean_symbol(callback_data.replace('info_metrics_', ''))
                self.logger.info(f"Info metrics button clicked for symbol: {symbol}")
                await self._handle_info_metrics_button(update, context, symbol)
            elif callback_data.startswith('info_ai_analysis_'):
                symbol = self.clean_symbol(callback_data.replace('info_ai_analysis_', ''))
                self.logger.info(f"Info AI analysis button clicked for symbol: {symbol}")
                await self._handle_info_ai_analysis_button(update, context, symbol)
            elif callback_data.startswith('info_compare_'):
                symbol = self.clean_symbol(callback_data.replace('info_compare_', ''))
                self.logger.info(f"Info compare button clicked for symbol: {symbol}")
                await self._handle_info_compare_button(update, context, symbol)
            elif callback_data.startswith('info_portfolio_'):
                symbol = self.clean_symbol(callback_data.replace('info_portfolio_', ''))
                self.logger.info(f"Info portfolio button clicked for symbol: {symbol}")
                await self._handle_info_portfolio_button(update, context, symbol)
            elif callback_data.startswith('info_dividends_'):
                symbol = self.clean_symbol(callback_data.replace('info_dividends_', ''))
                self.logger.info(f"Info dividends button clicked for symbol: {symbol}")
                await self._handle_single_dividends_button(update, context, symbol)
            elif callback_data.startswith('dividends_') and ',' not in callback_data:
                # Для одиночного актива (dividends_AAA)
                symbol = self.clean_symbol(callback_data.replace('dividends_', ''))
                self.logger.info(f"Dividends button clicked for symbol: {symbol}")
                await self._handle_single_dividends_button(update, context, symbol)
            elif callback_data.startswith('tushare_daily_chart_'):
                symbol = self.clean_symbol(callback_data.replace('tushare_daily_chart_', ''))
                self.logger.info(f"Tushare daily chart button clicked for symbol: {symbol}")
                await self._handle_tushare_daily_chart_button(update, context, symbol)
            elif callback_data.startswith('tushare_monthly_chart_'):
                symbol = self.clean_symbol(callback_data.replace('tushare_monthly_chart_', ''))
                self.logger.info(f"Tushare monthly chart button clicked for symbol: {symbol}")
                await self._handle_tushare_monthly_chart_button(update, context, symbol)
            elif callback_data.startswith('tushare_all_chart_'):
                symbol = self.clean_symbol(callback_data.replace('tushare_all_chart_', ''))
                self.logger.info(f"Tushare all chart button clicked for symbol: {symbol}")
                await self._handle_tushare_all_chart_button(update, context, symbol)
            elif callback_data.startswith('tushare_dividends_'):
                symbol = self.clean_symbol(callback_data.replace('tushare_dividends_', ''))
                self.logger.info(f"Tushare dividends button clicked for symbol: {symbol}")
                await self._handle_tushare_dividends_button(update, context, symbol)
            elif callback_data.startswith('portfolio_risk_metrics_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_risk_metrics_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio risk metrics button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_risk_metrics_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('risk_metrics_'):
                symbols = [self.clean_symbol(s) for s in callback_data.replace('risk_metrics_', '').split(',')]
                self.logger.info(f"Risk metrics button clicked for symbols: {symbols}")
                await self._handle_risk_metrics_button(update, context, symbols)
            elif callback_data.startswith('portfolio_monte_carlo_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_monte_carlo_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio monte carlo button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_monte_carlo_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('monte_carlo_'):
                symbols = [self.clean_symbol(s) for s in callback_data.replace('monte_carlo_', '').split(',')]
                self.logger.info(f"Monte Carlo button clicked for symbols: {symbols}")
                await self._handle_monte_carlo_button(update, context, symbols)
            elif callback_data.startswith('portfolio_forecast_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_forecast_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio forecast button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_forecast_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('forecast_'):
                symbols = [self.clean_symbol(s) for s in callback_data.replace('forecast_', '').split(',')]
                self.logger.info(f"Forecast button clicked for symbols: {symbols}")
                await self._handle_forecast_button(update, context, symbols)

            elif callback_data.startswith('portfolio_wealth_chart_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_wealth_chart_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio wealth chart button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_wealth_chart_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('wealth_chart_'):
                symbols = [self.clean_symbol(s) for s in callback_data.replace('wealth_chart_', '').split(',')]
                self.logger.info(f"Wealth chart button clicked for symbols: {symbols}")
                self.logger.info(f"Callback data: '{callback_data}'")
                self.logger.info(f"Parsed symbols: {[f'{s}' for s in symbols]}")
                self.logger.info(f"Symbol types: {[type(s) for s in symbols]}")
                self.logger.info(f"Symbol lengths: {[len(str(s)) if s else 'None' for s in symbols]}")
                await self._handle_portfolio_wealth_chart_button(update, context, symbols)
            elif callback_data.startswith('portfolio_returns_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_returns_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio returns button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_returns_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('returns_'):
                symbols = [self.clean_symbol(s) for s in callback_data.replace('returns_', '').split(',')]
                self.logger.info(f"Returns button clicked for symbols: {symbols}")
                await self._handle_portfolio_returns_button(update, context, symbols)
            elif callback_data.startswith('portfolio_rolling_cagr_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_rolling_cagr_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio rolling CAGR button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_rolling_cagr_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('rolling_cagr_'):
                symbols = [self.clean_symbol(s) for s in callback_data.replace('rolling_cagr_', '').split(',')]
                self.logger.info(f"Rolling CAGR button clicked for symbols: {symbols}")
                await self._handle_portfolio_rolling_cagr_button(update, context, symbols)
            elif callback_data.startswith('portfolio_compare_assets_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_compare_assets_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio compare assets button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_compare_assets_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('portfolio_ai_analysis_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_ai_analysis_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio AI analysis button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_ai_analysis_button(update, context, portfolio_symbol)
            elif callback_data.startswith('portfolio_compare_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_compare_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio compare button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_compare_button(update, context, portfolio_symbol)
            elif callback_data.startswith('portfolio_drawdowns_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_drawdowns_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio drawdowns button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_drawdowns_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('portfolio_dividends_'):
                portfolio_symbol_raw = callback_data.replace('portfolio_dividends_', '')
                # Don't apply clean_symbol to portfolio symbols that contain commas (okama portfolio symbols)
                if ',' in portfolio_symbol_raw:
                    portfolio_symbol = portfolio_symbol_raw
                else:
                    portfolio_symbol = self.clean_symbol(portfolio_symbol_raw)
                self.logger.info(f"Portfolio dividends button clicked for portfolio: {portfolio_symbol}")
                await self._handle_portfolio_dividends_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith('compare_assets_'):
                symbols = [self.clean_symbol(s) for s in callback_data.replace('compare_assets_', '').split(',')]
                self.logger.info(f"Compare assets button clicked for symbols: {symbols}")
                await self._handle_portfolio_compare_assets_button(update, context, symbols)
            elif callback_data == 'compare_risk_return':
                self.logger.info("Compare Risk / Return button clicked")
                await self._handle_risk_return_compare_button(update, context)
            elif callback_data == 'risk_return_compare':
                self.logger.info("Risk / Return button clicked")
                await self._handle_risk_return_compare_button(update, context)
            elif callback_data == 'data_analysis_compare':
                self.logger.info("Data analysis button clicked")
                await self._handle_data_analysis_compare_button(update, context)
            elif callback_data == 'yandexgpt_analysis_compare':
                self.logger.info("YandexGPT analysis button clicked")
                await self._handle_yandexgpt_analysis_compare_button(update, context)
            elif callback_data == 'metrics_compare':
                self.logger.info("Metrics button clicked")
                await self._handle_metrics_compare_button(update, context)
            elif callback_data == 'efficient_frontier_compare':
                self.logger.info("Efficient Frontier button clicked")
                await self._handle_efficient_frontier_compare_button(update, context)
            elif callback_data == 'compare_portfolio':
                # Get symbols from user context instead of callback_data to avoid size limit
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                symbols = user_context.get('compare_portfolio_symbols', [])
                self.logger.info(f"Compare portfolio button clicked for symbols: {symbols}")
                await self._handle_compare_portfolio_button(update, context, symbols)
            elif callback_data == 'namespace_home':
                self.logger.info("Namespace home button clicked")
                await self._handle_namespace_home_button(update, context)
            elif callback_data.startswith('namespace_') and callback_data not in ['namespace_analysis', 'namespace_compare', 'namespace_portfolio', 'namespace_home']:
                namespace = self.clean_symbol(callback_data.replace('namespace_', ''))
                self.logger.info(f"Namespace button clicked for: {namespace}")
                await self._handle_namespace_button(update, context, namespace)
            elif callback_data.startswith('excel_namespace_'):
                namespace = self.clean_symbol(callback_data.replace('excel_namespace_', ''))
                self.logger.info(f"Excel namespace button clicked for: {namespace}")
                await self._handle_excel_namespace_button(update, context, namespace)
            elif callback_data == 'namespace_analysis':
                self.logger.info("Namespace analysis button clicked")
                await self._handle_namespace_analysis_button(update, context)
            elif callback_data == 'namespace_compare':
                self.logger.info("Namespace compare button clicked")
                await self._handle_namespace_compare_button(update, context)
            elif callback_data == 'namespace_portfolio':
                self.logger.info("Namespace portfolio button clicked")
                await self._handle_namespace_portfolio_button(update, context)
            elif callback_data.startswith('nav_namespace_'):
                # Handle navigation for okama namespaces
                parts = callback_data.replace('nav_namespace_', '').split('_')
                if len(parts) >= 2:
                    namespace = self.clean_symbol(parts[0])
                    page = int(parts[1])
                    self.logger.info(f"Navigation button clicked for namespace: {namespace}, page: {page}")
                    await self._show_namespace_symbols(update, context, namespace, is_callback=True, page=page)
            elif callback_data.startswith('nav_tushare_'):
                # Handle navigation for tushare namespaces
                parts = callback_data.replace('nav_tushare_', '').split('_')
                if len(parts) >= 2:
                    namespace = self.clean_symbol(parts[0])
                    page = int(parts[1])
                    self.logger.info(f"Navigation button clicked for tushare namespace: {namespace}, page: {page}")
                    await self._show_tushare_namespace_symbols(update, context, namespace, is_callback=True, page=page)
            elif callback_data == 'noop':
                # Handle page number buttons - do nothing
                self.logger.info("Page number button clicked - no action needed")
                return
            else:
                self.logger.warning(f"Unknown button callback: {callback_data}")
                await self._send_callback_message(update, context, "❌ Неизвестная кнопка")
                
        except Exception as e:
            self.logger.error(f"Error in button callback: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_risk_return_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle Risk / Return (CAGR) button for all comparison types"""
        try:
            # Don't remove keyboard yet - wait for successful message creation
            
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            symbols = user_context.get('current_symbols', [])
            display_symbols = user_context.get('display_symbols', symbols)  # Use descriptive names for display
            currency = user_context.get('current_currency', 'USD')
            period = user_context.get('current_period', None)
            expanded_symbols = user_context.get('expanded_symbols', [])
            portfolio_contexts = user_context.get('portfolio_contexts', [])

            # Validate that we have symbols to compare
            if not expanded_symbols:
                await self._send_callback_message(update, context, "ℹ️ Нет данных для сравнения. Выполните команду /compare заново.")
                return

            await self._send_ephemeral_message(update, context, "📊 Создаю график Risk / Return (CAGR)…", delete_after=3)

            # Prepare assets for comparison
            asset_list_items = []
            asset_names = []
            
            # Use display_symbols for proper naming - they already contain descriptive names
            for i, symbol in enumerate(symbols):
                if i < len(expanded_symbols):
                    if isinstance(expanded_symbols[i], (pd.Series, pd.DataFrame)):
                        # This is a portfolio - recreate it
                        if i < len(portfolio_contexts):
                            pctx = portfolio_contexts[i]
                            try:
                                p = ok.Portfolio(
                                    pctx.get('portfolio_symbols', []),
                                    weights=pctx.get('portfolio_weights', []),
                                    ccy=pctx.get('portfolio_currency') or currency,
                                )
                                asset_list_items.append(p)
                                asset_names.append(display_symbols[i])  # Use descriptive name
                            except Exception as pe:
                                self.logger.warning(f"Failed to recreate portfolio for Risk/Return: {pe}")
                    else:
                        # This is a regular asset
                        asset_list_items.append(symbol)
                        asset_names.append(display_symbols[i])  # Use descriptive name

            if not asset_list_items:
                await self._send_callback_message(update, context, "❌ Не удалось подготовить активы для построения графика")
                return

            # Create AssetList with selected assets/portfolios
            img_buffer = None
            try:
                asset_list = ok.AssetList(asset_list_items, ccy=currency)
                
                # okama plotting
                asset_list.plot_assets(kind="cagr")
                current_fig = plt.gcf()
                
                # Apply styling
                if current_fig.axes:
                    ax = current_fig.axes[0]
                    chart_styles.apply_styling(
                        ax,
                        title=f"Risk / Return: CAGR\n{', '.join(asset_names)}",
                        ylabel='CAGR (%)',
                        grid=True,
                        legend=True,
                        copyright=True
                    )
                img_buffer = io.BytesIO()
                chart_styles.save_figure(current_fig, img_buffer)
                img_buffer.seek(0)
                chart_styles.cleanup_figure(current_fig)
            except Exception as plot_error:
                # Fallback: compute CAGR manually and plot as bar chart
                self.logger.warning(f"Risk/Return okama plot failed, falling back to manual bar: {plot_error}")
                try:
                    cagr_values = {}
                    
                    # Calculate CAGR for each asset
                    for i, asset in enumerate(asset_list_items):
                        asset_name = asset_names[i]
                        try:
                            if isinstance(asset, str):
                                # Individual asset
                                asset_obj = ok.Asset(asset)
                                cagr = asset_obj.get_cagr()
                            else:
                                # Portfolio
                                cagr = asset.get_cagr()
                            
                            if hasattr(cagr, 'iloc'):
                                cagr_val = float(cagr.iloc[0])
                            elif hasattr(cagr, '__iter__') and not isinstance(cagr, str):
                                cagr_val = float(list(cagr)[0])
                            else:
                                cagr_val = float(cagr)
                        except Exception:
                            # Manual CAGR calculation
                            try:
                                if isinstance(asset, str):
                                    asset_obj = ok.Asset(asset)
                                    wealth_index = asset_obj.wealth_index
                                else:
                                    wealth_index = asset.wealth_index
                                
                                wi = wealth_index.dropna()
                                periods = len(wi)
                                cagr_val = ((wi.iloc[-1] / wi.iloc[0]) ** (12.0 / max(periods, 1))) - 1 if periods > 1 else 0.0
                            except Exception:
                                cagr_val = 0.0
                        
                        cagr_values[asset_name] = cagr_val

                    cagr_df = pd.DataFrame.from_dict(cagr_values, orient='index')
                    cagr_df.columns = ['CAGR']
                    fig, ax = chart_styles.create_bar_chart(
                        cagr_df['CAGR'],
                        title=f"Risk / Return: CAGR\n{', '.join(asset_names)}",
                        ylabel='CAGR (%)'
                    )
                    img_buffer = io.BytesIO()
                    chart_styles.save_figure(fig, img_buffer)
                    img_buffer.seek(0)
                    chart_styles.cleanup_figure(fig)
                except Exception as fallback_error:
                    self.logger.error(f"Risk/Return manual bar failed: {fallback_error}")
                    await self._send_callback_message(update, context, "❌ Не удалось построить график Risk / Return (CAGR)")
                    return

            # Create keyboard for compare command
            keyboard = self._create_compare_command_keyboard(symbols, currency, update)
            
            # Remove keyboard from previous message before sending new message
            await self._remove_keyboard_before_new_message(update, context)
            
            # Send image with keyboard
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(f"⚖️ Risk / Return (CAGR) для сравнения: {', '.join(asset_names)}"),
                reply_markup=keyboard
            )

        except Exception as e:
            self.logger.error(f"Error handling Risk / Return button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при построении Risk / Return: {str(e)}")


    async def _handle_efficient_frontier_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle Efficient Frontier button for all comparison types"""
        try:
            # Don't remove keyboard yet - wait for successful message creation
            
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            symbols = user_context.get('current_symbols', [])
            display_symbols = user_context.get('display_symbols', symbols)  # Use descriptive names for display
            currency = user_context.get('current_currency', 'USD')
            period = user_context.get('current_period', None)
            expanded_symbols = user_context.get('expanded_symbols', [])
            portfolio_contexts = user_context.get('portfolio_contexts', [])

            # Validate that we have symbols to compare
            if not expanded_symbols:
                await self._send_callback_message(update, context, "ℹ️ Нет данных для сравнения. Выполните команду /compare заново.")
                return

            await self._send_ephemeral_message(update, context, "📈 Создаю график эффективной границы…", delete_after=3)

            # Prepare assets for comparison
            asset_list_items = []
            asset_names = []
            
            # Use display_symbols for proper naming - they already contain descriptive names
            for i, symbol in enumerate(symbols):
                if i < len(expanded_symbols):
                    if isinstance(expanded_symbols[i], (pd.Series, pd.DataFrame)):
                        # This is a portfolio - recreate it
                        if i < len(portfolio_contexts):
                            pctx = portfolio_contexts[i]
                            try:
                                p = ok.Portfolio(
                                    pctx.get('portfolio_symbols', []),
                                    weights=pctx.get('portfolio_weights', []),
                                    ccy=pctx.get('portfolio_currency') or currency,
                                )
                                asset_list_items.append(p)
                                asset_names.append(display_symbols[i])  # Use descriptive name
                            except Exception as pe:
                                self.logger.warning(f"Failed to recreate portfolio for Efficient Frontier: {pe}")
                    else:
                        # This is a regular asset
                        asset_list_items.append(symbol)
                        asset_names.append(display_symbols[i])  # Use descriptive name

            if not asset_list_items:
                await self._send_callback_message(update, context, "❌ Не удалось подготовить активы для построения графика")
                return

            # Create AssetList with selected assets/portfolios
            img_buffer = None
            try:
                asset_list = ok.AssetList(asset_list_items, ccy=currency)
                
                # Create Efficient Frontier
                ef = ok.EfficientFrontier(asset_list, ccy=currency)
                
                # Log debug information
                self.logger.info(f"Created EfficientFrontier with {len(asset_names)} assets: {asset_names}")
                
                # Create chart with proper styling using chart_styles
                current_fig, ax = chart_styles.create_efficient_frontier_chart(
                    ef, 
                    asset_names, 
                    data_source='okama'
                )
                
                # Check if chart creation was successful
                if current_fig is None:
                    raise Exception("Failed to create efficient frontier chart - chart creation returned None")
                
                img_buffer = io.BytesIO()
                chart_styles.save_figure(current_fig, img_buffer)
                img_buffer.seek(0)
                chart_styles.cleanup_figure(current_fig)
            except Exception as plot_error:
                self.logger.error(f"Efficient Frontier plot failed: {plot_error}")
                await self._send_callback_message(update, context, f"❌ Не удалось построить график эффективной границы: {str(plot_error)}")
                return

            # Send image without keyboard
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(f"📈 Эффективная граница для сравнения: {', '.join(asset_names)}")
            )

        except Exception as e:
            self.logger.error(f"Error handling Efficient Frontier button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при построении эффективной границы: {str(e)}")

    async def _handle_compare_portfolio_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle portfolio button for compare command - add compared assets to portfolio"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            # Get user context to check for portfolio contexts
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            portfolio_contexts = user_context.get('portfolio_contexts', [])
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            # Filter symbols to only include regular assets (not portfolios)
            # If there are both portfolios and regular assets in context, only offer regular assets
            regular_assets = []
            portfolio_symbols = []
            
            for i, symbol in enumerate(symbols):
                # Check if this symbol corresponds to a portfolio in the context
                is_portfolio = False
                if i < len(expanded_symbols) and isinstance(expanded_symbols[i], (pd.Series, pd.DataFrame)):
                    # This is a portfolio
                    is_portfolio = True
                    portfolio_symbols.append(symbol)
                elif i < len(portfolio_contexts):
                    # Check if this symbol has portfolio context
                    portfolio_context = portfolio_contexts[i]
                    if len(portfolio_context.get('portfolio_symbols', [])) > 1:
                        # This is a portfolio (has multiple symbols)
                        is_portfolio = True
                        portfolio_symbols.append(symbol)
                
                if not is_portfolio:
                    regular_assets.append(symbol)
            
            # If we have both portfolios and regular assets, only offer regular assets
            if portfolio_symbols and regular_assets:
                symbols_to_use = regular_assets
                portfolio_text = f"💼 **Добавить активы в портфель**\n\n"
                portfolio_text += f"Активы для добавления: `{' '.join(symbols_to_use)}`\n\n"
            else:
                # Use all symbols if no mixed context
                symbols_to_use = symbols
                portfolio_text = f"💼 **Добавить активы в портфель**\n\n"
                portfolio_text += f"Активы для добавления: `{' '.join(symbols_to_use)}`\n\n"
            
            # Set user context to wait for portfolio weights input
            self._update_user_context(user_id, 
                waiting_for_portfolio_weights=True,
                portfolio_base_symbols=symbols_to_use
            )
            
            portfolio_text += "**Примеры:**\n"
            if len(symbols_to_use) >= 2:
                portfolio_text += f"• `{symbols_to_use[0]}:0.6 {symbols_to_use[1]}:0.4`\n"
            if len(symbols_to_use) >= 3:
                portfolio_text += f"• `{symbols_to_use[0]}:0.5 {symbols_to_use[1]}:0.3 {symbols_to_use[2]}:0.2`\n"
            if len(symbols_to_use) >= 4:
                portfolio_text += f"• `{symbols_to_use[0]}:0.4 {symbols_to_use[1]}:0.3 {symbols_to_use[2]}:0.2 {symbols_to_use[3]}:0.1`\n"
            if len(symbols_to_use) >= 5:
                portfolio_text += f"• `{symbols_to_use[0]}:0.3 {symbols_to_use[1]}:0.25 {symbols_to_use[2]}:0.2 {symbols_to_use[3]}:0.15 {symbols_to_use[4]}:0.1`\n"
            if len(symbols_to_use) >= 6:
                portfolio_text += f"• `{symbols_to_use[0]}:0.25 {symbols_to_use[1]}:0.2 {symbols_to_use[2]}:0.15 {symbols_to_use[3]}:0.15 {symbols_to_use[4]}:0.15 {symbols_to_use[5]}:0.1`\n"
            portfolio_text += "\n"
            
            portfolio_text += "💡 Сумма долей должна равняться 1.0 (100%)\n"
            portfolio_text += "💡 Можно добавить дополнительные активы к сравниваемым\n\n"
            portfolio_text += "💬 Введите тикеры для включения в состав портфеля:"
            
            await self._send_callback_message(update, context, portfolio_text, parse_mode='Markdown')
            
        except Exception as e:
            self.logger.error(f"Error handling compare portfolio button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при подготовке портфеля: {str(e)}", parse_mode='Markdown')

    async def _handle_data_analysis_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle data analysis button click for comparison charts"""
        # Initialize variables at the beginning to ensure they're available in except block
        symbols = []
        currency = 'USD'
        
        try:
            # Don't remove keyboard yet - wait for successful message creation
            
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            symbols = user_context.get('current_symbols', [])
            currency = user_context.get('current_currency', 'USD')
            expanded_symbols = user_context.get('expanded_symbols', [])
            portfolio_contexts = user_context.get('portfolio_contexts', [])

            # Validate that we have symbols to compare
            if not expanded_symbols:
                await self._send_callback_message(update, context, "ℹ️ Нет данных для сравнения. Выполните команду /compare заново.")
                return

            # Check if Gemini service is available
            if not self.gemini_service or not self.gemini_service.is_available():
                await self._send_callback_message(update, context, "❌ Сервис анализа данных недоступен.", parse_mode='Markdown')
                return

            await self._send_ephemeral_message(update, context, "Анализирую данные", parse_mode='Markdown', delete_after=3)

            # Prepare data for analysis
            try:
                data_info = await self._prepare_data_for_analysis(symbols, currency, expanded_symbols, portfolio_contexts, user_id)
                
                # Analyze data with Gemini
                data_analysis = self.gemini_service.analyze_data(data_info)
                
                if data_analysis and data_analysis.get('success'):
                    analysis_text = data_analysis.get('analysis', '')
                    
                    if analysis_text:
                        # Get asset names from data_info for display
                        asset_names = data_info.get('asset_names', {})
                        
                        # Create list with asset names if available
                        assets_with_names = []
                        for symbol in symbols:
                            if symbol in asset_names and asset_names[symbol] != symbol:
                                assets_with_names.append(f"{symbol} ({asset_names[symbol]})")
                            else:
                                assets_with_names.append(symbol)
                        


                        
                        # Create keyboard for compare command
                        keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                        await self._send_callback_message_with_keyboard_removal(update, context, analysis_text, parse_mode='Markdown', reply_markup=keyboard)
                    else:
                        # Create keyboard for compare command
                        keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                        await self._send_callback_message_with_keyboard_removal(update, context, "🤖 Анализ данных выполнен, но результат пуст", parse_mode='Markdown', reply_markup=keyboard)
                        
                else:
                    error_msg = data_analysis.get('error', 'Неизвестная ошибка') if data_analysis else 'Анализ не выполнен'
                    # Create keyboard for compare command
                    keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                    await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка анализа данных: {error_msg}", parse_mode='Markdown', reply_markup=keyboard)
                    
            except Exception as data_error:
                self.logger.error(f"Error preparing data for analysis: {data_error}")
                # Create keyboard for compare command
                keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка при подготовке данных для анализа: {str(data_error)}", parse_mode='Markdown', reply_markup=keyboard)

        except Exception as e:
            self.logger.error(f"Error handling data analysis button: {e}")
            # Create keyboard for compare command
            keyboard = self._create_compare_command_keyboard(symbols, currency, update)
            await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка при анализе данных: {str(e)}", parse_mode='Markdown', reply_markup=keyboard)

    async def _handle_yandexgpt_analysis_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle Gemini analysis button click for comparison charts"""
        # Initialize variables at the beginning to ensure they're available in except block
        symbols = []
        currency = 'USD'
        
        try:
            # Don't remove keyboard yet - wait for successful message creation
            
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            symbols = user_context.get('current_symbols', [])
            currency = user_context.get('current_currency', 'USD')
            expanded_symbols = user_context.get('expanded_symbols', [])
            portfolio_contexts = user_context.get('portfolio_contexts', [])

            # Validate that we have symbols to compare
            if not expanded_symbols:
                await self._send_callback_message(update, context, "ℹ️ Нет данных для сравнения. Выполните команду /compare заново.", parse_mode='Markdown')
                return

            # Check if Gemini service is available
            if not self.gemini_service or not self.gemini_service.is_available():
                await self._send_callback_message(update, context, "❌ Сервис анализа данных недоступен.", parse_mode='Markdown')
                return

            await self._send_ephemeral_message(update, context, "Анализирую данные...", parse_mode='Markdown', delete_after=3)

            # Prepare data for analysis
            try:
                data_info = await self._prepare_data_for_analysis(symbols, currency, expanded_symbols, portfolio_contexts, user_id)
                
                if data_info:
                    # Perform Gemini analysis
                    gemini_analysis = self.gemini_service.analyze_data(data_info)
                    
                    if gemini_analysis and gemini_analysis.get('success'):
                        analysis_text = gemini_analysis.get('analysis', '')
                        
                        if analysis_text:
                            # Get asset names from data_info for display
                            asset_names = data_info.get('asset_names', {})
                            
                            # Create list with asset names if available
                            assets_with_names = []
                            for symbol in symbols:
                                if symbol in asset_names and asset_names[symbol] != symbol:
                                    assets_with_names.append(f"{symbol} ({asset_names[symbol]})")
                                else:
                                    assets_with_names.append(symbol)
                            

                            
                            # Create keyboard for compare command
                            keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                            await self._send_callback_message_with_keyboard_removal(update, context, analysis_text, parse_mode='Markdown', reply_markup=keyboard)
                        else:
                            # Create keyboard for compare command
                            keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                            await self._send_callback_message_with_keyboard_removal(update, context, "🤖 Анализ данных выполнен, но результат пуст", parse_mode='Markdown', reply_markup=keyboard)
                    else:
                        error_msg = gemini_analysis.get('error', 'Неизвестная ошибка') if gemini_analysis else 'Анализ не выполнен'
                        # Create keyboard for compare command
                        keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                        await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка анализа данных: {error_msg}", parse_mode='Markdown', reply_markup=keyboard)
                    
            except Exception as data_error:
                self.logger.error(f"Error preparing data for Gemini analysis: {data_error}")
                # Create keyboard for compare command
                keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка при подготовке данных для анализа: {str(data_error)}", parse_mode='Markdown', reply_markup=keyboard)

        except Exception as e:
            self.logger.error(f"Error handling Gemini analysis button: {e}")
            # Create keyboard for compare command
            keyboard = self._create_compare_command_keyboard(symbols, currency, update)
            await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка при анализе данных: {str(e)}", parse_mode='Markdown', reply_markup=keyboard)

    async def _handle_metrics_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle metrics button click for comparison charts - show summary metrics table"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            symbols = user_context.get('current_symbols', [])
            currency = user_context.get('current_currency', 'USD')
            expanded_symbols = user_context.get('expanded_symbols', [])
            portfolio_contexts = user_context.get('portfolio_contexts', [])
            specified_period = user_context.get('specified_period')

            # Validate that we have symbols to compare
            if not expanded_symbols:
                await self._send_callback_message(update, context, "ℹ️ Нет данных для сравнения. Выполните команду /compare заново.")
                return

            await self._send_ephemeral_message(update, context, "📊 Подготавливаю таблицу метрик...", parse_mode='Markdown', delete_after=3)

            # Create summary metrics table
            try:
                summary_table = self._create_summary_metrics_table(
                    symbols, currency, expanded_symbols, portfolio_contexts, specified_period
                )
                
                if summary_table and not summary_table.startswith("❌"):
                    # Create keyboard for compare command
                    keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                    
                    # Send table as message with keyboard
                    period_text = self._format_period_for_display(specified_period)
                    header_text = f"📊 **Сводная таблица ключевых метрик**"
                    if period_text:
                        header_text += f" {period_text}"
                    table_message = f"{header_text}\n\n```\n{summary_table}\n```"
                    await self._send_callback_message_with_keyboard_removal(update, context, table_message, parse_mode='Markdown', reply_markup=keyboard)
                else:
                    # Create keyboard for compare command
                    keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                    await self._send_callback_message_with_keyboard_removal(update, context, "❌ Не удалось создать таблицу метрик", reply_markup=keyboard)
                    
            except Exception as metrics_error:
                self.logger.error(f"Error creating metrics table: {metrics_error}")
                # Create keyboard for compare command
                keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка при создании таблицы метрик: {str(metrics_error)}", parse_mode='Markdown', reply_markup=keyboard)

        except Exception as e:
            self.logger.error(f"Error handling metrics button: {e}")
            # Create keyboard for compare command
            keyboard = self._create_compare_command_keyboard(symbols, currency, update)
            await self._send_callback_message_with_keyboard_removal(update, context, f"❌ Ошибка при обработке кнопки метрик: {str(e)}", parse_mode='Markdown', reply_markup=keyboard)

    def _get_current_timestamp(self) -> str:
        """Get current timestamp as string"""
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _generate_portfolio_name(self, symbols: list, weights: list) -> str:
        """Generate a meaningful portfolio name based on symbols and weights"""
        try:
            # Clean symbol names (remove .US, .RU, etc.)
            clean_symbols = []
            for symbol in symbols:
                clean_symbol = symbol.split('.')[0] if '.' in symbol else symbol
                clean_symbols.append(clean_symbol)
            
            # If portfolio has 1-2 assets, use their names
            if len(clean_symbols) <= 2:
                if len(clean_symbols) == 1:
                    return f"Портфель {clean_symbols[0]}"
                else:
                    return f"Портфель {clean_symbols[0]} + {clean_symbols[1]}"
            
            # If portfolio has 3+ assets, use top 2 by weight
            elif len(clean_symbols) >= 3:
                # Sort by weights (descending)
                symbol_weight_pairs = list(zip(clean_symbols, weights))
                symbol_weight_pairs.sort(key=lambda x: x[1], reverse=True)
                
                top_symbols = [pair[0] for pair in symbol_weight_pairs[:2]]
                return f"Портфель {top_symbols[0]} + {top_symbols[1]} + {len(clean_symbols)-2} др."
            
            # Fallback
            return f"Портфель из {len(clean_symbols)} активов"
            
        except Exception as e:
            self.logger.warning(f"Could not generate portfolio name: {e}")
            return f"Портфель из {len(symbols)} активов"

    def _get_portfolio_basic_metrics(self, portfolio, symbols: list, weights: list, currency: str) -> str:
        """Get basic portfolio metrics for creation message"""
        try:
            self.logger.info(f"=== STARTING PORTFOLIO BASIC METRICS CALCULATION ===")
            self.logger.info(f"Getting portfolio basic metrics for symbols: {symbols}")
            self.logger.info(f"Portfolio type: {type(portfolio)}")
            self.logger.info(f"Portfolio attributes: {dir(portfolio)}")
            
            # Check if pandas is available
            try:
                import pandas as pd
                self.logger.info("Pandas import successful")
            except ImportError as e:
                self.logger.error(f"Pandas import failed: {e}")
                raise e
            
            # Get metrics directly from portfolio object
            self.logger.info("Getting metrics directly from portfolio object...")
            
            # Get CAGR
            cagr_value = None
            if hasattr(portfolio, 'get_cagr'):
                try:
                    cagr = portfolio.get_cagr()
                    self.logger.info(f"Got CAGR: {cagr}")
                    # Handle Series - get first value
                    if hasattr(cagr, 'iloc'):
                        cagr_value = cagr.iloc[0]
                    elif hasattr(cagr, '__getitem__'):
                        cagr_value = cagr[0]
                    else:
                        cagr_value = cagr
                    self.logger.info(f"Extracted CAGR value: {cagr_value}")
                except Exception as e:
                    self.logger.warning(f"Could not get CAGR: {e}")
            
            # Get volatility (risk_annual)
            volatility_value = None
            if hasattr(portfolio, 'risk_annual'):
                try:
                    risk_annual = portfolio.risk_annual
                    self.logger.info(f"Got risk_annual: {risk_annual}")
                    # Handle Series - get last value (most recent)
                    if hasattr(risk_annual, 'iloc'):
                        volatility_value = risk_annual.iloc[-1]
                    elif hasattr(risk_annual, '__getitem__'):
                        volatility_value = risk_annual[-1]
                    else:
                        volatility_value = risk_annual
                    self.logger.info(f"Extracted volatility value: {volatility_value}")
                except Exception as e:
                    self.logger.warning(f"Could not get volatility: {e}")
            
            # Get Sharpe ratio
            sharpe_value = None
            if hasattr(portfolio, 'get_sharpe_ratio'):
                try:
                    sharpe = portfolio.get_sharpe_ratio()
                    self.logger.info(f"Got Sharpe ratio: {sharpe}")
                    # Handle Series if needed
                    if hasattr(sharpe, 'iloc'):
                        sharpe_value = sharpe.iloc[0]
                    elif hasattr(sharpe, '__getitem__'):
                        sharpe_value = sharpe[0]
                    else:
                        sharpe_value = sharpe
                    self.logger.info(f"Extracted Sharpe ratio value: {sharpe_value}")
                except Exception as e:
                    self.logger.warning(f"Could not get Sharpe ratio: {e}")
            
            # If Sharpe ratio is still None, try manual calculation
            if sharpe_value is None and cagr_value is not None and volatility_value is not None and volatility_value > 0:
                try:
                    # Use unified Sharpe ratio calculation
                    sharpe_value = self.calculate_sharpe_ratio(cagr_value, volatility_value, currency)
                    self.logger.info(f"Calculated Sharpe ratio using unified function: {sharpe_value}")
                except Exception as e:
                    self.logger.warning(f"Could not calculate Sharpe ratio manually: {e}")
            
            # Get max drawdown from wealth_index
            max_drawdown_value = None
            if hasattr(portfolio, 'wealth_index'):
                try:
                    wealth_index = portfolio.wealth_index
                    self.logger.info(f"Got wealth_index: {len(wealth_index)} data points")
                    
                    # Calculate max drawdown from wealth_index
                    if hasattr(wealth_index, 'iloc'):
                        # If it's a DataFrame, get the first column (portfolio value)
                        if hasattr(wealth_index, 'columns'):
                            portfolio_values = wealth_index.iloc[:, 0]
                        else:
                            portfolio_values = wealth_index
                    else:
                        portfolio_values = wealth_index
                    
                    running_max = portfolio_values.expanding().max()
                    drawdown = (portfolio_values - running_max) / running_max
                    max_drawdown_value = drawdown.min()
                    self.logger.info(f"Calculated max drawdown: {max_drawdown_value:.4f}")
                except Exception as e:
                    self.logger.warning(f"Could not calculate max drawdown: {e}")
            
            # Build symbols with weights
            symbols_with_weights = []
            for i, symbol in enumerate(symbols):
                symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                weight = weights[i] if i < len(weights) else 0.0
                symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
            
            # Get rebalancing period - determine based on portfolio data
            rebalancing_period = "Ежегодно"  # Default fallback
            try:
                # Try to get actual rebalancing period from portfolio
                if hasattr(portfolio, 'rebalancing_period'):
                    rebalancing_period = portfolio.rebalancing_period
                elif hasattr(portfolio, 'rebalance_freq'):
                    rebalancing_period = portfolio.rebalance_freq
                # If no specific rebalancing period is set, determine from data frequency
                elif hasattr(portfolio, 'wealth_index') and portfolio.wealth_index is not None:
                    # Check data frequency to determine rebalancing
                    if hasattr(portfolio.wealth_index, 'index'):
                        freq = portfolio.wealth_index.index.freq
                        if freq:
                            if 'D' in str(freq) or 'day' in str(freq).lower():
                                rebalancing_period = "Ежедневно"
                            elif 'W' in str(freq) or 'week' in str(freq).lower():
                                rebalancing_period = "Еженедельно"
                            elif 'M' in str(freq) or 'month' in str(freq).lower():
                                rebalancing_period = "Ежемесячно"
                            elif 'Q' in str(freq) or 'quarter' in str(freq).lower():
                                rebalancing_period = "Ежеквартально"
                            elif 'Y' in str(freq) or 'year' in str(freq).lower():
                                rebalancing_period = "Ежегодно"
            except Exception as e:
                self.logger.warning(f"Could not determine rebalancing period: {e}")
                rebalancing_period = "Ежегодно"
            
            # Get period info with duration calculation
            period_info = ""
            if hasattr(portfolio, 'first_date') and hasattr(portfolio, 'last_date'):
                if portfolio.first_date and portfolio.last_date:
                    try:
                        # Convert to datetime if needed and format as date only
                        from datetime import datetime
                        
                        # Handle different date formats
                        if isinstance(portfolio.first_date, str):
                            start_date = datetime.strptime(portfolio.first_date.split()[0], '%Y-%m-%d').date()
                        else:
                            start_date = portfolio.first_date.date() if hasattr(portfolio.first_date, 'date') else portfolio.first_date
                            
                        if isinstance(portfolio.last_date, str):
                            end_date = datetime.strptime(portfolio.last_date.split()[0], '%Y-%m-%d').date()
                        else:
                            end_date = portfolio.last_date.date() if hasattr(portfolio.last_date, 'date') else portfolio.last_date
                        
                        # Calculate duration
                        duration_days = (end_date - start_date).days
                        duration_years = duration_days / 365.25
                        
                        # Format duration
                        if duration_years >= 1:
                            years = int(duration_years)
                            months = int((duration_years - years) * 12)
                            if months > 0:
                                duration_str = f"{years} г. {months} мес."
                            else:
                                duration_str = f"{years} г."
                        else:
                            months = int(duration_days / 30.44)
                            duration_str = f"{months} мес."
                        
                        # Format period info with duration
                        period_info = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')} ({duration_str})"
                        
                    except Exception as e:
                        self.logger.warning(f"Could not format period info: {e}")
                        # Fallback to simple format
                        period_info = f"{portfolio.first_date} - {portfolio.last_date}"
            
            # Build metrics text
            metrics_text = f"• 📊 {', '.join(symbols_with_weights)}\n"
            metrics_text += f"• **Базовая валюта:** {currency}\n"
            metrics_text += f"• **Период ребалансировки:** {rebalancing_period}\n"
            if period_info:
                metrics_text += f"• **Период:** {period_info}\n"
            
            # Format metrics with proper handling of None values
            if cagr_value is not None:
                metrics_text += f"• **CAGR (Среднегодовая доходность):** {cagr_value:.2%}\n"
            else:
                metrics_text += f"• **CAGR (Среднегодовая доходность):** Недоступно\n"
                
            if volatility_value is not None:
                metrics_text += f"• **Волатильность:** {volatility_value:.2%}\n"
            else:
                metrics_text += f"• **Волатильность:** Недоступно\n"
                
            if sharpe_value is not None:
                metrics_text += f"• **Коэфф. Шарпа:** {sharpe_value:.2f}\n"
            else:
                metrics_text += f"• **Коэфф. Шарпа:** Недоступно\n"
                
            if max_drawdown_value is not None:
                metrics_text += f"• **Макс. просадка:** {max_drawdown_value:.2%}\n"
            else:
                metrics_text += f"• **Макс. просадка:** Недоступно\n"
            
            self.logger.info(f"=== PORTFOLIO BASIC METRICS CALCULATION COMPLETED ===")
            self.logger.info(f"Final metrics: CAGR={cagr_value}, Volatility={volatility_value}, Sharpe={sharpe_value}, MaxDD={max_drawdown_value}")
            return metrics_text
            
        except Exception as e:
            self.logger.error(f"Could not get portfolio basic metrics: {e}")
            self.logger.error(f"Error type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            # Fallback to basic info
            symbols_with_weights = []
            for i, symbol in enumerate(symbols):
                symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                weight = weights[i] if i < len(weights) else 0.0
                symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
            
            return f"\n\n📊 **Основные метрики:**\n• **Состав:** {', '.join(symbols_with_weights)}\n• **Валюта:** {currency}\n• **Период ребалансировки:** Ежегодно\n• **CAGR (Среднегодовая доходность):** Недоступно\n• **Волатильность:** Недоступно\n• **Коэфф. Шарпа:** Недоступно\n• **Макс. просадка:** Недоступно\n"

    async def _prepare_portfolio_data_for_analysis(self, portfolio, symbols: list, weights: list, currency: str, user_id: int) -> Dict[str, Any]:
        """Prepare comprehensive portfolio data for Gemini portfolio analysis"""
        try:
            # Create portfolio metrics table using the specialized function
            portfolio_metrics_table = ""
            try:
                portfolio_metrics_table = self._create_portfolio_summary_metrics_table(
                    portfolio=portfolio,
                    symbols=symbols,
                    weights=weights,
                    currency=currency
                )
            except Exception as e:
                self.logger.warning(f"Failed to create portfolio metrics table: {e}")
            
            # Create individual assets metrics table for comparison
            individual_assets_metrics = ""
            try:
                individual_assets_metrics = self._create_summary_metrics_table(
                    symbols=symbols,
                    currency=currency,
                    expanded_symbols=symbols,  # Individual assets
                    portfolio_contexts=[],  # No portfolio contexts for individual assets
                    specified_period=None
                )
            except Exception as e:
                self.logger.warning(f"Failed to create individual assets metrics table: {e}")
            
            # Calculate correlations
            correlations = []
            try:
                if len(symbols) > 1:
                    asset_list = ok.AssetList(symbols, ccy=currency)
                    corr_matrix = asset_list.assets_ror.corr()
                    if corr_matrix is not None and not corr_matrix.empty:
                        correlations = corr_matrix.values.tolist()
                        self.logger.info(f"Portfolio correlation matrix calculated successfully, shape: {corr_matrix.shape}")
                    else:
                        self.logger.warning("Portfolio correlation matrix is empty")
            except Exception as e:
                self.logger.warning(f"Failed to calculate correlations: {e}")
            
            # Calculate efficient frontier data
            efficient_frontier_data = None
            try:
                if len(symbols) > 1:
                    asset_list = ok.AssetList(symbols, ccy=currency)
                    ef = ok.EfficientFrontier(asset_list, ccy=currency)
                    
                    # Extract key points from efficient frontier
                    efficient_frontier_data = {
                        'min_risk_portfolio': {
                            'risk': ef.gmv_annualized[0] if hasattr(ef, 'gmv_annualized') and ef.gmv_annualized is not None else None,
                            'return': ef.gmv_annualized[1] if hasattr(ef, 'gmv_annualized') and ef.gmv_annualized is not None else None,
                            'weights': ef.gmv_weights.tolist() if hasattr(ef, 'gmv_weights') and ef.gmv_weights is not None else None
                        },
                        'max_return_portfolio': {
                            'risk': None,  # Will be calculated from optimize_return
                            'return': None,  # Will be calculated from optimize_return
                            'weights': None  # Will be calculated from optimize_return
                        },
                        'max_sharpe_portfolio': {
                            'risk': None,  # Will be calculated from get_tangency_portfolio
                            'return': None,  # Will be calculated from get_tangency_portfolio
                            'weights': None,  # Will be calculated from get_tangency_portfolio
                            'sharpe_ratio': None
                        },
                        'asset_names': symbols,
                        'currency': currency
                    }
                    
                    # Try to get max return portfolio
                    try:
                        max_return_result = ef.optimize_return()
                        if max_return_result is not None and len(max_return_result) >= 2:
                            efficient_frontier_data['max_return_portfolio']['risk'] = max_return_result[0]
                            efficient_frontier_data['max_return_portfolio']['return'] = max_return_result[1]
                            if hasattr(ef, 'weights') and ef.weights is not None:
                                efficient_frontier_data['max_return_portfolio']['weights'] = ef.weights.tolist()
                    except Exception as e:
                        self.logger.warning(f"Failed to get max return portfolio: {e}")
                    
                    # Try to get tangency portfolio (max Sharpe)
                    try:
                        tangency_result = ef.get_tangency_portfolio()
                        if tangency_result is not None and len(tangency_result) >= 2:
                            efficient_frontier_data['max_sharpe_portfolio']['risk'] = tangency_result[0]
                            efficient_frontier_data['max_sharpe_portfolio']['return'] = tangency_result[1]
                            if hasattr(ef, 'weights') and ef.weights is not None:
                                efficient_frontier_data['max_sharpe_portfolio']['weights'] = ef.weights.tolist()
                                # Calculate Sharpe ratio
                                risk_free_rate = self.get_risk_free_rate(currency, 5.0)
                                if tangency_result[0] > 0:
                                    efficient_frontier_data['max_sharpe_portfolio']['sharpe_ratio'] = (tangency_result[1] - risk_free_rate) / tangency_result[0]
                    except Exception as e:
                        self.logger.warning(f"Failed to get tangency portfolio: {e}")
                        
            except Exception as e:
                self.logger.warning(f"Failed to calculate efficient frontier: {e}")
            
            # Get asset names
            asset_names = {}
            try:
                for symbol in symbols:
                    try:
                        asset = ok.Asset(symbol)
                        if hasattr(asset, 'name') and asset.name:
                            asset_names[symbol] = asset.name
                        else:
                            asset_names[symbol] = symbol
                    except Exception as e:
                        asset_names[symbol] = symbol
                        self.logger.warning(f"Could not get name for asset {symbol}: {e}")
            except Exception as e:
                self.logger.warning(f"Failed to get asset names: {e}")
            
            portfolio_data = {
                'symbols': symbols,
                'weights': weights,
                'currency': currency,
                'period': 'полный доступный период данных',
                'portfolio_metrics_table': portfolio_metrics_table,
                'individual_assets_metrics': individual_assets_metrics,
                'correlations': correlations,
                'efficient_frontier': efficient_frontier_data,
                'asset_names': asset_names,
                'analysis_type': 'portfolio_analysis',
                'additional_info': f'Портфель состоит из {len(symbols)} активов с весами: {", ".join([f"{symbol}: {weight:.1%}" for symbol, weight in zip(symbols, weights)])}'
            }
            
            return portfolio_data
            
        except Exception as e:
            self.logger.error(f"Error preparing portfolio data for analysis: {e}")
            return {
                'symbols': symbols,
                'weights': weights,
                'currency': currency,
                'period': 'полный доступный период данных',
                'portfolio_metrics_table': '❌ Ошибка при создании таблицы метрик портфеля',
                'individual_assets_metrics': '❌ Ошибка при создании таблицы метрик активов',
                'correlations': [],
                'efficient_frontier': None,
                'asset_names': {symbol: symbol for symbol in symbols},
                'analysis_type': 'portfolio_analysis',
                'additional_info': f'Ошибка при подготовке данных: {str(e)}'
            }

    async def _prepare_data_for_analysis(self, symbols: list, currency: str, expanded_symbols: list, portfolio_contexts: list, user_id: int) -> Dict[str, Any]:
        """Prepare comprehensive financial data for Gemini analysis"""
        try:
            # Get user context to access describe table
            describe_table = ""
            if user_id:
                user_context = self._get_user_context(user_id)
                describe_table = user_context.get('describe_table', '')
            
            # Create summary metrics table for Gemini analysis (using the same optimized table)
            summary_metrics_table = ""
            try:
                summary_metrics_table = self._create_summary_metrics_table(
                    symbols=symbols,
                    currency=currency,
                    expanded_symbols=expanded_symbols,
                    portfolio_contexts=portfolio_contexts,
                    specified_period=None
                )
            except Exception as e:
                self.logger.warning(f"Failed to create summary metrics table: {e}")
            
            data_info = {
                'symbols': symbols,
                'currency': currency,
                'period': 'полный доступный период данных',
                'performance': {},
                'correlations': [],
                'additional_info': '',
                'summary_metrics_table': summary_metrics_table,  # This now contains the optimized table
                'describe_table': '',  # No longer needed since summary_metrics_table contains the same data
                'asset_count': len(symbols),
                'analysis_type': 'asset_comparison',
                'asset_names': {}  # Dictionary to store asset names
            }
            
            # Calculate detailed performance metrics for each symbol
            for i, symbol in enumerate(symbols):
                asset_data = None
                
                if i < len(expanded_symbols):
                    expanded_item = expanded_symbols[i]
                    
                    # Check if this is a portfolio (DataFrame/Series) or a regular asset (string)
                    if isinstance(expanded_item, (pd.Series, pd.DataFrame)):
                        # This is a portfolio - use the portfolio object from context
                        if i < len(portfolio_contexts):
                            portfolio_context = portfolio_contexts[i]
                            asset_data = portfolio_context.get('portfolio_object')
                    else:
                        # This is a regular asset symbol - create Asset object
                        try:
                            asset_data = ok.Asset(symbol)
                        except Exception as e:
                            self.logger.warning(f"Failed to create Asset object for {symbol}: {e}")
                            asset_data = None
                
                # Fallback: try to create Asset from symbol if we don't have asset_data
                if asset_data is None:
                    try:
                        asset_data = ok.Asset(symbol)
                    except Exception as e:
                        self.logger.warning(f"Failed to create Asset object for {symbol}: {e}")
                        asset_data = None
                
                # Get asset name
                asset_name = symbol  # Default to symbol
                if asset_data is not None:
                    try:
                        if hasattr(asset_data, 'name') and asset_data.name:
                            asset_name = asset_data.name
                        elif hasattr(asset_data, 'symbol') and asset_data.symbol:
                            asset_name = asset_data.symbol
                    except Exception as e:
                        self.logger.warning(f"Failed to get asset name for {symbol}: {e}")
                
                data_info['asset_names'][symbol] = asset_name
                
                # Get comprehensive performance metrics
                performance_metrics = {}
                
                # Calculate metrics from price data
                try:
                    # Get price data for calculations (синхронизировано с AssetList - приоритет месячным данным)
                    if asset_data is not None:
                        if hasattr(asset_data, 'close_monthly') and asset_data.close_monthly is not None:
                            prices = asset_data.close_monthly
                            data_type = "monthly"
                        elif hasattr(asset_data, 'wealth_index') and asset_data.wealth_index is not None:
                            # For Portfolio objects
                            wealth_index = asset_data.wealth_index
                            if hasattr(wealth_index, 'iloc') and len(wealth_index) > 1:
                                if hasattr(wealth_index, 'columns'):
                                    prices = wealth_index.iloc[:, 0]
                                else:
                                    prices = wealth_index
                            data_type = "wealth_index"
                        elif hasattr(asset_data, 'adj_close') and asset_data.adj_close is not None:
                            prices = asset_data.adj_close
                            data_type = "adjusted"
                        elif hasattr(asset_data, 'close_daily') and asset_data.close_daily is not None:
                            prices = asset_data.close_daily
                            data_type = "daily"
                        else:
                            prices = None
                            data_type = "none"
                    else:
                        prices = None
                        data_type = "none"
                    
                    self.logger.info(f"Data preparation for {symbol}: type={data_type}, prices_length={len(prices) if prices is not None else 0}")
                    
                    if prices is not None and len(prices) > 1:
                        # Calculate returns from prices
                        returns = prices.pct_change().dropna()
                        
                        # Basic metrics
                        if asset_data is not None and hasattr(asset_data, 'total_return'):
                            performance_metrics['total_return'] = asset_data.total_return
                        else:
                            # Calculate total return from first and last price
                            total_return = (prices.iloc[-1] / prices.iloc[0]) - 1
                            performance_metrics['total_return'] = total_return
                        
                        # Annual return (CAGR)
                        if asset_data is not None and hasattr(asset_data, 'annual_return'):
                            performance_metrics['annual_return'] = asset_data.annual_return
                        else:
                            # Calculate CAGR based on data frequency
                            periods = len(prices)
                            
                            # Determine data frequency and calculate years accordingly
                            if asset_data is not None and hasattr(asset_data, 'close_monthly') and asset_data.close_monthly is not None:
                                # Monthly data
                                years = periods / 12.0
                            elif asset_data is not None and hasattr(asset_data, 'close_daily') and asset_data.close_daily is not None:
                                # Daily data - use 252 trading days per year
                                years = periods / 252.0
                            else:
                                # Default to monthly assumption
                                years = periods / 12.0
                            
                            if years > 0:
                                cagr = ((prices.iloc[-1] / prices.iloc[0]) ** (1.0 / years)) - 1
                                performance_metrics['annual_return'] = cagr
                                self.logger.info(f"CAGR calculation for {symbol}: periods={periods}, years={years:.2f}, cagr={cagr:.4f}")
                            else:
                                performance_metrics['annual_return'] = 0.0
                                self.logger.warning(f"CAGR calculation failed for {symbol}: years={years}")
                        
                        # Volatility
                        if asset_data is not None and hasattr(asset_data, 'volatility'):
                            performance_metrics['volatility'] = asset_data.volatility
                        else:
                            # Calculate annualized volatility based on data frequency
                            if asset_data is not None and hasattr(asset_data, 'close_monthly') and asset_data.close_monthly is not None:
                                # Monthly data - annualize by sqrt(12)
                                volatility = returns.std() * (12 ** 0.5)
                            elif asset_data is not None and hasattr(asset_data, 'close_daily') and asset_data.close_daily is not None:
                                # Daily data - annualize by sqrt(252)
                                volatility = returns.std() * (252 ** 0.5)
                            else:
                                # Default to monthly assumption
                                volatility = returns.std() * (12 ** 0.5)
                            
                            performance_metrics['volatility'] = volatility
                            self.logger.info(f"Volatility calculation for {symbol}: data_type={data_type}, volatility={volatility:.4f}")
                        
                        # Max drawdown
                        if asset_data is not None and hasattr(asset_data, 'max_drawdown'):
                            performance_metrics['max_drawdown'] = asset_data.max_drawdown
                        else:
                            # Calculate max drawdown
                            cumulative = (1 + returns).cumprod()
                            running_max = cumulative.expanding().max()
                            drawdown = (cumulative - running_max) / running_max
                            max_drawdown = drawdown.min()
                            performance_metrics['max_drawdown'] = max_drawdown
                        
                        # Store returns for Sortino calculation
                        performance_metrics['_returns'] = returns
                    
                    else:
                        # Fallback values if no price data
                        performance_metrics['total_return'] = 0.0
                        performance_metrics['annual_return'] = 0.0
                        performance_metrics['volatility'] = 0.0
                        performance_metrics['max_drawdown'] = 0.0
                        performance_metrics['_returns'] = None
                
                except Exception as e:
                    self.logger.warning(f"Failed to calculate basic metrics for {symbol}: {e}")
                    performance_metrics['total_return'] = 0.0
                    performance_metrics['annual_return'] = 0.0
                    performance_metrics['volatility'] = 0.0
                    performance_metrics['max_drawdown'] = 0.0
                    performance_metrics['_returns'] = None
                
                # Sharpe ratio calculation
                try:
                    annual_return = performance_metrics.get('annual_return', 0)
                    volatility = performance_metrics.get('volatility', 0)
                    sharpe_ratio = self.calculate_sharpe_ratio(annual_return, volatility, currency, asset_data=asset_data)
                    performance_metrics['sharpe_ratio'] = sharpe_ratio
                except Exception as e:
                    self.logger.warning(f"Failed to calculate Sharpe ratio for {symbol}: {e}")
                    performance_metrics['sharpe_ratio'] = 0.0
                
                # Sortino ratio calculation
                try:
                    if asset_data is not None and hasattr(asset_data, 'sortino_ratio'):
                        performance_metrics['sortino_ratio'] = asset_data.sortino_ratio
                    else:
                        # Manual Sortino ratio calculation
                        annual_return = performance_metrics.get('annual_return', 0)
                        returns = performance_metrics.get('_returns')
                        
                        if returns is not None and len(returns) > 0:
                            # Calculate downside deviation (only negative returns)
                            negative_returns = returns[returns < 0]
                            if len(negative_returns) > 0:
                                # Annualize downside deviation based on data frequency
                                if asset_data is not None and hasattr(asset_data, 'close_monthly') and asset_data.close_monthly is not None:
                                    # Monthly data - annualize by sqrt(12)
                                    downside_deviation = negative_returns.std() * (12 ** 0.5)
                                elif asset_data is not None and hasattr(asset_data, 'close_daily') and asset_data.close_daily is not None:
                                    # Daily data - annualize by sqrt(252)
                                    downside_deviation = negative_returns.std() * (252 ** 0.5)
                                else:
                                    # Default to monthly assumption
                                    downside_deviation = negative_returns.std() * (12 ** 0.5)
                                
                                if downside_deviation > 0:
                                    # Use proper risk-free rate based on currency
                                    risk_free_rate = self.get_risk_free_rate(currency, years)
                                    sortino_ratio = (annual_return - risk_free_rate) / downside_deviation
                                    performance_metrics['sortino_ratio'] = sortino_ratio
                                else:
                                    performance_metrics['sortino_ratio'] = 0.0
                            else:
                                # No negative returns, use volatility as fallback
                                volatility = performance_metrics.get('volatility', 0)
                                if volatility > 0:
                                    # Use proper risk-free rate based on currency
                                    risk_free_rate = self.get_risk_free_rate(currency, years)
                                    sortino_ratio = (annual_return - risk_free_rate) / volatility
                                    performance_metrics['sortino_ratio'] = sortino_ratio
                                else:
                                    performance_metrics['sortino_ratio'] = 0.0
                        else:
                            # Fallback to Sharpe ratio if no returns data
                            performance_metrics['sortino_ratio'] = performance_metrics.get('sharpe_ratio', 0.0)
                except Exception as e:
                    self.logger.warning(f"Failed to calculate Sortino ratio for {symbol}: {e}")
                    performance_metrics['sortino_ratio'] = 0.0
                
                # Additional metrics calculation
                try:
                    # Calmar Ratio = Annual Return / Max Drawdown (absolute value)
                    annual_return = performance_metrics.get('annual_return', 0)
                    max_drawdown = performance_metrics.get('max_drawdown', 0)
                    if max_drawdown != 0:
                        calmar_ratio = annual_return / abs(max_drawdown)
                        performance_metrics['calmar_ratio'] = calmar_ratio
                        self.logger.info(f"Calmar ratio for {symbol}: {calmar_ratio:.4f}")
                    else:
                        performance_metrics['calmar_ratio'] = 0.0
                    
                    # VaR 95% and CVaR 95% calculation
                    returns = performance_metrics.get('_returns')
                    if returns is not None and len(returns) > 0:
                        # VaR 95% - 5th percentile of returns (worst 5% of returns)
                        var_95 = returns.quantile(0.05)
                        performance_metrics['var_95'] = var_95
                        
                        # CVaR 95% - Expected value of returns below VaR 95%
                        returns_below_var = returns[returns <= var_95]
                        if len(returns_below_var) > 0:
                            cvar_95 = returns_below_var.mean()
                            performance_metrics['cvar_95'] = cvar_95
                        else:
                            performance_metrics['cvar_95'] = var_95
                        
                        self.logger.info(f"VaR 95% for {symbol}: {var_95:.4f}, CVaR 95%: {performance_metrics['cvar_95']:.4f}")
                    else:
                        performance_metrics['var_95'] = 0.0
                        performance_metrics['cvar_95'] = 0.0
                        
                except Exception as e:
                    self.logger.warning(f"Failed to calculate additional metrics for {symbol}: {e}")
                    performance_metrics['calmar_ratio'] = 0.0
                    performance_metrics['var_95'] = 0.0
                    performance_metrics['cvar_95'] = 0.0
                
                # Clean up temporary data
                if '_returns' in performance_metrics:
                    del performance_metrics['_returns']
                
                data_info['performance'][symbol] = performance_metrics
            
            # Calculate correlation matrix if we have multiple assets
            if len(expanded_symbols) > 1:
                try:
                    correlation_matrix = []
                    
                    # Get user context to access portfolio information
                    user_context = self._get_user_context(user_id)
                    portfolio_contexts = user_context.get('portfolio_contexts', [])
                    
                    # Separate portfolios and individual assets using expanded_symbols
                    portfolio_data = []
                    asset_symbols = []
                    
                    for i, expanded_symbol in enumerate(expanded_symbols):
                        if isinstance(expanded_symbol, (pd.Series, pd.DataFrame)):
                            # This is a portfolio wealth index
                            portfolio_data.append(expanded_symbol)
                        else:
                            # This is an individual asset symbol
                            asset_symbols.append(expanded_symbol)
                    
                    # Calculate correlation data for all items (same method as in _create_mixed_comparison_correlation_matrix)
                    correlation_data = {}
                    
                    # Process portfolios separately
                    for i, portfolio_context in enumerate(portfolio_contexts):
                        if i < len(portfolio_data):
                            try:
                                # Get portfolio details from context
                                assets = portfolio_context.get('portfolio_symbols', [])
                                weights = portfolio_context.get('portfolio_weights', [])
                                symbol = portfolio_context.get('symbol', f'Portfolio_{i+1}')
                                
                                if assets and weights and len(assets) == len(weights):
                                    # Create portfolio using ok.Portfolio
                                    portfolio = ok.Portfolio(
                                        assets=assets,
                                        weights=weights,
                                        rebalancing_strategy=ok.Rebalance(period="year"),
                                        symbol=symbol
                                    )
                                    
                                    # Calculate returns for portfolio
                                    wealth_series = portfolio.wealth_index
                                    returns = wealth_series.pct_change().dropna()
                                    if len(returns) > 0:
                                        correlation_data[symbol] = returns
                            except Exception as portfolio_error:
                                self.logger.warning(f"Could not process portfolio {i} for correlation: {portfolio_error}")
                                continue
                    
                    # Process individual assets separately
                    if asset_symbols:
                        try:
                            asset_asset_list = ok.AssetList(asset_symbols, ccy=currency)
                            
                            for symbol in asset_symbols:
                                if symbol in asset_asset_list.wealth_indexes.columns:
                                    # Calculate returns for individual asset
                                    wealth_series = asset_asset_list.wealth_indexes[symbol]
                                    returns = wealth_series.pct_change().dropna()
                                    if len(returns) > 0:
                                        correlation_data[symbol] = returns
                        except Exception as asset_error:
                            self.logger.warning(f"Could not process individual assets: {asset_error}")
                    
                    # Calculate correlation matrix using the same method as in correlation button
                    if len(correlation_data) >= 2:
                        # Combine all returns into a DataFrame
                        returns_df = pd.DataFrame(correlation_data)
                        
                        # Calculate correlation matrix (same as in _create_mixed_comparison_correlation_matrix)
                        correlation_matrix_df = returns_df.corr()
                        correlation_matrix = correlation_matrix_df.values.tolist()
                        
                        self.logger.info(f"AI analysis correlation matrix calculated successfully, shape: {correlation_matrix_df.shape}")
                    else:
                        # Fallback: create identity matrix
                        correlation_matrix = []
                        for i in range(len(symbols)):
                            row = []
                            for j in range(len(symbols)):
                                row.append(1.0 if i == j else 0.0)
                            correlation_matrix.append(row)
                        self.logger.warning(f"Not enough correlation data for AI analysis: {len(correlation_data)}")
                    
                    data_info['correlations'] = correlation_matrix
                    
                except Exception as e:
                    self.logger.warning(f"Failed to calculate correlations for AI analysis: {e}")
                    # Create identity matrix as fallback
                    correlation_matrix = []
                    for i in range(len(symbols)):
                        row = []
                        for j in range(len(symbols)):
                            row.append(1.0 if i == j else 0.0)
                        correlation_matrix.append(row)
                    data_info['correlations'] = correlation_matrix
            
            # Add portfolio context information
            if portfolio_contexts:
                portfolio_info = []
                for pctx in portfolio_contexts:
                    portfolio_info.append(f"Портфель {pctx.get('symbol', 'Unknown')}: {len(pctx.get('portfolio_symbols', []))} активов")
                data_info['additional_info'] = f"Включает портфели: {'; '.join(portfolio_info)}"
            
            # Add efficient frontier data if we have multiple assets
            if len(expanded_symbols) > 1:
                try:
                    # Prepare assets for efficient frontier calculation
                    asset_list_items = []
                    asset_names = []
                    
                    for i, symbol in enumerate(symbols):
                        if i < len(expanded_symbols):
                            if isinstance(expanded_symbols[i], (pd.Series, pd.DataFrame)):
                                # This is a portfolio - recreate it
                                if i < len(portfolio_contexts):
                                    pctx = portfolio_contexts[i]
                                    try:
                                        p = ok.Portfolio(
                                            pctx.get('portfolio_symbols', []),
                                            weights=pctx.get('portfolio_weights', []),
                                            ccy=pctx.get('portfolio_currency') or currency,
                                        )
                                        asset_list_items.append(p)
                                        asset_names.append(symbol)
                                    except Exception as pe:
                                        self.logger.warning(f"Failed to recreate portfolio for Efficient Frontier: {pe}")
                            else:
                                # This is a regular asset
                                asset_list_items.append(symbol)
                                asset_names.append(symbol)
                    
                    if len(asset_list_items) > 1:
                        # Create AssetList and EfficientFrontier
                        asset_list = ok.AssetList(asset_list_items, ccy=currency)
                        ef = ok.EfficientFrontier(asset_list, ccy=currency)
                        
                        # Get efficient frontier data
                        try:
                            # Extract key points from efficient frontier using correct okama methods
                            efficient_frontier_data = {
                                'min_risk_portfolio': {
                                    'risk': ef.gmv_annualized[0] if hasattr(ef, 'gmv_annualized') and ef.gmv_annualized is not None else None,
                                    'return': ef.gmv_annualized[1] if hasattr(ef, 'gmv_annualized') and ef.gmv_annualized is not None else None,
                                    'weights': ef.gmv_weights.tolist() if hasattr(ef, 'gmv_weights') and ef.gmv_weights is not None else None
                                },
                                'max_return_portfolio': {
                                    'risk': None,  # Will be calculated from optimize_return
                                    'return': None,  # Will be calculated from optimize_return
                                    'weights': None  # Will be calculated from optimize_return
                                },
                                'max_sharpe_portfolio': {
                                    'risk': None,  # Will be calculated from get_tangency_portfolio
                                    'return': None,  # Will be calculated from get_tangency_portfolio
                                    'weights': None,  # Will be calculated from get_tangency_portfolio
                                    'sharpe_ratio': None
                                },
                                'asset_names': asset_names,
                                'currency': currency
                            }
                            
                            # Get max return portfolio
                            try:
                                max_return_result = ef.optimize_return()
                                if max_return_result and 'Weights' in max_return_result:
                                    efficient_frontier_data['max_return_portfolio']['weights'] = max_return_result['Weights'].tolist()
                                    efficient_frontier_data['max_return_portfolio']['return'] = max_return_result.get('Mean_return_monthly', 0) * 12  # Annualize
                                    efficient_frontier_data['max_return_portfolio']['risk'] = max_return_result.get('Risk_monthly', 0) * (12 ** 0.5)  # Annualize
                            except Exception as e:
                                self.logger.warning(f"Failed to get max return portfolio: {e}")
                            
                            # Get max Sharpe portfolio (tangency portfolio)
                            try:
                                tangency_result = ef.get_tangency_portfolio()
                                if tangency_result and 'Weights' in tangency_result:
                                    efficient_frontier_data['max_sharpe_portfolio']['weights'] = tangency_result['Weights'].tolist()
                                    efficient_frontier_data['max_sharpe_portfolio']['return'] = tangency_result.get('Rate_of_return', 0)
                                    efficient_frontier_data['max_sharpe_portfolio']['risk'] = tangency_result.get('Risk', 0)
                                    # Calculate Sharpe ratio
                                    if efficient_frontier_data['max_sharpe_portfolio']['risk'] > 0:
                                        efficient_frontier_data['max_sharpe_portfolio']['sharpe_ratio'] = (
                                            efficient_frontier_data['max_sharpe_portfolio']['return'] / 
                                            efficient_frontier_data['max_sharpe_portfolio']['risk']
                                        )
                            except Exception as e:
                                self.logger.warning(f"Failed to get tangency portfolio: {e}")
                            
                            # Log the extracted data for debugging
                            self.logger.info(f"Efficient frontier data extracted:")
                            self.logger.info(f"  Min risk: risk={efficient_frontier_data['min_risk_portfolio']['risk']}, return={efficient_frontier_data['min_risk_portfolio']['return']}")
                            self.logger.info(f"  Max return: risk={efficient_frontier_data['max_return_portfolio']['risk']}, return={efficient_frontier_data['max_return_portfolio']['return']}")
                            self.logger.info(f"  Max Sharpe: risk={efficient_frontier_data['max_sharpe_portfolio']['risk']}, return={efficient_frontier_data['max_sharpe_portfolio']['return']}")
                            
                            data_info['efficient_frontier'] = efficient_frontier_data
                            self.logger.info(f"Efficient frontier data prepared for {len(asset_names)} assets")
                            
                        except Exception as ef_error:
                            self.logger.warning(f"Failed to extract efficient frontier data: {ef_error}")
                            data_info['efficient_frontier'] = None
                    
                except Exception as e:
                    self.logger.warning(f"Failed to calculate efficient frontier: {e}")
                    data_info['efficient_frontier'] = None
            
            # Add analysis metadata
            data_info['analysis_metadata'] = {
                'timestamp': self._get_current_timestamp(),
                'data_source': 'okama.AssetList.describe()',
                'analysis_depth': 'comprehensive',
                'includes_correlations': len(data_info['correlations']) > 0,
                'includes_describe_table': bool(describe_table),
                'includes_efficient_frontier': 'efficient_frontier' in data_info and data_info['efficient_frontier'] is not None
            }
            
            return data_info
            
        except Exception as e:
            self.logger.error(f"Error preparing data for analysis: {e}")
            return {
                'symbols': symbols,
                'currency': currency,
                'period': 'полный доступный период данных',
                'performance': {},
                'correlations': [],
                'additional_info': f'Ошибка подготовки данных: {str(e)}',
                'describe_table': '',
                'asset_count': len(symbols),
                'analysis_type': 'asset_comparison',
                'analysis_metadata': {
                    'timestamp': self._get_current_timestamp(),
                    'data_source': 'error_fallback',
                    'analysis_depth': 'basic',
                    'includes_correlations': False,
                    'includes_describe_table': False
                }
            }

    def _prepare_comprehensive_metrics(self, symbols: list, currency: str, expanded_symbols: list, portfolio_contexts: list, user_id: int) -> Dict[str, Any]:
        """Prepare comprehensive metrics data for Excel export"""
        try:
            # Get user context to access describe table
            describe_table = ""
            if user_id:
                user_context = self._get_user_context(user_id)
                describe_table = user_context.get('describe_table', '')
            
            metrics_data = {
                'symbols': symbols,
                'currency': currency,
                'period': 'полный доступный период данных',
                'performance': {},
                'detailed_metrics': {},
                'correlations': [],
                'describe_table': describe_table,
                'asset_count': len(symbols),
                'analysis_type': 'metrics_export',
                'timestamp': self._get_current_timestamp(),
                'asset_names': {}  # Dictionary to store asset names
            }
            
            # Calculate detailed performance metrics for each symbol
            for i, symbol in enumerate(symbols):
                try:
                    # Get the actual asset/portfolio object from portfolio_contexts
                    asset_data = None
                    if i < len(portfolio_contexts):
                        portfolio_context = portfolio_contexts[i]
                        portfolio_object = portfolio_context.get('portfolio_object')
                        
                        if portfolio_object is not None:
                            # This is a portfolio - use portfolio object
                            asset_data = portfolio_object
                        else:
                            # This is a regular asset - create Asset object
                            portfolio_symbols = portfolio_context.get('portfolio_symbols', [symbol])
                            if portfolio_symbols and len(portfolio_symbols) > 0:
                                asset_symbol = portfolio_symbols[0]
                            else:
                                asset_symbol = symbol  # Fallback to the symbol from the loop
                            
                            try:
                                asset_data = ok.Asset(asset_symbol)
                            except Exception as e:
                                self.logger.warning(f"Failed to create Asset object for {asset_symbol}: {e}")
                                asset_data = None
                    
                    if asset_data is None:
                        # Fallback: try to create Asset from symbol
                        try:
                            asset_data = ok.Asset(symbol)
                        except Exception as e:
                            self.logger.warning(f"Failed to create Asset object for {symbol}: {e}")
                            asset_data = None
                    
                    # Get asset name
                    asset_name = symbol  # Default to symbol
                    try:
                        if asset_data is not None:
                            if hasattr(asset_data, 'name') and asset_data.name:
                                asset_name = asset_data.name
                            elif hasattr(asset_data, 'symbol') and asset_data.symbol:
                                asset_name = asset_data.symbol
                    except Exception as e:
                        self.logger.warning(f"Failed to get asset name for {symbol}: {e}")
                    
                    metrics_data['asset_names'][symbol] = asset_name
                    
                    # Get comprehensive performance metrics
                    detailed_metrics = {}
                    
                    if asset_data is not None:
                        # Calculate metrics from asset/portfolio data
                        try:
                            # Get price data for calculations
                            if hasattr(asset_data, 'close_monthly') and asset_data.close_monthly is not None:
                                prices = asset_data.close_monthly
                            elif hasattr(asset_data, 'close_daily') and asset_data.close_daily is not None:
                                prices = asset_data.close_daily
                            elif hasattr(asset_data, 'adj_close') and asset_data.adj_close is not None:
                                prices = asset_data.adj_close
                            elif hasattr(asset_data, 'wealth_index') and asset_data.wealth_index is not None:
                                prices = asset_data.wealth_index
                            else:
                                prices = None
                            
                            if prices is not None and len(prices) > 1:
                                # Calculate returns from prices
                                returns = prices.pct_change().dropna()
                                
                                # Basic metrics
                                if hasattr(asset_data, 'total_return'):
                                    detailed_metrics['total_return'] = asset_data.total_return
                                else:
                                    # Calculate total return from first and last price
                                    total_return = (prices.iloc[-1] / prices.iloc[0]) - 1
                                    detailed_metrics['total_return'] = total_return
                                
                                # Annual return (CAGR)
                                if hasattr(asset_data, 'annual_return'):
                                    detailed_metrics['annual_return'] = asset_data.annual_return
                                else:
                                    # Calculate CAGR
                                    periods = len(prices)
                                    years = periods / 12.0  # Assuming monthly data
                                    if years > 0:
                                        cagr = ((prices.iloc[-1] / prices.iloc[0]) ** (1.0 / years)) - 1
                                        detailed_metrics['annual_return'] = cagr
                                    else:
                                        detailed_metrics['annual_return'] = 0.0
                                
                                # Volatility
                                if hasattr(asset_data, 'volatility'):
                                    detailed_metrics['volatility'] = asset_data.volatility
                                else:
                                    # Calculate annualized volatility
                                    volatility = returns.std() * (12 ** 0.5)  # Annualized for monthly data
                                    detailed_metrics['volatility'] = volatility
                                
                                # Max drawdown
                                if hasattr(asset_data, 'max_drawdown'):
                                    detailed_metrics['max_drawdown'] = asset_data.max_drawdown
                                else:
                                    # Calculate max drawdown
                                    cumulative = (1 + returns).cumprod()
                                    running_max = cumulative.expanding().max()
                                    drawdown = (cumulative - running_max) / running_max
                                    max_drawdown = drawdown.min()
                                    detailed_metrics['max_drawdown'] = max_drawdown
                                
                                # Store returns for Sharpe and Sortino calculations
                                detailed_metrics['_returns'] = returns
                                
                            else:
                                # Fallback values if no price data
                                detailed_metrics['total_return'] = 0.0
                                detailed_metrics['annual_return'] = 0.0
                                detailed_metrics['volatility'] = 0.0
                                detailed_metrics['max_drawdown'] = 0.0
                                detailed_metrics['_returns'] = None
                        
                        except Exception as e:
                            self.logger.warning(f"Failed to calculate basic metrics for {symbol}: {e}")
                            detailed_metrics['total_return'] = 0.0
                            detailed_metrics['annual_return'] = 0.0
                            detailed_metrics['volatility'] = 0.0
                            detailed_metrics['max_drawdown'] = 0.0
                            detailed_metrics['_returns'] = None
                    else:
                        # No asset data available
                        detailed_metrics['total_return'] = 0.0
                        detailed_metrics['annual_return'] = 0.0
                        detailed_metrics['volatility'] = 0.0
                        detailed_metrics['max_drawdown'] = 0.0
                        detailed_metrics['_returns'] = None
                        
                    # Sharpe ratio calculation
                    try:
                        annual_return = detailed_metrics.get('annual_return', 0)
                        volatility = detailed_metrics.get('volatility', 0)
                        sharpe_ratio = self.calculate_sharpe_ratio(annual_return, volatility, currency, asset_data=asset_data)
                        detailed_metrics['sharpe_ratio'] = sharpe_ratio
                    except Exception as e:
                        self.logger.warning(f"Failed to calculate Sharpe ratio for {symbol}: {e}")
                        detailed_metrics['sharpe_ratio'] = 0.0
                    
                    # Sortino ratio calculation
                    try:
                        if asset_data is not None and hasattr(asset_data, 'sortino_ratio'):
                            detailed_metrics['sortino_ratio'] = asset_data.sortino_ratio
                        else:
                            # Manual Sortino ratio calculation
                            annual_return = detailed_metrics.get('annual_return', 0)
                            returns = detailed_metrics.get('_returns')
                            
                            if returns is not None and len(returns) > 0:
                                # Calculate downside deviation (only negative returns)
                                negative_returns = returns[returns < 0]
                                if len(negative_returns) > 0:
                                    downside_deviation = negative_returns.std() * (12 ** 0.5)  # Annualized
                                    if downside_deviation > 0:
                                        # Use proper risk-free rate based on currency
                                        risk_free_rate = self.get_risk_free_rate(currency, years)
                                        sortino_ratio = (annual_return - risk_free_rate) / downside_deviation
                                        detailed_metrics['sortino_ratio'] = sortino_ratio
                                    else:
                                        detailed_metrics['sortino_ratio'] = 0.0
                                else:
                                    # No negative returns, use volatility as fallback
                                    volatility = detailed_metrics.get('volatility', 0)
                                    if volatility > 0:
                                        # Use proper risk-free rate based on currency
                                        risk_free_rate = self.get_risk_free_rate(currency, years)
                                        sortino_ratio = (annual_return - risk_free_rate) / volatility
                                        detailed_metrics['sortino_ratio'] = sortino_ratio
                                    else:
                                        detailed_metrics['sortino_ratio'] = 0.0
                            else:
                                # Fallback to Sharpe ratio if no returns data
                                detailed_metrics['sortino_ratio'] = detailed_metrics.get('sharpe_ratio', 0.0)
                    except Exception as e:
                        self.logger.warning(f"Failed to calculate Sortino ratio for {symbol}: {e}")
                        detailed_metrics['sortino_ratio'] = 0.0
                    
                    # Additional metrics calculation
                    try:
                        # Calmar Ratio = Annual Return / Max Drawdown (absolute value)
                        annual_return = detailed_metrics.get('annual_return', 0)
                        max_drawdown = detailed_metrics.get('max_drawdown', 0)
                        if max_drawdown != 0:
                            calmar_ratio = annual_return / abs(max_drawdown)
                            detailed_metrics['calmar_ratio'] = calmar_ratio
                        else:
                            detailed_metrics['calmar_ratio'] = 0.0
                        
                        # VaR 95% and CVaR 95% calculation
                        returns = detailed_metrics.get('_returns')
                        if returns is not None and len(returns) > 0:
                            # VaR 95% - 5th percentile of returns (worst 5% of returns)
                            var_95 = returns.quantile(0.05)
                            detailed_metrics['var_95'] = var_95
                            
                            # CVaR 95% - Expected value of returns below VaR 95%
                            returns_below_var = returns[returns <= var_95]
                            if len(returns_below_var) > 0:
                                cvar_95 = returns_below_var.mean()
                                detailed_metrics['cvar_95'] = cvar_95
                            else:
                                detailed_metrics['cvar_95'] = var_95
                        else:
                            detailed_metrics['var_95'] = 0.0
                            detailed_metrics['cvar_95'] = 0.0
                            
                    except Exception as e:
                        self.logger.warning(f"Failed to calculate additional metrics for {symbol}: {e}")
                        detailed_metrics['calmar_ratio'] = 0.0
                        detailed_metrics['var_95'] = 0.0
                        detailed_metrics['cvar_95'] = 0.0
                    
                    # Clean up temporary data
                    if '_returns' in detailed_metrics:
                        del detailed_metrics['_returns']
                    
                    metrics_data['detailed_metrics'][symbol] = detailed_metrics
                        
                except Exception as e:
                    self.logger.warning(f"Failed to get detailed metrics for {symbol}: {e}")
                    metrics_data['detailed_metrics'][symbol] = {
                        'total_return': 0.0,
                        'annual_return': 0.0,
                        'volatility': 0.0,
                        'sharpe_ratio': 0.0,
                        'sortino_ratio': 0.0,
                        'max_drawdown': 0.0,
                        'calmar_ratio': 0.0,
                        'var_95': 0.0,
                        'cvar_95': 0.0
                    }
            
            # Calculate correlations if we have multiple assets
            if len(symbols) > 1:
                try:
                    # Try to create AssetList for correlation calculation
                    correlation_matrix = []
                    
                    # Create AssetList from symbols for correlation calculation
                    try:
                        # Get clean symbols for correlation calculation
                        clean_symbols = []
                        for i, symbol in enumerate(symbols):
                            if i < len(portfolio_contexts):
                                portfolio_context = portfolio_contexts[i]
                                portfolio_symbols = portfolio_context.get('portfolio_symbols', [symbol])
                                clean_symbols.extend(portfolio_symbols)
                            else:
                                clean_symbols.append(symbol)
                        
                        # Remove duplicates while preserving order
                        clean_symbols = list(dict.fromkeys(clean_symbols))
                        
                        if len(clean_symbols) > 1:
                            # Create AssetList for correlation calculation
                            asset_list = ok.AssetList(clean_symbols)
                            if hasattr(asset_list, 'correlation_matrix'):
                                correlation_matrix = asset_list.correlation_matrix.tolist()
                            else:
                                # Fallback correlation matrix
                                correlation_matrix = []
                                for i in range(len(clean_symbols)):
                                    row = []
                                    for j in range(len(clean_symbols)):
                                        if i == j:
                                            row.append(1.0)
                                        else:
                                            row.append(0.3)  # Conservative estimate
                                    correlation_matrix.append(row)
                        else:
                            # Single asset - identity matrix
                            correlation_matrix = [[1.0]]
                    
                    except Exception as e:
                        self.logger.warning(f"Failed to calculate correlations with AssetList: {e}")
                        # Create a simple correlation matrix as fallback
                        correlation_matrix = []
                        for i in range(len(symbols)):
                            row = []
                            for j in range(len(symbols)):
                                if i == j:
                                    row.append(1.0)
                                else:
                                    # Estimate correlation based on asset types
                                    row.append(0.3)  # Conservative estimate
                            correlation_matrix.append(row)
                    
                    metrics_data['correlations'] = correlation_matrix
                    
                except Exception as e:
                    self.logger.warning(f"Failed to calculate correlations: {e}")
                    # Create identity matrix as fallback
                    correlation_matrix = []
                    for i in range(len(symbols)):
                        row = []
                        for j in range(len(symbols)):
                            row.append(1.0 if i == j else 0.0)
                        correlation_matrix.append(row)
                    metrics_data['correlations'] = correlation_matrix
            
            # Add portfolio context information
            if portfolio_contexts:
                portfolio_info = []
                for portfolio_context in portfolio_contexts:
                    portfolio_symbol = portfolio_context.get('symbol', 'Unknown')
                    portfolio_info.append(f"{portfolio_symbol}")
                metrics_data['additional_info'] = f"Включает портфели: {'; '.join(portfolio_info)}"
            
            return metrics_data
            
        except Exception as e:
            self.logger.error(f"Error preparing comprehensive metrics: {e}")
            return None

    def _create_summary_metrics_table(self, symbols: list, currency: str, expanded_symbols: list, portfolio_contexts: list, specified_period: str = None) -> str:
        """Create optimized metrics table using only okama.AssetList.describe data with additional metrics"""
        try:
            # Create AssetList for describe data
            asset_list = ok.AssetList(symbols, ccy=currency)
            describe_data = asset_list.describe()
            
            if describe_data is None or describe_data.empty:
                return "❌ Не удалось получить данные для анализа"
            
            # Prepare table data
            table_data = []
            headers = ["Метрика"] + symbols
            
            # Convert describe data to table format
            for idx in describe_data.index:
                property_name = describe_data.loc[idx, 'property']
                period = describe_data.loc[idx, 'period']
                
                # Create metric name with period information
                if pd.isna(period) or period == 'None' or period == '':
                    metric_name = str(property_name)
                else:
                    metric_name = f"{property_name} ({period})"
                
                row = [metric_name]
                for symbol in symbols:
                    if symbol in describe_data.columns:
                        value = describe_data.loc[idx, symbol]
                        if pd.isna(value):
                            row.append("N/A")
                        else:
                            # Format based on metric type
                            if isinstance(value, (int, float)):
                                if 'return' in str(property_name).lower() or 'cagr' in str(property_name).lower():
                                    row.append(f"{value*100:.2f}%")
                                elif 'volatility' in str(property_name).lower() or 'risk' in str(property_name).lower():
                                    row.append(f"{value*100:.2f}%")
                                elif 'ratio' in str(property_name).lower():
                                    row.append(f"{value:.3f}")
                                elif 'drawdown' in str(property_name).lower():
                                    row.append(f"{value*100:.2f}%")
                                elif 'yield' in str(property_name).lower():
                                    row.append(f"{value*100:.2f}%")
                                else:
                                    row.append(f"{value:.4f}")
                            else:
                                row.append(str(value))
                    else:
                        row.append("N/A")
                table_data.append(row)
            
            # Add additional metrics at the end
            self._add_risk_free_rate_row(table_data, symbols, currency)
            self._add_sharpe_ratio_row(table_data, symbols, currency, asset_list)
            self._add_sortino_ratio_row(table_data, symbols, currency)
            self._add_calmar_ratio_row(table_data, symbols, currency)
            
            # Create markdown table
            table_markdown = self._create_enhanced_markdown_table(table_data, headers)
            
            return f"{table_markdown}"
            
        except Exception as e:
            self.logger.error(f"Error creating summary metrics table: {e}")
            return "❌ Ошибка при создании таблицы метрик"

    def _add_risk_free_rate_row(self, table_data: list, symbols: list, currency: str):
        """Add risk-free rate row to table"""
        try:
            risk_free_rate = self.get_risk_free_rate(currency, 5.0)  # Use 5-year period
            risk_free_row = ["Risk free rate"]
            for symbol in symbols:
                risk_free_row.append(f"{risk_free_rate*100:.2f}%")
            table_data.append(risk_free_row)
        except Exception as e:
            self.logger.warning(f"Could not add risk-free rate row: {e}")
            risk_free_row = ["Risk free rate"] + ["N/A"] * len(symbols)
            table_data.append(risk_free_row)

    def _add_sharpe_ratio_row(self, table_data: list, symbols: list, currency: str, asset_list):
        """Add Sharpe ratio row using manual calculation with okama data"""
        try:
            risk_free_rate = self.get_risk_free_rate(currency, 5.0)
            sharpe_row = ["Sharpe Ratio"]
            
            for symbol in symbols:
                try:
                    # Calculate Sharpe ratio manually using CAGR and Risk from describe data
                    describe_data = asset_list.describe()
                    
                    # Find CAGR and Risk values for the symbol
                    cagr_value = None
                    risk_value = None
                    
                    # First, try to find CAGR for 5 years period
                    for idx in describe_data.index:
                        property_name = describe_data.loc[idx, 'property']
                        period = describe_data.loc[idx, 'period']
                        
                        if symbol in describe_data.columns:
                            value = describe_data.loc[idx, symbol]
                            if not pd.isna(value):
                                if property_name == 'CAGR' and period == '5 years':
                                    cagr_value = value
                                elif property_name == 'Risk' and period == '5 years':
                                    risk_value = value
                    
                    # Fallback: if no 5-year data, try to find any available CAGR and Risk
                    if cagr_value is None:
                        for idx in describe_data.index:
                            property_name = describe_data.loc[idx, 'property']
                            if symbol in describe_data.columns:
                                value = describe_data.loc[idx, symbol]
                                if not pd.isna(value) and property_name == 'CAGR':
                                    cagr_value = value
                                    break
                    
                    if risk_value is None:
                        for idx in describe_data.index:
                            property_name = describe_data.loc[idx, 'property']
                            if symbol in describe_data.columns:
                                value = describe_data.loc[idx, symbol]
                                if not pd.isna(value) and property_name == 'Risk':
                                    risk_value = value
                                    break
                    
                    if cagr_value is not None and risk_value is not None and risk_value > 0:
                        sharpe = (cagr_value - risk_free_rate) / risk_value
                        sharpe_row.append(f"{sharpe:.3f}")
                    else:
                        sharpe_row.append("N/A")
                        
                except Exception as e:
                    self.logger.warning(f"Could not calculate Sharpe ratio for {symbol}: {e}")
                    sharpe_row.append("N/A")
            
            table_data.append(sharpe_row)
        except Exception as e:
            self.logger.warning(f"Could not add Sharpe ratio row: {e}")
            sharpe_row = ["Sharpe Ratio"] + ["N/A"] * len(symbols)
            table_data.append(sharpe_row)

    def _add_sortino_ratio_row(self, table_data: list, symbols: list, currency: str):
        """Add Sortino ratio row"""
        try:
            risk_free_rate = self.get_risk_free_rate(currency, 5.0)
            sortino_row = ["Sortino Ratio"]
            
            for symbol in symbols:
                try:
                    asset = ok.Asset(symbol)
                    if hasattr(asset, 'close_monthly') and asset.close_monthly is not None:
                        prices = asset.close_monthly
                        returns = prices.pct_change().dropna()
                        
                        if len(returns) > 1:
                            # Calculate downside deviation (only negative returns)
                            downside_returns = returns[returns < 0]
                            if len(downside_returns) > 1:
                                downside_deviation = downside_returns.std() * np.sqrt(12)  # Annualized
                                if downside_deviation > 0:
                                    # Get CAGR from asset
                                    cagr = asset.annual_return if hasattr(asset, 'annual_return') else 0
                                    sortino = (cagr - risk_free_rate) / downside_deviation
                                    sortino_row.append(f"{sortino:.3f}")
                                else:
                                    sortino_row.append("N/A")
                            else:
                                sortino_row.append("N/A")
                        else:
                            sortino_row.append("N/A")
                    else:
                        sortino_row.append("N/A")
                except Exception as e:
                    self.logger.warning(f"Could not calculate Sortino ratio for {symbol}: {e}")
                    sortino_row.append("N/A")
            
            table_data.append(sortino_row)
        except Exception as e:
            self.logger.warning(f"Could not add Sortino ratio row: {e}")
            sortino_row = ["Sortino Ratio"] + ["N/A"] * len(symbols)
            table_data.append(sortino_row)

    def _add_calmar_ratio_row(self, table_data: list, symbols: list, currency: str):
        """Add Calmar ratio row using describe data"""
        try:
            calmar_row = ["Calmar Ratio"]
            
            # Create AssetList to get describe data
            asset_list = ok.AssetList(symbols, ccy=currency)
            describe_data = asset_list.describe()
            
            for symbol in symbols:
                try:
                    # Find CAGR and Max drawdown values for the symbol
                    cagr_value = None
                    max_drawdown_value = None
                    
                    # Try to find 5-year period first, then fallback to available periods
                    for idx in describe_data.index:
                        property_name = describe_data.loc[idx, 'property']
                        period = describe_data.loc[idx, 'period']
                        
                        if symbol in describe_data.columns:
                            value = describe_data.loc[idx, symbol]
                            if not pd.isna(value):
                                # Look for 5-year CAGR first
                                if property_name == 'CAGR' and ('5 years' in str(period) or period == '5 years'):
                                    cagr_value = value
                                # Look for Max drawdowns - try 5-year first, then any available
                                elif property_name == 'Max drawdowns' and ('5 years' in str(period) or period == '5 years'):
                                    max_drawdown_value = value
                    
                    # If we didn't find 5-year Max drawdowns, try to find any Max drawdowns period
                    if max_drawdown_value is None:
                        for idx in describe_data.index:
                            property_name = describe_data.loc[idx, 'property']
                            period = describe_data.loc[idx, 'period']
                            
                            if symbol in describe_data.columns:
                                value = describe_data.loc[idx, symbol]
                                if not pd.isna(value) and property_name == 'Max drawdowns':
                                    max_drawdown_value = value
                                    break
                    
                    # If we didn't find 5-year CAGR, try to find any CAGR period
                    if cagr_value is None:
                        for idx in describe_data.index:
                            property_name = describe_data.loc[idx, 'property']
                            period = describe_data.loc[idx, 'period']
                            
                            if symbol in describe_data.columns:
                                value = describe_data.loc[idx, symbol]
                                if not pd.isna(value) and property_name == 'CAGR':
                                    cagr_value = value
                                    break
                    
                    if cagr_value is not None and max_drawdown_value is not None and max_drawdown_value < 0:
                        calmar = cagr_value / abs(max_drawdown_value)
                        calmar_row.append(f"{calmar:.3f}")
                    else:
                        calmar_row.append("N/A")
                        
                except Exception as e:
                    self.logger.warning(f"Could not calculate Calmar ratio for {symbol}: {e}")
                    calmar_row.append("N/A")
            
            table_data.append(calmar_row)
        except Exception as e:
            self.logger.warning(f"Could not add Calmar ratio row: {e}")
            calmar_row = ["Calmar Ratio"] + ["N/A"] * len(symbols)
            table_data.append(calmar_row)


    def _create_portfolio_summary_metrics_table(self, portfolio, symbols: list, weights: list, currency: str) -> str:
        """Create summary metrics table for a single portfolio using ok.Portfolio.describe() (same format as _create_summary_metrics_table)"""
        try:
            # Use portfolio.describe() to get the same metrics as AssetList.describe()
            describe_data = portfolio.describe()
            
            if describe_data is None or describe_data.empty:
                return "❌ Не удалось получить данные для анализа"
            
            # Prepare table data - single column for portfolio
            table_data = []
            headers = ["Метрика", "Значение"]
            
            # Convert describe data to table format (same logic as _create_summary_metrics_table)
            for idx in describe_data.index:
                property_name = describe_data.loc[idx, 'property']
                period = describe_data.loc[idx, 'period']
                
                # Create metric name with period information
                if pd.isna(period) or period == 'None' or period == '':
                    metric_name = str(property_name)
                else:
                    metric_name = f"{property_name} ({period})"
                
                # Get value from describe data - find the portfolio column
                portfolio_column = None
                for col in describe_data.columns:
                    if col.startswith('portfolio_') and col.endswith('.PF'):
                        portfolio_column = col
                        break
                
                if portfolio_column:
                    value = describe_data.loc[idx, portfolio_column]
                else:
                    # Fallback: use first non-property, non-period column
                    value_cols = [col for col in describe_data.columns if col not in ['property', 'period', 'inflation']]
                    if value_cols:
                        value = describe_data.loc[idx, value_cols[0]]
                    else:
                        value = None
                
                if pd.isna(value):
                    formatted_value = "N/A"
                else:
                    # Format based on metric type (same logic as _create_summary_metrics_table)
                    if isinstance(value, (int, float)):
                        if 'return' in str(property_name).lower() or 'cagr' in str(property_name).lower():
                            formatted_value = f"{value*100:.2f}%"
                        elif 'volatility' in str(property_name).lower() or 'risk' in str(property_name).lower():
                            formatted_value = f"{value*100:.2f}%"
                        elif 'ratio' in str(property_name).lower():
                            formatted_value = f"{value:.3f}"
                        elif 'drawdown' in str(property_name).lower():
                            formatted_value = f"{value*100:.2f}%"
                        elif 'yield' in str(property_name).lower():
                            formatted_value = f"{value*100:.2f}%"
                        else:
                            formatted_value = f"{value:.4f}"
                    else:
                        formatted_value = str(value)
                
                table_data.append([metric_name, formatted_value])
            
            # Add additional metrics at the end (same as _create_summary_metrics_table)
            self._add_portfolio_risk_free_rate_row(table_data, currency)
            self._add_portfolio_sharpe_ratio_row(table_data, currency, portfolio)
            self._add_portfolio_sortino_ratio_row(table_data, currency, portfolio)
            self._add_portfolio_calmar_ratio_row(table_data, currency, portfolio)
            
            # Create markdown table using the same function as _create_summary_metrics_table
            table_markdown = self._create_enhanced_markdown_table(table_data, headers)
            
            return f"{table_markdown}"
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio summary metrics table: {e}")
            return "❌ Ошибка при создании таблицы метрик"

    def _add_portfolio_risk_free_rate_row(self, table_data: list, currency: str):
        """Add risk-free rate row to portfolio table"""
        try:
            risk_free_rate = self.get_risk_free_rate(currency, 5.0)  # Use 5-year period
            table_data.append(["Risk free rate", f"{risk_free_rate*100:.2f}%"])
        except Exception as e:
            self.logger.warning(f"Could not add risk-free rate row: {e}")
            table_data.append(["Risk free rate", "N/A"])

    def _add_portfolio_sharpe_ratio_row(self, table_data: list, currency: str, portfolio):
        """Add Sharpe ratio row for portfolio using describe data"""
        try:
            risk_free_rate = self.get_risk_free_rate(currency, 5.0)
            describe_data = portfolio.describe()
            
            # Find CAGR and Risk values from describe data
            cagr_value = None
            risk_value = None
            
            # Find the portfolio column
            portfolio_column = None
            for col in describe_data.columns:
                if col.startswith('portfolio_') and col.endswith('.PF'):
                    portfolio_column = col
                    break
            
            if portfolio_column:
                for idx in describe_data.index:
                    property_name = describe_data.loc[idx, 'property']
                    period = describe_data.loc[idx, 'period']
                    
                    value = describe_data.loc[idx, portfolio_column]
                    if not pd.isna(value):
                        if property_name == 'CAGR' and period == '5 years':
                            cagr_value = value
                        elif property_name == 'Risk':
                            risk_value = value
            
            if cagr_value is not None and risk_value is not None and risk_value > 0:
                sharpe = (cagr_value - risk_free_rate) / risk_value
                table_data.append(["Sharpe Ratio", f"{sharpe:.3f}"])
            else:
                table_data.append(["Sharpe Ratio", "N/A"])
                
        except Exception as e:
            self.logger.warning(f"Could not add Sharpe ratio row: {e}")
            table_data.append(["Sharpe Ratio", "N/A"])

    def _add_portfolio_sortino_ratio_row(self, table_data: list, currency: str, portfolio):
        """Add Sortino ratio row for portfolio"""
        try:
            risk_free_rate = self.get_risk_free_rate(currency, 5.0)
            
            # Get price data from portfolio
            if hasattr(portfolio, 'wealth_index') and portfolio.wealth_index is not None:
                prices = portfolio.wealth_index
                if hasattr(prices, 'iloc') and len(prices) > 1:
                    if hasattr(prices, 'columns'):
                        prices = prices.iloc[:, 0]
                    
                    returns = prices.pct_change().dropna()
                    
                    if len(returns) > 1:
                        # Calculate downside deviation (only negative returns)
                        downside_returns = returns[returns < 0]
                        if len(downside_returns) > 1:
                            downside_deviation = downside_returns.std() * np.sqrt(12)  # Annualized
                            if downside_deviation > 0:
                                # Get CAGR from describe data
                                describe_data = portfolio.describe()
                                cagr_value = None
                                
                                # Find the portfolio column
                                portfolio_column = None
                                for col in describe_data.columns:
                                    if col.startswith('portfolio_') and col.endswith('.PF'):
                                        portfolio_column = col
                                        break
                                
                                if portfolio_column:
                                    for idx in describe_data.index:
                                        property_name = describe_data.loc[idx, 'property']
                                        period = describe_data.loc[idx, 'period']
                                        if property_name == 'CAGR' and period == '5 years':
                                            cagr_value = describe_data.loc[idx, portfolio_column]
                                            break
                                
                                if cagr_value is not None:
                                    sortino = (cagr_value - risk_free_rate) / downside_deviation
                                    table_data.append(["Sortino Ratio", f"{sortino:.3f}"])
                                else:
                                    table_data.append(["Sortino Ratio", "N/A"])
                            else:
                                table_data.append(["Sortino Ratio", "N/A"])
                        else:
                            table_data.append(["Sortino Ratio", "N/A"])
                    else:
                        table_data.append(["Sortino Ratio", "N/A"])
                else:
                    table_data.append(["Sortino Ratio", "N/A"])
            else:
                table_data.append(["Sortino Ratio", "N/A"])
                
        except Exception as e:
            self.logger.warning(f"Could not add Sortino ratio row: {e}")
            table_data.append(["Sortino Ratio", "N/A"])

    def _add_portfolio_calmar_ratio_row(self, table_data: list, currency: str, portfolio):
        """Add Calmar ratio row for portfolio using describe data"""
        try:
            describe_data = portfolio.describe()
            
            # Find CAGR and Max drawdown values from describe data
            cagr_value = None
            max_drawdown_value = None
            
            # Find the portfolio column
            portfolio_column = None
            for col in describe_data.columns:
                if col.startswith('portfolio_') and col.endswith('.PF'):
                    portfolio_column = col
                    break
            
            if portfolio_column:
                for idx in describe_data.index:
                    property_name = describe_data.loc[idx, 'property']
                    period = describe_data.loc[idx, 'period']
                    
                    value = describe_data.loc[idx, portfolio_column]
                    if not pd.isna(value):
                        if property_name == 'CAGR' and period == '5 years':
                            cagr_value = value
                        elif property_name == 'Max drawdown':
                            max_drawdown_value = value
            
            if cagr_value is not None and max_drawdown_value is not None and max_drawdown_value < 0:
                calmar = cagr_value / abs(max_drawdown_value)
                table_data.append(["Calmar Ratio", f"{calmar:.3f}"])
            else:
                table_data.append(["Calmar Ratio", "N/A"])
                
        except Exception as e:
            self.logger.warning(f"Could not add Calmar ratio row: {e}")
            table_data.append(["Calmar Ratio", "N/A"])

    def _create_portfolio_metrics_table(self, portfolio_symbol: str, symbols: list, weights: list, currency: str, portfolio_object) -> str:
        """Create metrics table for a single portfolio (similar to _create_summary_metrics_table but for one portfolio)"""
        try:
            # Prepare table data
            table_data = []
            headers = ["Метрика", "Значение"]
            
            # Calculate key metrics for the portfolio
            portfolio_metrics = {}
            
            if portfolio_object is not None:
                # Get price data from portfolio
                prices = None
                
                # Try to get price data from different sources
                if hasattr(portfolio_object, 'wealth_index') and portfolio_object.wealth_index is not None:
                    wealth_index = portfolio_object.wealth_index
                    if hasattr(wealth_index, 'iloc') and len(wealth_index) > 1:
                        if hasattr(wealth_index, 'columns'):
                            prices = wealth_index.iloc[:, 0]
                        else:
                            prices = wealth_index
                
                if prices is not None and len(prices) > 1:
                    # CAGR (Annual Return)
                    try:
                        start_date = prices.index[0]
                        end_date = prices.index[-1]
                        
                        # Calculate years based on actual date range
                        try:
                            # Handle different date types (Period, Timestamp, etc.)
                            if hasattr(start_date, 'to_timestamp'):
                                start_date = start_date.to_timestamp()
                            if hasattr(end_date, 'to_timestamp'):
                                end_date = end_date.to_timestamp()
                            
                            if hasattr(start_date, 'year') and hasattr(end_date, 'year'):
                                years = (end_date - start_date).days / 365.25
                            else:
                                years = len(prices) / 12  # Fallback: assuming monthly data
                        except Exception:
                            years = len(prices) / 12  # Fallback: assuming monthly data
                        
                        if years > 0:
                            total_return = (prices.iloc[-1] / prices.iloc[0]) - 1
                            portfolio_metrics['cagr'] = (1 + total_return) ** (1 / years) - 1
                            
                    except Exception as e:
                        self.logger.warning(f"Could not calculate CAGR for portfolio: {e}")
                        portfolio_metrics['cagr'] = None
                    
                    # Volatility
                    try:
                        returns = prices.pct_change().dropna()
                        if len(returns) > 1:
                            # Annualize volatility based on data frequency
                            if hasattr(prices.index, 'freq') and prices.index.freq:
                                freq_str = str(prices.index.freq)
                                if 'D' in freq_str:  # Daily data
                                    portfolio_metrics['volatility'] = returns.std() * (252 ** 0.5)  # Annualized
                                elif 'M' in freq_str:  # Monthly data
                                    portfolio_metrics['volatility'] = returns.std() * (12 ** 0.5)  # Annualized
                                else:
                                    portfolio_metrics['volatility'] = returns.std() * (12 ** 0.5)  # Default to monthly
                            else:
                                # Fallback: assume monthly data
                                portfolio_metrics['volatility'] = returns.std() * (12 ** 0.5)
                    except Exception as e:
                        self.logger.warning(f"Could not calculate volatility for portfolio: {e}")
                        portfolio_metrics['volatility'] = None
                    
                    # Sharpe Ratio
                    try:
                        # Calculate Sharpe ratio manually using CAGR and volatility
                        if portfolio_metrics.get('cagr') is not None and portfolio_metrics.get('volatility') is not None and portfolio_metrics['volatility'] > 0:
                            # Calculate years for period-based rate selection
                            years = None
                            if prices is not None and len(prices) > 1:
                                start_date = prices.index[0]
                                end_date = prices.index[-1]
                                if hasattr(start_date, 'to_timestamp'):
                                    start_date = start_date.to_timestamp()
                                if hasattr(end_date, 'to_timestamp'):
                                    end_date = end_date.to_timestamp()
                                years = (end_date - start_date).days / 365.25
                            
                            # Use proper risk-free rate based on currency
                            risk_free_rate = self.get_risk_free_rate(currency, years)
                            portfolio_metrics['sharpe'] = (portfolio_metrics['cagr'] - risk_free_rate) / portfolio_metrics['volatility']
                            portfolio_metrics['risk_free_rate'] = risk_free_rate
                    except Exception as e:
                        self.logger.warning(f"Could not calculate Sharpe ratio for portfolio: {e}")
                        portfolio_metrics['sharpe'] = None
                    
                    # Max Drawdown
                    try:
                        # Calculate max drawdown from price data
                        running_max = prices.expanding().max()
                        drawdown = (prices - running_max) / running_max
                        portfolio_metrics['max_drawdown'] = drawdown.min()
                    except Exception as e:
                        self.logger.warning(f"Could not calculate max drawdown for portfolio: {e}")
                        portfolio_metrics['max_drawdown'] = None
                    
                    # Sortino Ratio
                    try:
                        # Calculate Sortino ratio manually using CAGR and downside deviation
                        if prices is not None and len(prices) > 1:
                            returns = prices.pct_change().dropna()
                            if len(returns) > 1:
                                # Calculate downside deviation (only negative returns)
                                downside_returns = returns[returns < 0]
                                if len(downside_returns) > 1:
                                    downside_deviation = downside_returns.std()
                                    if downside_deviation > 0 and portfolio_metrics.get('cagr') is not None:
                                        # Use the same risk-free rate as calculated for Sharpe ratio
                                        risk_free_rate = portfolio_metrics.get('risk_free_rate', self.get_risk_free_rate(currency))
                                        portfolio_metrics['sortino'] = (portfolio_metrics['cagr'] - risk_free_rate) / downside_deviation
                    except Exception as e:
                        self.logger.warning(f"Could not calculate Sortino ratio for portfolio: {e}")
                        portfolio_metrics['sortino'] = None
                    
                    # Calmar Ratio
                    try:
                        # Calculate Calmar ratio using CAGR and max drawdown
                        if portfolio_metrics.get('cagr') is not None and portfolio_metrics.get('max_drawdown') is not None and portfolio_metrics['max_drawdown'] < 0:
                            portfolio_metrics['calmar'] = portfolio_metrics['cagr'] / abs(portfolio_metrics['max_drawdown'])
                    except Exception as e:
                        self.logger.warning(f"Could not calculate Calmar ratio for portfolio: {e}")
                        portfolio_metrics['calmar'] = None
                    
                    # VaR 95% and CVaR 95%
                    try:
                        # Use the same price data we found for CAGR calculation
                        if prices is not None and len(prices) > 1:
                            returns = prices.pct_change().dropna()
                            if len(returns) > 1:
                                # Calculate VaR 95% (5th percentile of returns)
                                var_95 = returns.quantile(0.05)
                                portfolio_metrics['var_95'] = var_95
                                
                                # Calculate CVaR 95% (expected value of returns below VaR 95%)
                                cvar_95 = returns[returns <= var_95].mean()
                                portfolio_metrics['cvar_95'] = cvar_95
                    except Exception as e:
                        self.logger.warning(f"Could not calculate VaR/CVaR for portfolio: {e}")
                        portfolio_metrics['var_95'] = None
                        portfolio_metrics['cvar_95'] = None
            
            # Create table rows
            # CAGR row
            cagr_row = ["**CAGR**"]
            if portfolio_metrics.get('cagr') is not None:
                cagr_row.append(f"{portfolio_metrics['cagr']*100:.2f}%")
            else:
                cagr_row.append("N/A")
            table_data.append(cagr_row)
            
            # Volatility row
            volatility_row = ["**Волатильность**"]
            if portfolio_metrics.get('volatility') is not None:
                volatility_row.append(f"{portfolio_metrics['volatility']*100:.2f}%")
            else:
                volatility_row.append("N/A")
            table_data.append(volatility_row)
            
            # Sharpe Ratio row
            sharpe_row = ["**Коэф. Шарпа**"]
            if portfolio_metrics.get('sharpe') is not None:
                sharpe_row.append(f"{portfolio_metrics['sharpe']:.3f}")
            else:
                sharpe_row.append("N/A")
            table_data.append(sharpe_row)
            
            # Max Drawdown row
            max_drawdown_row = ["**Макс. просадка**"]
            if portfolio_metrics.get('max_drawdown') is not None:
                max_drawdown_row.append(f"{portfolio_metrics['max_drawdown']*100:.2f}%")
            else:
                max_drawdown_row.append("N/A")
            table_data.append(max_drawdown_row)
            
            # Sortino Ratio row
            sortino_row = ["**Коэф. Сортино**"]
            if portfolio_metrics.get('sortino') is not None:
                sortino_row.append(f"{portfolio_metrics['sortino']:.3f}")
            else:
                sortino_row.append("N/A")
            table_data.append(sortino_row)
            
            # Calmar Ratio row
            calmar_row = ["**Коэф. Кальмара**"]
            if portfolio_metrics.get('calmar') is not None:
                calmar_row.append(f"{portfolio_metrics['calmar']:.3f}")
            else:
                calmar_row.append("N/A")
            table_data.append(calmar_row)
            
            # VaR 95% row
            var_row = ["**VaR 95%**"]
            if portfolio_metrics.get('var_95') is not None:
                var_row.append(f"{portfolio_metrics['var_95']*100:.2f}%")
            else:
                var_row.append("N/A")
            table_data.append(var_row)
            
            # CVaR 95% row
            cvar_row = ["**CVaR 95%**"]
            if portfolio_metrics.get('cvar_95') is not None:
                cvar_row.append(f"{portfolio_metrics['cvar_95']*100:.2f}%")
            else:
                cvar_row.append("N/A")
            table_data.append(cvar_row)
            
            # Create enhanced markdown table with better formatting
            table_markdown = self._create_enhanced_markdown_table(table_data, headers)
            
            return table_markdown
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio metrics table: {e}")
            return "❌ Ошибка при создании таблицы метрик портфеля"

    def _format_period_for_display(self, specified_period: str = None) -> str:
        """Format period string for display in headers"""
        if not specified_period:
            return ""
        
        # Handle different period formats
        if specified_period.upper().endswith('Y'):
            years = specified_period[:-1]
            try:
                years_int = int(years)
                if years_int == 1:
                    return "за 1 год"
                elif years_int in [2, 3, 4]:
                    return f"за {years_int} года"
                else:
                    return f"за {years_int} лет"
            except ValueError:
                return f"за {specified_period}"
        
        return f"за {specified_period}"

    def _create_enhanced_markdown_table(self, table_data: list, headers: list) -> str:
        """Create Telegram-compatible markdown table with better formatting"""
        try:
            # Use tabulate with pipe format for Telegram compatibility
            # This ensures proper markdown table format without HTML elements
            
            # Format the data for better readability
            formatted_data = []
            for row in table_data:
                formatted_row = []
                # First column: metric name (keep as is)
                formatted_row.append(str(row[0]))
                # Other columns: values with proper spacing
                for i in range(1, len(row)):
                    formatted_row.append(str(row[i]))
                formatted_data.append(formatted_row)
            
            # Create markdown table using tabulate
            table_markdown = tabulate.tabulate(formatted_data, headers=headers, tablefmt="pipe")
            
            return table_markdown
            
        except Exception as e:
            self.logger.error(f"Error creating enhanced markdown table: {e}")
            # Fallback to simple tabulate
            return tabulate.tabulate(table_data, headers=headers, tablefmt="pipe")

    def _create_compare_command_keyboard(self, symbols: list, currency: str, update: Update = None, specified_period: str = None) -> InlineKeyboardMarkup:
        """Create keyboard for compare command button responses"""
        try:
            keyboard = []
            
            # Remove inline keyboards for comparison analysis buttons as requested
            # Only keep AI analysis button if available
            if self.gemini_service and self.gemini_service.is_available():
                keyboard.append([
                    InlineKeyboardButton("🧠 AI-анализ", callback_data="data_analysis_compare")
                ])
            
            # Add Portfolio button - store symbols in context to avoid callback_data size limit
            # Store symbols in user context for portfolio button
            if update:
                user_id = update.effective_user.id
                user_context = self._get_user_context(user_id)
                user_context['compare_portfolio_symbols'] = symbols
            
            keyboard.append([
                InlineKeyboardButton("💼 В Портфель", callback_data="compare_portfolio")
            ])
            
            return InlineKeyboardMarkup(keyboard)
            
        except Exception as e:
            self.logger.error(f"Error creating compare command keyboard: {e}")
            # Return empty keyboard as fallback
            return InlineKeyboardMarkup([])

    def _create_portfolio_command_keyboard(self, portfolio_symbol: str) -> ReplyKeyboardMarkup:
        """Create Reply Keyboard for portfolio command button responses"""
        try:
            # Use the existing reply keyboard function
            return self._create_portfolio_reply_keyboard()
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio command keyboard: {e}")
            # Return empty keyboard as fallback
            return ReplyKeyboardMarkup([])

    def _create_portfolio_reply_keyboard(self) -> ReplyKeyboardMarkup:
        """Create Reply Keyboard for portfolio command with three rows of buttons"""
        try:
            keyboard = [
                # Первый ряд
                [
                    KeyboardButton("▫️ Накоп. доходность"),
                    KeyboardButton("▫️ Доходность ГГ"),
                    KeyboardButton("▫️ Динамика дох."),
                    KeyboardButton("▫️ Дивиденды")
                ],
                # Второй ряд
                [
                    KeyboardButton("▫️ Метрики"),
                    KeyboardButton("▫️ Монте-Карло"),
                    KeyboardButton("▫️ Процентили (10/50/90)"),
                    KeyboardButton("▫️ Просадки")
                ],
                # Третий ряд
                [
                    KeyboardButton("▫️ Нейроанализ"),
                    KeyboardButton("▫️ Портфель vs Активы"),
                    KeyboardButton("▫️ Сравнить")
                ]
            ]
            
            return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio reply keyboard: {e}")
            # Return empty keyboard as fallback
            return ReplyKeyboardMarkup([])

    def _create_compare_reply_keyboard(self) -> ReplyKeyboardMarkup:
        """Create Reply Keyboard for compare command with three rows of buttons"""
        try:
            keyboard = [
                # Первый ряд
                [
                    KeyboardButton("▫️ Доходность"),
                    KeyboardButton("▫️ Дивиденды"),
                    KeyboardButton("▫️ Просадки")
                ],
                # Второй ряд
                [
                    KeyboardButton("▫️ Метрики"),
                    KeyboardButton("▫️ Корреляция"),
                    KeyboardButton("▫️ Эффективная граница")
                ],
                # Третий ряд
                [
                    KeyboardButton("▫️ Нейроанализ"),
                    KeyboardButton("▫️ В Портфель")
                ]
            ]
            
            return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
            
        except Exception as e:
            self.logger.error(f"Error creating compare reply keyboard: {e}")
            # Return empty keyboard as fallback
            return ReplyKeyboardMarkup([])

    def _create_list_namespace_reply_keyboard(self, namespace: str, current_page: int, total_pages: int, total_symbols: int) -> ReplyKeyboardMarkup:
        """Create Reply Keyboard for /list <код> command with navigation and action buttons"""
        try:
            keyboard = []
            
            # Navigation buttons (only if more than one page)
            if total_pages > 1:
                nav_buttons = []
                
                # Previous button
                if current_page > 0:
                    nav_buttons.append(KeyboardButton("⬅️ Назад"))
                
                # Page indicator
                nav_buttons.append(KeyboardButton(f"{current_page + 1}/{total_pages}"))
                
                # Next button
                if current_page < total_pages - 1:
                    nav_buttons.append(KeyboardButton("➡️ Вперед"))
                
                keyboard.append(nav_buttons)
            
            # Action buttons
            keyboard.append([
                KeyboardButton("📊 Excel"),
                KeyboardButton("🔍 Анализ"),
                KeyboardButton("⚖️ Сравнить")
            ])
            
            keyboard.append([
                KeyboardButton("💼 В портфель"),
                KeyboardButton("📚 База данных")
            ])
            
            return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
            
        except Exception as e:
            self.logger.error(f"Error creating list namespace reply keyboard: {e}")
            # Return empty keyboard as fallback
            return ReplyKeyboardMarkup([])

    def _create_namespace_reply_keyboard(self) -> ReplyKeyboardMarkup:
        """Create Reply Keyboard for /list command with namespace buttons"""
        try:
            keyboard = []
            
            # Основные биржи
            keyboard.append([
                KeyboardButton("🇺🇸 US"),
                KeyboardButton("🇷🇺 MOEX"),
                KeyboardButton("🇬🇧 LSE")
            ])
            
            # Европейские биржи
            keyboard.append([
                KeyboardButton("🇩🇪 XETR"),
                KeyboardButton("🇫🇷 XFRA"),
                KeyboardButton("🇳🇱 XAMS")
            ])
            
            # Китайские биржи
            keyboard.append([
                KeyboardButton("🇨🇳 SSE"),
                KeyboardButton("🇨🇳 SZSE"),
                KeyboardButton("🇨🇳 BSE")
            ])
            
            keyboard.append([
                KeyboardButton("🇭🇰 HKEX")
            ])
            
            # Индексы и валюты
            keyboard.append([
                KeyboardButton("📊 INDX"),
                KeyboardButton("💱 FX"),
                KeyboardButton("🏦 CBR")
            ])
            
            # Товары и криптовалюты
            keyboard.append([
                KeyboardButton("🛢️ COMM"),
                KeyboardButton("₿ CC"),
                KeyboardButton("🏠 RE")
            ])
            
            # Инфляция и депозиты
            keyboard.append([
                KeyboardButton("📈 INFL"),
                KeyboardButton("💰 PIF"),
                KeyboardButton("🏦 RATE")
            ])
            
            return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
            
        except Exception as e:
            self.logger.error(f"Error creating namespace reply keyboard: {e}")
            # Return empty keyboard as fallback
            return ReplyKeyboardMarkup([])

    def _create_info_reply_keyboard(self) -> ReplyKeyboardMarkup:
        """Create Reply Keyboard for /info command with period and action buttons"""
        try:
            keyboard = []
            
            # Row 1: Period buttons
            keyboard.append([
                KeyboardButton("1 год"),
                KeyboardButton("5 лет"),
                KeyboardButton("Макс. срок"),
                KeyboardButton("Дивиденды")
            ])
            
            # Row 2: Action buttons
            keyboard.append([
                KeyboardButton("⚖️ Сравнение"),
                KeyboardButton("💼 В Портфель")
            ])
            
            return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
            
        except Exception as e:
            self.logger.error(f"Error creating info reply keyboard: {e}")
            # Return empty keyboard as fallback
            return ReplyKeyboardMarkup([])

    def _create_start_reply_keyboard(self) -> ReplyKeyboardMarkup:
        """Create Reply Keyboard for /start command with main action buttons"""
        try:
            keyboard = []
            
            # Row 1: Main actions
            keyboard.append([
                KeyboardButton("Анализ"),
                KeyboardButton("Сравнение"),
                KeyboardButton("Портфель")
            ])
            
            # Row 2: Secondary actions
            keyboard.append([
                KeyboardButton("База данных"),
                KeyboardButton("Справка")
            ])
            
            return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
            
        except Exception as e:
            self.logger.error(f"Error creating start reply keyboard: {e}")
            # Return empty keyboard as fallback
            return ReplyKeyboardMarkup([])

    async def _show_portfolio_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show Reply Keyboard for portfolio management"""
        try:
            portfolio_reply_keyboard = self._create_portfolio_reply_keyboard()
            # Send persistent message with keyboard
            await self._send_message_safe(
                update, 
                "", 
                reply_markup=portfolio_reply_keyboard
            )
        except Exception as e:
            self.logger.error(f"Error showing portfolio reply keyboard: {e}")

    async def _show_compare_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Show Reply Keyboard for compare management"""
        try:
            compare_reply_keyboard = self._create_compare_reply_keyboard()
            # Send persistent message with keyboard
            await self._send_message_safe(
                update, 
                "⚖️ Выберите действие для сравнения:", 
                reply_markup=compare_reply_keyboard
            )
        except Exception as e:
            self.logger.error(f"Error showing compare reply keyboard: {e}")

    def _is_portfolio_reply_keyboard_button(self, text: str) -> bool:
        """Check if the text is a portfolio Reply Keyboard button"""
        portfolio_buttons = [
            "▫️ Накоп. доходность",
            "▫️ Доходность ГГ",
            "▫️ Динамика дох.",
            "▫️ Дивиденды",
            "▫️ Метрики",
            "▫️ Монте-Карло",
            "▫️ Процентили (10/50/90)",
            "▫️ Просадки",
            "▫️ Нейроанализ",
            "▫️ Портфель vs Активы",
            "▫️ Сравнить"
        ]
        return text in portfolio_buttons

    def _is_compare_reply_keyboard_button(self, text: str) -> bool:
        """Check if the text is a compare Reply Keyboard button"""
        compare_buttons = [
            "▫️ Доходность",
            "▫️ Дивиденды",
            "▫️ Просадки",
            "▫️ Метрики",
            "▫️ Корреляция",
            "▫️ Эффективная граница",
            "▫️ Нейроанализ",
            "▫️ В Портфель"
        ]
        return text in compare_buttons

    def _is_list_reply_keyboard_button(self, text: str) -> bool:
        """Check if the text is a list Reply Keyboard button"""
        list_buttons = [
            "⬅️ Назад",
            "➡️ Вперед",
            "📊 Excel",
            "🔍 Анализ",
            "⚖️ Сравнить",
            "💼 В портфель",
            "📚 База данных"
        ]
        # Also check for page indicators like "1/5", "2/5", etc.
        import re
        page_pattern = r'^\d+/\d+$'
        return text in list_buttons or bool(re.match(page_pattern, text))

    def _is_namespace_reply_keyboard_button(self, text: str) -> bool:
        """Check if the text is a namespace Reply Keyboard button"""
        namespace_buttons = [
            "🇺🇸 US", "🇷🇺 MOEX", "🇬🇧 LSE",
            "🇩🇪 XETR", "🇫🇷 XFRA", "🇳🇱 XAMS",
            "🇨🇳 SSE", "🇨🇳 SZSE", "🇨🇳 BSE", "🇭🇰 HKEX",
            "📊 INDX", "💱 FX", "🏦 CBR",
            "🛢️ COMM", "₿ CC", "🏠 RE",
            "📈 INFL", "💰 PIF", "🏦 RATE"
        ]
        return text in namespace_buttons

    def _is_info_reply_keyboard_button(self, text: str) -> bool:
        """Check if the text is an info Reply Keyboard button"""
        info_buttons = [
            "1 год",
            "5 лет", 
            "Макс. срок",
            "Дивиденды",
            "⚖️ Сравнение",
            "💼 В Портфель"
        ]
        return text in info_buttons

    def _is_start_reply_keyboard_button(self, text: str) -> bool:
        """Check if the text is a start Reply Keyboard button"""
        start_buttons = [
            "Анализ",
            "Сравнение",
            "Портфель",
            "База данных",
            "Справка"
        ]
        return text in start_buttons

    def _is_reply_keyboard_button(self, text: str) -> bool:
        """Check if the text is any Reply Keyboard button (portfolio, compare, list, namespace, info, or start)"""
        return (self._is_portfolio_reply_keyboard_button(text) or 
                self._is_compare_reply_keyboard_button(text) or 
                self._is_list_reply_keyboard_button(text) or
                self._is_namespace_reply_keyboard_button(text) or
                self._is_info_reply_keyboard_button(text) or
                self._is_start_reply_keyboard_button(text))

    async def _handle_reply_keyboard_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle Reply Keyboard button presses - determine context and call appropriate handler"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Determine context based on user's last activity and available data
            last_assets = user_context.get('last_assets', [])
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Check if button exists in different contexts
            is_compare_button = self._is_compare_reply_keyboard_button(text)
            is_portfolio_button = self._is_portfolio_reply_keyboard_button(text)
            is_list_button = self._is_list_reply_keyboard_button(text)
            is_namespace_button = self._is_namespace_reply_keyboard_button(text)
            is_info_button = self._is_info_reply_keyboard_button(text)
            is_start_button = self._is_start_reply_keyboard_button(text)
            
            if is_start_button:
                # Handle start buttons (from /start command)
                await self._handle_start_reply_keyboard_button(update, context, text)
            elif is_info_button:
                # Handle info buttons (from /info command)
                await self._handle_info_reply_keyboard_button(update, context, text)
            elif is_namespace_button:
                # Handle namespace buttons (from /list command)
                await self._handle_namespace_reply_keyboard_button(update, context, text)
            elif is_list_button:
                # Handle list namespace buttons
                await self._handle_list_reply_keyboard_button(update, context, text)
            elif is_compare_button and is_portfolio_button:
                # Button exists in both contexts - determine by last analysis type and data availability
                last_analysis_type = user_context.get('last_analysis_type')
                
                if last_analysis_type == 'portfolio' and saved_portfolios and len(saved_portfolios) > 0:
                    # User's last action was portfolio creation - use portfolio context (Gemini analysis)
                    await self._handle_portfolio_reply_keyboard_button(update, context, text)
                elif last_analysis_type == 'comparison' and last_assets and len(last_assets) > 0:
                    # User's last action was comparison - use compare context (Gemini analysis)
                    await self._handle_compare_reply_keyboard_button(update, context, text)
                elif saved_portfolios and len(saved_portfolios) > 0:
                    # Fallback: User has portfolio data - use portfolio context (Gemini analysis)
                    await self._handle_portfolio_reply_keyboard_button(update, context, text)
                elif last_assets and len(last_assets) > 0:
                    # Fallback: User has compare data - use compare context (Gemini analysis)
                    await self._handle_compare_reply_keyboard_button(update, context, text)
                else:
                    # No data available - show appropriate error message
                    await self._send_message_safe(update, f"❌ Нет данных для анализа. Создайте сравнение командой `/compare` или портфель командой `/portfolio`")
            elif is_compare_button:
                # Button only exists in compare context
                await self._handle_compare_reply_keyboard_button(update, context, text)
            elif is_portfolio_button:
                # Button only exists in portfolio context
                await self._handle_portfolio_reply_keyboard_button(update, context, text)
            else:
                await self._send_message_safe(update, f"❌ Неизвестная кнопка: {text}")
                
        except Exception as e:
            self.logger.error(f"Error handling reply keyboard button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_portfolio_reply_keyboard_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle portfolio Reply Keyboard button presses"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Get the last portfolio symbol from user context
            saved_portfolios = user_context.get('saved_portfolios', {})
            if not saved_portfolios:
                await self._send_message_safe(update, "❌ Нет сохраненных портфелей. Создайте портфель командой `/portfolio`")
                return
            
            # Get the most recent portfolio
            portfolio_symbols = list(saved_portfolios.keys())
            if not portfolio_symbols:
                await self._send_message_safe(update, "❌ Нет сохраненных портфелей. Создайте портфель командой `/portfolio`")
                return
            
            # Use the most recent portfolio (last in the list)
            portfolio_symbol = portfolio_symbols[-1]
            
            # Map button text to callback data
            button_mapping = {
                "▫️ Накоп. доходность": f"portfolio_wealth_chart_{portfolio_symbol}",
                "▫️ Доходность ГГ": f"portfolio_returns_{portfolio_symbol}",
                "▫️ Динамика дох.": f"portfolio_rolling_cagr_{portfolio_symbol}",
                "▫️ Дивиденды": f"portfolio_dividends_{portfolio_symbol}",
                "▫️ Метрики": f"portfolio_risk_metrics_{portfolio_symbol}",
                "▫️ Монте-Карло": f"portfolio_monte_carlo_{portfolio_symbol}",
                "▫️ Процентили (10/50/90)": f"portfolio_forecast_{portfolio_symbol}",
                "▫️ Просадки": f"portfolio_drawdowns_{portfolio_symbol}",
                "▫️ Нейроанализ": f"portfolio_ai_analysis_{portfolio_symbol}",
                "▫️ Портфель vs Активы": f"portfolio_compare_assets_{portfolio_symbol}",
                "▫️ Сравнить": f"portfolio_compare_{portfolio_symbol}"
            }
            
            callback_data = button_mapping.get(text)
            if not callback_data:
                await self._send_message_safe(update, f"❌ Неизвестная кнопка: {text}")
                return
            
            # No processing message needed - direct execution
            
            # Call the appropriate function directly based on callback_data
            if callback_data.startswith("portfolio_wealth_chart_"):
                await self._handle_portfolio_wealth_chart_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_returns_"):
                await self._handle_portfolio_returns_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_rolling_cagr_"):
                await self._handle_portfolio_rolling_cagr_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_dividends_"):
                await self._handle_portfolio_dividends_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_risk_metrics_"):
                await self._handle_portfolio_risk_metrics_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_monte_carlo_"):
                await self._handle_portfolio_monte_carlo_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_forecast_"):
                await self._handle_portfolio_forecast_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_drawdowns_"):
                await self._handle_portfolio_drawdowns_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_ai_analysis_"):
                await self._handle_portfolio_ai_analysis_button(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_compare_assets_"):
                await self._handle_portfolio_compare_assets_by_symbol(update, context, portfolio_symbol)
            elif callback_data.startswith("portfolio_compare_"):
                await self._handle_portfolio_compare_button(update, context, portfolio_symbol)
            else:
                await self._send_message_safe(update, f"❌ Неизвестная функция: {callback_data}")
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio reply keyboard button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_compare_reply_keyboard_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle compare Reply Keyboard button presses"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Get the last compare symbols from user context
            last_symbols = user_context.get('last_assets', [])
            if not last_symbols:
                await self._send_message_safe(update, "❌ Нет данных для сравнения. Создайте сравнение командой `/compare`")
                return
            
            # Map button text to function calls
            if text == "▫️ Доходность":
                # Show default comparison chart (wealth index)
                await self._create_comparison_wealth_chart(update, context, last_symbols)
            elif text == "▫️ Дивиденды":
                await self._handle_dividends_button(update, context, last_symbols)
            elif text == "▫️ Просадки":
                await self._handle_drawdowns_button(update, context, last_symbols)
            elif text == "▫️ Метрики":
                await self._handle_metrics_compare_button(update, context)
            elif text == "▫️ Корреляция":
                await self._handle_correlation_button(update, context, last_symbols)
            elif text == "▫️ Эффективная граница":
                await self._handle_efficient_frontier_compare_button(update, context)
            elif text == "▫️ Нейроанализ":
                await self._handle_yandexgpt_analysis_compare_button(update, context)
            elif text == "▫️ В Портфель":
                await self._handle_compare_portfolio_button(update, context, last_symbols)
            else:
                await self._send_message_safe(update, f"❌ Неизвестная кнопка: {text}")
            
        except Exception as e:
            self.logger.error(f"Error handling compare reply keyboard button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_list_reply_keyboard_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle list Reply Keyboard button presses"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Get the current namespace context
            current_namespace = user_context.get('current_namespace')
            current_page = user_context.get('current_namespace_page', 0)
            
            if not current_namespace:
                await self._send_message_safe(update, "❌ Нет активного пространства имен. Используйте команду `/list <код>`")
                return
            
            # Handle navigation buttons
            if text == "⬅️ Назад":
                if current_page > 0:
                    new_page = current_page - 1
                    self._update_user_context(user_id, current_namespace_page=new_page)
                    # Check if it's a Chinese exchange
                    chinese_exchanges = ['SSE', 'SZSE', 'BSE', 'HKEX']
                    if current_namespace in chinese_exchanges:
                        await self._show_tushare_namespace_symbols(update, context, current_namespace, is_callback=False, page=new_page)
                    else:
                        await self._show_namespace_symbols(update, context, current_namespace, is_callback=False, page=new_page)
                else:
                    await self._send_message_safe(update, "❌ Вы уже на первой странице")
                    
            elif text == "➡️ Вперед":
                # We need to get total pages to check if we can go forward
                try:
                    chinese_exchanges = ['SSE', 'SZSE', 'BSE', 'HKEX']
                    if current_namespace in chinese_exchanges:
                        # For Chinese exchanges, get count from tushare
                        if self.tushare_service:
                            total_count = self.tushare_service.get_exchange_symbols_count(current_namespace)
                        else:
                            total_count = 0
                    else:
                        # For regular exchanges, get count from okama
                        import okama as ok
                        symbols_df = ok.symbols_in_namespace(current_namespace)
                        total_count = len(symbols_df)
                    
                    symbols_per_page = 20
                    total_pages = (total_count + symbols_per_page - 1) // symbols_per_page
                    
                    if current_page < total_pages - 1:
                        new_page = current_page + 1
                        self._update_user_context(user_id, current_namespace_page=new_page)
                        if current_namespace in chinese_exchanges:
                            await self._show_tushare_namespace_symbols(update, context, current_namespace, is_callback=False, page=new_page)
                        else:
                            await self._show_namespace_symbols(update, context, current_namespace, is_callback=False, page=new_page)
                    else:
                        await self._send_message_safe(update, "❌ Вы уже на последней странице")
                except Exception as e:
                    await self._send_message_safe(update, f"❌ Ошибка при навигации: {str(e)}")
                    
            elif text == "📊 Excel":
                # Handle Excel export
                await self._handle_excel_namespace_button(update, context, current_namespace)
                
            elif text == "🔍 Анализ":
                # Handle analysis button
                await self._handle_namespace_analysis_button(update, context)
                
            elif text == "⚖️ Сравнить":
                # Handle compare button
                await self._handle_namespace_compare_button(update, context)
                
            elif text == "💼 В портфель":
                # Handle portfolio button
                await self._handle_namespace_portfolio_button(update, context)
                
            elif text == "📚 База данных":
                # Return to namespace list
                await self.namespace_command(update, context)
                
            else:
                # Check if it's a page indicator (like "1/5")
                import re
                page_pattern = r'^(\d+)/(\d+)$'
                match = re.match(page_pattern, text)
                if match:
                    # It's a page indicator - do nothing, just acknowledge
                    await self._send_message_safe(update, f"📄 Страница {text}")
                else:
                    await self._send_message_safe(update, f"❌ Неизвестная кнопка: {text}")
            
        except Exception as e:
            self.logger.error(f"Error handling list reply keyboard button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_namespace_reply_keyboard_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle namespace Reply Keyboard button presses (from /list command)"""
        try:
            # Extract namespace code from button text
            namespace_mapping = {
                "🇺🇸 US": "US",
                "🇷🇺 MOEX": "MOEX", 
                "🇬🇧 LSE": "LSE",
                "🇩🇪 XETR": "XETR",
                "🇫🇷 XFRA": "XFRA",
                "🇳🇱 XAMS": "XAMS",
                "🇨🇳 SSE": "SSE",
                "🇨🇳 SZSE": "SZSE",
                "🇨🇳 BSE": "BSE",
                "🇭🇰 HKEX": "HKEX",
                "📊 INDX": "INDX",
                "💱 FX": "FX",
                "🏦 CBR": "CBR",
                "🛢️ COMM": "COMM",
                "₿ CC": "CC",
                "🏠 RE": "RE",
                "📈 INFL": "INFL",
                "💰 PIF": "PIF",
                "🏦 RATE": "RATE"
            }
            
            namespace = namespace_mapping.get(text)
            if not namespace:
                await self._send_message_safe(update, f"❌ Неизвестное пространство имен: {text}")
                return
            
            self.logger.info(f"Handling namespace reply keyboard button: {text} -> {namespace}")
            
            # Check if it's a Chinese exchange
            chinese_exchanges = ['SSE', 'SZSE', 'BSE', 'HKEX']
            if namespace in chinese_exchanges:
                await self._show_tushare_namespace_symbols(update, context, namespace, is_callback=False, page=0)
            else:
                await self._show_namespace_symbols(update, context, namespace, is_callback=False, page=0)
                
        except Exception as e:
            self.logger.error(f"Error handling namespace reply keyboard button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_info_reply_keyboard_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle info Reply Keyboard button presses (from /info command)"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Get current symbol from context
            current_symbol = user_context.get('current_info_symbol')
            if not current_symbol:
                await self._send_message_safe(update, "❌ Нет активного символа для анализа. Используйте команду `/info <символ>`")
                return
            
            self.logger.info(f"Handling info reply keyboard button: {text} for symbol: {current_symbol}")
            
            # Handle period buttons
            if text == "1 год":
                await self._handle_info_period_reply_button(update, context, current_symbol, "1Y")
            elif text == "5 лет":
                await self._handle_info_period_reply_button(update, context, current_symbol, "5Y")
            elif text == "Макс. срок":
                await self._handle_info_period_reply_button(update, context, current_symbol, "MAX")
            elif text == "Дивиденды":
                await self._handle_info_dividends_reply_button(update, context, current_symbol)
            elif text == "⚖️ Сравнение":
                await self._handle_info_compare_redirect_button(update, context, current_symbol)
            elif text == "💼 В Портфель":
                await self._handle_info_portfolio_redirect_button(update, context, current_symbol)
            else:
                await self._send_message_safe(update, f"❌ Неизвестная кнопка: {text}")
                
        except Exception as e:
            self.logger.error(f"Error handling info reply keyboard button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_info_compare_redirect_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle compare redirect button for info command - redirect to /compare"""
        try:
            self.logger.info(f"Redirecting to compare command with symbol: {symbol}")
            
            # Don't set context.args - symbol is already saved in context
            # Execute compare command without arguments
            context.args = []
            await self.compare_command(update, context)
                
        except Exception as e:
            self.logger.error(f"Error handling info compare redirect button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при перенаправлении на сравнение: {str(e)}")

    async def _handle_info_portfolio_redirect_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle portfolio redirect button for info command - redirect to /portfolio"""
        try:
            self.logger.info(f"Redirecting to portfolio command with symbol: {symbol}")
            
            # Don't set context.args - symbol is already saved in context
            # Execute portfolio command without arguments
            context.args = []
            await self.portfolio_command(update, context)
                
        except Exception as e:
            self.logger.error(f"Error handling info portfolio redirect button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при перенаправлении на портфель: {str(e)}")

    async def _handle_start_reply_keyboard_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
        """Handle start Reply Keyboard button presses (from /start command)"""
        try:
            self.logger.info(f"Handling start reply keyboard button: {text}")
            
            # Handle start buttons
            if text == "Анализ":
                # Execute info command without parameters
                context.args = []
                await self.info_command(update, context)
            elif text == "Сравнение":
                # Execute compare command without parameters
                context.args = []
                await self.compare_command(update, context)
            elif text == "Портфель":
                # Execute portfolio command without parameters
                context.args = []
                await self.portfolio_command(update, context)
            elif text == "База данных":
                # Execute list command without parameters
                context.args = []
                await self.namespace_command(update, context)
            elif text == "Справка":
                # Execute help command
                await self.help_command(update, context)
            else:
                await self._send_message_safe(update, f"❌ Неизвестная кнопка: {text}")
                
        except Exception as e:
            self.logger.error(f"Error handling start reply keyboard button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обработке кнопки: {str(e)}")

    async def _handle_info_period_reply_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str, period: str):
        """Handle period switching for info command via reply keyboard"""
        try:
            await self._send_ephemeral_message(update, context, f"📊 Обновляю данные за {period}...", delete_after=2)
            
            # Determine data source
            data_source = self.determine_data_source(symbol)
            
            if data_source == 'tushare':
                # Handle Tushare assets
                await self._handle_tushare_info_period_reply_button(update, context, symbol, period)
            else:
                # Handle Okama assets
                await self._handle_okama_info_period_reply_button(update, context, symbol, period)
                
        except Exception as e:
            self.logger.error(f"Error handling info period reply button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обновлении данных: {str(e)}")

    async def _handle_info_dividends_reply_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle dividends button for info command via reply keyboard"""
        try:
            await self._handle_single_dividends_button(update, context, symbol)
        except Exception as e:
            self.logger.error(f"Error handling info dividends reply button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при получении дивидендов: {str(e)}")

    async def _handle_info_compare_reply_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle compare button for info command via reply keyboard"""
        try:
            await self._handle_info_compare_button(update, context, symbol)
        except Exception as e:
            self.logger.error(f"Error handling info compare reply button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при сравнении: {str(e)}")

    async def _handle_info_portfolio_reply_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle portfolio button for info command via reply keyboard"""
        try:
            await self._handle_info_portfolio_button(update, context, symbol)
        except Exception as e:
            self.logger.error(f"Error handling info portfolio reply button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при добавлении в портфель: {str(e)}")

    async def _handle_tushare_info_period_reply_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str, period: str):
        """Handle period switching for Tushare assets via reply keyboard"""
        try:
            # Get symbol info
            symbol_info = self.tushare_service.get_symbol_info(symbol)
            if not symbol_info:
                await self._send_message_safe(update, f"❌ Информация о символе {symbol} не найдена")
                return
            
            # Format information according to new structure
            info_text = self._format_tushare_info_response(symbol_info, symbol)
            
            # Create reply keyboard for management
            reply_markup = self._create_info_reply_keyboard()
            
            # Save current symbol context for reply keyboard handling
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                current_info_symbol=symbol
            )
            
            # Try to get chart data
            chart_data = await self._get_tushare_chart(symbol)
            
            if chart_data:
                # Send chart with info text
                chart_caption = self._format_tushare_chart_caption(symbol_info, symbol, period)
                await self._send_photo_safe(update, chart_data, caption=chart_caption, reply_markup=reply_markup, context=context)
            else:
                # Send only text
                await self._send_message_safe(update, info_text, reply_markup=reply_markup)
                
        except Exception as e:
            self.logger.error(f"Error handling Tushare info period reply button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обновлении данных: {str(e)}")

    async def _handle_okama_info_period_reply_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str, period: str):
        """Handle period switching for Okama assets via reply keyboard"""
        try:
            # Get asset
            asset = ok.Asset(symbol)
            
            # Get key metrics for the period
            key_metrics = await self._get_asset_key_metrics(asset, symbol, period=period)
            
            # Format information
            info_text = self._format_asset_info_response(asset, symbol, key_metrics)
            
            # Create reply keyboard for management
            reply_markup = self._create_info_reply_keyboard()
            
            # Save current symbol context for reply keyboard handling
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                current_info_symbol=symbol
            )
            
            # Get chart data
            chart_data = await self._get_daily_chart(symbol)
            
            if chart_data:
                # Send chart with info text
                caption = f"📈 График доходности за {period}\n\n{info_text}"
                await self._send_photo_safe(update, chart_data, caption=caption, reply_markup=reply_markup, context=context)
            else:
                # Send only text
                await self._send_message_safe(update, info_text, reply_markup=reply_markup)
                
        except Exception as e:
            self.logger.error(f"Error handling Okama info period reply button: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при обновлении данных: {str(e)}")

    async def _remove_portfolio_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Remove portfolio Reply Keyboard if it exists - DEPRECATED: Use _manage_reply_keyboard instead"""
        try:
            await self._manage_reply_keyboard(update, context, keyboard_type=None)
        except Exception as e:
            self.logger.warning(f"Could not remove portfolio reply keyboard: {e}")

    async def _remove_compare_reply_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Remove compare Reply Keyboard if it exists - DEPRECATED: Use _manage_reply_keyboard instead"""
        try:
            await self._manage_reply_keyboard(update, context, keyboard_type=None)
        except Exception as e:
            self.logger.warning(f"Could not remove compare reply keyboard: {e}")

    async def _create_comparison_wealth_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Create and send comparison wealth chart"""
        try:
            # Get user context for currency and period
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            currency = user_context.get('last_currency', 'USD')
            specified_period = user_context.get('last_period')
            
            # Check if this is a mixed comparison (portfolios + assets)
            last_analysis_type = user_context.get('last_analysis_type', 'comparison')
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            if last_analysis_type == 'comparison' and any(isinstance(s, (pd.Series, pd.DataFrame)) for s in expanded_symbols):
                # This is a mixed comparison, handle differently
                await self._send_ephemeral_message(update, context, "📈 Создаю график накопленной доходности для смешанного сравнения...", delete_after=3)
                await self._create_mixed_comparison_wealth_chart(update, context, symbols, currency)
            else:
                # Regular comparison, create AssetList
                comparison = ok.AssetList(symbols, ccy=currency)
                
                # Create chart
                fig, ax = chart_styles.create_comparison_chart(
                    comparison.wealth_indexes, symbols, currency, title="Сравнение накопленной доходности"
                )
                
                # Save chart to bytes
                img_buffer = io.BytesIO()
                chart_styles.save_figure(fig, img_buffer)
                img_buffer.seek(0)
                img_bytes = img_buffer.getvalue()
                
                # Clear matplotlib cache
                chart_styles.cleanup_figure(fig)
                
                # Create caption
                caption = f"⚖️ Сравнение накопленной доходности: {', '.join(symbols)}\n\n"
                caption += f"💵 Валюта: {currency}\n"
                if specified_period:
                    caption += f"📅 Период: {specified_period}\n"
                
                # Create compare reply keyboard
                compare_reply_keyboard = self._create_compare_reply_keyboard()
                
                # Send chart with reply keyboard
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id,
                    photo=img_buffer,
                    caption=self._truncate_caption(caption),
                    reply_markup=compare_reply_keyboard
                )
                
                # Update user context to track active keyboard
                self._update_user_context(user_id, active_reply_keyboard="compare")
                self.logger.info("Compare reply keyboard set with comparison chart")
            
        except Exception as e:
            self.logger.error(f"Error creating comparison wealth chart: {e}")
            await self._send_message_safe(update, f"❌ Ошибка при создании графика сравнения: {str(e)}")



    async def _send_message_with_keyboard_management(self, update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                                   message_type: str, content: any, caption: str = None, 
                                                   keyboard: InlineKeyboardMarkup = None, parse_mode: str = None):
        """
        Универсальная функция для отправки сообщений с правильным управлением клавиатурой
        
        Args:
            update: Telegram update object
            context: Telegram context object
            message_type: Тип сообщения ('photo', 'text', 'document')
            content: Содержимое сообщения (фото, текст, документ)
            caption: Подпись для фото/документа
            keyboard: Клавиатура для добавления к сообщению
            parse_mode: Режим парсинга (Markdown, HTML)
        """
        try:
            # Удаляем клавиатуру с предыдущего сообщения
            await self._remove_keyboard_before_new_message(update, context)
            
            # Отправляем новое сообщение с клавиатурой
            if message_type == 'photo':
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id,
                    photo=content,
                    caption=caption,
                    reply_markup=keyboard,
                    parse_mode=parse_mode
                )
            elif message_type == 'text':
                await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=content,
                    reply_markup=keyboard,
                    parse_mode=parse_mode
                )
            elif message_type == 'document':
                await context.bot.send_document(
                    chat_id=update.effective_chat.id,
                    document=content,
                    caption=caption,
                    reply_markup=keyboard,
                    parse_mode=parse_mode
                )
            else:
                self.logger.error(f"Unsupported message type: {message_type}")
                await self._send_callback_message(update, context, f"❌ Неподдерживаемый тип сообщения: {message_type}")
                
        except Exception as e:
            self.logger.error(f"Error in _send_message_with_keyboard_management: {e}")
            # Fallback: отправляем сообщение без клавиатуры
            await self._send_callback_message(update, context, f"❌ Ошибка при отправке сообщения: {str(e)}")

    def _create_enhanced_chart_caption(self, symbols: list, currency: str, specified_period: str) -> str:
        """Create enhanced chart caption with HTML formatting for better Telegram compatibility"""
        try:
            # Create chart title section
            chart_title = f"📈 <b>График накопленной доходности</b>"
            
            # Create assets info section
            assets_info = f"<b>Активы:</b> {', '.join(symbols)}"
            
            # Create currency info section
            currency_info = f"<b>Валюта:</b> {currency}"
            
            # Create period info section if specified
            period_info = ""
            if specified_period:
                period_info = f"<b>Период:</b> {specified_period}"
            
            # Combine all sections with proper HTML formatting
            caption_parts = [
                chart_title,
                "",
                assets_info,
                currency_info
            ]
            
            if period_info:
                caption_parts.append(period_info)
            
            return "\n".join(caption_parts)
            
        except Exception as e:
            self.logger.error(f"Error creating enhanced chart caption: {e}")
            # Fallback to simple caption
            return f"📈 <b>График накопленной доходности</b>\n\n<b>Активы:</b> {', '.join(symbols)}\n<b>Валюта:</b> {currency}"


    def _create_metrics_excel(self, metrics_data: Dict[str, Any], symbols: list, currency: str) -> io.BytesIO:
        """Create Excel file with comprehensive metrics"""
        try:
            buffer = io.BytesIO()
            
            if EXCEL_AVAILABLE:
                # Create Excel file with openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                wb = Workbook()
                
                # Remove default sheet
                wb.remove(wb.active)
                
                # Create Summary sheet
                ws_summary = wb.create_sheet("Summary", 0)
                
                # Summary data
                asset_names = metrics_data.get('asset_names', {})
                
                # Create assets list with names if available
                assets_with_names = []
                for symbol in symbols:
                    if symbol in asset_names and asset_names[symbol] != symbol:
                        assets_with_names.append(f"{symbol} ({asset_names[symbol]})")
                    else:
                        assets_with_names.append(symbol)
                
                summary_data = [
                    ["Metric", "Value"],
                    ["Analysis Date", metrics_data['timestamp']],
                    ["Currency", currency],
                    ["Assets Count", len(symbols)],
                    ["Assets", ", ".join(assets_with_names)],
                    ["Period", metrics_data['period']]
                ]
                
                for row in summary_data:
                    ws_summary.append(row)
                
                # Style summary sheet
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws_summary[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Create Detailed Metrics sheet
                ws_metrics = wb.create_sheet("Detailed Metrics", 1)
                
                # Prepare detailed metrics data
                detailed_metrics = metrics_data.get('detailed_metrics', {})
                
                # Create headers with asset names
                headers = ["Metric"]
                for symbol in symbols:
                    if symbol in asset_names and asset_names[symbol] != symbol:
                        headers.append(f"{symbol} ({asset_names[symbol]})")
                    else:
                        headers.append(symbol)
                
                ws_metrics.append(headers)
                
                # Define metrics to include
                metric_names = [
                    ("Total Return", "total_return"),
                    ("Annual Return (CAGR)", "annual_return"),
                    ("Volatility", "volatility"),
                    ("Sharpe Ratio", "sharpe_ratio"),
                    ("Sortino Ratio", "sortino_ratio"),
                    ("Max Drawdown", "max_drawdown"),
                    ("Calmar Ratio", "calmar_ratio"),
                    ("VaR 95%", "var_95"),
                    ("CVaR 95%", "cvar_95")
                ]
                
                # Add data rows
                for metric_name, metric_key in metric_names:
                    row = [metric_name]
                    for symbol in symbols:
                        value = detailed_metrics.get(symbol, {}).get(metric_key, 0.0)
                        if isinstance(value, (int, float)):
                            row.append(round(value, 4))
                        elif hasattr(value, '__class__') and 'Mock' in str(value.__class__):
                            # Handle Mock objects
                            row.append(0.0)
                        else:
                            row.append(value)
                    ws_metrics.append(row)
                
                # Style metrics sheet
                for cell in ws_metrics[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths
                for column in ws_metrics.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    ws_metrics.column_dimensions[column_letter].width = adjusted_width
                
                # Create Correlation Matrix sheet
                if metrics_data.get('correlations'):
                    ws_corr = wb.create_sheet("Correlation Matrix", 2)
                    
                    # Add headers with asset names
                    corr_headers = [""]
                    for symbol in symbols:
                        if symbol in asset_names and asset_names[symbol] != symbol:
                            corr_headers.append(f"{symbol} ({asset_names[symbol]})")
                        else:
                            corr_headers.append(symbol)
                    ws_corr.append(corr_headers)
                    
                    # Add correlation data
                    correlations = metrics_data['correlations']
                    for i, symbol in enumerate(symbols):
                        # Use asset name for row header
                        row_header = symbol
                        if symbol in asset_names and asset_names[symbol] != symbol:
                            row_header = f"{symbol} ({asset_names[symbol]})"
                        
                        row = [row_header]
                        for j in range(len(symbols)):
                            try:
                                corr_value = correlations[i][j]
                                if isinstance(corr_value, (int, float)):
                                    row.append(round(corr_value, 4))
                                else:
                                    row.append(0.0)
                            except (IndexError, TypeError):
                                row.append(0.0)
                        ws_corr.append(row)
                    
                    # Style correlation sheet
                    for cell in ws_corr[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    # Auto-adjust column widths
                    for column in ws_corr.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 15)
                        ws_corr.column_dimensions[column_letter].width = adjusted_width
                
                # Save to buffer
                wb.save(buffer)
                buffer.seek(0)
                
            else:
                # Fallback to CSV format
                import csv
                import io as csv_io
                
                # Create CSV content
                csv_content = []
                
                # Summary
                csv_content.append(["SUMMARY"])
                csv_content.append(["Metric", "Value"])
                csv_content.append(["Analysis Date", metrics_data['timestamp']])
                csv_content.append(["Currency", currency])
                csv_content.append(["Assets Count", len(symbols)])
                csv_content.append(["Assets", ", ".join(symbols)])
                csv_content.append([])
                
                # Detailed metrics
                csv_content.append(["DETAILED METRICS"])
                detailed_metrics = metrics_data.get('detailed_metrics', {})
                
                headers = ["Metric"] + symbols
                csv_content.append(headers)
                
                metric_names = [
                    ("Total Return", "total_return"),
                    ("Annual Return (CAGR)", "annual_return"),
                    ("Volatility", "volatility"),
                    ("Sharpe Ratio", "sharpe_ratio"),
                    ("Sortino Ratio", "sortino_ratio"),
                    ("Max Drawdown", "max_drawdown"),
                    ("Calmar Ratio", "calmar_ratio"),
                    ("VaR 95%", "var_95"),
                    ("CVaR 95%", "cvar_95")
                ]
                
                for metric_name, metric_key in metric_names:
                    row = [metric_name]
                    for symbol in symbols:
                        value = detailed_metrics.get(symbol, {}).get(metric_key, 0.0)
                        if isinstance(value, (int, float)):
                            row.append(round(value, 4))
                        elif hasattr(value, '__class__') and 'Mock' in str(value.__class__):
                            # Handle Mock objects
                            row.append(0.0)
                        else:
                            row.append(value)
                    csv_content.append(row)
                
                # Write CSV to buffer
                csv_buffer = csv_io.StringIO()
                writer = csv.writer(csv_buffer)
                for row in csv_content:
                    writer.writerow(row)
                
                buffer.write(csv_buffer.getvalue().encode('utf-8'))
                buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            self.logger.error(f"Error creating metrics Excel: {e}")
            return None

    async def _handle_drawdowns_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle drawdowns button click"""
        try:
            # Don\'t remove keyboard yet - wait for successful message creation
            
            
            user_id = update.effective_user.id
            self.logger.info(f"Handling drawdowns button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context keys: {list(user_context.keys())}")
            
            if 'current_symbols' not in user_context:
                self.logger.warning(f"current_symbols not found in user context for user {user_id}")
                await self._send_callback_message(update, context, "❌ Данные о сравнении не найдены. Выполните команду /compare заново.")
                return
            
            symbols = user_context['current_symbols']
            currency = user_context.get('current_currency', 'USD')
            period = user_context.get('current_period', None)
            
            self.logger.info(f"Creating drawdowns chart for symbols: {symbols}, currency: {currency}, period: {period}")
            await self._send_ephemeral_message(update, context, "📉 Создаю график...", delete_after=3)
            
            # Check if this is a mixed comparison (portfolios + assets)
            user_context = self._get_user_context(user_id)
            last_analysis_type = user_context.get('last_analysis_type', 'comparison')
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            if last_analysis_type == 'comparison' and any(isinstance(s, (pd.Series, pd.DataFrame)) for s in expanded_symbols):
                # This is a mixed comparison, handle differently
                await self._send_ephemeral_message(update, context, "📉 Создаю график для смешанного сравнения...", delete_after=3)
                await self._create_mixed_comparison_drawdowns_chart(update, context, symbols, currency)
            else:
                # Regular comparison, create AssetList with period support
                if period:
                    years = int(period[:-1])  # Extract number from '5Y'
                    from datetime import timedelta
                    end_date = datetime.now()
                    start_date = end_date - timedelta(days=years * 365)
                    asset_list = ok.AssetList(symbols, ccy=currency, 
                                            first_date=start_date.strftime('%Y-%m-%d'), 
                                            last_date=end_date.strftime('%Y-%m-%d'))
                else:
                    asset_list = ok.AssetList(symbols, ccy=currency)
                await self._create_drawdowns_chart(update, context, asset_list, symbols, currency)
        
        except Exception as e:
            self.logger.error(f"Error handling drawdowns button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика drawdowns: {str(e)}")

    async def _create_mixed_comparison_drawdowns_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list, currency: str):
        """Create drawdowns chart for mixed comparison (portfolios + assets)"""
        try:
            self.logger.info(f"Creating mixed comparison drawdowns chart for symbols: {symbols}")
            
            # Get user context to restore portfolio information
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            portfolio_contexts = user_context.get('portfolio_contexts', [])
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            # Separate portfolios and individual assets using expanded_symbols
            portfolio_data = []
            asset_symbols = []
            
            for i, expanded_symbol in enumerate(expanded_symbols):
                if isinstance(expanded_symbol, (pd.Series, pd.DataFrame)):
                    # This is a portfolio wealth index
                    portfolio_data.append(expanded_symbol)
                else:
                    # This is an individual asset symbol
                    asset_symbols.append(expanded_symbol)
            
            # Calculate drawdowns for all items
            drawdowns_data = {}
            
            # Process portfolios separately to avoid AssetList creation issues
            for i, portfolio_context in enumerate(portfolio_contexts):
                if i < len(portfolio_data):
                    try:
                        self.logger.info(f"Processing portfolio {i} for drawdowns")
                        
                        # Get portfolio details from context
                        assets = portfolio_context.get('portfolio_symbols', [])
                        weights = portfolio_context.get('portfolio_weights', [])
                        symbol = portfolio_context.get('symbol', f'Portfolio_{i+1}')
                        
                        if assets and weights and len(assets) == len(weights):
                            self.logger.info(f"Portfolio {i} assets: {assets}, weights: {weights}")
                            
                            # Create portfolio using ok.Portfolio
                            import okama as ok
                            portfolio = ok.Portfolio(
                                assets=assets,
                                weights=weights,
                                rebalancing_strategy=ok.Rebalance(period="year"),
                                symbol=symbol
                            )
                            
                            # Calculate drawdowns for portfolio
                            wealth_series = portfolio.wealth_index
                            self.logger.info(f"Portfolio {symbol} wealth_series length: {len(wealth_series)}, dtype: {wealth_series.dtype}")
                            
                            returns = wealth_series.pct_change().dropna()
                            if len(returns) > 0:
                                cumulative = (1 + returns).cumprod()
                                running_max = cumulative.expanding().max()
                                drawdowns = (cumulative - running_max) / running_max
                                drawdowns_data[symbol] = drawdowns
                                self.logger.info(f"Successfully created drawdowns for {symbol}: {len(drawdowns)} points")
                            else:
                                self.logger.warning(f"Portfolio {symbol}: No returns data after pct_change")
                        else:
                            self.logger.warning(f"Portfolio {i} missing valid assets/weights data")
                    except Exception as portfolio_error:
                        self.logger.warning(f"Could not process portfolio {i}: {portfolio_error}")
                        continue
            
            # Process individual assets separately
            if asset_symbols:
                try:
                    asset_asset_list = ok.AssetList(asset_symbols, ccy=currency)
                    
                    for symbol in asset_symbols:
                        if symbol in asset_asset_list.wealth_indexes.columns:
                            # Calculate drawdowns for individual asset
                            wealth_series = asset_asset_list.wealth_indexes[symbol]
                            self.logger.info(f"Asset {symbol} wealth_series length: {len(wealth_series)}, dtype: {wealth_series.dtype}")
                            
                            returns = wealth_series.pct_change().dropna()
                            if len(returns) > 0:
                                cumulative = (1 + returns).cumprod()
                                running_max = cumulative.expanding().max()
                                drawdowns = (cumulative - running_max) / running_max
                                drawdowns_data[symbol] = drawdowns
                                self.logger.info(f"Successfully created drawdowns for {symbol}: {len(drawdowns)} points")
                            else:
                                self.logger.warning(f"Asset {symbol}: No returns data after pct_change")
                        else:
                            self.logger.warning(f"Asset {symbol} not found in wealth_indexes columns")
                except Exception as asset_error:
                    self.logger.warning(f"Could not process individual assets: {asset_error}")
            
            if not drawdowns_data:
                await self._send_callback_message(update, context, "❌ Не удалось создать данные для графика просадок")
                return
            
            # Create chart using chart_styles
            try:
                # Combine all drawdowns into a DataFrame
                drawdowns_df = pd.DataFrame(drawdowns_data)
                
                fig, ax = chart_styles.create_drawdowns_chart(
                    drawdowns_df, list(drawdowns_data.keys()), currency
                )
                
                # Save chart to bytes with memory optimization
                img_buffer = io.BytesIO()
                chart_styles.save_figure(fig, img_buffer)
                img_buffer.seek(0)
                img_bytes = img_buffer.getvalue()
                
                # Clear matplotlib cache to free memory
                chart_styles.cleanup_figure(fig)
                
                # Create keyboard for compare command
                keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                
                # Remove keyboard from previous message before sending new message
                await self._remove_keyboard_before_new_message(update, context)
                
                # Send drawdowns chart with keyboard
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id, 
                    photo=io.BytesIO(img_bytes),
                    caption=self._truncate_caption(f"📉 График просадок для смешанного сравнения\n\nПоказывает просадки портфелей и активов"),
                    reply_markup=keyboard
                )
                
            except Exception as chart_error:
                self.logger.error(f"Error creating drawdowns chart: {chart_error}")
                await self._send_callback_message(update, context, f"❌ Ошибка при создании графика просадок: {str(chart_error)}")
                
        except Exception as e:
            self.logger.error(f"Error in mixed comparison drawdowns chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика просадок: {str(e)}")
            
            if not drawdowns_data:
                await self._send_callback_message(update, context, "❌ Не удалось создать данные для графика просадок")
                return
            
            # Clean drawdowns data to handle Period indices
            cleaned_drawdowns_data = {}
            for key, series in drawdowns_data.items():
                if isinstance(series, pd.Series):
                    # Convert Period index to datetime if needed
                    if hasattr(series.index, 'dtype') and str(series.index.dtype).startswith('period'):
                        series.index = series.index.to_timestamp()
                    cleaned_drawdowns_data[key] = series
            
            # Create drawdowns chart
            fig, ax = chart_styles.create_drawdowns_chart(
                data=pd.DataFrame(cleaned_drawdowns_data),
                symbols=list(cleaned_drawdowns_data.keys()),
                ccy=currency
            )
            
            # Save chart
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            img_buffer.seek(0)
            
            # Clear matplotlib cache
            chart_styles.cleanup_figure(fig)
            
            # Create caption
            portfolio_count = len(portfolio_data)
            asset_count = len(asset_symbols)
            
            # Get portfolio names from context
            portfolio_names = []
            for i, portfolio_series in enumerate(portfolio_data):
                if i < len(portfolio_contexts):
                    portfolio_names.append(portfolio_contexts[i]['symbol'])
                else:
                    portfolio_names.append(f'Portfolio_{i+1}')
            
            caption = f"📉 Просадки\n\n"
            caption += f"📊 Состав:\n"
            if portfolio_count > 0:
                caption += f"• Портфели: {', '.join(portfolio_names)}\n"
            if asset_count > 0:
                caption += f"• Индивидуальные активы: {', '.join(asset_symbols)}\n"
            caption += f"• Валюта: {currency}\n"
            
            # Create keyboard for compare command
            keyboard = self._create_compare_command_keyboard(symbols, currency, update)
            
            # Remove keyboard from previous message before sending new message
            await self._remove_keyboard_before_new_message(update, context)
            
            # Send chart with keyboard
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(caption),
                reply_markup=keyboard
            )
            
        except Exception as e:
            self.logger.error(f"Error creating mixed comparison drawdowns chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика просадок: {str(e)}")

    async def _handle_dividends_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle dividends button click"""
        try:
            # Don\'t remove keyboard yet - wait for successful message creation
            
            
            user_id = update.effective_user.id
            self.logger.info(f"Handling dividends button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context keys: {list(user_context.keys())}")
            
            if 'current_symbols' not in user_context:
                self.logger.warning(f"current_symbols not found in user context for user {user_id}")
                await self._send_callback_message(update, context, "❌ Данные о сравнении не найдены. Выполните команду /compare заново.")
                return
            
            symbols = user_context['current_symbols']
            currency = user_context.get('current_currency', 'USD')
            period = user_context.get('current_period', None)
            
            self.logger.info(f"Creating dividends chart for symbols: {symbols}, currency: {currency}, period: {period}")
            await self._send_ephemeral_message(update, context, "Создаю график дивидендной доходности...", delete_after=3)
            
            # Check if this is a mixed comparison (portfolios + assets)
            user_context = self._get_user_context(user_id)
            last_analysis_type = user_context.get('last_analysis_type', 'comparison')
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            if last_analysis_type == 'comparison' and any(isinstance(s, (pd.Series, pd.DataFrame)) for s in expanded_symbols):
                # This is a mixed comparison, handle differently
                await self._send_ephemeral_message(update, context, "Создаю график дивидендной доходности...", delete_after=3)
                await self._create_mixed_comparison_dividends_chart(update, context, symbols, currency)
            else:
                # Regular comparison, create AssetList with period support
                if period:
                    years = int(period[:-1])  # Extract number from '5Y'
                    from datetime import timedelta
                    end_date = datetime.now()
                    start_date = end_date - timedelta(days=years * 365)
                    asset_list = ok.AssetList(symbols, ccy=currency, 
                                            first_date=start_date.strftime('%Y-%m-%d'), 
                                            last_date=end_date.strftime('%Y-%m-%d'))
                else:
                    asset_list = ok.AssetList(symbols, ccy=currency)
                await self._create_dividend_yield_chart(update, context, asset_list, symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error handling dividends button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика дивидендной доходности: {str(e)}")

    async def _create_mixed_comparison_dividends_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list, currency: str):
        """Create dividends chart for mixed comparison (portfolios + assets)"""
        try:
            self.logger.info(f"Creating mixed comparison dividends chart for symbols: {symbols}")
            
            # Get user context to restore portfolio information
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            portfolio_contexts = user_context.get('portfolio_contexts', [])
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            # Separate portfolios and individual assets using expanded_symbols
            portfolio_data = []
            asset_symbols = []
            
            for i, expanded_symbol in enumerate(expanded_symbols):
                if isinstance(expanded_symbol, (pd.Series, pd.DataFrame)):
                    # This is a portfolio wealth index
                    portfolio_data.append(expanded_symbol)
                else:
                    # This is an individual asset symbol
                    asset_symbols.append(expanded_symbol)
            
            # Create dividends data for both portfolios and assets
            dividends_data = {}
            
            # Process portfolios separately to avoid AssetList creation issues
            for i, portfolio_context in enumerate(portfolio_contexts):
                if i < len(portfolio_data):
                    try:
                        self.logger.info(f"Processing portfolio {i} for dividends")
                        
                        # Get portfolio details from context
                        assets = portfolio_context.get('portfolio_symbols', [])
                        weights = portfolio_context.get('portfolio_weights', [])
                        symbol = portfolio_context.get('symbol', f'Portfolio_{i+1}')
                        
                        if assets and weights and len(assets) == len(weights):
                            self.logger.info(f"Portfolio {i} assets: {assets}, weights: {weights}")
                            
                            # Create separate AssetList for portfolio assets
                            try:
                                portfolio_asset_list = ok.AssetList(assets, ccy=currency)
                                
                                if hasattr(portfolio_asset_list, 'dividend_yields'):
                                    # Calculate weighted dividend yield
                                    total_dividend_yield = 0
                                    for asset, weight in zip(assets, weights):
                                        if asset in portfolio_asset_list.dividend_yields.columns:
                                            dividend_yield = portfolio_asset_list.dividend_yields[asset].iloc[-1] if not portfolio_asset_list.dividend_yields[asset].empty else 0
                                            total_dividend_yield += dividend_yield * weight
                                            self.logger.info(f"Asset {asset}: dividend_yield={dividend_yield}, weight={weight}")
                                        else:
                                            self.logger.warning(f"Asset {asset} not found in dividend_yields columns")
                                    
                                    dividends_data[symbol] = total_dividend_yield
                                    self.logger.info(f"Successfully calculated weighted dividend yield for {symbol}: {total_dividend_yield}")
                                else:
                                    self.logger.warning(f"Portfolio asset list does not have dividend_yields attribute for {symbol}")
                                    dividends_data[symbol] = 0  # Default value
                            except Exception as asset_list_error:
                                self.logger.warning(f"Could not create AssetList for portfolio {symbol}: {asset_list_error}")
                                dividends_data[symbol] = 0  # Default value
                        else:
                            self.logger.warning(f"Portfolio {symbol} missing valid assets/weights data for dividends")
                            dividends_data[symbol] = 0  # Default value
                    except Exception as portfolio_error:
                        self.logger.warning(f"Could not process portfolio {i} for dividends: {portfolio_error}")
                        dividends_data[symbol] = 0  # Default value
                        continue
            
            # Process individual assets separately
            if asset_symbols:
                try:
                    asset_asset_list = ok.AssetList(asset_symbols, ccy=currency)
                    
                    if hasattr(asset_asset_list, 'dividend_yields'):
                        for symbol in asset_symbols:
                            if symbol in asset_asset_list.dividend_yields.columns:
                                dividend_yield = asset_asset_list.dividend_yields[symbol].iloc[-1] if not asset_asset_list.dividend_yields[symbol].empty else 0
                                dividends_data[symbol] = dividend_yield
                                self.logger.info(f"Successfully got dividend yield for {symbol}: {dividend_yield}")
                            else:
                                self.logger.warning(f"Asset {symbol} not found in dividend_yields columns")
                                dividends_data[symbol] = 0  # Default value
                    else:
                        self.logger.warning("Asset list does not have dividend_yields attribute")
                        for symbol in asset_symbols:
                            dividends_data[symbol] = 0  # Default value
                except Exception as asset_error:
                    self.logger.warning(f"Could not process individual assets: {asset_error}")
                    for symbol in asset_symbols:
                        dividends_data[symbol] = 0  # Default value
            
            if not dividends_data:
                await self._send_callback_message(update, context, "❌ Не удалось создать данные для графика дивидендной доходности")
                return
            
            # Validate dividends data before creating chart
            valid_dividends_data = {}
            for symbol, dividend_yield in dividends_data.items():
                if dividend_yield is not None and not pd.isna(dividend_yield):
                    valid_dividends_data[symbol] = dividend_yield
                else:
                    self.logger.warning(f"Invalid dividend yield for {symbol}: {dividend_yield}")
                    valid_dividends_data[symbol] = 0  # Default to 0
            
            if not valid_dividends_data:
                await self._send_callback_message(update, context, "❌ Не удалось создать валидные данные для графика дивидендной доходности")
                return
            
            # Create chart using chart_styles
            try:
                # Convert Series to DataFrame for chart creation
                dividends_df = pd.DataFrame(valid_dividends_data, index=[0]).T
                fig, ax = chart_styles.create_dividend_yield_chart(
                    dividends_df, list(valid_dividends_data.keys())
                )
                
                # Save chart to bytes with memory optimization
                img_buffer = io.BytesIO()
                chart_styles.save_figure(fig, img_buffer)
                img_buffer.seek(0)
                img_bytes = img_buffer.getvalue()
                
                # Clear matplotlib cache to free memory
                chart_styles.cleanup_figure(fig)
                
                # Create keyboard for compare command
                keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                
                # Remove keyboard from previous message before sending new message
                await self._remove_keyboard_before_new_message(update, context)
                
                # Send dividend yield chart with keyboard
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id, 
                    photo=io.BytesIO(img_bytes),
                    caption=self._truncate_caption(f"💰 График дивидендной доходности для смешанного сравнения\n\nПоказывает дивидендную доходность портфелей и активов"),
                    reply_markup=keyboard
                )
                
            except Exception as chart_error:
                self.logger.error(f"Error creating dividend yield chart: {chart_error}")
                await self._send_callback_message(update, context, f"❌ Ошибка при создании графика дивидендной доходности: {str(chart_error)}")
                
        except Exception as e:
            self.logger.error(f"Error in mixed comparison dividends chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика дивидендной доходности: {str(e)}")

    async def _handle_correlation_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle correlation matrix button click"""
        try:
            # Don\'t remove keyboard yet - wait for successful message creation
            
            
            user_id = update.effective_user.id
            self.logger.info(f"Handling correlation button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context keys: {list(user_context.keys())}")
            
            if 'current_symbols' not in user_context:
                self.logger.warning(f"current_symbols not found in user context for user {user_id}")
                await self._send_callback_message(update, context, "❌ Данные о сравнении не найдены. Выполните команду /compare заново.")
                return
            
            symbols = user_context['current_symbols']
            currency = user_context.get('current_currency', 'USD')
            period = user_context.get('current_period', None)
            
            self.logger.info(f"Creating correlation matrix for symbols: {symbols}, currency: {currency}, period: {period}")
            await self._send_ephemeral_message(update, context, "🔗 Создаю корреляционную матрицу...", delete_after=3)
            
            # Check if this is a mixed comparison (portfolios + assets)
            user_context = self._get_user_context(user_id)
            last_analysis_type = user_context.get('last_analysis_type', 'comparison')
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            if last_analysis_type == 'comparison' and any(isinstance(s, (pd.Series, pd.DataFrame)) for s in expanded_symbols):
                # This is a mixed comparison, handle differently
                await self._send_ephemeral_message(update, context, "🔗 Создаю корреляционную матрицу для смешанного сравнения...", delete_after=3)
                await self._create_mixed_comparison_correlation_matrix(update, context, symbols, currency)
            else:
                # Regular comparison, create AssetList with period support
                if period:
                    years = int(period[:-1])  # Extract number from '5Y'
                    from datetime import timedelta
                    end_date = datetime.now()
                    start_date = end_date - timedelta(days=years * 365)
                    asset_list = ok.AssetList(symbols, ccy=currency, 
                                            first_date=start_date.strftime('%Y-%m-%d'), 
                                            last_date=end_date.strftime('%Y-%m-%d'))
                else:
                    asset_list = ok.AssetList(symbols, ccy=currency)
                await self._create_correlation_matrix(update, context, asset_list, symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error handling correlation button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании корреляционной матрицы: {str(e)}")

    async def _create_mixed_comparison_correlation_matrix(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list, currency: str):
        """Create correlation matrix for mixed comparison (portfolios + assets)"""
        try:
            self.logger.info(f"Creating mixed comparison correlation matrix for symbols: {symbols}")
            
            # Get user context to restore portfolio information
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            portfolio_contexts = user_context.get('portfolio_contexts', [])
            expanded_symbols = user_context.get('expanded_symbols', [])
            
            self.logger.info(f"Portfolio contexts: {len(portfolio_contexts)}")
            self.logger.info(f"Expanded symbols: {len(expanded_symbols)}")
            self.logger.info(f"Expanded symbols types: {[type(s).__name__ for s in expanded_symbols]}")
            
            # Collect all individual assets from portfolios and simple assets
            all_assets = set()
            
            # Add assets from portfolios
            for portfolio_context in portfolio_contexts:
                portfolio_assets = portfolio_context.get('portfolio_symbols', [])
                all_assets.update(portfolio_assets)
                self.logger.info(f"Added {len(portfolio_assets)} assets from portfolio: {portfolio_assets}")
            
            # Add simple assets (not portfolios)
            for expanded_symbol in expanded_symbols:
                if not isinstance(expanded_symbol, (pd.Series, pd.DataFrame)):
                    # This is an individual asset symbol
                    all_assets.add(expanded_symbol)
            
            all_assets = list(all_assets)
            self.logger.info(f"Total unique assets for correlation: {len(all_assets)} - {all_assets}")
            
            if len(all_assets) < 2:
                error_msg = f"❌ Недостаточно данных для корреляционной матрицы (только {len(all_assets)} актив). Нужно минимум 2 актива."
                await self._send_callback_message(update, context, error_msg)
                return
            
            # Create AssetList with all assets for correlation matrix
            try:
                import okama as ok
                asset_list = ok.AssetList(all_assets, ccy=currency)
                
                # Check if assets_ror data is available
                if not hasattr(asset_list, 'assets_ror') or asset_list.assets_ror is None or asset_list.assets_ror.empty:
                    self.logger.warning("assets_ror data not available for mixed comparison")
                    await self._send_callback_message(update, context, "ℹ️ Данные о доходности активов недоступны для создания корреляционной матрицы")
                    return
                
                # Get correlation matrix
                correlation_matrix = asset_list.assets_ror.corr()
                
                self.logger.info(f"Correlation matrix created successfully, shape: {correlation_matrix.shape}")
                
                if correlation_matrix.empty:
                    self.logger.warning("Correlation matrix is empty")
                    await self._send_callback_message(update, context, "❌ Не удалось вычислить корреляционную матрицу")
                    return
                
                # Create correlation matrix visualization using chart_styles
                fig, ax = chart_styles.create_correlation_matrix_chart(
                    correlation_matrix, data_source='okama'
                )
                
                # Save chart to bytes with memory optimization
                img_buffer = io.BytesIO()
                chart_styles.save_figure(fig, img_buffer)
                img_buffer.seek(0)
                img_bytes = img_buffer.getvalue()
                
                # Clear matplotlib cache to free memory
                chart_styles.cleanup_figure(fig)
                
                # Prepare correlation values text for caption
                correlation_values_text = self._format_correlation_values(correlation_matrix)
                
                # Create keyboard for compare command
                keyboard = self._create_compare_command_keyboard(symbols, currency, update)
                
                # Remove keyboard from previous message before sending new message
                await self._remove_keyboard_before_new_message(update, context)
                
                # Send correlation matrix with keyboard
                self.logger.info("Sending correlation matrix image...")
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id, 
                    photo=io.BytesIO(img_bytes),
                    caption=self._truncate_caption(f"🔗 Корреляционная матрица для смешанного сравнения\n\nПоказывает корреляцию между доходностями всех активов (от -1 до +1)\n\n• +1: полная положительная корреляция\n• 0: отсутствие корреляции\n• -1: полная отрицательная корреляция\n\n{correlation_values_text}"),
                    reply_markup=keyboard
                )
                self.logger.info("Correlation matrix image sent successfully")
                
                plt.close(fig)
                
            except Exception as chart_error:
                self.logger.error(f"Error creating correlation matrix chart: {chart_error}")
                await self._send_callback_message(update, context, f"❌ Ошибка при создании корреляционной матрицы: {str(chart_error)}")
                
        except Exception as e:
            self.logger.error(f"Error in mixed comparison correlation matrix: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании корреляционной матрицы: {str(e)}")

    async def _handle_daily_chart_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle daily chart button click for single asset"""
        try:
            await self._send_ephemeral_message(update, context, "📈 Создаю график за 1 год...", delete_after=3)
            
            # Получаем ежедневный график за 1 год
            daily_chart = await self._get_daily_chart(symbol)
            
            if daily_chart:
                caption = f"📈 График за 1 год {symbol}\n\n"
                
                await update.callback_query.message.reply_photo(
                    photo=daily_chart,
                    caption=self._truncate_caption(caption)
                )
            else:
                await self._send_callback_message(update, context, "❌ Не удалось получить график за 1 год")
                
        except Exception as e:
            self.logger.error(f"Error handling daily chart button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика за 1 год: {str(e)}")

    async def _handle_monthly_chart_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle monthly chart button click for single asset"""
        try:
            await self._send_ephemeral_message(update, context, "📅 Создаю график за 5 лет...", delete_after=3)
            
            # Получаем месячный график за 5 лет
            monthly_chart = await self._get_monthly_chart(symbol)
            
            if monthly_chart:
                caption = f"📅 График за 5 лет {symbol}\n\n"
                
                await update.callback_query.message.reply_photo(
                    photo=monthly_chart,
                    caption=self._truncate_caption(caption)
                )
            else:
                await self._send_callback_message(update, context, "❌ Не удалось получить график за 5 лет")
                
        except Exception as e:
            self.logger.error(f"Error handling monthly chart button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика за 5 лет: {str(e)}")

    async def _handle_all_chart_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle all chart button click for single asset"""
        try:
            await self._send_ephemeral_message(update, context, "📊 Создаю график за весь период...", delete_after=3)
            
            # Получаем график за весь период
            all_chart = await self._get_all_chart(symbol)
            
            if all_chart:
                caption = f"📊 График за весь период {symbol}\n\n"
                
                await update.callback_query.message.reply_photo(
                    photo=all_chart,
                    caption=self._truncate_caption(caption)
                )
            else:
                await self._send_callback_message(update, context, "❌ Не удалось получить график за весь период")
                
        except Exception as e:
            self.logger.error(f"Error handling all chart button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика за весь период: {str(e)}")

    async def _handle_info_period_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str, period: str):
        """Handle period switching for info command"""
        try:
            await self._send_ephemeral_message(update, context, f"📊 Обновляю данные за {period}...", delete_after=2)
            
            # Determine data source
            data_source = self.determine_data_source(symbol)
            
            if data_source == 'tushare':
                # Handle Tushare assets
                await self._handle_tushare_info_period_button(update, context, symbol, period)
            else:
                # Handle Okama assets
                await self._handle_okama_info_period_button(update, context, symbol, period)
                
        except Exception as e:
            self.logger.error(f"Error handling info period button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при обновлении данных: {str(e)}")

    async def _handle_okama_info_period_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str, period: str):
        """Handle period switching for Okama assets"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            # Get asset and metrics for the new period
            asset = ok.Asset(symbol)
            key_metrics = await self._get_asset_key_metrics(asset, symbol, period)
            
            # Format response with new period
            info_text = self._format_asset_info_response(asset, symbol, key_metrics)
            info_text = info_text.replace("(за 1 год)", f"(за {period})")
            
            # Create updated keyboard with new period selected
            keyboard = self._create_info_interactive_keyboard_with_period(symbol, period)
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            # Get chart for the new period
            chart_data = await self._get_chart_for_period(symbol, period)
            
            if chart_data:
                caption = f"📈 График доходности за {period}\n\n{info_text}"
                # Send new message with chart and info
                await self._send_photo_safe(update, chart_data, caption=caption, reply_markup=reply_markup, context=context)
            else:
                # If no chart, send text only
                await self._send_message_safe(update, info_text, reply_markup=reply_markup)
                
        except Exception as e:
            self.logger.error(f"Error handling Okama info period button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при обновлении данных: {str(e)}")

    async def _handle_tushare_info_period_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str, period: str):
        """Handle period switching for Tushare assets"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            if not self.tushare_service:
                await self._send_callback_message(update, context, "❌ Сервис Tushare недоступен")
                return
            
            # Get symbol information from Tushare
            symbol_info = self.tushare_service.get_symbol_info(symbol)
            
            if 'error' in symbol_info:
                await self._send_callback_message(update, context, f"❌ Ошибка: {symbol_info['error']}")
                return
            
            # Format information according to new structure
            info_text = self._format_tushare_info_response(symbol_info, symbol, period)
            
            # Create updated keyboard with new period selected
            keyboard = self._create_info_interactive_keyboard_with_period(symbol, period)
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            # Try to get chart data for the period
            chart_data = await self._get_tushare_chart_for_period(symbol, period)
            
            if chart_data:
                period_text = {
                    '1Y': '1 год',
                    '5Y': '5 лет', 
                    'MAX': 'MAX'
                }.get(period, period)
                # Create enhanced caption with English information
                chart_caption = self._format_tushare_chart_caption(symbol_info, symbol, period_text)
                caption = f"{chart_caption}\n\n{info_text}"
                await self._send_photo_safe(update, chart_data, caption=caption, reply_markup=reply_markup, context=context)
            else:
                await self._send_message_safe(update, info_text, reply_markup=reply_markup)
                
        except Exception as e:
            self.logger.error(f"Error handling Tushare info period button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при обновлении данных: {str(e)}")

    async def _handle_info_risks_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle risks and drawdowns button for info command"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            await self._send_ephemeral_message(update, context, "📉 Анализирую риски и просадки...", delete_after=2)
            
            asset = ok.Asset(symbol)
            
            # Get risk metrics
            risk_text = f"📉 **Анализ рисков для {symbol}**\n\n"
            
            # Volatility
            if hasattr(asset, 'volatility_annual'):
                volatility = asset.volatility_annual
                if hasattr(volatility, 'iloc'):
                    vol_value = volatility.iloc[-1]
                else:
                    vol_value = volatility
                risk_text += f"**Волатильность:** {vol_value:.2%}\n"
            
            # Max drawdown
            if hasattr(asset, 'max_drawdown'):
                max_dd = asset.max_drawdown
                if hasattr(max_dd, 'iloc'):
                    dd_value = max_dd.iloc[-1]
                else:
                    dd_value = max_dd
                risk_text += f"**Максимальная просадка:** {dd_value:.2%}\n"
            
            # VaR 95%
            if hasattr(asset, 'var_95'):
                var_95 = asset.var_95
                if hasattr(var_95, 'iloc'):
                    var_value = var_95.iloc[-1]
                else:
                    var_value = var_95
                risk_text += f"**VaR 95%:** {var_value:.2%}\n"
            
            # Sharpe ratio
            if hasattr(asset, 'get_sharpe_ratio'):
                sharpe = asset.get_sharpe_ratio()
                if hasattr(sharpe, 'iloc'):
                    sharpe_value = sharpe.iloc[0]
                else:
                    sharpe_value = sharpe
                risk_text += f"**Коэффициент Шарпа:** {sharpe_value:.2f}\n"
            
            await self._send_callback_message(update, context, risk_text)
            
        except Exception as e:
            self.logger.error(f"Error handling info risks button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при анализе рисков: {str(e)}")

    async def _handle_info_metrics_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle all metrics button for info command"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            await self._send_ephemeral_message(update, context, "🔍 Получаю все метрики...", delete_after=2)
            
            asset = ok.Asset(symbol)
            
            # Get comprehensive metrics
            metrics_text = f"📊 **Все метрики для {symbol}**\n\n"
            
            # Basic info
            metrics_text += f"**Основная информация:**\n"
            metrics_text += f"• Название: {getattr(asset, 'name', 'N/A')}\n"
            metrics_text += f"• Страна: {getattr(asset, 'country', 'N/A')}\n"
            metrics_text += f"• Тип: {getattr(asset, 'asset_type', 'N/A')}\n"
            metrics_text += f"• Биржа: {getattr(asset, 'exchange', 'N/A')}\n"
            metrics_text += f"• Валюта: {getattr(asset, 'currency', 'N/A')}\n\n"
            
            # Performance metrics
            metrics_text += f"**Метрики производительности:**\n"
            
            # CAGR
            if hasattr(asset, 'get_cagr'):
                cagr = asset.get_cagr()
                if hasattr(cagr, 'iloc'):
                    cagr_value = cagr.iloc[0]
                else:
                    cagr_value = cagr
                metrics_text += f"• CAGR: {cagr_value:.2%}\n"
            
            # Volatility
            if hasattr(asset, 'volatility_annual'):
                volatility = asset.volatility_annual
                if hasattr(volatility, 'iloc'):
                    vol_value = volatility.iloc[-1]
                else:
                    vol_value = volatility
                metrics_text += f"• Волатильность: {vol_value:.2%}\n"
            
            # Sharpe ratio
            if hasattr(asset, 'get_sharpe_ratio'):
                sharpe = asset.get_sharpe_ratio()
                if hasattr(sharpe, 'iloc'):
                    sharpe_value = sharpe.iloc[0]
                else:
                    sharpe_value = sharpe
                metrics_text += f"• Коэффициент Шарпа: {sharpe_value:.2f}\n"
            
            # Max drawdown
            if hasattr(asset, 'max_drawdown'):
                max_dd = asset.max_drawdown
                if hasattr(max_dd, 'iloc'):
                    dd_value = max_dd.iloc[-1]
                else:
                    dd_value = max_dd
                metrics_text += f"• Максимальная просадка: {dd_value:.2%}\n"
            
            await self._send_callback_message(update, context, metrics_text)
            
        except Exception as e:
            self.logger.error(f"Error handling info metrics button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при получении метрик: {str(e)}")

    async def _handle_info_ai_analysis_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle AI analysis button for info command"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            await self._send_ephemeral_message(update, context, "🧠 Анализирую график с помощью AI...", delete_after=3)
            
            # Get asset data for analysis
            asset = ok.Asset(symbol)
            
            # Prepare data for AI analysis
            data_info = {
                'symbols': [symbol],
                'currency': getattr(asset, 'currency', 'USD'),
                'period': 'полный доступный период данных',
                'performance': {},
                'analysis_type': 'single_asset_analysis',
                'asset_name': getattr(asset, 'name', symbol)
            }
            
            # Get AI analysis
            if self.gemini_service:
                analysis_result = self.gemini_service.analyze_data(data_info)
                if analysis_result and 'analysis' in analysis_result:
                    analysis_text = f"🧠 **AI-анализ графика {symbol}**\n\n{analysis_result['analysis']}"
                else:
                    analysis_text = f"🧠 **AI-анализ графика {symbol}**\n\n❌ Не удалось получить AI-анализ"
            else:
                analysis_text = f"🧠 **AI-анализ графика {symbol}**\n\n❌ AI-сервис недоступен"
            
            await self._send_callback_message(update, context, analysis_text, parse_mode='Markdown')
            
        except Exception as e:
            self.logger.error(f"Error handling info AI analysis button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при AI-анализе: {str(e)}", parse_mode='Markdown')

    async def _handle_info_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle compare button for info command"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            # Set user context to wait for comparison input
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                waiting_for_compare=True,
                compare_base_symbol=symbol
            )
            
            # Suggest popular alternatives
            suggestions = self._get_popular_alternatives(symbol)
            
            compare_text = f"➡️ **Сравнить {symbol} с:**\n\n"
            compare_text += "Отправьте название актива для сравнения или выберите из популярных альтернатив:\n\n"
            
            for suggestion in suggestions:
                compare_text += f"• `{suggestion}`\n"
            
            compare_text += f"\nИли напишите любой другой тикер для сравнения с {symbol}"
            
            await self._send_callback_message(update, context, compare_text, parse_mode='Markdown')
            
        except Exception as e:
            self.logger.error(f"Error handling info compare button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при подготовке сравнения: {str(e)}")

    async def _handle_info_portfolio_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle portfolio button for info command"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            # Set user context to wait for portfolio input
            user_id = update.effective_user.id
            self._update_user_context(user_id, 
                waiting_for_portfolio=True,
                portfolio_base_symbol=symbol
            )
            
            portfolio_text = f"💼 **Добавить {symbol} в портфель**\n\n"
            portfolio_text += f"Введите состав портфеля, включая {symbol}.\n\n"
            portfolio_text += "**Примеры команд:**\n"
            portfolio_text += f"• `{symbol}:0.6 QQQ.US:0.4`\n"
            portfolio_text += f"• `{symbol}:0.5 BND.US:0.3 GC.COMM:0.2`\n"
            portfolio_text += f"• `{symbol}:0.7 VTI.US:0.3`\n\n"
            portfolio_text += f"Или введите любой другой состав портфеля с {symbol}"
            
            await self._send_callback_message(update, context, portfolio_text, parse_mode='Markdown')
            
        except Exception as e:
            self.logger.error(f"Error handling info portfolio button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при подготовке портфеля: {str(e)}", parse_mode='Markdown')

    def _get_popular_alternatives(self, symbol: str) -> List[str]:
        """Get popular alternatives for comparison"""
        alternatives = {
            'SPY.US': ['QQQ.US', 'VOO.US', 'IVV.US', 'VTI.US'],
            'QQQ.US': ['SPY.US', 'VTI.US', 'VUG.US', 'ARKK.US'],
            'AAPL.US': ['MSFT.US', 'GOOGL.US', 'AMZN.US', 'TSLA.US'],
            'SBER.MOEX': ['GAZP.MOEX', 'LKOH.MOEX', 'NVTK.MOEX', 'ROSN.MOEX'],
            'VOO.US': ['SPY.US', 'IVV.US', 'VTI.US', 'VTSAX.US']
        }
        
        return alternatives.get(symbol, ['SPY.US', 'QQQ.US', 'VTI.US', 'BND.US'])

    async def _update_message_with_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, chart_data: bytes, caption: str, reply_markup):
        """Update existing message with new chart and caption"""
        try:
            import io
            from telegram import InputMediaPhoto
            
            # Create media object
            media = InputMediaPhoto(
                media=io.BytesIO(chart_data),
                caption=caption,
                parse_mode='Markdown'
            )
            
            # Update the message
            await context.bot.edit_message_media(
                chat_id=update.callback_query.message.chat_id,
                message_id=update.callback_query.message.message_id,
                media=media,
                reply_markup=reply_markup
            )
            
        except Exception as e:
            self.logger.error(f"Error updating message with chart: {e}")
            # Fallback: send new message
            await self._send_photo_safe(update, chart_data, caption=caption, reply_markup=reply_markup, context=context)

    async def _update_message_with_text(self, update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, reply_markup):
        """Update existing message with new text"""
        try:
            await context.bot.edit_message_text(
                chat_id=update.callback_query.message.chat_id,
                message_id=update.callback_query.message.message_id,
                text=text,
                parse_mode='Markdown',
                reply_markup=reply_markup
            )
            
        except Exception as e:
            self.logger.error(f"Error updating message with text: {e}")
            # Fallback: send new message
            await self._send_message_safe(update, text, reply_markup=reply_markup)

    async def _get_chart_for_period(self, symbol: str, period: str) -> Optional[bytes]:
        """Get chart for specific period using daily data with proper filtering"""
        try:
            import io
            
            def create_period_chart():
                # Устанавливаем backend для headless режима
                import matplotlib
                matplotlib.use('Agg')
                
                asset = ok.Asset(symbol)
                
                # Получаем дневные данные
                daily_data = asset.close_daily
                
                # Определяем количество торговых дней для периода
                if period == '1Y':
                    trading_days = 252  # ~1 год торговых дней
                elif period == '5Y':
                    trading_days = 1260  # ~5 лет торговых дней
                elif period == 'MAX':
                    trading_days = len(daily_data)  # Все доступные данные
                else:
                    trading_days = 252  # По умолчанию 1 год
                
                # Фильтруем данные по периоду
                if trading_days < len(daily_data):
                    filtered_data = daily_data.tail(trading_days)
                else:
                    filtered_data = daily_data
                
                # Получаем информацию об активе для заголовка
                asset_name = getattr(asset, 'name', symbol)
                currency = getattr(asset, 'currency', '')
                
                # Используем ChartStyles для создания графика
                fig, ax = chart_styles.create_price_chart(
                    data=filtered_data,
                    symbol=symbol,
                    currency=currency,
                    period=period,
                    data_source='okama'
                )
                
                # Создаем заголовок
                title = f"{symbol} | {asset_name} | {currency} | {period}"
                ax.set_title(title)
                
                # Убираем подписи осей
                ax.set_xlabel('')
                ax.set_ylabel('')
                
                # Конвертируем в bytes
                buffer = io.BytesIO()
                fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
                buffer.seek(0)
                chart_bytes = buffer.getvalue()
                buffer.close()
                
                return chart_bytes
            
            # Выполняем создание графика в отдельном потоке
            import asyncio
            loop = asyncio.get_event_loop()
            chart_bytes = await loop.run_in_executor(None, create_period_chart)
            
            return chart_bytes
            
        except Exception as e:
            self.logger.error(f"Error getting chart for period {period}: {e}")
            self.logger.error(f"Error type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None

    async def _handle_single_dividends_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle dividends button click for single asset"""
        try:
            # Remove buttons from the old message (only for callback queries)
            if hasattr(update, 'callback_query') and update.callback_query is not None:
                try:
                    await update.callback_query.edit_message_reply_markup(reply_markup=None)
                except Exception as e:
                    self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            await self._send_callback_message(update, context, "💵 Получаю информацию о дивидендах...")
            
            # Получаем информацию о дивидендах
            try:
                asset = ok.Asset(symbol)
                if hasattr(asset, 'dividends') and asset.dividends is not None:
                    dividend_info = {'dividends': asset.dividends, 'currency': getattr(asset, 'currency', '')}
                else:
                    dividend_info = {'error': 'No dividends data'}
            except Exception as e:
                dividend_info = {'error': str(e)}
            
            if 'error' not in dividend_info:
                dividends = dividend_info.get('dividends')
                currency = dividend_info.get('currency', '')
                
                # Проверяем, что дивиденды не пустые (исправляем проблему с pandas Series)
                has_dividends = False
                if dividends is not None:
                    if isinstance(dividends, pd.Series):
                        has_dividends = not dividends.empty and dividends.size > 0
                    else:
                        has_dividends = bool(dividends) and len(dividends) > 0
                
                if has_dividends:
                    # Получаем график дивидендов
                    dividend_chart = await self._get_dividend_chart(symbol)
                    
                    if dividend_chart:
                        # Send photo - handle both callback query and regular message
                        if hasattr(update, 'callback_query') and update.callback_query is not None:
                            await update.callback_query.message.reply_photo(
                                photo=dividend_chart,
                                caption=f"💵 Дивиденды {symbol}"
                            )
                        else:
                            await self._send_photo_safe(update, dividend_chart, f"💵 Дивиденды {symbol}", context=context)
                    else:
                        await self._send_callback_message(update, context, f"💵 Дивиденды {symbol} - график недоступен")
                else:
                    await self._send_callback_message(update, context, f"💵 Дивиденды по активу {symbol} не найдены")
            else:
                await self._send_callback_message(update, context, f"💵 Информация о дивидендах по активу {symbol} недоступна")
                
        except Exception as e:
            self.logger.error(f"Error handling dividends button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при получении дивидендов: {str(e)}")

    async def _handle_tushare_daily_chart_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle Tushare daily chart button click"""
        try:
            await self._send_ephemeral_message(update, context, "📈 Создаю график за 1 год...", delete_after=3)
            
            if not self.tushare_service:
                await self._send_callback_message(update, context, "❌ Сервис Tushare недоступен")
                return
            
            # Get daily chart from Tushare
            chart_bytes = await self._get_tushare_daily_chart(symbol)
            
            if chart_bytes:
                # Get symbol info for enhanced caption
                symbol_info = self.tushare_service.get_symbol_info(symbol)
                chart_caption = self._format_tushare_chart_caption(symbol_info, symbol, "1 год")
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id,
                    photo=io.BytesIO(chart_bytes),
                    caption=self._truncate_caption(chart_caption)
                )
            else:
                await self._send_callback_message(update, context, "❌ Не удалось создать график")
                
        except Exception as e:
            self.logger.error(f"Error handling Tushare daily chart button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика: {str(e)}")

    async def _handle_tushare_monthly_chart_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle Tushare monthly chart button click"""
        try:
            await self._send_ephemeral_message(update, context, "📅 Создаю график за 5 лет...", delete_after=3)
            
            if not self.tushare_service:
                await self._send_callback_message(update, context, "❌ Сервис Tushare недоступен")
                return
            
            # Get monthly chart from Tushare
            chart_bytes = await self._get_tushare_monthly_chart(symbol)
            
            if chart_bytes:
                # Get symbol info for enhanced caption
                symbol_info = self.tushare_service.get_symbol_info(symbol)
                chart_caption = self._format_tushare_chart_caption(symbol_info, symbol, "5 лет")
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id,
                    photo=io.BytesIO(chart_bytes),
                    caption=self._truncate_caption(chart_caption)
                )
            else:
                await self._send_callback_message(update, context, "❌ Не удалось создать график")
                
        except Exception as e:
            self.logger.error(f"Error handling Tushare monthly chart button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика: {str(e)}")

    async def _handle_tushare_all_chart_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle Tushare all chart button click"""
        try:
            await self._send_ephemeral_message(update, context, "📊 Создаю график за весь период...", delete_after=3)
            
            if not self.tushare_service:
                await self._send_callback_message(update, context, "❌ Сервис Tushare недоступен")
                return
            
            # Get all chart from Tushare
            chart_bytes = await self._get_tushare_all_chart(symbol)
            
            if chart_bytes:
                await context.bot.send_photo(
                    chat_id=update.effective_chat.id,
                    photo=io.BytesIO(chart_bytes),
                    caption=self._truncate_caption(f"📊 График за весь период {symbol}")
                )
            else:
                await self._send_callback_message(update, context, "❌ Не удалось создать график")
                
        except Exception as e:
            self.logger.error(f"Error handling Tushare all chart button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика: {str(e)}")

    async def _get_tushare_all_chart(self, symbol: str) -> Optional[bytes]:
        """Get chart from Tushare data for all available period"""
        try:
            import io
            
            def create_tushare_all_chart():
                try:
                    # Set backend for headless mode
                    import matplotlib
                    matplotlib.use('Agg')
                    
                    # Get monthly data from Tushare (for all period)
                    monthly_data = self.tushare_service.get_monthly_data(symbol)
                    
                    if monthly_data.empty:
                        self.logger.warning(f"Monthly data is empty for {symbol}")
                        return None
                    
                    # Prepare data for chart - set trade_date as index
                    chart_data = monthly_data.set_index('trade_date')['close']
                    
                    # Determine currency based on exchange
                    currency = 'HKD' if symbol.endswith('.HK') else 'CNY'
                    
                    # Get asset name from symbol
                    asset_name = symbol.split('.')[0]
                    
                    # Create chart using ChartStyles
                    fig, ax = chart_styles.create_price_chart(
                        data=chart_data,
                        symbol=symbol,
                        currency=currency,
                        period='All',
                        data_source='okama'
                    )
                    
                    # Обновляем заголовок с нужным форматом
                    title = f"{symbol} | {asset_name} | {currency} | All"
                    ax.set_title(title, **chart_styles.title)
                    
                    # Убираем подписи осей
                    ax.set_xlabel('')
                    ax.set_ylabel('')
                    
                    # Save to bytes
                    output = io.BytesIO()
                    chart_styles.save_figure(fig, output)
                    output.seek(0)
                    
                    # Cleanup
                    chart_styles.cleanup_figure(fig)
                    
                    return output.getvalue()
                    
                except Exception as e:
                    self.logger.error(f"Error in create_tushare_all_chart for {symbol}: {e}")
                    import traceback
                    self.logger.error(f"Traceback: {traceback.format_exc()}")
                    return None
            
            # Run chart creation in thread to avoid blocking
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(create_tushare_all_chart)
                chart_bytes = future.result(timeout=30)
            
            return chart_bytes
            
        except Exception as e:
            self.logger.error(f"Error getting Tushare all chart for {symbol}: {e}")
            return None

    async def _handle_tushare_dividends_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbol: str):
        """Handle Tushare dividends button click with enhanced information"""
        try:
            await self._send_callback_message(update, context, "💵 Получаю информацию о дивидендах и рейтингах...")
            
            if not self.tushare_service:
                await self._send_callback_message(update, context, "❌ Сервис Tushare недоступен")
                return
            
            # Get symbol info for English name
            symbol_info = self.tushare_service.get_symbol_info(symbol)
            english_name = symbol_info.get('name', symbol.split('.')[0])
            
            # Get dividend data from Tushare
            dividend_data = self.tushare_service.get_dividend_data(symbol)
            
            # Get rating data if available
            rating_data = self.tushare_service.get_rating_data(symbol)
            
            # Format dividend information
            info_text = f"💵 Дивиденды {symbol}\n"
            info_text += f"🏢 {english_name}\n\n"
            
            if dividend_data.empty:
                info_text += "❌ Дивиденды не найдены\n"
                info_text += "💡 Возможные причины:\n"
                info_text += "   • Компания не выплачивает дивиденды\n"
                info_text += "   • Данные о дивидендах недоступны в Tushare\n"
                info_text += "   • Компания выплачивает дивиденды нерегулярно\n\n"
            else:
                # Calculate dividend statistics
                total_dividends = len(dividend_data)
                recent_dividends = dividend_data.head(5)  # Most recent 5
                
                info_text += f"📊 Всего выплат: {total_dividends}\n"
                
                # Calculate average dividend
                if 'cash_div_tax' in dividend_data.columns:
                    avg_dividend = dividend_data['cash_div_tax'].mean()
                    info_text += f"💰 Средний дивиденд: {avg_dividend:.4f}\n\n"
                
                info_text += "📅 Последние дивиденды:\n"
                
                for _, row in recent_dividends.iterrows():
                    ann_date = row['ann_date'].strftime('%Y-%m-%d') if pd.notna(row['ann_date']) else 'N/A'
                    div_proc_date = row['div_proc_date'].strftime('%Y-%m-%d') if pd.notna(row['div_proc_date']) else 'N/A'
                    stk_div_date = row['stk_div_date'].strftime('%Y-%m-%d') if pd.notna(row['stk_div_date']) else 'N/A'
                    
                    # Get dividend amount (prefer after-tax, fallback to before-tax, then tax)
                    dividend_amount = 0
                    if 'cash_div_after_tax' in row and row['cash_div_after_tax'] > 0:
                        dividend_amount = row['cash_div_after_tax']
                    elif 'cash_div_before_tax' in row and row['cash_div_before_tax'] > 0:
                        dividend_amount = row['cash_div_before_tax']
                    elif 'cash_div_tax' in row and row['cash_div_tax'] > 0:
                        dividend_amount = row['cash_div_tax']
                    
                    info_text += f"📅 {ann_date}: {dividend_amount:.4f}\n"
                    if div_proc_date != 'N/A':
                        info_text += f"   💰 Выплата: {div_proc_date}\n"
                    if stk_div_date != 'N/A':
                        info_text += f"   📊 Регистрация: {stk_div_date}\n"
                    info_text += "\n"
            
            # Add rating information if available
            if not rating_data.empty:
                info_text += "⭐ Рейтинги:\n"
                recent_ratings = rating_data.head(3)  # Most recent 3 ratings
                
                for _, row in recent_ratings.iterrows():
                    rating_date = row['rating_date'].strftime('%Y-%m-%d') if pd.notna(row['rating_date']) else 'N/A'
                    
                    # Get rating information (field names may vary)
                    rating_info = []
                    if 'rating_agency' in row and pd.notna(row['rating_agency']):
                        rating_info.append(f"Агентство: {row['rating_agency']}")
                    if 'rating_type' in row and pd.notna(row['rating_type']):
                        rating_info.append(f"Тип: {row['rating_type']}")
                    if 'rating_value' in row and pd.notna(row['rating_value']):
                        rating_info.append(f"Рейтинг: {row['rating_value']}")
                    
                    info_text += f"📅 {rating_date}\n"
                    if rating_info:
                        info_text += f"   {' | '.join(rating_info)}\n"
                    info_text += "\n"
            else:
                info_text += "⭐ Рейтинги: Недоступны\n"
            
            # Try to create dividend chart if we have data
            if not dividend_data.empty:
                try:
                    chart_bytes = await self._create_tushare_dividend_chart(symbol, dividend_data, symbol_info)
                    if chart_bytes:
                        caption = f"💵 График дивидендов {symbol}\n🏢 {english_name}"
                        await context.bot.send_photo(
                            chat_id=update.effective_chat.id,
                            photo=io.BytesIO(chart_bytes),
                            caption=self._truncate_caption(caption)
                        )
                except Exception as e:
                    self.logger.warning(f"Could not create dividend chart: {e}")
            
            await self._send_callback_message(update, context, info_text)
                
        except Exception as e:
            self.logger.error(f"Error handling Tushare dividends button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при получении дивидендов: {str(e)}")

    async def _create_tushare_dividend_chart(self, symbol: str, dividend_data, symbol_info: Dict[str, Any]) -> Optional[bytes]:
        """Create dividend chart for Tushare asset"""
        try:
            import io
            import pandas as pd
            from services.chart_styles import ChartStyles
            
            # Prepare dividend data for chart
            dividend_data = dividend_data.copy()
            
            # Get dividend amounts (prefer after-tax, fallback to before-tax, then tax)
            dividend_amounts = []
            for _, row in dividend_data.iterrows():
                amount = 0
                if 'cash_div_after_tax' in row and row['cash_div_after_tax'] > 0:
                    amount = row['cash_div_after_tax']
                elif 'cash_div_before_tax' in row and row['cash_div_before_tax'] > 0:
                    amount = row['cash_div_before_tax']
                elif 'cash_div_tax' in row and row['cash_div_tax'] > 0:
                    amount = row['cash_div_tax']
                dividend_amounts.append(amount)
            
            dividend_data['dividend_amount'] = dividend_amounts
            
            # Filter out zero dividends
            dividend_data = dividend_data[dividend_data['dividend_amount'] > 0]
            
            if dividend_data.empty:
                return None
            
            # Group by year and sum dividends
            dividend_data['year'] = dividend_data['ann_date'].dt.year
            yearly_dividends = dividend_data.groupby('year')['dividend_amount'].sum()
            
            # Take last 10 years
            yearly_dividends = yearly_dividends.tail(10)
            
            if yearly_dividends.empty:
                return None
            
            # Create chart using ChartStyles
            chart_styles = ChartStyles()
            fig, ax = chart_styles.create_dividends_chart(
                data=yearly_dividends,
                symbol=symbol,
                currency='HKD' if symbol.endswith('.HK') else 'CNY',
                asset_name=symbol_info.get('name', symbol.split('.')[0])
            )
            
            # Set title with proper format
            english_name = symbol_info.get('name', symbol.split('.')[0])
            currency = 'HKD' if symbol.endswith('.HK') else 'CNY'
            title = f"{symbol} | {english_name} | {currency} | Дивиденды"
            ax.set_title(title, **chart_styles.title)
            
            # Убираем подписи осей
            ax.set_xlabel('')
            ax.set_ylabel('')
            
            # Convert to bytes
            buffer = io.BytesIO()
            fig.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            chart_bytes = buffer.getvalue()
            buffer.close()
            
            return chart_bytes
            
        except Exception as e:
            self.logger.error(f"Error creating Tushare dividend chart for {symbol}: {e}")
            return None

    async def _get_monthly_chart(self, symbol: str) -> Optional[bytes]:
        """Получить месячный график за последние 5 лет используя ChartStyles"""
        try:
            import io
            
            def create_monthly_chart():
                # Устанавливаем backend для headless режима
                import matplotlib
                matplotlib.use('Agg')
                
                asset = ok.Asset(symbol)
                
                # Получаем месячные данные
                monthly_data = asset.close_monthly
                
                # Берем последние 60 месяцев (5 лет)
                filtered_data = monthly_data.tail(60)
                
                # Получаем информацию об активе для заголовка
                asset_name = getattr(asset, 'name', symbol)
                currency = getattr(asset, 'currency', '')
                
                # Используем ChartStyles для создания графика
                fig, ax = chart_styles.create_price_chart(
                    data=filtered_data,
                    symbol=symbol,
                    currency=currency,
                    period='5Y',
                    data_source='okama'
                )
                
                # Обновляем заголовок с нужным форматом
                title = f"{symbol} | {asset_name} | {currency} | 5Y"
                ax.set_title(title, **chart_styles.title)
                
                # Убираем подписи осей
                ax.set_xlabel('')
                ax.set_ylabel('')
                
                # Сохраняем в bytes
                output = io.BytesIO()
                chart_styles.save_figure(fig, output)
                output.seek(0)
                
                # Очистка
                chart_styles.cleanup_figure(fig)
                
                return output.getvalue()
            
            # Выполняем с таймаутом
            chart_data = await asyncio.wait_for(
                asyncio.to_thread(create_monthly_chart),
                timeout=30.0
            )
            
            return chart_data
            
        except Exception as e:
            self.logger.error(f"Error getting monthly chart for {symbol}: {e}")
            self.logger.error(f"Error type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None

    async def _get_all_chart(self, symbol: str) -> Optional[bytes]:
        """Получить график за весь доступный период используя ChartStyles"""
        try:
            import io
            
            def create_all_chart():
                # Устанавливаем backend для headless режима
                import matplotlib
                matplotlib.use('Agg')
                
                asset = ok.Asset(symbol)
                
                # Получаем месячные данные за весь период
                monthly_data = asset.close_monthly
                
                # Получаем информацию об активе для заголовка
                asset_name = getattr(asset, 'name', symbol)
                currency = getattr(asset, 'currency', '')
                
                # Используем ChartStyles для создания графика
                fig, ax = chart_styles.create_price_chart(
                    data=monthly_data,
                    symbol=symbol,
                    currency=currency,
                    period='All',
                    data_source='okama'
                )
                
                # Обновляем заголовок с нужным форматом
                title = f"{symbol} | {asset_name} | {currency} | All"
                ax.set_title(title, **chart_styles.title)
                
                # Убираем подписи осей
                ax.set_xlabel('')
                ax.set_ylabel('')
                
                # Сохраняем в bytes
                output = io.BytesIO()
                chart_styles.save_figure(fig, output)
                output.seek(0)
                
                # Очистка
                chart_styles.cleanup_figure(fig)
                
                return output.getvalue()
            
            # Выполняем с таймаутом
            chart_data = await asyncio.wait_for(
                asyncio.to_thread(create_all_chart),
                timeout=30.0
            )
            
            return chart_data
            
        except Exception as e:
            self.logger.error(f"Error getting all chart for {symbol}: {e}")
            self.logger.error(f"Error type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None

    async def _get_dividend_chart(self, symbol: str) -> Optional[bytes]:
        """Получить график дивидендов с копирайтом"""
        try:
            # Получаем данные о дивидендах
            try:
                asset = ok.Asset(symbol)
                if hasattr(asset, 'dividends') and asset.dividends is not None:
                    dividend_info = {'dividends': asset.dividends, 'currency': getattr(asset, 'currency', '')}
                else:
                    dividend_info = {'error': 'No dividends data'}
            except Exception as e:
                dividend_info = {'error': str(e)}
            
            if 'error' in dividend_info:
                return None
            
            dividends = dividend_info.get('dividends')
            if dividends is None:
                return None
            
            # Проверяем, что дивиденды не пустые (исправляем проблему с pandas Series)
            if isinstance(dividends, pd.Series):
                if dividends.empty or dividends.size == 0:
                    self.logger.info(f"No dividends data found for {symbol} (empty Series)")
                    return None
            else:
                if not dividends or len(dividends) == 0:
                    self.logger.info(f"No dividends data found for {symbol} (empty data)")
                    return None
            
            # Получаем название компании
            asset_name = symbol  # Default to symbol
            try:
                if hasattr(asset, 'name') and asset.name:
                    asset_name = asset.name
                elif hasattr(asset, 'symbol') and asset.symbol:
                    asset_name = asset.symbol
            except Exception as e:
                self.logger.warning(f"Failed to get asset name for {symbol}: {e}")
            
            # Создаем график дивидендов
            dividend_chart = self._create_dividend_chart(symbol, dividend_info['dividends'], dividend_info.get('currency', ''), asset_name)
            
            if dividend_chart:
                # Копирайт уже добавлен в _create_dividend_chart
                return dividend_chart
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error getting dividend chart for {symbol}: {e}")
            return None

    async def _get_tushare_daily_chart(self, symbol: str) -> Optional[bytes]:
        """Get daily chart from Tushare data"""
        try:
            import io
            
            def create_tushare_daily_chart():
                try:
                    # Set backend for headless mode
                    import matplotlib
                    matplotlib.use('Agg')
                    
                    # Get daily data from Tushare
                    daily_data = self.tushare_service.get_daily_data(symbol)
                    
                    if daily_data.empty:
                        self.logger.warning(f"Daily data is empty for {symbol}")
                        return None
                    
                    # Prepare data for chart - set trade_date as index
                    chart_data = daily_data.set_index('trade_date')['close']
                    
                    # Take last 252 trading days (approximately 1 year)
                    filtered_data = chart_data.tail(252)
                    
                    # Determine currency based on exchange
                    currency = 'HKD' if symbol.endswith('.HK') else 'CNY'
                    
                    # Get asset information including English name
                    symbol_info = self.tushare_service.get_symbol_info(symbol)
                    english_name = symbol_info.get('name', symbol.split('.')[0])
                    
                    # Create chart using ChartStyles
                    fig, ax = chart_styles.create_price_chart(
                        data=filtered_data,
                        symbol=symbol,
                        currency=currency,
                        period='1Y',
                        data_source='tushare'
                    )
                    
                    # Обновляем заголовок с нужным форматом: Тикер | Английское название | Валюта | Срок
                    title = f"{symbol} | {english_name} | {currency} | 1Y"
                    ax.set_title(title, **chart_styles.title)
                    
                    # Убираем подписи осей
                    ax.set_xlabel('')
                    ax.set_ylabel('')
                    
                    # Save to bytes
                    output = io.BytesIO()
                    chart_styles.save_figure(fig, output)
                    output.seek(0)
                    
                    # Cleanup
                    chart_styles.cleanup_figure(fig)
                    
                    return output.getvalue()
                    
                except Exception as e:
                    self.logger.error(f"Error in create_tushare_daily_chart for {symbol}: {e}")
                    import traceback
                    self.logger.error(f"Traceback: {traceback.format_exc()}")
                    return None
            
            # Run chart creation in thread to avoid blocking
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(create_tushare_daily_chart)
                chart_bytes = future.result(timeout=30)
                
            return chart_bytes
            
        except Exception as e:
            self.logger.error(f"Error getting Tushare daily chart for {symbol}: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None

    async def _get_tushare_monthly_chart(self, symbol: str) -> Optional[bytes]:
        """Get monthly chart from Tushare data"""
        try:
            import io
            
            def create_tushare_monthly_chart():
                try:
                    # Set backend for headless mode
                    import matplotlib
                    matplotlib.use('Agg')
                    
                    # Get monthly data from Tushare
                    monthly_data = self.tushare_service.get_monthly_data(symbol)
                    
                    if monthly_data.empty:
                        self.logger.warning(f"Monthly data is empty for {symbol}")
                        return None
                    
                    # Prepare data for chart - set trade_date as index
                    chart_data = monthly_data.set_index('trade_date')['close']
                    
                    # Take last 60 months (5 years)
                    filtered_data = chart_data.tail(60)
                    
                    # Determine currency based on exchange
                    currency = 'HKD' if symbol.endswith('.HK') else 'CNY'
                    
                    # Get asset information including English name
                    symbol_info = self.tushare_service.get_symbol_info(symbol)
                    english_name = symbol_info.get('name', symbol.split('.')[0])
                    
                    # Create chart using ChartStyles
                    fig, ax = chart_styles.create_price_chart(
                        data=filtered_data,
                        symbol=symbol,
                        currency=currency,
                        period='5Y',
                        data_source='tushare'
                    )
                    
                    # Обновляем заголовок с нужным форматом: Тикер | Английское название | Валюта | Срок
                    title = f"{symbol} | {english_name} | {currency} | 5Y"
                    ax.set_title(title, **chart_styles.title)
                    
                    # Убираем подписи осей
                    ax.set_xlabel('')
                    ax.set_ylabel('')
                    
                    # Save to bytes
                    output = io.BytesIO()
                    chart_styles.save_figure(fig, output)
                    output.seek(0)
                    
                    # Cleanup
                    chart_styles.cleanup_figure(fig)
                    
                    return output.getvalue()
                    
                except Exception as e:
                    self.logger.error(f"Error in create_tushare_monthly_chart for {symbol}: {e}")
                    import traceback
                    self.logger.error(f"Traceback: {traceback.format_exc()}")
                    return None
            
            # Run chart creation in thread to avoid blocking
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(create_tushare_monthly_chart)
                chart_bytes = future.result(timeout=30)
                
            return chart_bytes
            
        except Exception as e:
            self.logger.error(f"Error getting Tushare monthly chart for {symbol}: {e}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None

    async def _get_dividend_table_image(self, symbol: str) -> Optional[bytes]:
        """Получить изображение таблицы дивидендов с копирайтом"""
        try:
            # Получаем данные о дивидендах
            try:
                asset = ok.Asset(symbol)
                if hasattr(asset, 'dividends') and asset.dividends is not None:
                    dividend_info = {'dividends': asset.dividends, 'currency': getattr(asset, 'currency', '')}
                else:
                    dividend_info = {'error': 'No dividends data'}
            except Exception as e:
                dividend_info = {'error': str(e)}
            
            if 'error' in dividend_info:
                return None
            
            dividends = dividend_info.get('dividends')
            if dividends is None:
                return None
            
            # Проверяем, что дивиденды не пустые (исправляем проблему с pandas Series)
            if isinstance(dividends, pd.Series):
                if dividends.empty or dividends.size == 0:
                    return None
            else:
                if not dividends or len(dividends) == 0:
                    return None
            
            # Создаем изображение таблицы дивидендов
            dividend_table = self._create_dividend_table_image(symbol, dividend_info['dividends'], dividend_info.get('currency', ''))
            
            if dividend_table:
                # Копирайт уже добавлен в _create_dividend_table_image
                return dividend_table
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error getting dividend table image for {symbol}: {e}")
            return None

    def _create_dividend_chart(self, symbol: str, dividends: dict, currency: str, asset_name: str = None) -> Optional[bytes]:
        """Создать график дивидендов используя ChartStyles"""
        try:
            import io
            
            # Конвертируем дивиденды в pandas Series
            dividend_series = pd.Series(dividends)
            
            # Сортируем по дате
            dividend_series = dividend_series.sort_index()
            
            # Используем ChartStyles для создания графика дивидендов
            fig, ax = chart_styles.create_dividends_chart(
                data=dividend_series,
                symbol=symbol,
                currency=currency,
                asset_name=asset_name
            )
            
            # Сохраняем в bytes
            output = io.BytesIO()
            chart_styles.save_figure(fig, output)
            output.seek(0)
            
            # Очистка
            chart_styles.cleanup_figure(fig)
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error creating dividend chart: {e}")
            return None

    def _create_dividend_table_image(self, symbol: str, dividends: dict, currency: str) -> Optional[bytes]:
        """Создать отдельное изображение с таблицей дивидендов используя ChartStyles"""
        try:
            import io
            
            # Конвертируем дивиденды в pandas Series
            dividend_series = pd.Series(dividends)
            
            # Сортируем по дате (новые сверху)
            dividend_series = dividend_series.sort_index(ascending=False)
            
            # Берем последние 15 выплат для таблицы
            recent_dividends = dividend_series.head(15)
            
            # Создаем таблицу
            table_data = []
            table_headers = ['Дата', f'Сумма ({currency})']
            
            for date, amount in recent_dividends.items():
                # Форматируем дату для лучшей читаемости
                if hasattr(date, 'strftime'):
                    formatted_date = date.strftime('%Y-%m-%d')
                else:
                    formatted_date = str(date)[:10]
                table_data.append([formatted_date, f'{amount:.2f}'])
            
            # Создаем фигуру для таблицы используя chart_styles
            fig, ax, table = chart_styles.create_dividend_table_chart(
                table_data=table_data,
                headers=table_headers,
                title=f'Таблица дивидендов {symbol}\nПоследние {len(table_data)} выплат'
            )
            
            # Стилизуем таблицу
            table.auto_set_font_size(False)
            table.set_fontsize(11)
            table.scale(1, 2.0)
            
            # Цвета для заголовков
            for i in range(len(table_headers)):
                table[(0, i)].set_facecolor(chart_styles.colors['success'])
                table[(0, i)].set_text_props(weight='bold', color='white')
                table[(0, i)].set_height(0.12)
            
            # Цвета для строк данных (чередование)
            for i in range(1, len(table_data) + 1):
                for j in range(len(table_headers)):
                    if i % 2 == 0:
                        table[(i, j)].set_facecolor(chart_styles.colors['neutral'])
                    else:
                        table[(i, j)].set_facecolor('white')
                    table[(i, j)].set_text_props(color=chart_styles.colors['text'])
                    table[(i, j)].set_height(0.08)
            
            # Добавляем общую статистику внизу
            total_dividends = dividend_series.sum()
            avg_dividend = dividend_series.mean()
            max_dividend = dividend_series.max()
            
            stats_text = f'Общая сумма: {total_dividends:.2f} {currency} | '
            stats_text += f'Средняя выплата: {avg_dividend:.2f} {currency} | '
            stats_text += f'Максимальная выплата: {max_dividend:.2f} {currency}'
            
            ax.text(0.5, 0.02, stats_text, transform=ax.transAxes, 
                   fontsize=10, ha='center', color=chart_styles.colors['text'],
                   bbox=dict(boxstyle='round,pad=0.5', facecolor=chart_styles.colors['neutral'], alpha=0.8))
            
            # Сохраняем в bytes
            output = io.BytesIO()
            chart_styles.save_figure(fig, output)
            output.seek(0)
            
            # Очистка
            chart_styles.cleanup_figure(fig)
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error creating dividend table image: {e}")
            return None

    async def _handle_risk_metrics_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle risk metrics button click for portfolio"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling risk metrics button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context keys: {list(user_context.keys())}")
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            self.logger.info(f"Creating risk metrics for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "💼 Расчет метрик портфеля...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_risk_metrics_report(update, context, portfolio, final_symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error handling risk metrics button: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при анализе рисков: {str(e)}", parse_mode='Markdown')

    async def _handle_portfolio_risk_metrics_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio risk metrics button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio risk metrics by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            await self._send_ephemeral_message(update, context, "💼 Анализирую метрики портфеля...", delete_after=3)
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            # Create portfolio metrics table using portfolio-specific logic
            try:
                # Create portfolio object for metrics calculation
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency)
                
                # Create portfolio-specific metrics table
                summary_table = self._create_portfolio_summary_metrics_table(
                    portfolio, valid_symbols, valid_weights, currency
                )
                
                if summary_table and not summary_table.startswith("❌"):
                    # Send table as message with reply keyboard
                    header_text = f"📊 **Сводная таблица ключевых метрик**"
                    table_message = f"{header_text}\n\n```\n{summary_table}\n```"
                    await self._send_portfolio_message_with_reply_keyboard(update, context, table_message, parse_mode='Markdown')
                else:
                    await self._send_portfolio_message_with_reply_keyboard(update, context, "❌ Не удалось создать таблицу метрик")
                    
            except Exception as metrics_error:
                self.logger.error(f"Error creating summary metrics table: {metrics_error}")
                await self._send_portfolio_message_with_reply_keyboard(update, context, f"❌ Ошибка при создании таблицы метрик: {str(metrics_error)}", parse_mode='Markdown')
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio risk metrics by symbol: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при анализе рисков: {str(e)}", parse_mode='Markdown')

    async def _handle_monte_carlo_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle Monte Carlo button click for portfolio"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling Monte Carlo button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context keys: {list(user_context.keys())}")
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            self.logger.info(f"Creating Monte Carlo forecast for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "🎲 Создаю прогноз Monte Carlo...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_monte_carlo_forecast(update, context, portfolio, final_symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error handling Monte Carlo button: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании прогноза Monte Carlo: {str(e)}")

    async def _handle_portfolio_monte_carlo_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio Monte Carlo button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio Monte Carlo by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            await self._send_ephemeral_message(update, context, "🎲 Создаю прогноз Monte Carlo...", delete_after=3)
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_monte_carlo_forecast(update, context, portfolio, final_symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio Monte Carlo by symbol: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании прогноза Monte Carlo: {str(e)}")

    async def _handle_forecast_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle forecast button click for portfolio"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling forecast button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context keys: {list(user_context.keys())}")
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            self.logger.info(f"Creating forecast for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "📈 Создаю прогноз с процентилями...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_forecast_chart(update, context, portfolio, final_symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error handling forecast button: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании прогноза: {str(e)}")

    async def _handle_portfolio_forecast_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio forecast button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio forecast by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            await self._send_ephemeral_message(update, context, "📈 Создаю прогноз с процентилями...", delete_after=3)
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_forecast_chart(update, context, portfolio, final_symbols, currency)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio forecast by symbol: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании прогноза: {str(e)}")

    async def _create_risk_metrics_report(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str):
        """Create and send risk metrics report for portfolio as Excel file"""
        try:
            await self._send_ephemeral_message(update, context, "📊 Подготавливаю детальную статистику...", parse_mode='Markdown', delete_after=3)

            # Prepare comprehensive metrics data
            metrics_data = self._prepare_portfolio_metrics_data(portfolio, symbols, currency)
            
            if metrics_data:
                # Create Excel file
                excel_buffer = self._create_portfolio_metrics_excel(metrics_data, symbols, currency)
                
                if excel_buffer:
                    # Ensure portfolio keyboard is shown
                    await self._manage_reply_keyboard(update, context, "portfolio")
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=io.BytesIO(excel_buffer.getvalue()),
                        filename=f"portfolio_risk_metrics_{'_'.join(symbols[:3])}_{currency}.xlsx",
                        caption=f"💼 **Детальная статистика портфеля**\n\n"
                               f"🔍 **Анализируемые активы:** {', '.join(symbols)}\n"
                               f"💰 **Валюта:** {currency}\n"
                               f"📅 **Дата создания:** {self._get_current_timestamp()}\n\n"
                               f"📋 **Содержит:**\n"
                               f"• Основные метрики производительности\n"
                               f"• Коэффициенты Шарпа и Сортино\n"
                               f"• Анализ рисков и доходности\n"
                               f"• Детальная статистика портфеля",
                    )
                else:
                    await self._send_callback_message(update, context, "❌ Ошибка при создании Excel файла")
                    
            else:
                await self._send_callback_message(update, context, "❌ Не удалось подготовить данные для экспорта")
                
        except Exception as e:
            self.logger.error(f"Error creating risk metrics report: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании отчета о рисках: {str(e)}")

    def _calculate_portfolio_years(self, portfolio, returns) -> float:
        """Calculate the actual number of years for portfolio CAGR calculation"""
        try:
            # Handle empty returns
            if len(returns) == 0:
                return 0
            # Try to get actual dates from portfolio object
            if hasattr(portfolio, 'first_date') and hasattr(portfolio, 'last_date'):
                from datetime import datetime
                try:
                    # Parse dates from portfolio object
                    if isinstance(portfolio.first_date, str):
                        start_date = datetime.strptime(portfolio.first_date, '%Y-%m-%d')
                    else:
                        start_date = portfolio.first_date
                    
                    if isinstance(portfolio.last_date, str):
                        end_date = datetime.strptime(portfolio.last_date, '%Y-%m-%d')
                    else:
                        end_date = portfolio.last_date
                    
                    # Calculate years difference
                    days_diff = (end_date - start_date).days
                    years = days_diff / 365.25  # Use 365.25 for leap years
                    
                    if years > 0:
                        return years
                except Exception as e:
                    self.logger.warning(f"Could not parse portfolio dates: {e}")
            
            # Fallback: estimate from returns data frequency
            if len(returns) > 0:
                # Try to detect data frequency from index
                if hasattr(returns, 'index'):
                    index = returns.index
                    if len(index) > 1:
                        # Calculate average time difference between data points
                        time_diffs = []
                        for i in range(1, min(len(index), 10)):  # Check first 10 differences
                            try:
                                diff = (index[i] - index[i-1]).days
                                time_diffs.append(diff)
                            except:
                                continue
                        
                        if time_diffs:
                            avg_days = sum(time_diffs) / len(time_diffs)
                            total_days = (index[-1] - index[0]).days
                            years = total_days / 365.25
                            return years
            
            # Final fallback: assume monthly data
            return len(returns) / 12 if len(returns) > 0 else 0
            
        except Exception as e:
            self.logger.warning(f"Error calculating portfolio years: {e}")
            # Ultimate fallback: assume monthly data
            return len(returns) / 12 if len(returns) > 0 else 0

    def _calculate_asset_years(self, asset, returns) -> float:
        """Calculate the actual number of years for asset CAGR calculation"""
        try:
            # Handle empty returns
            if len(returns) == 0:
                return 0
            # Try to get actual dates from asset object
            if hasattr(asset, 'first_date') and hasattr(asset, 'last_date'):
                from datetime import datetime
                try:
                    # Parse dates from asset object
                    if isinstance(asset.first_date, str):
                        start_date = datetime.strptime(asset.first_date, '%Y-%m-%d')
                    else:
                        start_date = asset.first_date
                    
                    if isinstance(asset.last_date, str):
                        end_date = datetime.strptime(asset.last_date, '%Y-%m-%d')
                    else:
                        end_date = asset.last_date
                    
                    # Calculate years difference
                    days_diff = (end_date - start_date).days
                    years = days_diff / 365.25  # Use 365.25 for leap years
                    
                    if years > 0:
                        return years
                except Exception as e:
                    self.logger.warning(f"Could not parse asset dates: {e}")
            
            # Fallback: estimate from returns data frequency
            if len(returns) > 0:
                # Try to detect data frequency from index
                if hasattr(returns, 'index'):
                    index = returns.index
                    if len(index) > 1:
                        # Calculate average time difference between data points
                        time_diffs = []
                        for i in range(1, min(len(index), 10)):  # Check first 10 differences
                            try:
                                diff = (index[i] - index[i-1]).days
                                time_diffs.append(diff)
                            except:
                                continue
                        
                        if time_diffs:
                            avg_days = sum(time_diffs) / len(time_diffs)
                            total_days = (index[-1] - index[0]).days
                            years = total_days / 365.25
                            return years
            
            # Final fallback: assume monthly data
            return len(returns) / 12 if len(returns) > 0 else 0
            
        except Exception as e:
            self.logger.warning(f"Error calculating asset years: {e}")
            # Ultimate fallback: assume monthly data
            return len(returns) / 12 if len(returns) > 0 else 0

    def _prepare_portfolio_metrics_data(self, portfolio, symbols: list, currency: str) -> Dict[str, Any]:
        """Prepare comprehensive metrics data for portfolio Excel export"""
        try:
            metrics_data = {
                'timestamp': self._get_current_timestamp(),
                'currency': currency,
                'symbols': symbols,
                'period': 'MAX',
                'detailed_metrics': {},
                'portfolio_metrics': {}
            }
            
            # Portfolio-level metrics - calculate manually to ensure proper values
            portfolio_metrics = {}
            
            try:
                # Get portfolio returns data
                if hasattr(portfolio, 'returns'):
                    returns = portfolio.returns
                elif hasattr(portfolio, 'get_returns'):
                    returns = portfolio.get_returns()
                else:
                    # Fallback: calculate returns from price data
                    if hasattr(portfolio, 'prices'):
                        prices = portfolio.prices
                        returns = prices.pct_change().dropna()
                    else:
                        returns = None
                
                if returns is not None and len(returns) > 0:
                    # Annual Return (CAGR)
                    if hasattr(portfolio, 'mean_return_annual'):
                        try:
                            portfolio_metrics['annual_return'] = float(portfolio.mean_return_annual) * 100
                        except:
                            # Calculate manually with real portfolio period
                            total_return = (1 + returns).prod() - 1
                            years = self._calculate_portfolio_years(portfolio, returns)
                            cagr = (1 + total_return) ** (1 / years) - 1
                            portfolio_metrics['annual_return'] = cagr * 100
                    else:
                        # Calculate manually with real portfolio period
                        total_return = (1 + returns).prod() - 1
                        years = self._calculate_portfolio_years(portfolio, returns)
                        cagr = (1 + total_return) ** (1 / years) - 1
                        portfolio_metrics['annual_return'] = cagr * 100
                    
                    # Volatility
                    if hasattr(portfolio, 'volatility_annual'):
                        try:
                            portfolio_metrics['volatility'] = float(portfolio.volatility_annual) * 100
                        except:
                            # Calculate manually
                            volatility = returns.std() * (12 ** 0.5)  # Annualized for monthly data
                            portfolio_metrics['volatility'] = volatility * 100
                    else:
                        # Calculate manually
                        volatility = returns.std() * (12 ** 0.5)  # Annualized for monthly data
                        portfolio_metrics['volatility'] = volatility * 100
                    
                    # Sharpe Ratio
                    if hasattr(portfolio, 'sharpe_ratio'):
                        try:
                            portfolio_metrics['sharpe_ratio'] = float(portfolio.sharpe_ratio)
                        except:
                            # Calculate using unified function
                            annual_return = portfolio_metrics['annual_return'] / 100
                            volatility = portfolio_metrics['volatility'] / 100
                            sharpe_ratio = self.calculate_sharpe_ratio(annual_return, volatility, currency, asset_data=portfolio)
                            portfolio_metrics['sharpe_ratio'] = sharpe_ratio
                    else:
                        # Calculate using unified function
                        annual_return = portfolio_metrics['annual_return'] / 100
                        volatility = portfolio_metrics['volatility'] / 100
                        sharpe_ratio = self.calculate_sharpe_ratio(annual_return, volatility, currency, asset_data=portfolio)
                        portfolio_metrics['sharpe_ratio'] = sharpe_ratio
                    
                    # Sortino Ratio
                    if hasattr(portfolio, 'sortino_ratio'):
                        try:
                            portfolio_metrics['sortino_ratio'] = float(portfolio.sortino_ratio)
                        except:
                            # Calculate manually
                            annual_return = portfolio_metrics['annual_return'] / 100
                            negative_returns = returns[returns < 0]
                            if len(negative_returns) > 0:
                                downside_deviation = negative_returns.std() * (12 ** 0.5)  # Annualized
                                if downside_deviation > 0:
                                    sortino_ratio = (annual_return - 0.02) / downside_deviation
                                    portfolio_metrics['sortino_ratio'] = sortino_ratio
                                else:
                                    portfolio_metrics['sortino_ratio'] = 0.0
                            else:
                                # No negative returns, use volatility as fallback
                                volatility = portfolio_metrics['volatility'] / 100
                                if volatility > 0:
                                    sortino_ratio = (annual_return - 0.02) / volatility
                                    portfolio_metrics['sortino_ratio'] = sortino_ratio
                                else:
                                    portfolio_metrics['sortino_ratio'] = 0.0
                    else:
                        # Calculate manually
                        annual_return = portfolio_metrics['annual_return'] / 100
                        negative_returns = returns[returns < 0]
                        if len(negative_returns) > 0:
                            downside_deviation = negative_returns.std() * (12 ** 0.5)  # Annualized
                            if downside_deviation > 0:
                                sortino_ratio = (annual_return - 0.02) / downside_deviation
                                portfolio_metrics['sortino_ratio'] = sortino_ratio
                            else:
                                portfolio_metrics['sortino_ratio'] = 0.0
                        else:
                            # No negative returns, use volatility as fallback
                            volatility = portfolio_metrics['volatility'] / 100
                            if volatility > 0:
                                sortino_ratio = (annual_return - 0.02) / volatility
                                portfolio_metrics['sortino_ratio'] = sortino_ratio
                            else:
                                portfolio_metrics['sortino_ratio'] = 0.0
                    
                    # Max Drawdown
                    if hasattr(portfolio, 'max_drawdown'):
                        try:
                            portfolio_metrics['max_drawdown'] = float(portfolio.max_drawdown) * 100
                        except:
                            # Calculate manually
                            cumulative = (1 + returns).cumprod()
                            running_max = cumulative.expanding().max()
                            drawdown = (cumulative - running_max) / running_max
                            max_drawdown = drawdown.min()
                            portfolio_metrics['max_drawdown'] = max_drawdown * 100
                    else:
                        # Calculate manually
                        cumulative = (1 + returns).cumprod()
                        running_max = cumulative.expanding().max()
                        drawdown = (cumulative - running_max) / running_max
                        max_drawdown = drawdown.min()
                        portfolio_metrics['max_drawdown'] = max_drawdown * 100
                    
                    # Calmar Ratio
                    annual_return = portfolio_metrics['annual_return'] / 100
                    max_drawdown = portfolio_metrics['max_drawdown'] / 100
                    if max_drawdown != 0:
                        calmar_ratio = annual_return / abs(max_drawdown)
                        portfolio_metrics['calmar_ratio'] = calmar_ratio
                    else:
                        portfolio_metrics['calmar_ratio'] = 0.0
                    
                    # VaR 95% and CVaR 95%
                    var_95 = returns.quantile(0.05)
                    portfolio_metrics['var_95'] = var_95 * 100
                    
                    returns_below_var = returns[returns <= var_95]
                    if len(returns_below_var) > 0:
                        cvar_95 = returns_below_var.mean()
                        portfolio_metrics['cvar_95'] = cvar_95 * 100
                    else:
                        portfolio_metrics['cvar_95'] = var_95 * 100
                        
                else:
                    # No returns data available
                    portfolio_metrics = {
                        'annual_return': 0.0,
                        'volatility': 0.0,
                        'sharpe_ratio': 0.0,
                        'sortino_ratio': 0.0,
                        'max_drawdown': 0.0,
                        'calmar_ratio': 0.0,
                        'var_95': 0.0,
                        'cvar_95': 0.0
                    }
                    
            except Exception as e:
                self.logger.warning(f"Could not calculate portfolio metrics: {e}")
                portfolio_metrics = {
                    'annual_return': 0.0,
                    'volatility': 0.0,
                    'sharpe_ratio': 0.0,
                    'sortino_ratio': 0.0,
                    'max_drawdown': 0.0,
                    'calmar_ratio': 0.0,
                    'var_95': 0.0,
                    'cvar_95': 0.0
                }
            
            metrics_data['portfolio_metrics'] = portfolio_metrics
            
            # Individual asset metrics - use the same calculation logic as in _prepare_comprehensive_metrics
            for symbol in symbols:
                try:
                    asset = ok.Asset(symbol)
                    asset_metrics = {}
                    
                    # Get asset returns data
                    if hasattr(asset, 'returns'):
                        returns = asset.returns
                    elif hasattr(asset, 'get_returns'):
                        returns = asset.get_returns()
                    else:
                        # Fallback: calculate returns from price data
                        if hasattr(asset, 'prices'):
                            prices = asset.prices
                            returns = prices.pct_change().dropna()
                        else:
                            returns = None
                    
                    if returns is not None and len(returns) > 0:
                        # Annual Return (CAGR)
                        if hasattr(asset, 'mean_return_annual'):
                            try:
                                asset_metrics['annual_return'] = float(asset.mean_return_annual) * 100
                            except:
                                # Calculate manually with real asset period
                                total_return = (1 + returns).prod() - 1
                                years = self._calculate_asset_years(asset, returns)
                                cagr = (1 + total_return) ** (1 / years) - 1
                                asset_metrics['annual_return'] = cagr * 100
                        else:
                            # Calculate manually with real asset period
                            total_return = (1 + returns).prod() - 1
                            years = self._calculate_asset_years(asset, returns)
                            cagr = (1 + total_return) ** (1 / years) - 1
                            asset_metrics['annual_return'] = cagr * 100
                        
                        # Volatility
                        if hasattr(asset, 'volatility_annual'):
                            try:
                                asset_metrics['volatility'] = float(asset.volatility_annual) * 100
                            except:
                                # Calculate manually
                                volatility = returns.std() * (12 ** 0.5)  # Annualized for monthly data
                                asset_metrics['volatility'] = volatility * 100
                        else:
                            # Calculate manually
                            volatility = returns.std() * (12 ** 0.5)  # Annualized for monthly data
                            asset_metrics['volatility'] = volatility * 100
                        
                        # Sharpe Ratio
                        if hasattr(asset, 'sharpe_ratio'):
                            try:
                                asset_metrics['sharpe_ratio'] = float(asset.sharpe_ratio)
                            except:
                                # Calculate using unified function
                                annual_return = asset_metrics['annual_return'] / 100
                                volatility = asset_metrics['volatility'] / 100
                                sharpe_ratio = self.calculate_sharpe_ratio(annual_return, volatility, currency, asset_data=asset)
                                asset_metrics['sharpe_ratio'] = sharpe_ratio
                        else:
                            # Calculate using unified function
                            annual_return = asset_metrics['annual_return'] / 100
                            volatility = asset_metrics['volatility'] / 100
                            sharpe_ratio = self.calculate_sharpe_ratio(annual_return, volatility, currency, asset_data=asset)
                            asset_metrics['sharpe_ratio'] = sharpe_ratio
                        
                        # Sortino Ratio
                        if hasattr(asset, 'sortino_ratio'):
                            try:
                                asset_metrics['sortino_ratio'] = float(asset.sortino_ratio)
                            except:
                                # Calculate manually
                                annual_return = asset_metrics['annual_return'] / 100
                                negative_returns = returns[returns < 0]
                                if len(negative_returns) > 0:
                                    downside_deviation = negative_returns.std() * (12 ** 0.5)  # Annualized
                                    if downside_deviation > 0:
                                        sortino_ratio = (annual_return - 0.02) / downside_deviation
                                        asset_metrics['sortino_ratio'] = sortino_ratio
                                    else:
                                        asset_metrics['sortino_ratio'] = 0.0
                                else:
                                    # No negative returns, use volatility as fallback
                                    volatility = asset_metrics['volatility'] / 100
                                    if volatility > 0:
                                        sortino_ratio = (annual_return - 0.02) / volatility
                                        asset_metrics['sortino_ratio'] = sortino_ratio
                                    else:
                                        asset_metrics['sortino_ratio'] = 0.0
                        else:
                            # Calculate manually
                            annual_return = asset_metrics['annual_return'] / 100
                            negative_returns = returns[returns < 0]
                            if len(negative_returns) > 0:
                                downside_deviation = negative_returns.std() * (12 ** 0.5)  # Annualized
                                if downside_deviation > 0:
                                    sortino_ratio = (annual_return - 0.02) / downside_deviation
                                    asset_metrics['sortino_ratio'] = sortino_ratio
                                else:
                                    asset_metrics['sortino_ratio'] = 0.0
                            else:
                                # No negative returns, use volatility as fallback
                                volatility = asset_metrics['volatility'] / 100
                                if volatility > 0:
                                    sortino_ratio = (annual_return - 0.02) / volatility
                                    asset_metrics['sortino_ratio'] = sortino_ratio
                                else:
                                    asset_metrics['sortino_ratio'] = 0.0
                        
                        # Max Drawdown
                        if hasattr(asset, 'max_drawdown'):
                            try:
                                asset_metrics['max_drawdown'] = float(asset.max_drawdown) * 100
                            except:
                                # Calculate manually
                                cumulative = (1 + returns).cumprod()
                                running_max = cumulative.expanding().max()
                                drawdown = (cumulative - running_max) / running_max
                                max_drawdown = drawdown.min()
                                asset_metrics['max_drawdown'] = max_drawdown * 100
                        else:
                            # Calculate manually
                            cumulative = (1 + returns).cumprod()
                            running_max = cumulative.expanding().max()
                            drawdown = (cumulative - running_max) / running_max
                            max_drawdown = drawdown.min()
                            asset_metrics['max_drawdown'] = max_drawdown * 100
                        
                        # Calmar Ratio
                        annual_return = asset_metrics['annual_return'] / 100
                        max_drawdown = asset_metrics['max_drawdown'] / 100
                        if max_drawdown != 0:
                            calmar_ratio = annual_return / abs(max_drawdown)
                            asset_metrics['calmar_ratio'] = calmar_ratio
                        else:
                            asset_metrics['calmar_ratio'] = 0.0
                        
                        # VaR 95% and CVaR 95%
                        var_95 = returns.quantile(0.05)
                        asset_metrics['var_95'] = var_95 * 100
                        
                        returns_below_var = returns[returns <= var_95]
                        if len(returns_below_var) > 0:
                            cvar_95 = returns_below_var.mean()
                            asset_metrics['cvar_95'] = cvar_95 * 100
                        else:
                            asset_metrics['cvar_95'] = var_95 * 100
                            
                    else:
                        # No returns data available
                        asset_metrics = {
                            'annual_return': 0.0,
                            'volatility': 0.0,
                            'sharpe_ratio': 0.0,
                            'sortino_ratio': 0.0,
                            'max_drawdown': 0.0,
                            'calmar_ratio': 0.0,
                            'var_95': 0.0,
                            'cvar_95': 0.0
                        }
                    
                    metrics_data['detailed_metrics'][symbol] = asset_metrics
                    
                except Exception as e:
                    self.logger.warning(f"Could not get metrics for {symbol}: {e}")
                    metrics_data['detailed_metrics'][symbol] = {
                        'annual_return': 0.0,
                        'volatility': 0.0,
                        'sharpe_ratio': 0.0,
                        'sortino_ratio': 0.0,
                        'max_drawdown': 0.0,
                        'calmar_ratio': 0.0,
                        'var_95': 0.0,
                        'cvar_95': 0.0
                    }
            
            return metrics_data
            
        except Exception as e:
            self.logger.error(f"Error preparing portfolio metrics data: {e}")
            return None

    def _create_portfolio_metrics_excel(self, metrics_data: Dict[str, Any], symbols: list, currency: str) -> io.BytesIO:
        """Create Excel file with portfolio metrics"""
        try:
            buffer = io.BytesIO()
            
            if EXCEL_AVAILABLE:
                # Create Excel file with openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                wb = Workbook()
                
                # Remove default sheet
                wb.remove(wb.active)
                
                # Create Summary sheet
                ws_summary = wb.create_sheet("Summary", 0)
                
                # Summary data
                summary_data = [
                    ["Metric", "Value"],
                    ["Analysis Date", metrics_data['timestamp']],
                    ["Currency", currency],
                    ["Assets Count", len(symbols)],
                    ["Assets", ", ".join(symbols)],
                    ["Period", metrics_data['period']]
                ]
                
                for row in summary_data:
                    ws_summary.append(row)
                
                # Style summary sheet
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws_summary[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Create Portfolio Metrics sheet
                ws_portfolio = wb.create_sheet("Portfolio Metrics", 1)
                
                # Portfolio metrics data
                portfolio_metrics = metrics_data.get('portfolio_metrics', {})
                
                portfolio_data = [
                    ["Metric", "Value"],
                    ["Annual Return (%)", portfolio_metrics.get('annual_return', 0.0)],
                    ["Volatility (%)", portfolio_metrics.get('volatility', 0.0)],
                    ["Sharpe Ratio", portfolio_metrics.get('sharpe_ratio', 0.0)],
                    ["Sortino Ratio", portfolio_metrics.get('sortino_ratio', 0.0)],
                    ["Max Drawdown (%)", portfolio_metrics.get('max_drawdown', 0.0)],
                    ["Calmar Ratio", portfolio_metrics.get('calmar_ratio', 0.0)],
                    ["VaR 95% (%)", portfolio_metrics.get('var_95', 0.0)],
                    ["CVaR 95% (%)", portfolio_metrics.get('cvar_95', 0.0)]
                ]
                
                for row in portfolio_data:
                    ws_portfolio.append(row)
                
                # Style portfolio sheet
                for cell in ws_portfolio[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Create Detailed Metrics sheet
                ws_metrics = wb.create_sheet("Asset Metrics", 2)
                
                # Prepare detailed metrics data
                detailed_metrics = metrics_data.get('detailed_metrics', {})
                
                # Create headers
                headers = ["Metric"] + symbols
                ws_metrics.append(headers)
                
                # Define metrics to include
                metric_names = [
                    ("Annual Return (%)", "annual_return"),
                    ("Volatility (%)", "volatility"),
                    ("Sharpe Ratio", "sharpe_ratio"),
                    ("Sortino Ratio", "sortino_ratio"),
                    ("Max Drawdown (%)", "max_drawdown"),
                    ("Calmar Ratio", "calmar_ratio"),
                    ("VaR 95% (%)", "var_95"),
                    ("CVaR 95% (%)", "cvar_95")
                ]
                
                # Add data rows
                for metric_name, metric_key in metric_names:
                    row = [metric_name]
                    for symbol in symbols:
                        value = detailed_metrics.get(symbol, {}).get(metric_key, 0.0)
                        if isinstance(value, (int, float)):
                            row.append(round(value, 4))
                        else:
                            row.append(value)
                    ws_metrics.append(row)
                
                # Style metrics sheet
                for cell in ws_metrics[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths
                for ws in [ws_summary, ws_portfolio, ws_metrics]:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 20)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to buffer
                wb.save(buffer)
                buffer.seek(0)
                
            else:
                # Fallback to CSV format
                import csv
                import io as csv_io
                
                # Create CSV content
                csv_content = []
                
                # Summary
                csv_content.append(["SUMMARY"])
                csv_content.append(["Metric", "Value"])
                csv_content.append(["Analysis Date", metrics_data['timestamp']])
                csv_content.append(["Currency", currency])
                csv_content.append(["Assets Count", len(symbols)])
                csv_content.append(["Assets", ", ".join(symbols)])
                csv_content.append([])
                
                # Portfolio metrics
                csv_content.append(["PORTFOLIO METRICS"])
                portfolio_metrics = metrics_data.get('portfolio_metrics', {})
                csv_content.append(["Metric", "Value"])
                for metric_name, metric_key in [
                    ("Annual Return (%)", "annual_return"),
                    ("Volatility (%)", "volatility"),
                    ("Sharpe Ratio", "sharpe_ratio"),
                    ("Sortino Ratio", "sortino_ratio"),
                    ("Max Drawdown (%)", "max_drawdown"),
                    ("Calmar Ratio", "calmar_ratio"),
                    ("VaR 95% (%)", "var_95"),
                    ("CVaR 95% (%)", "cvar_95")
                ]:
                    value = portfolio_metrics.get(metric_key, 0.0)
                    csv_content.append([metric_name, value])
                csv_content.append([])
                
                # Detailed metrics
                csv_content.append(["ASSET METRICS"])
                detailed_metrics = metrics_data.get('detailed_metrics', {})
                
                headers = ["Metric"] + symbols
                csv_content.append(headers)
                
                metric_names = [
                    ("Annual Return (%)", "annual_return"),
                    ("Volatility (%)", "volatility"),
                    ("Sharpe Ratio", "sharpe_ratio"),
                    ("Sortino Ratio", "sortino_ratio"),
                    ("Max Drawdown (%)", "max_drawdown"),
                    ("Calmar Ratio", "calmar_ratio"),
                    ("VaR 95% (%)", "var_95"),
                    ("CVaR 95% (%)", "cvar_95")
                ]
                
                for metric_name, metric_key in metric_names:
                    row = [metric_name]
                    for symbol in symbols:
                        value = detailed_metrics.get(symbol, {}).get(metric_key, 0.0)
                        if isinstance(value, (int, float)):
                            row.append(round(value, 4))
                        else:
                            row.append(value)
                    csv_content.append(row)
                
                # Write CSV to buffer
                csv_buffer = csv_io.StringIO()
                writer = csv.writer(csv_buffer)
                for row in csv_content:
                    writer.writerow(row)
                
                buffer.write(csv_buffer.getvalue().encode('utf-8'))
                buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio metrics Excel: {e}")
            return None

    def _explain_risk_metrics(self, portfolio_description, portfolio) -> dict:
        """Explain each risk metric in detail"""
        explanations = {}
        
        try:
            # 1. Годовая волатильность
            if hasattr(portfolio, 'risk_annual'):
                risk_annual = portfolio.risk_annual
                if hasattr(risk_annual, 'tail'):
                    risk_value = risk_annual.tail(1).iloc[0] if not risk_annual.empty else None
                else:
                    risk_value = risk_annual.iloc[-1] if hasattr(risk_annual, 'iloc') else risk_annual
                
                if risk_value is not None:
                    risk_pct = float(risk_value) * 100
                    # Определяем уровень волатильности
                    if risk_pct <= 5:
                        volatility_level = "Очень низкая"
                        volatility_emoji = "🟢"
                        asset_type = "Типично для облигаций и депозитов"
                    elif risk_pct <= 10:
                        volatility_level = "Низкая"
                        volatility_emoji = "🟢"
                        asset_type = "Типично для консервативных портфелей"
                    elif risk_pct <= 15:
                        volatility_level = "Умеренная"
                        volatility_emoji = "🟡"
                        asset_type = "Типично для сбалансированных портфелей"
                    elif risk_pct <= 25:
                        volatility_level = "Высокая"
                        volatility_emoji = "🟠"
                        asset_type = "Типично для акций"
                    else:
                        volatility_level = "Очень высокая"
                        volatility_emoji = "🔴"
                        asset_type = "Типично для спекулятивных активов"
                    
                    explanations["1. Годовая волатильность"] = (
                        f"{volatility_emoji} **{volatility_level}** ({risk_pct:.1f}% годовых)\n\n"
                        f"💡 **Что это значит:**\n"
                        f"• Портфель может колебаться в пределах ±{risk_pct:.1f}% в год\n"
                        f"• {asset_type}\n\n"
                        f"💡 **Как использовать:**\n"
                        f"• Сравнивайте с другими портфелями\n"
                        f"• При одинаковой доходности выбирайте с меньшей волатильностью\n"
                        f"• Учитывайте свой риск-профиль"
                    )
            
            # 2. Полуотклонение (риск только вниз)
            if hasattr(portfolio, 'semideviation_annual'):
                semidev = portfolio.semideviation_annual
                if hasattr(semidev, 'iloc'):
                    semidev_value = semidev.iloc[-1] if not semidev.empty else None
                else:
                    semidev_value = semidev
                
                if semidev_value is not None:
                    semidev_pct = float(semidev_value) * 100
                    # Сравниваем с общей волатильностью
                    if hasattr(portfolio, 'risk_annual'):
                        risk_annual = portfolio.risk_annual
                        if hasattr(risk_annual, 'tail'):
                            risk_value = risk_annual.tail(1).iloc[0] if not risk_annual.empty else None
                        else:
                            risk_value = risk_annual.iloc[-1] if hasattr(risk_annual, 'iloc') else risk_annual
                        
                        if risk_value is not None:
                            risk_pct = float(risk_value) * 100
                            ratio = semidev_pct / risk_pct if risk_pct > 0 else 1
                            
                            if ratio < 0.7:
                                downside_characteristic = "Портфель 'гладкий' вниз, но может резко расти"
                            elif ratio > 0.9:
                                downside_characteristic = "Портфель одинаково колеблется вверх и вниз"
                            else:
                                downside_characteristic = "Сбалансированное поведение портфеля"
                        else:
                            downside_characteristic = "Стандартное поведение"
                    else:
                        downside_characteristic = "Стандартное поведение"
                    
                    explanations["2. Риск просадок (полуотклонение)"] = (
                        f"📉 **{semidev_pct:.1f}%** годовых\n\n"
                        f"💡 **Что это значит:**\n"
                        f"• Учитывает только отрицательные колебания\n"
                        f"• Более точно отражает страх инвестора\n\n"
                        f"💡 **Характеристика:**\n"
                        f"• {downside_characteristic}\n"
                        f"• Помогает понять, насколько 'комфортно' падать портфелю"
                    )
            
            # 3. VaR (Value at Risk)
            try:
                var_1 = portfolio.get_var_historic(level=1)
                var_5 = portfolio.get_var_historic(level=5)
                
                if var_1 is not None and var_5 is not None:
                    var_1_pct = float(var_1) * 100
                    var_5_pct = float(var_5) * 100
                    
                    # Оценка риска по VaR
                    if abs(var_5_pct) <= 3:
                        var_assessment = "Очень низкий риск потерь"
                        var_emoji = "🟢"
                    elif abs(var_5_pct) <= 8:
                        var_assessment = "Низкий риск потерь"
                        var_emoji = "🟢"
                    elif abs(var_5_pct) <= 15:
                        var_assessment = "Умеренный риск потерь"
                        var_emoji = "🟡"
                    elif abs(var_5_pct) <= 25:
                        var_assessment = "Высокий риск потерь"
                        var_emoji = "🟠"
                    else:
                        var_assessment = "Очень высокий риск потерь"
                        var_emoji = "🔴"
                    
                    explanations["3. VaR (риск потерь)"] = (
                        f"{var_emoji} **{var_assessment}**\n\n"
                        f"📊 **Значения:**\n"
                        f"• 1% VaR: {var_1_pct:.1f}% (99% вероятность, что потери не превысят)\n"
                        f"• 5% VaR: {var_5_pct:.1f}% (95% вероятность, что потери не превысят)\n\n"
                        f"💡 **Как использовать:**\n"
                        f"• Оцените, готовы ли потерять {abs(var_5_pct):.1f}% в месяц\n"
                        f"• Если нет — портфель слишком агрессивный для вас"
                    )
            except Exception as e:
                self.logger.warning(f"Could not get VaR: {e}")
            
            # 4. CVaR (Conditional Value at Risk)
            try:
                cvar_5 = portfolio.get_cvar_historic(level=5)
                
                if cvar_5 is not None:
                    cvar_5_pct = float(cvar_5) * 100
                    
                    # Оценка по CVaR
                    if abs(cvar_5_pct) <= 5:
                        cvar_assessment = "Очень низкие ожидаемые потери в кризис"
                        cvar_emoji = "🟢"
                    elif abs(cvar_5_pct) <= 12:
                        cvar_assessment = "Низкие ожидаемые потери в кризис"
                        cvar_emoji = "🟢"
                    elif abs(cvar_5_pct) <= 20:
                        cvar_assessment = "Умеренные ожидаемые потери в кризис"
                        cvar_emoji = "🟡"
                    elif abs(cvar_5_pct) <= 30:
                        cvar_assessment = "Высокие ожидаемые потери в кризис"
                        cvar_emoji = "🟠"
                    else:
                        cvar_assessment = "Очень высокие ожидаемые потери в кризис"
                        cvar_emoji = "🔴"
                    
                    explanations["4. CVaR (ожидаемые потери в кризис)"] = (
                        f"{cvar_emoji} **{cvar_assessment}**\n\n"
                        f"📊 **Значение:** {cvar_5_pct:.1f}%\n\n"
                        f"💡 **Что это значит:**\n"
                        f"• В худших 5% месяцев ожидайте потери около {abs(cvar_5_pct):.1f}%\n"
                        f"• Более консервативный показатель, чем VaR\n"
                        f"• Помогает понять глубину возможного 'провала'"
                    )
            except Exception as e:
                self.logger.warning(f"Could not get CVaR: {e}")
            
            # 5. Максимальная просадка
            if hasattr(portfolio, 'drawdowns'):
                drawdowns = portfolio.drawdowns
                if hasattr(drawdowns, 'min'):
                    max_dd = drawdowns.min()
                    if max_dd is not None:
                        max_dd_pct = float(max_dd) * 100
                        
                        # Оценка просадки
                        if max_dd_pct <= 5:
                            dd_assessment = "Очень низкая просадка"
                            dd_emoji = "🟢"
                            dd_advice = "Портфель очень стабильный, подходит для консервативных инвесторов"
                        elif max_dd_pct <= 15:
                            dd_assessment = "Низкая просадка"
                            dd_emoji = "🟢"
                            dd_advice = "Портфель стабильный, подходит для большинства инвесторов"
                        elif max_dd_pct <= 30:
                            dd_assessment = "Умеренная просадка"
                            dd_emoji = "🟡"
                            dd_advice = "Требует психологической устойчивости"
                        elif max_dd_pct <= 50:
                            dd_assessment = "Высокая просадка"
                            dd_emoji = "🟠"
                            dd_advice = "Убедитесь, что готовы к таким потерям"
                        else:
                            dd_assessment = "Очень высокая просадка"
                            dd_emoji = "🔴"
                            dd_advice = "Портфель очень агрессивный"
                        
                        explanations["5. Максимальная просадка"] = (
                            f"{dd_emoji} **{dd_assessment}** ({max_dd_pct:.1f}%)\n\n"
                            f"💡 **Что это значит:**\n"
                            f"• Самая большая потеря от пика до дна\n"
                            f"• Критично для психологической устойчивости\n\n"
                            f"💡 **Рекомендация:**\n"
                            f"• {dd_advice}\n"
                            f"• Если не выдержите просадку в {max_dd_pct:.1f}%, пересмотрите состав"
                        )
            
            # 6. Период восстановления
            if hasattr(portfolio, 'recovery_period'):
                recovery = portfolio.recovery_period
                if hasattr(recovery, 'max'):
                    max_recovery = recovery.max()
                    if max_recovery is not None:
                        recovery_years = float(max_recovery) / 12  # в годах
                        
                        # Оценка периода восстановления
                        if recovery_years <= 0.5:
                            recovery_assessment = "Очень быстрое восстановление"
                            recovery_emoji = "🟢"
                        elif recovery_years <= 1:
                            recovery_assessment = "Быстрое восстановление"
                            recovery_emoji = "🟢"
                        elif recovery_years <= 2:
                            recovery_assessment = "Умеренное восстановление"
                            recovery_emoji = "🟡"
                        elif recovery_years <= 4:
                            recovery_assessment = "Медленное восстановление"
                            recovery_emoji = "🟠"
                        else:
                            recovery_assessment = "Очень медленное восстановление"
                            recovery_emoji = "🔴"
                        
                        explanations["6. Период восстановления"] = (
                            f"{recovery_emoji} **{recovery_assessment}** ({recovery_years:.1f} года)\n\n"
                            f"💡 **Что это значит:**\n"
                            f"• Время возврата к предыдущему максимуму\n"
                            f"• Критично при планировании снятия денег\n\n"
                            f"💡 **Как использовать:**\n"
                            f"• Если планируете снимать деньги, убедитесь, что период восстановления приемлем\n"
                            f"• Иначе есть риск 'обнулиться' в неподходящий момент"
                        )
            
        except Exception as e:
            self.logger.error(f"Error explaining risk metrics: {e}")
            explanations["Ошибка анализа"] = f"Не удалось проанализировать метрики риска: {str(e)}"
        
        return explanations

    def _assess_portfolio_risk(self, portfolio_description, portfolio) -> str:
        """Assess overall portfolio risk level with improved logic for conservative assets"""
        try:
            risk_level = "Средний"
            risk_color = "🟡"
            risk_score = 0
            recommendations = []
            
            # Check volatility (weight: 40%)
            if hasattr(portfolio, 'risk_annual'):
                risk_annual = portfolio.risk_annual
                if hasattr(risk_annual, 'tail'):
                    risk_value = risk_annual.tail(1).iloc[0] if not risk_annual.empty else None
                else:
                    risk_value = risk_annual.iloc[-1] if hasattr(risk_annual, 'iloc') else risk_annual
                
                if risk_value is not None:
                    risk_pct = float(risk_value) * 100
                    
                    if risk_pct <= 5:
                        risk_score += 0  # Очень низкий риск
                        volatility_assessment = "Очень низкая"
                        volatility_emoji = "🟢"
                    elif risk_pct <= 10:
                        risk_score += 1  # Низкий риск
                        volatility_assessment = "Низкая"
                        volatility_emoji = "🟢"
                    elif risk_pct <= 15:
                        risk_score += 2  # Умеренный риск
                        volatility_assessment = "Умеренная"
                        volatility_emoji = "🟡"
                    elif risk_pct <= 25:
                        risk_score += 3  # Высокий риск
                        volatility_assessment = "Высокая"
                        volatility_emoji = "🟠"
                    else:
                        risk_score += 4  # Очень высокий риск
                        volatility_assessment = "Очень высокая"
                        volatility_emoji = "🔴"
                    
                    # Добавляем рекомендации по волатильности
                    if risk_pct <= 5:
                        recommendations.append("• Портфель очень консервативный, идеален для сохранения капитала")
                        recommendations.append("• Подходит для инвесторов с низкой толерантностью к риску")
                    elif risk_pct <= 10:
                        recommendations.append("• Портфель консервативный, подходит для большинства инвесторов")
                        recommendations.append("• Рассмотрите добавление акций для роста доходности")
                    elif risk_pct <= 15:
                        recommendations.append("• Портфель сбалансированный, подходит для долгосрочных целей")
                    elif risk_pct <= 25:
                        recommendations.append("• Портфель агрессивный, требует психологической устойчивости")
                        recommendations.append("• Рассмотрите добавление облигаций для снижения волатильности")
                    else:
                        recommendations.append("• Портфель очень агрессивный, подходит только для опытных инвесторов")
                        recommendations.append("• Увеличьте долю защитных активов (облигации, золото)")
            
            # Check max drawdown (weight: 30%)
            if hasattr(portfolio, 'drawdowns'):
                drawdowns = portfolio.drawdowns
                if hasattr(drawdowns, 'min'):
                    max_dd = drawdowns.min()
                    if max_dd is not None:
                        max_dd_pct = abs(float(max_dd) * 100)
                        
                        if max_dd_pct <= 5:
                            risk_score += 0  # Очень низкий риск
                            dd_assessment = "Очень низкая"
                        elif max_dd_pct <= 15:
                            risk_score += 1  # Низкий риск
                            dd_assessment = "Низкая"
                        elif max_dd_pct <= 30:
                            risk_score += 2  # Умеренный риск
                            dd_assessment = "Умеренная"
                        elif max_dd_pct <= 50:
                            risk_score += 3  # Высокий риск
                            dd_assessment = "Высокая"
                        else:
                            risk_score += 4  # Очень высокий риск
                            dd_assessment = "Очень высокая"
                        
                        # Добавляем рекомендации по просадке
                        if max_dd_pct <= 5:
                            recommendations.append("• Просадка очень низкая, портфель очень стабильный")
                        elif max_dd_pct <= 15:
                            recommendations.append("• Просадка низкая, подходит для большинства инвесторов")
                        elif max_dd_pct <= 30:
                            recommendations.append("• Просадка умеренная, требует психологической устойчивости")
                        else:
                            recommendations.append("• Просадка высокая, убедитесь, что готовы к таким потерям")
            
            # Check VaR (weight: 20%)
            try:
                var_5 = portfolio.get_var_historic(level=5)
                if var_5 is not None:
                    var_5_pct = abs(float(var_5) * 100)
                    
                    if var_5_pct <= 3:
                        risk_score += 0  # Очень низкий риск
                    elif var_5_pct <= 8:
                        risk_score += 1  # Низкий риск
                    elif var_5_pct <= 15:
                        risk_score += 2  # Умеренный риск
                    elif var_5_pct <= 25:
                        risk_score += 3  # Высокий риск
                    else:
                        risk_score += 4  # Очень высокий риск
            except Exception:
                pass
            
            # Check CVaR (weight: 10%)
            try:
                cvar_5 = portfolio.get_cvar_historic(level=5)
                if cvar_5 is not None:
                    cvar_5_pct = abs(float(cvar_5) * 100)
                    
                    if cvar_5_pct <= 5:
                        risk_score += 0  # Очень низкий риск
                    elif cvar_5_pct <= 12:
                        risk_score += 1  # Низкий риск
                    elif cvar_5_pct <= 20:
                        risk_score += 2  # Умеренный риск
                    elif cvar_5_pct <= 30:
                        risk_score += 3  # Высокий риск
                    else:
                        risk_score += 4  # Очень высокий риск
            except Exception:
                pass
            
            # Определяем общий уровень риска на основе score
            max_possible_score = 4  # Максимальный score для одного показателя
            normalized_score = risk_score / max_possible_score
            
            if normalized_score <= 0.25:
                risk_level = "Очень низкий"
                risk_color = "🟢"
            elif normalized_score <= 0.5:
                risk_level = "Низкий"
                risk_color = "🟢"
            elif normalized_score <= 0.75:
                risk_level = "Умеренный"
                risk_color = "🟡"
            elif normalized_score <= 1.0:
                risk_level = "Высокий"
                risk_color = "🟠"
            else:
                risk_level = "Очень высокий"
                risk_color = "🔴"
            
            # Формируем итоговую оценку
            assessment = f"{risk_color} **Уровень риска: {risk_level}**\n\n"
            
            if hasattr(portfolio, 'risk_annual'):
                risk_annual = portfolio.risk_annual
                if hasattr(risk_annual, 'tail'):
                    risk_value = risk_annual.tail(1).iloc[0] if not risk_annual.empty else None
                else:
                    risk_value = risk_annual.iloc[-1] if hasattr(risk_annual, 'iloc') else risk_annual
                
                if risk_value is not None:
                    risk_pct = float(risk_value) * 100
                    assessment += f"📊 **Ключевые показатели:**\n"
                    assessment += f"• Волатильность: {volatility_emoji} {volatility_assessment} ({risk_pct:.1f}%)\n"
                    
                    if hasattr(portfolio, 'drawdowns'):
                        drawdowns = portfolio.drawdowns
                        if hasattr(drawdowns, 'min'):
                            max_dd = drawdowns.min()
                            if max_dd is not None:
                                max_dd_pct = abs(float(max_dd) * 100)
                                assessment += f"• Макс. просадка: {dd_assessment} ({max_dd_pct:.1f}%)\n"
            
            assessment += f"\n📋 **Рекомендации:**\n"
            if recommendations:
                for rec in recommendations:
                    assessment += f"{rec}\n"
            else:
                assessment += "• Портфель сбалансированный, специальных рекомендаций не требуется."
            
            return assessment
            
        except Exception as e:
            self.logger.error(f"Error assessing portfolio risk: {e}")
            return "Не удалось оценить общий уровень риска портфеля."

    async def _create_monte_carlo_forecast(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str):
        """Create and send Monte Carlo forecast chart for portfolio"""
        try:
            self.logger.info(f"Creating Monte Carlo forecast chart for portfolio: {symbols}")
            
            # Generate Monte Carlo forecast (okama creates the figure)
            forecast_data = portfolio.plot_forecast_monte_carlo(distr="norm", years=10, n=20)

            # Get the current figure from matplotlib (created by okama)
            current_fig = plt.gcf()

            # Apply chart styles to the current figure using chart_styles
            if current_fig.axes:
                ax = current_fig.axes[0]
                
                # Get portfolio weights
                weights = portfolio.weights if hasattr(portfolio, 'weights') else None
                
                # Apply Monte Carlo chart styling using chart_styles
                chart_styles.create_monte_carlo_chart(
                    current_fig, ax, symbols, currency, weights, forecast_data=forecast_data
                )
            
            # Save the figure using chart_styles
            img_buffer = io.BytesIO()
            chart_styles.save_figure(current_fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(current_fig)
            
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(
                    f"💡 Возможные траектории роста портфеля на основе исторической волатильности и доходности."
                ),
            )
            
        except Exception as e:
            self.logger.error(f"Error creating Monte Carlo forecast chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика Monte Carlo: {str(e)}")

    async def _create_forecast_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str):
        """Create and send forecast chart with percentiles for portfolio"""
        try:
            self.logger.info(f"Creating forecast chart with percentiles for portfolio: {symbols}")
            
            # Generate forecast chart using okama
            # y.plot_forecast(years=10, today_value=1000, percentiles=[10, 50, 90])
            forecast_data = portfolio.plot_forecast(
                years=10, 
                today_value=1000, 
                percentiles=[10, 50, 90]
            )
            
            # Get the current figure from matplotlib (created by okama)
            current_fig = plt.gcf()
            
            # Apply chart styles to the current figure using the new unified method
            if current_fig.axes:
                ax = current_fig.axes[0]  # Get the first (and usually only) axes
                
                # Apply unified percentile forecast chart styling
                chart_styles.create_percentile_forecast_chart(
                    current_fig,
                    ax,
                    symbols=symbols,
                    currency=currency,
                    data_source='okama'
                )
            
            # Save the figure using chart_styles
            img_buffer = io.BytesIO()
            chart_styles.save_figure(current_fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(current_fig)
            
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(
                    f"📈 Прогноз с процентилями для портфеля: {', '.join(symbols)}\n\n"
                    f"• Период: 10 лет\n"
                    f"• Начальная стоимость: 1000 {currency}\n"
                    f"• 10% процентиль: пессимистичный сценарий\n"
                    f"• 50% процентиль: средний сценарий\n"
                    f"• 90% процентиль: оптимистичный сценарий"
                ),
            )
            
        except Exception as e:
            self.logger.error(f"Error creating forecast chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика прогноза: {str(e)}")

    async def _handle_portfolio_drawdowns_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio drawdowns button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio drawdowns by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            period = portfolio_info.get('period')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}, period={period}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            await self._send_ephemeral_message(update, context, "📉 Создаю график просадок...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    if i < len(weights):
                        valid_weights.append(weights[i])
                    else:
                        valid_weights.append(1.0 / len(final_symbols))
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            if period:
                years = int(period[:-1])  # Extract number from '5Y'
                from datetime import timedelta
                end_date = datetime.now()
                start_date = end_date - timedelta(days=years * 365)
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency,
                                       first_date=start_date.strftime('%Y-%m-%d'), 
                                       last_date=end_date.strftime('%Y-%m-%d'))
                self.logger.info(f"Created portfolio with period {period}")
            else:
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency)
                self.logger.info(f"Created portfolio with maximum available period")
            await self._create_portfolio_drawdowns_chart(update, context, portfolio, final_symbols, currency, weights, portfolio_symbol)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio drawdowns by symbol: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика просадок: {str(e)}")

    async def _handle_portfolio_drawdowns_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle portfolio drawdowns button click"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio drawdowns button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            self.logger.info(f"Button symbols: {button_symbols}")
            self.logger.info(f"Final symbols: {final_symbols}")
            self.logger.info(f"Current symbols from context: {user_context.get('current_symbols')}")
            self.logger.info(f"Last assets from context: {user_context.get('last_assets')}")
            
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            self.logger.info(f"Creating drawdowns chart for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "📉 Создаю график просадок...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_portfolio_drawdowns_chart(update, context, portfolio, final_symbols, currency, weights, "Портфель")
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio drawdowns button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика просадок: {str(e)}")

    async def _create_portfolio_drawdowns_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str, weights: list, portfolio_name: str = None):
        """Create and send portfolio drawdowns chart"""
        try:
            self.logger.info(f"Creating portfolio drawdowns chart for portfolio: {symbols}")
            
            # Get drawdowns data from portfolio
            drawdowns_data = portfolio.drawdowns
            
            # Create drawdowns chart using chart_styles
            fig, ax = chart_styles.create_portfolio_drawdowns_chart(
                data=drawdowns_data, symbols=symbols, currency=currency, weights=weights, portfolio_name=portfolio_name
            )
            
            # Save the figure using chart_styles
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            chart_styles.cleanup_figure(fig)
            img_buffer.seek(0)
            
            # Get drawdowns statistics
            try:
                # Get 5 largest drawdowns
                largest_drawdowns = portfolio.drawdowns.nsmallest(5)
                
                # Get longest recovery periods (convert to years)
                longest_recoveries = portfolio.recovery_period.nlargest(5) / 12
                
                # Build enhanced caption with weights in title
                symbols_with_weights = []
                for i, symbol in enumerate(symbols):
                    symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                    weight = weights[i] if i < len(weights) else 0.0
                    symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
                
                caption = f"📉 Просадки портфеля: {', '.join(symbols_with_weights)}\n\n"
                caption += f"📊 Параметры:\n"
                caption += f"• Валюта: {currency}\n\n"
                
                # Add largest drawdowns
                caption += f"📉 5 самых больших просадок:\n"
                for i, (date, drawdown) in enumerate(largest_drawdowns.items(), 1):
                    date_str = date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date)
                    drawdown_pct = drawdown * 100
                    caption += f"{i}. {date_str}: {drawdown_pct:.2f}%\n"
                
                caption += f"\n⏱️ Самые долгие периоды восстановления:\n"
                for i, (date, recovery_years) in enumerate(longest_recoveries.items(), 1):
                    date_str = date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date)
                    caption += f"{i}. {date_str}: {recovery_years:.1f} лет\n"

                
            except Exception as e:
                self.logger.warning(f"Could not get drawdowns statistics: {e}")
                # Fallback to basic caption
                caption = f"📉 Просадки портфеля: {', '.join(symbols)}\n\n"
                caption += f"📊 Параметры:\n"
                caption += f"• Валюта: {currency}\n"
                caption += f"• Веса: {', '.join([f'{w:.1%}' for w in weights])}\n\n"
            
            # Send the chart
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(caption)
            )
            
            # Show Reply Keyboard for portfolio management
            await self._manage_reply_keyboard(update, context, "portfolio")
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio drawdowns chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика просадок: {str(e)}")

    async def _create_portfolio_dividends_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str, weights: list, portfolio_name: str = None):
        """Create and send portfolio dividends chart"""
        try:
            self.logger.info(f"Creating portfolio dividends chart for portfolio: {symbols}")
            
            # Try to get dividend yield data from portfolio first (respects period)
            try:
                # First try portfolio dividend yield with assets (shows individual assets)
                if hasattr(portfolio, 'dividend_yield_with_assets'):
                    dividend_yield_data = portfolio.dividend_yield_with_assets
                    self.logger.info("Using portfolio.dividend_yield_with_assets (respects period)")
                else:
                    # Fallback to portfolio dividend yield (aggregated)
                    dividend_yield_data = portfolio.dividend_yield
                    self.logger.info("Using portfolio.dividend_yield (respects period)")
                
                if dividend_yield_data is None or dividend_yield_data.empty:
                    await self._send_callback_message(update, context, "❌ Данные о дивидендах не содержат информацию для отображения.")
                    return
                    
            except Exception as portfolio_error:
                self.logger.warning(f"Could not get dividend yield data from portfolio: {portfolio_error}")
                # Fallback to creating AssetList with period from user context
                try:
                    user_id = update.effective_user.id
                    user_context = self._get_user_context(user_id)
                    current_period = user_context.get('current_period')
                    
                    import okama as ok
                    if current_period:
                        years = int(current_period[:-1])
                        from datetime import datetime, timedelta
                        end_date = datetime.now()
                        start_date = end_date - timedelta(days=years * 365)
                        asset_list = ok.AssetList(symbols, ccy=currency,
                                                first_date=start_date.strftime('%Y-%m-%d'), 
                                                last_date=end_date.strftime('%Y-%m-%d'))
                        self.logger.info(f"Created AssetList with period {current_period}")
                    else:
                        asset_list = ok.AssetList(symbols, ccy=currency)
                        self.logger.info("Created AssetList without period (MAX)")
                    
                    if hasattr(asset_list, 'dividend_yields') and not asset_list.dividend_yields.empty:
                        dividend_yield_data = asset_list.dividend_yields
                        self.logger.info("Using AssetList dividend_yields as fallback")
                    else:
                        raise ValueError("No dividend yield data available")
                        
                except Exception as e:
                    self.logger.error(f"Could not get dividend yield data: {e}")
                    await self._send_callback_message(update, context, "❌ Данные о дивидендах не содержат информацию для отображения.")
                    return
            
            # Create dividends chart using chart_styles
            fig, ax = chart_styles.create_dividend_yield_chart(
                data=dividend_yield_data, symbols=symbols, weights=weights, portfolio_name=portfolio_name
            )
            
            # Save the figure using chart_styles
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            chart_styles.cleanup_figure(fig)
            img_buffer.seek(0)
            
            # Build caption with weights in title
            symbols_with_weights = []
            for i, symbol in enumerate(symbols):
                symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                weight = weights[i] if i < len(weights) else 0.0
                symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
            
            caption = f"Дивидендная доходность портфеля: {', '.join(symbols_with_weights)}\n\n"
            
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(caption),
            )
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio dividends chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика дивидендов: {str(e)}")

    async def _handle_portfolio_returns_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle portfolio returns button click"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio returns button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            self.logger.info(f"Creating returns chart for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "💰 Создаю график доходности...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_portfolio_returns_chart(update, context, portfolio, final_symbols, currency, weights)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio returns button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика доходности: {str(e)}")

    async def _handle_portfolio_dividends_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio dividends button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio dividends by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            period = portfolio_info.get('period')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}, period={period}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            await self._send_ephemeral_message(update, context, "Создаю график дивидендной доходности...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    if i < len(weights):
                        valid_weights.append(weights[i])
                    else:
                        valid_weights.append(1.0 / len(final_symbols))
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            if period:
                years = int(period[:-1])  # Extract number from '5Y'
                from datetime import timedelta
                end_date = datetime.now()
                start_date = end_date - timedelta(days=years * 365)
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency,
                                       first_date=start_date.strftime('%Y-%m-%d'), 
                                       last_date=end_date.strftime('%Y-%m-%d'))
                self.logger.info(f"Created portfolio with period {period}")
            else:
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency)
                self.logger.info(f"Created portfolio with maximum available period")
            await self._create_portfolio_dividends_chart(update, context, portfolio, final_symbols, currency, weights, portfolio_symbol)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio dividends by symbol: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика дивидендов: {str(e)}")

    async def _handle_portfolio_returns_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio returns button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio returns by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            period = portfolio_info.get('period')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}, period={period}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            await self._send_ephemeral_message(update, context, "💰 Создаю график доходности...", delete_after=3)
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            if period:
                years = int(period[:-1])  # Extract number from '5Y'
                from datetime import timedelta
                end_date = datetime.now()
                start_date = end_date - timedelta(days=years * 365)
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency,
                                       first_date=start_date.strftime('%Y-%m-%d'), 
                                       last_date=end_date.strftime('%Y-%m-%d'))
                self.logger.info(f"Created portfolio with period {period}")
            else:
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency)
                self.logger.info(f"Created portfolio with maximum available period")
            
            await self._create_portfolio_returns_chart(update, context, portfolio, final_symbols, currency, weights)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio returns by symbol: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика доходности: {str(e)}")

    async def _create_portfolio_returns_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str, weights: list):
        """Create and send portfolio returns chart"""
        try:
            self.logger.info(f"Creating portfolio returns chart for portfolio: {symbols}")
            
            # Generate annual returns data
            returns_data = portfolio.annual_return_ts
            
            # Create portfolio returns chart with chart_styles
            fig, ax = chart_styles.create_portfolio_returns_chart(
                data=returns_data, symbols=symbols, currency=currency, weights=weights
            )
            
            # Save chart to bytes with memory optimization
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(fig)
            
            # Get returns statistics
            try:
                # Get returns statistics
                mean_return_monthly = portfolio.mean_return_monthly
                mean_return_annual = portfolio.mean_return_annual
                cagr = portfolio.get_cagr()
                
                # Handle CAGR which might be a Series
                if hasattr(cagr, '__iter__') and not isinstance(cagr, str):
                    # If it's a Series or array-like, get the first value
                    if hasattr(cagr, 'iloc'):
                        cagr_value = cagr.iloc[0]
                    elif hasattr(cagr, '__getitem__'):
                        cagr_value = cagr[0]
                    else:
                        cagr_value = list(cagr)[0]
                else:
                    cagr_value = cagr
                
                # Build enhanced caption with weights in title
                symbols_with_weights = []
                for i, symbol in enumerate(symbols):
                    symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                    weight = weights[i] if i < len(weights) else 0.0
                    symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
                
                caption += f"• Средняя годовая доходность: {mean_return_annual:.2%}\n"
                caption += f"• CAGR (Compound Annual Growth Rate): {cagr_value:.2%}\n\n"
                
            except Exception as e:
                self.logger.warning(f"Could not get returns statistics: {e}")
                # Fallback to basic caption with weights in title
                symbols_with_weights = []
                for i, symbol in enumerate(symbols):
                    symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                    weight = weights[i] if i < len(weights) else 0.0
                    symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
                
                caption = f"Динамика доходности портфеля\n\n"
            
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(caption),
            )
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio returns chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика доходности: {str(e)}")

    async def _handle_portfolio_wealth_chart_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle portfolio wealth chart button click"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio wealth chart button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            self.logger.info(f"Creating wealth chart for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "📈 Создаю график накопленной доходности...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database - be more lenient
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    # Don't check price data length as it might be empty but symbol still valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Get period from user context
            current_period = user_context.get('current_period')
            
            # Create Portfolio with validated symbols and period
            if current_period:
                years = int(current_period[:-1])  # Extract number from '5Y'
                from datetime import timedelta
                end_date = datetime.now()
                start_date = end_date - timedelta(days=years * 365)
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency,
                                       first_date=start_date.strftime('%Y-%m-%d'), 
                                       last_date=end_date.strftime('%Y-%m-%d'))
                self.logger.info(f"Created portfolio with period {current_period}")
            else:
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency)
                self.logger.info(f"Created portfolio with maximum available period")
            
            await self._create_portfolio_wealth_chart(update, context, portfolio, final_symbols, currency, weights, "Портфель")
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio wealth chart button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика накопленной доходности: {str(e)}")

    async def _create_portfolio_wealth_chart_with_info(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str, weights: list, portfolio_symbol: str, portfolio_text: str, reply_markup):
        """Create and send portfolio wealth chart with portfolio info in caption and buttons"""
        try:
            self.logger.info(f"Creating portfolio wealth chart with info for portfolio: {symbols}")
            
            # Generate wealth chart using chart_styles
            wealth_index = portfolio.wealth_index
            
            # Create portfolio chart with chart_styles using optimized method
            fig, ax = chart_styles.create_portfolio_wealth_chart(
                data=wealth_index, symbols=symbols, currency=currency, weights=weights, portfolio_name=portfolio_symbol
            )
            
            # Save chart to bytes with memory optimization
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(fig)
            
            # Get final portfolio value safely
            try:
                final_value = portfolio.wealth_index.iloc[-1]
                
                # Handle different types of final_value
                if hasattr(final_value, '__iter__') and not isinstance(final_value, str):
                    if hasattr(final_value, 'iloc'):
                        final_value = final_value.iloc[0]
                    elif hasattr(final_value, '__getitem__'):
                        final_value = final_value[0]
                    else:
                        final_value = list(final_value)[0]
                
                # Convert to float safely
                if isinstance(final_value, (int, float)):
                    final_value = float(final_value)
                else:
                    final_value_str = str(final_value)
                    try:
                        final_value = float(final_value_str)
                    except (ValueError, TypeError):
                        import re
                        numeric_match = re.search(r'[\d.]+', final_value_str)
                        if numeric_match:
                            final_value = float(numeric_match.group())
                        else:
                            raise ValueError(f"Cannot convert {final_value} to float")
                
                # Add period information
                try:
                    period_length = portfolio.period_length
                except Exception as e:
                    self.logger.warning(f"Could not get period length: {e}")
                    period_length = "неизвестный период"
            except Exception as e:
                self.logger.warning(f"Could not get final portfolio value: {e}")
                final_value = 0.0
                period_length = "неизвестный период"
            
            # Create comprehensive caption with portfolio info in new format
            # Get portfolio metrics for caption
            try:
                # Get CAGR
                cagr_value = None
                if hasattr(portfolio, 'get_cagr'):
                    try:
                        cagr = portfolio.get_cagr()
                        if hasattr(cagr, 'iloc'):
                            cagr_value = cagr.iloc[0]
                        elif hasattr(cagr, '__getitem__'):
                            cagr_value = cagr[0]
                        else:
                            cagr_value = cagr
                    except Exception as e:
                        self.logger.warning(f"Could not get CAGR: {e}")
                
                # Get volatility (risk_annual)
                volatility_value = None
                if hasattr(portfolio, 'risk_annual'):
                    try:
                        risk_annual = portfolio.risk_annual
                        if hasattr(risk_annual, 'iloc'):
                            volatility_value = risk_annual.iloc[-1]
                        elif hasattr(risk_annual, '__getitem__'):
                            volatility_value = risk_annual[-1]
                        else:
                            volatility_value = risk_annual
                    except Exception as e:
                        self.logger.warning(f"Could not get volatility: {e}")
                
                # Get Sharpe ratio
                sharpe_value = None
                if hasattr(portfolio, 'get_sharpe_ratio'):
                    try:
                        sharpe = portfolio.get_sharpe_ratio()
                        if hasattr(sharpe, 'iloc'):
                            sharpe_value = sharpe.iloc[0]
                        elif hasattr(sharpe, '__getitem__'):
                            sharpe_value = sharpe[0]
                        else:
                            sharpe_value = sharpe
                    except Exception as e:
                        self.logger.warning(f"Could not get Sharpe ratio: {e}")
                
                # Get max drawdown from wealth_index
                max_drawdown_value = None
                if hasattr(portfolio, 'wealth_index'):
                    try:
                        wealth_index = portfolio.wealth_index
                        if hasattr(wealth_index, 'iloc'):
                            if hasattr(wealth_index, 'columns'):
                                portfolio_values = wealth_index.iloc[:, 0]
                            else:
                                portfolio_values = wealth_index
                        else:
                            portfolio_values = wealth_index
                        
                        running_max = portfolio_values.expanding().max()
                        drawdown = (portfolio_values - running_max) / running_max
                        max_drawdown_value = drawdown.min()
                    except Exception as e:
                        self.logger.warning(f"Could not calculate max drawdown: {e}")
                        
            except Exception as e:
                self.logger.warning(f"Could not get portfolio metrics for caption: {e}")
            
            # Create caption in new format
            chart_caption = f"💼 Имя: {portfolio_symbol}\n"
            chart_caption += f"💵 Базовая валюта: {currency}\n"
            chart_caption += f"📊 Состав:\n"
            
            # Add portfolio composition
            for i, symbol in enumerate(symbols):
                symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                weight = weights[i] if i < len(weights) else 0.0
                chart_caption += f"• {symbol_name} — {weight:.1%}\n"
            
            # Add key metrics
            chart_caption += f"📊 Ключевые показатели:\n"
            
            if cagr_value is not None:
                chart_caption += f"• CAGR: {cagr_value:+.1%}\n"
            else:
                chart_caption += f"• CAGR: Недоступно\n"
                
            if volatility_value is not None:
                chart_caption += f"• Волатильность: {volatility_value:.1%}\n"
            else:
                chart_caption += f"• Волатильность: Недоступно\n"
                
            if sharpe_value is not None:
                chart_caption += f"• Sharpe: {sharpe_value:.2f}\n"
            else:
                chart_caption += f"• Sharpe: Недоступно\n"
                
            if max_drawdown_value is not None:
                chart_caption += f"• Макс. просадка: {max_drawdown_value:.1%}\n"
            else:
                chart_caption += f"• Макс. просадка: Недоступно\n"


            # Create portfolio reply keyboard
            portfolio_reply_keyboard = self._create_portfolio_reply_keyboard()
            
            # Send the chart with caption and portfolio keyboard
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=io.BytesIO(img_bytes),
                caption=self._truncate_caption(chart_caption),
                reply_markup=portfolio_reply_keyboard
            )
            
            # Update user context to track active keyboard
            user_id = update.effective_user.id
            self._update_user_context(user_id, active_reply_keyboard="portfolio")
            self.logger.info("Portfolio reply keyboard set with chart")
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio wealth chart with info: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика накопленной доходности: {str(e)}")

    async def _create_portfolio_wealth_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str, weights: list, portfolio_name: str = None):
        """Create and send portfolio wealth chart"""
        try:
            self.logger.info(f"Creating portfolio wealth chart for portfolio: {symbols}")
            
            # Generate wealth chart using chart_styles
            wealth_index = portfolio.wealth_index
            
            # Create portfolio chart with chart_styles using optimized method
            fig, ax = chart_styles.create_portfolio_wealth_chart(
                data=wealth_index, symbols=symbols, currency=currency, weights=weights, portfolio_name=portfolio_name
            )
            
            # Save chart to bytes with memory optimization
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            img_buffer.seek(0)
            img_bytes = img_buffer.getvalue()
            
            # Clear matplotlib cache to free memory
            chart_styles.cleanup_figure(fig)
            
            # Build caption with weights in title
            symbols_with_weights = []
            for i, symbol in enumerate(symbols):
                symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                weight = weights[i] if i < len(weights) else 0.0
                symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
            

            
            # Get final portfolio value safely
            try:
                final_value = portfolio.wealth_index.iloc[-1]
                
                # Handle different types of final_value
                if hasattr(final_value, '__iter__') and not isinstance(final_value, str):
                    if hasattr(final_value, 'iloc'):
                        final_value = final_value.iloc[0]
                    elif hasattr(final_value, '__getitem__'):
                        final_value = final_value[0]
                    else:
                        final_value = list(final_value)[0]
                
                # Convert to float safely
                if isinstance(final_value, (int, float)):
                    final_value = float(final_value)
                else:
                    final_value_str = str(final_value)
                    try:
                        final_value = float(final_value_str)
                    except (ValueError, TypeError):
                        import re
                        numeric_match = re.search(r'[\d.]+', final_value_str)
                        if numeric_match:
                            final_value = float(numeric_match.group())
                        else:
                            raise ValueError(f"Cannot convert {final_value} to float")
                
                # Add period information
                try:
                    period_length = portfolio.period_length
                except Exception as e:
                    self.logger.warning(f"Could not get period length: {e}")
            except Exception as e:
                self.logger.warning(f"Could not get final portfolio value: {e}")
            
            caption = f"При условии инвестирования 1000 {currency} за {period_length} лет накопленная доходность составила: {final_value:.2f} {currency}"

            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=io.BytesIO(img_bytes),
                caption=self._truncate_caption(caption),
            )
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio wealth chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика накопленной доходности: {str(e)}")

    async def _handle_portfolio_wealth_chart_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio wealth chart button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio wealth chart by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                error_msg = f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново."
                await self._send_callback_message(update, context, error_msg)
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            period = portfolio_info.get('period')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}, period={period}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            await self._send_ephemeral_message(update, context, "📈 Создаю график накопленной доходности...", delete_after=3)
            
            # Create portfolio with period if specified
            if period:
                years = int(period[:-1])  # Extract number from '5Y'
                from datetime import timedelta
                end_date = datetime.now()
                start_date = end_date - timedelta(days=years * 365)
                portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency,
                                       first_date=start_date.strftime('%Y-%m-%d'), 
                                       last_date=end_date.strftime('%Y-%m-%d'))
                self.logger.info(f"Created portfolio with period {period}")
            else:
                portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency)
                self.logger.info(f"Created portfolio with maximum available period")
            
            await self._create_portfolio_wealth_chart(update, context, portfolio, symbols, currency, weights, portfolio_symbol)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio wealth chart by symbol: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика накопленной доходности: {str(e)}")

    async def _handle_portfolio_rolling_cagr_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle portfolio rolling CAGR button click"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio rolling CAGR button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            self.logger.info(f"Creating rolling CAGR chart for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "📈 Создаю график Rolling CAGR...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_portfolio_rolling_cagr_chart(update, context, portfolio, final_symbols, currency, weights)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio rolling CAGR button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика Rolling CAGR: {str(e)}")

    async def _handle_portfolio_rolling_cagr_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio rolling CAGR button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio rolling CAGR by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            await self._send_ephemeral_message(update, context, "📈 Создаю график Rolling CAGR...", delete_after=3)
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_portfolio_rolling_cagr_chart(update, context, portfolio, final_symbols, currency, weights)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio rolling CAGR by symbol: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика Rolling CAGR: {str(e)}")

    async def _create_portfolio_rolling_cagr_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str, weights: list):
        """Create and send portfolio rolling CAGR chart"""
        try:
            self.logger.info(f"Creating portfolio rolling CAGR chart for portfolio: {symbols}")
            
            # Get rolling CAGR data
            rolling_cagr_data = portfolio.get_rolling_cagr()
            
            # Create standardized rolling CAGR chart using chart_styles
            fig, ax = chart_styles.create_portfolio_rolling_cagr_chart(
                data=rolling_cagr_data, symbols=symbols, currency=currency, weights=weights
            )
            
            # Save the figure using standardized method
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            chart_styles.cleanup_figure(fig)
            img_buffer.seek(0)
            
            # Get rolling CAGR statistics
            try:
                # Get rolling CAGR data for statistics
                rolling_cagr_series = portfolio.get_rolling_cagr()
                
                # Calculate statistics
                mean_rolling_cagr = rolling_cagr_series.mean()
                std_rolling_cagr = rolling_cagr_series.std()
                min_rolling_cagr = rolling_cagr_series.min()
                max_rolling_cagr = rolling_cagr_series.max()
                current_rolling_cagr = rolling_cagr_series.iloc[-1] if not rolling_cagr_series.empty else None
                
                # Build enhanced caption with weights in title
                symbols_with_weights = []
                for i, symbol in enumerate(symbols):
                    symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                    weight = weights[i] if i < len(weights) else 0.0
                    symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
                
                caption = f"📈 Rolling CAGR (MAX период) портфеля: {', '.join(symbols_with_weights)}\n\n"
                caption += f"📊 Параметры:\n"
                caption += f"• Валюта: {currency}\n"
                caption += f"• Окно: макс. период (весь доступный период)\n\n"
                
                # Add rolling CAGR statistics
                caption += f"📈 Статистика Rolling CAGR:\n"
                if current_rolling_cagr is not None:
                    caption += f"• Текущий Rolling CAGR: {current_rolling_cagr:.2%}\n"
                caption += f"• Средний Rolling CAGR: {mean_rolling_cagr:.2%}\n"
                caption += f"• Стандартное отклонение: {std_rolling_cagr:.2%}\n"
                caption += f"• Минимальный: {min_rolling_cagr:.2%}\n"
                caption += f"• Максимальный: {max_rolling_cagr:.2%}\n\n"
                
            except Exception as e:
                self.logger.warning(f"Could not get rolling CAGR statistics: {e}")
                # Fallback to basic caption
                caption = f"📈 Rolling CAGR (MAX период) портфеля: {', '.join(symbols)}\n\n"
                caption += f"📊 Параметры:\n"
                caption += f"• Валюта: {currency}\n"
                caption += f"• Веса: {', '.join([f'{w:.1%}' for w in weights])}\n"
                caption += f"• Окно: макс. период (весь доступный период)\n\n"
            
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(caption),
            )
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio rolling CAGR chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика Rolling CAGR: {str(e)}")

    async def _handle_portfolio_compare_assets_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, symbols: list):
        """Handle portfolio compare assets button click"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio compare assets button for user {user_id}")
            
            user_context = self._get_user_context(user_id)
            self.logger.info(f"User context content: {user_context}")
            
            # Prefer symbols passed from the button payload; fallback to context
            button_symbols = symbols
            final_symbols = button_symbols or user_context.get('current_symbols') or user_context.get('last_assets')
            self.logger.info(f"Available keys in user context: {list(user_context.keys())}")
            self.logger.info(f"Button symbols: {button_symbols}")
            self.logger.info(f"Final symbols: {final_symbols}")
            self.logger.info(f"Current symbols from context: {user_context.get('current_symbols')}")
            self.logger.info(f"Last assets from context: {user_context.get('last_assets')}")
            
            if not final_symbols:
                self.logger.warning("No symbols provided by button and none found in context")
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены. Выполните команду /portfolio заново.")
                return
            
            # Check if we have portfolio-specific data
            portfolio_weights = user_context.get('portfolio_weights', [])
            portfolio_currency = user_context.get('current_currency', 'USD')
            
            # Try to find portfolio name from saved portfolios
            portfolio_name = None
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Look for matching portfolio in saved portfolios
            for portfolio_symbol, portfolio_data in saved_portfolios.items():
                if portfolio_data.get('symbols') == final_symbols:
                    portfolio_name = portfolio_data.get('portfolio_name')
                    self.logger.info(f"Found matching portfolio: {portfolio_symbol} with name: {portfolio_name}")
                    break
            
            # If we have portfolio weights, use them; otherwise use equal weights
            if portfolio_weights and len(portfolio_weights) == len(final_symbols):
                weights = portfolio_weights
                currency = portfolio_currency
                self.logger.info(f"Using stored portfolio weights: {weights}")
            else:
                # Fallback to equal weights if no portfolio weights found
                weights = self._normalize_or_equalize_weights(final_symbols, [])
                currency = user_context.get('current_currency', 'USD')
                self.logger.info(f"Using equal weights as fallback: {weights}")
            
            # Filter out None values and empty strings
            final_symbols = [s for s in final_symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            self.logger.info(f"Creating compare assets chart for portfolio: {final_symbols}, currency: {currency}, weights: {weights}")
            await self._send_ephemeral_message(update, context, "⚖️ Создаю график сравнения с активами...", delete_after=3)
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_portfolio_compare_assets_chart(update, context, portfolio, final_symbols, currency, weights, portfolio_name)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio compare assets button: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика сравнения: {str(e)}")

    async def _handle_portfolio_compare_assets_by_symbol(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio compare assets button click by portfolio symbol"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio compare assets by symbol for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            portfolio_name = portfolio_info.get('portfolio_name')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}, name={portfolio_name}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            await self._send_ephemeral_message(update, context, "⚖️ Создаю график сравнения с активами...", delete_after=3)
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            self.logger.info(f"Filtered symbols: {final_symbols}")
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Debug logging
                    self.logger.info(f"Validating symbol {i}: '{symbol}' (type: {type(symbol)})")
                    
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create Portfolio with validated symbols and period
            portfolio = self._create_portfolio_with_period(valid_symbols, valid_weights, currency, user_context)
            
            await self._create_portfolio_compare_assets_chart(update, context, portfolio, final_symbols, currency, weights, portfolio_name)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio compare assets by symbol: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика сравнения: {str(e)}")

    async def _create_portfolio_compare_assets_chart(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio, symbols: list, currency: str, weights: list, portfolio_name: str = None):
        """Create and send portfolio compare assets chart"""
        try:
            self.logger.info(f"Creating portfolio compare assets chart for portfolio: {symbols}")
            
            # Get wealth index with assets data
            compare_data = portfolio.wealth_index_with_assets
            
            # Create standardized comparison chart using chart_styles
            fig, ax = chart_styles.create_portfolio_compare_assets_chart(
                data=compare_data, symbols=symbols, currency=currency, weights=weights
            )
            
            # Save the figure using standardized method
            img_buffer = io.BytesIO()
            chart_styles.save_figure(fig, img_buffer)
            chart_styles.cleanup_figure(fig)
            img_buffer.seek(0)
            
            # Get portfolio comparison statistics
            try:
                # Build enhanced caption with weights in title
                symbols_with_weights = []
                for i, symbol in enumerate(symbols):
                    symbol_name = symbol.split('.')[0] if '.' in symbol else symbol
                    weight = weights[i] if i < len(weights) else 0.0
                    symbols_with_weights.append(f"{symbol_name} ({weight:.1%})")
                
                caption = f"📊 Портфель vs Активы: {', '.join(symbols_with_weights)}\n\n"
                caption += f"📊 Параметры:\n"
                caption += f"• Валюта: {currency}\n\n"
                
                # Add portfolio performance vs individual assets
                portfolio_final = portfolio.wealth_index.iloc[-1]
                caption += f"📈 Итоговые значения (накопленная доходность):\n"
                caption += f"• Портфель: {portfolio_final:.2f}\n"
                
                # Get individual asset final values
                for symbol in symbols:
                    try:
                        # Validate symbol before creating Asset
                        if not symbol or symbol.strip() == '':
                            self.logger.warning(f"Empty symbol: '{symbol}'")
                            caption += f"• {symbol}: недоступно\n"
                            continue
                        
                        # Check for invalid characters
                        invalid_chars = ['(', ')', ',']
                        if any(char in symbol for char in invalid_chars):
                            self.logger.warning(f"Invalid symbol contains brackets: '{symbol}'")
                            caption += f"• {symbol}: недоступно\n"
                            continue
                        
                        # Check for proper format
                        if '.' not in symbol:
                            self.logger.warning(f"Symbol missing namespace separator: '{symbol}'")
                            caption += f"• {symbol}: недоступно\n"
                            continue
                        
                        # Get individual asset
                        asset = ok.Asset(symbol, ccy=currency)
                        
                        # Calculate wealth index from price data
                        price_data = asset.price
                        self.logger.info(f"DEBUG: Price data type for {symbol}: {type(price_data)}")
                        
                        # Handle different types of price data
                        if price_data is None:
                            caption += f"• {symbol}: недоступно\n"
                        elif isinstance(price_data, (int, float)):
                            # Single price value - use it directly
                            self.logger.info(f"DEBUG: Single price value for {symbol}: {price_data}")
                            asset_final = float(price_data)
                            caption += f"• {symbol}: {asset_final:.2f}\n"
                        elif hasattr(price_data, '__len__') and len(price_data) > 0:
                            # Time series data - calculate cumulative returns
                            self.logger.info(f"DEBUG: Time series data for {symbol}, length: {len(price_data)}")
                            returns = price_data.pct_change().dropna()
                            wealth_index = (1 + returns).cumprod()
                            asset_final = wealth_index.iloc[-1]
                            caption += f"• {symbol}: {asset_final:.2f}\n"
                        else:
                            caption += f"• {symbol}: недоступно\n"
                    except Exception as e:
                        self.logger.warning(f"Could not get final value for {symbol}: {e}")
                        caption += f"• {symbol}: недоступно\n"
                
            except Exception as e:
                self.logger.warning(f"Could not get comparison statistics: {e}")
                # Fallback to basic caption
                caption = f"📊 Портфель vs Активы: {', '.join(symbols)}\n\n"
                caption += f"📊 Параметры:\n"
                caption += f"• Валюта: {currency}\n"
                caption += f"• Веса: {', '.join([f'{w:.1%}' for w in weights])}\n\n"
            
            # Ensure portfolio keyboard is shown
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo=img_buffer,
                caption=self._truncate_caption(caption),
            )
            
        except Exception as e:
            self.logger.error(f"Error creating portfolio compare assets chart: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании графика сравнения: {str(e)}")

    async def _handle_portfolio_ai_analysis_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio AI analysis button click"""
        try:
            user_id = update.effective_user.id
            self.logger.info(f"Handling portfolio AI analysis for user {user_id}, portfolio: {portfolio_symbol}")
            
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель '{portfolio_symbol}' не найден. Создайте портфель заново.", parse_mode='Markdown')
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            
            self.logger.info(f"Retrieved portfolio data: symbols={symbols}, weights={weights}, currency={currency}")
            
            if not symbols:
                await self._send_callback_message(update, context, "❌ Данные о портфеле не найдены.")
                return
            
            # Check if Gemini service is available
            if not self.gemini_service or not self.gemini_service.is_available():
                await self._send_callback_message(update, context, "❌ Сервис анализа данных недоступен.", parse_mode='Markdown')
                return
            
            await self._send_ephemeral_message(update, context, "🤖 Анализирую портфель...", parse_mode='Markdown', delete_after=3)
            
            # Filter out None values and empty strings
            final_symbols = [s for s in symbols if s is not None and str(s).strip()]
            if not final_symbols:
                self.logger.warning("All symbols were None or empty after filtering")
                await self._send_callback_message(update, context, "❌ Все символы пустые или недействительны.")
                return
            
            # Validate symbols before creating portfolio
            valid_symbols = []
            valid_weights = []
            invalid_symbols = []
            
            for i, symbol in enumerate(final_symbols):
                try:
                    # Test if symbol exists in database
                    test_asset = ok.Asset(symbol)
                    # If asset was created successfully, consider it valid
                    valid_symbols.append(symbol)
                    valid_weights.append(weights[i])
                    self.logger.info(f"Symbol {symbol} validated successfully")
                except Exception as e:
                    invalid_symbols.append(symbol)
                    self.logger.warning(f"Symbol {symbol} is invalid: {e}")
            
            if not valid_symbols:
                error_msg = f"❌ Все символы недействительны: {', '.join(invalid_symbols)}"
                if any('.FX' in s for s in invalid_symbols):
                    error_msg += "\n\n💡 Валютные пары (.FX) могут быть недоступны в базе данных okama."
                await self._send_callback_message(update, context, error_msg)
                return
            
            if invalid_symbols:
                await self._send_callback_message(update, context, f"⚠️ Некоторые символы недоступны: {', '.join(invalid_symbols)}")
            
            # Normalize weights for valid symbols
            if valid_weights:
                total_weight = sum(valid_weights)
                if total_weight > 0:
                    valid_weights = [w / total_weight for w in valid_weights]
                else:
                    valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            else:
                valid_weights = [1.0 / len(valid_symbols)] * len(valid_symbols)
            
            # Create portfolio object for analysis
            try:
                portfolio = ok.Portfolio(valid_symbols, weights=valid_weights, ccy=currency)
            except Exception as e:
                self.logger.error(f"Failed to create portfolio object: {e}")
                await self._send_callback_message(update, context, f"❌ Ошибка создания портфеля: {str(e)}", parse_mode='Markdown')
                return
            
            # Prepare portfolio data for analysis using the new specialized function
            try:
                portfolio_data = await self._prepare_portfolio_data_for_analysis(
                    portfolio=portfolio,
                    symbols=valid_symbols,
                    weights=valid_weights,
                    currency=currency,
                    user_id=user_id
                )
                
                # Analyze portfolio with Gemini using the new portfolio analysis method
                portfolio_analysis = self.gemini_service.analyze_portfolio(portfolio_data)
                
                if portfolio_analysis and portfolio_analysis.get('success'):
                    analysis_text = portfolio_analysis.get('analysis', '')
                    
                    if analysis_text:
                        await self._send_portfolio_ai_analysis_with_keyboard(update, context, analysis_text, parse_mode='Markdown')
                    else:
                        await self._send_portfolio_message_with_reply_keyboard(update, context, "🤖 Анализ портфеля выполнен, но результат пуст", parse_mode='Markdown')
                        
                else:
                    error_msg = portfolio_analysis.get('error', 'Неизвестная ошибка') if portfolio_analysis else 'Анализ не выполнен'
                    await self._send_portfolio_message_with_reply_keyboard(update, context, f"❌ Ошибка анализа портфеля: {error_msg}", parse_mode='Markdown')
                    
            except Exception as data_error:
                self.logger.error(f"Error preparing data for portfolio analysis: {data_error}")
                await self._send_portfolio_message_with_reply_keyboard(update, context, f"❌ Ошибка при подготовке данных для анализа: {str(data_error)}", parse_mode='Markdown')
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio AI analysis: {e}")
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            await self._send_callback_message(update, context, f"❌ Ошибка при анализе портфеля: {str(e)}", parse_mode='Markdown')

    async def _handle_portfolio_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio compare button click - redirect to /compare command without arguments"""
        try:
            # Remove buttons from the old message
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception as e:
                self.logger.warning(f"Could not remove buttons from old message: {e}")
            
            # Call the compare command without arguments
            await self.compare_command(update, context)
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio compare button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при переходе к сравнению: {str(e)}")

    async def _handle_portfolio_main_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, portfolio_symbol: str):
        """Handle portfolio main button click - show portfolio main information with keyboard"""
        try:
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            
            # Get portfolio data from saved portfolios
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Use the new portfolio finder function
            found_portfolio_key = self._find_portfolio_by_symbol(portfolio_symbol, saved_portfolios, user_id)
            
            if not found_portfolio_key:
                await self._send_callback_message(update, context, f"❌ Портфель {portfolio_symbol} не найден в сохраненных портфелях")
                return
            
            # Use the found portfolio key
            portfolio_symbol = found_portfolio_key
            
            portfolio_info = saved_portfolios[portfolio_symbol]
            symbols = portfolio_info.get('symbols', [])
            weights = portfolio_info.get('weights', [])
            currency = portfolio_info.get('currency', 'USD')
            
            # Create portfolio information text
            portfolio_text = f"💼 **Портфель {portfolio_symbol}**\n\n"
            
            # Add portfolio composition
            portfolio_text += "💼 **Состав портфеля:**\n"
            for i, (symbol, weight) in enumerate(zip(symbols, weights)):
                portfolio_text += f"• {symbol}: {weight:.1%}\n"
            
            portfolio_text += f"\n💰 **Валюта:** {currency}\n"
            
            # Add basic metrics if available
            try:
                # Create portfolio for metrics calculation
                portfolio = ok.Portfolio(symbols, weights=weights, ccy=currency)
                metrics_text = self._get_portfolio_basic_metrics(portfolio, symbols, weights, currency)
                portfolio_text += metrics_text
            except Exception as e:
                self.logger.warning(f"Could not add metrics to portfolio text: {e}")
            
            portfolio_text += f"\n💼 Сравнить портфель с другими активами: `/compare {portfolio_symbol}`\n"
            
            # Ensure portfolio keyboard is shown and send message
            await self._manage_reply_keyboard(update, context, "portfolio")
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text=portfolio_text,
                parse_mode='Markdown',
            )
            
        except Exception as e:
            self.logger.error(f"Error handling portfolio main button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при отображении портфеля: {str(e)}")

    async def _handle_namespace_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, namespace: str):
        """Handle namespace button click - show symbols in specific namespace with reply keyboard"""
        try:
            self.logger.info(f"Handling namespace button for: {namespace}")
            
            # Check if it's a Chinese exchange
            chinese_exchanges = ['SSE', 'SZSE', 'BSE', 'HKEX']
            if namespace in chinese_exchanges:
                await self._show_tushare_namespace_symbols_with_reply_keyboard(update, context, namespace, page=0)
            else:
                await self._show_namespace_symbols_with_reply_keyboard(update, context, namespace, page=0)
                
        except ImportError:
            await self._send_callback_message(update, context, "❌ Библиотека okama не установлена")
        except Exception as e:
            self.logger.error(f"Error in namespace button handler: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка: {str(e)}")

    async def _handle_excel_namespace_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE, namespace: str):
        """Handle Excel export button click for namespace"""
        try:
            self.logger.info(f"Handling Excel export for namespace: {namespace}")
            
            # Check if it's a Chinese exchange
            chinese_exchanges = ['SSE', 'SZSE', 'BSE', 'HKEX']
            if namespace in chinese_exchanges:
                await self._handle_tushare_excel_export(update, context, namespace)
                return
            
            # Get symbols in namespace for non-Chinese exchanges
            try:
                symbols_df = ok.symbols_in_namespace(namespace)
                
                # Check if DataFrame is empty
                if symbols_df.empty:
                    await self._send_callback_message(update, context, f"❌ Пространство имен '{namespace}' не найдено или пусто")
                    return
                
                total_symbols = len(symbols_df)
                
                # Show progress message
                await self._send_ephemeral_message(update, context, f"📊 Создаю Excel файл...", delete_after=3)
                
                # Create Excel file in memory
                excel_buffer = io.BytesIO()
                symbols_df.to_excel(excel_buffer, index=False, sheet_name=f'{namespace}_Symbols')
                excel_buffer.seek(0)
                
                # Send Excel file
                await context.bot.send_document(
                    chat_id=update.effective_chat.id,
                    document=excel_buffer,
                    filename=f"{namespace}_symbols.xlsx",
                    caption=self._truncate_caption(f"📊 Полный список символов в пространстве {namespace} ({total_symbols})")
                )
                
                excel_buffer.close()
                
            except Exception as e:
                await self._send_callback_message(update, context, f"❌ Ошибка при получении символов для '{namespace}': {str(e)}")
                
        except ImportError:
            await self._send_callback_message(update, context, "❌ Библиотека okama не установлена")
        except Exception as e:
            self.logger.error(f"Error in Excel namespace button handler: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка: {str(e)}")

    async def _handle_tushare_excel_export(self, update: Update, context: ContextTypes.DEFAULT_TYPE, namespace: str):
        """Handle Excel export for Chinese exchanges using Tushare"""
        try:
            if not self.tushare_service:
                await self._send_callback_message(update, context, "❌ Сервис Tushare недоступен")
                return
            
            # Show progress message
            await self._send_ephemeral_message(update, context, f"📊 Создаю Excel файл для {namespace}...", delete_after=3)
            
            # Get ALL symbols data from Tushare (no limit for Excel export)
            symbols_data = self.tushare_service.get_exchange_symbols_full(namespace)
            total_count = len(symbols_data)
            
            if not symbols_data:
                await self._send_callback_message(update, context, f"❌ Символы для биржи '{namespace}' не найдены")
                return
            
            # Create DataFrame from symbols data
            df = pd.DataFrame(symbols_data)
            
            # Add additional columns for better Excel formatting
            df['Exchange'] = namespace
            df['Exchange_Name'] = {
                'SSE': 'Shanghai Stock Exchange',
                'SZSE': 'Shenzhen Stock Exchange',
                'BSE': 'Beijing Stock Exchange',
                'HKEX': 'Hong Kong Stock Exchange'
            }.get(namespace, namespace)
            
            # Reorder columns
            df = df[['symbol', 'name', 'currency', 'list_date', 'Exchange', 'Exchange_Name']]
            
            # Rename columns for better readability
            df.columns = ['Symbol', 'Company Name', 'Currency', 'List Date', 'Exchange Code', 'Exchange Name']
            
            # Create Excel file in memory
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=f'{namespace}_Symbols')
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets[f'{namespace}_Symbols']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            
            # Send Excel file
            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=excel_buffer,
                filename=f"{namespace}_symbols.xlsx",
                caption=self._truncate_caption(f"📊 Полный список символов биржи {namespace} ({total_count:,} символов)")
            )
            
            excel_buffer.close()
            
        except Exception as e:
            self.logger.error(f"Error in Tushare Excel export for {namespace}: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка при создании Excel файла: {str(e)}")

    async def _handle_namespace_home_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle namespace home button click - show main namespace list"""
        try:
            self.logger.info("Handling namespace home button")
            
            # Show available namespaces (same as /list command without args)
            import okama as ok
            namespaces = ok.namespaces
            
            # Prepare data for tabulate
            headers = ["Код", "Описание", "Категория"]
            namespace_data = []
            
            # Categorize namespaces for better organization
            categories = {
                'Биржи': ['MOEX', 'US', 'LSE', 'XAMS', 'XETR', 'XFRA', 'XSTU', 'XTAE', 'SSE', 'SZSE', 'BSE', 'HKEX'],
                'Индексы': ['INDX'],
                'Валюты': ['FX', 'CBR'],
                'Товары': ['COMM'],
                'Криптовалюты': ['CC'],
                'Инфляция': ['INFL'],
                'Недвижимость': ['RE'],
                'Портфели': ['PF', 'PIF'],
                'Депозиты': ['RATE'],
                'Коэффициенты': ['RATIO']
            }
            
            # Create categorized data
            for namespace, description in namespaces.items():
                category = "Другое"
                for cat_name, cat_namespaces in categories.items():
                    if namespace in cat_namespaces:
                        category = cat_name
                        break
                
                namespace_data.append([namespace, description, category])
            
            # Add Chinese exchanges manually (not in ok.namespaces)
            chinese_exchanges = {
                'SSE': 'Shanghai Stock Exchange',
                'SZSE': 'Shenzhen Stock Exchange', 
                'BSE': 'Beijing Stock Exchange',
                'HKEX': 'Hong Kong Stock Exchange'
            }
            
            for exchange_code, exchange_name in chinese_exchanges.items():
                namespace_data.append([exchange_code, exchange_name, 'Биржи'])
            
            # Sort by category and then by namespace
            namespace_data.sort(key=lambda x: (x[2], x[0]))
            response = f"📚 Доступные пространства имен (namespaces): {len(namespaces)}\n\n"
            
            # Create table using tabulate or fallback to simple format
            if TABULATE_AVAILABLE:
                # Use plain format for best Telegram display
                table = tabulate.tabulate(namespace_data, headers=headers, tablefmt="plain")
                response += f"```\n{table}\n```\n\n"
            else:
                # Fallback to simple text format
                response += "Код | Описание | Категория\n"
                response += "--- | --- | ---\n"
                for row in namespace_data:
                    response += f"`{row[0]}` | {row[1]} | {row[2]}\n"
                response += "\n"
            
            
            # Создаем кнопки для основных пространств имен
                keyboard = []
            
            # Основные биржи
            keyboard.append([
                InlineKeyboardButton("🇷🇺 MOEX", callback_data="namespace_MOEX"),
                InlineKeyboardButton("🇷🇺 PIF", callback_data="namespace_PIF"),
                InlineKeyboardButton("💱 FX", callback_data="namespace_FX")
            ])

            keyboard.append([
                InlineKeyboardButton("🇺🇸 US", callback_data="namespace_US"),
                InlineKeyboardButton("🇬🇧 LSE", callback_data="namespace_LSE"),                
                InlineKeyboardButton("🇭🇰 HKEX", callback_data="namespace_HKEX")
            ])

            # Китайские биржи
            keyboard.append([
                InlineKeyboardButton("🇨🇳 SSE", callback_data="namespace_SSE"),
                InlineKeyboardButton("🇨🇳 SZSE", callback_data="namespace_SZSE"),
                InlineKeyboardButton("🇨🇳 BSE", callback_data="namespace_BSE")
            ])

            keyboard.append([
                InlineKeyboardButton("🇩🇪 XETR", callback_data="namespace_XETR"),
                InlineKeyboardButton("🇩🇪 XFRA", callback_data="namespace_XFRA"),
                InlineKeyboardButton("🇩🇪 XSTU", callback_data="namespace_XSTU")
            ])

            keyboard.append([
                InlineKeyboardButton("🇮🇱 XTAE", callback_data="namespace_XTAE"),
                InlineKeyboardButton("🇳🇱 XAMS", callback_data="namespace_XAMS"),
                InlineKeyboardButton("📊 INDX", callback_data="namespace_INDX")
            ])
            
            # Товары и криптовалюты
            keyboard.append([
                InlineKeyboardButton("🛢️ COMM", callback_data="namespace_COMM"),
                InlineKeyboardButton("₿ CC", callback_data="namespace_CC"),
                InlineKeyboardButton("🏠 RE", callback_data="namespace_RE")
            ])
            
            # Инфляция и депозиты
            keyboard.append([
                InlineKeyboardButton("📈 INFL", callback_data="namespace_INFL"),
                InlineKeyboardButton("🏦 RATE", callback_data="namespace_RATE"),
                InlineKeyboardButton("🏦 CBR", callback_data="namespace_CBR")     
            ])
            
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            # Send message via callback with reply_markup
            await context.bot.send_message(
                chat_id=update.callback_query.message.chat_id,
                text=response,
                parse_mode='Markdown',
                reply_markup=reply_markup
            )
                
        except Exception as e:
            self.logger.error(f"Error handling namespace home button: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка: {str(e)}")

    async def _handle_namespace_analysis_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle namespace analysis button click - call info command directly"""
        try:
            self.logger.info("Handling namespace analysis button")
            
            # Call info command without arguments directly
            context.args = []
            await self.info_command(update, context)
                
        except Exception as e:
            self.logger.error(f"Error in namespace analysis button handler: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка: {str(e)}")

    async def _handle_namespace_compare_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle namespace compare button click - show compare command help"""
        try:
            self.logger.info("Handling namespace compare button")
            
            # Get user's saved portfolios for help message (same as compare command)
            user_id = update.effective_user.id
            user_context = self._get_user_context(user_id)
            saved_portfolios = user_context.get('saved_portfolios', {})
            
            # Clear any existing compare context when starting fresh
            self._update_user_context(user_id, compare_first_symbol=None, waiting_for_compare=False)
            
            # Get random examples for user
            examples = self.get_random_examples(3)
            examples_text = ", ".join(examples)
            
            help_text = "⚖️ Сравнение\n\n"
            help_text += f"Примеры команд: {examples_text}\n\n"

            # Add saved portfolios information
            if saved_portfolios:
                help_text += "💼 *Сохраненные портфели:*\n"
                for name, portfolio_data in saved_portfolios.items():
                    symbols = portfolio_data.get('symbols', [])
                    weights = portfolio_data.get('weights', [])
                    if symbols and weights:
                        portfolio_str = " ".join([f"{s}:{w:.1f}" for s, w in zip(symbols, weights)])
                        help_text += f"• `{name}`: {portfolio_str}\n"
                help_text += "\n"

            help_text += "💡 *Способы сравнения:*\n"
            help_text += "• `/compare AAPL.US MSFT.US` - сравнение активов\n"
            help_text += "• `/compare Портфель1 Портфель2` - сравнение портфелей\n"
            help_text += "• `/compare AAPL.US Портфель1` - актив vs портфель\n\n"
            help_text += "💬 Введите тикеры для сравнения:"
            
            # Send the same message as compare command without arguments
            await self._send_callback_message(update, context, help_text)
                
        except Exception as e:
            self.logger.error(f"Error in namespace compare button handler: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка: {str(e)}")

    async def _handle_namespace_portfolio_button(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle namespace portfolio button click - call portfolio command directly"""
        try:
            self.logger.info("Handling namespace portfolio button")
            
            # Call portfolio command without arguments directly
            context.args = []
            await self.portfolio_command(update, context)
                
        except Exception as e:
            self.logger.error(f"Error in namespace portfolio button handler: {e}")
            await self._send_callback_message(update, context, f"❌ Ошибка: {str(e)}")



    def run(self):
        """Run the bot"""
        # Create application
        application = Application.builder().token(Config.TELEGRAM_BOT_TOKEN).build()
        
        # Add handlers
        application.add_handler(CommandHandler("start", self.start_command))
        application.add_handler(CommandHandler("help", self.help_command))
        application.add_handler(CommandHandler("info", self.info_command))
        application.add_handler(CommandHandler("list", self.namespace_command))
        application.add_handler(CommandHandler("search", self.search_command))
        application.add_handler(CommandHandler("compare", self.compare_command))
        application.add_handler(CommandHandler("portfolio", self.portfolio_command))
        
        # Add callback query handler for buttons
        application.add_handler(CallbackQueryHandler(self.button_callback))
        
        # Add message handler for waiting user input after empty /info
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, self.handle_message))
        
        # Start the bot
        logger.info("Starting Okama Finance Bot...")
        application.run_polling()

    def _check_existing_portfolio(self, symbols: List[str], weights: List[float], saved_portfolios: Dict) -> Optional[str]:
        """
        Проверяет, существует ли портфель с такими же активами и пропорциями.
        
        Args:
            symbols: Список символов активов
            weights: Список весов активов
            saved_portfolios: Словарь сохраненных портфелей
            
        Returns:
            Символ существующего портфеля или None, если не найден
        """
        # Нормализуем веса для сравнения (сумма = 1.0)
        total_weight = sum(weights)
        normalized_weights = [w / total_weight for w in weights]
        
        for portfolio_symbol, portfolio_info in saved_portfolios.items():
            existing_symbols = portfolio_info.get('symbols', [])
            existing_weights = portfolio_info.get('weights', [])
            
            # Проверяем количество активов
            if len(symbols) != len(existing_symbols):
                continue
                
            # Проверяем, что все символы совпадают (с учетом регистра)
            if set(symbol.upper() for symbol in symbols) != set(symbol.upper() for symbol in existing_symbols):
                continue
                
            # Нормализуем существующие веса
            existing_total = sum(existing_weights)
            if existing_total == 0:
                continue
            normalized_existing_weights = [w / existing_total for w in existing_weights]
            
            # Проверяем, что веса совпадают с точностью до 0.001
            weight_matches = True
            for i, (new_weight, existing_weight) in enumerate(zip(normalized_weights, normalized_existing_weights)):
                if abs(new_weight - existing_weight) > 0.001:
                    weight_matches = False
                    break
                    
            if weight_matches:
                return portfolio_symbol
                
        return None

if __name__ == "__main__":
    try:
        logger.info(f"Starting Finance Bot with Python {sys.version}")
        logger.info(f"Python version info: {sys.version_info}")
        
        # Perform health check
        health_check()
        
        # Optional HTTP health server for platforms expecting an open PORT
        port_env = os.getenv('PORT')
        if port_env:
            try:
                bind_port = int(port_env)
                class HealthHandler(BaseHTTPRequestHandler):
                    def do_GET(self):
                        payload = {
                            "status": "ok",
                            "service": "okama-finance-bot",
                            "environment": "RENDER" if os.getenv('RENDER') else "LOCAL"
                        }
                        self.send_response(200)
                        self.send_header('Content-Type', 'application/json')
                        self.end_headers()
                        self.wfile.write(json.dumps(payload).encode('utf-8'))
                    def log_message(self, format, *args):
                        return
                def serve_health():
                    server = HTTPServer(('0.0.0.0', bind_port), HealthHandler)
                    logger.info(f"Health server listening on 0.0.0.0:{bind_port}")
                    server.serve_forever()
                threading.Thread(target=serve_health, daemon=True).start()
            except Exception as e:
                logger.warning(f"Failed to start health server on PORT={port_env}: {e}")
        
        if sys.version_info >= (3, 13):
            logger.info("✅ Running on Python 3.13+ with latest python-telegram-bot")
        elif sys.version_info >= (3, 12):
            logger.info("✅ Running on Python 3.12+ with latest python-telegram-bot")
        
        logger.info("🚀 Initializing bot services...")
        bot = ShansAi()
        logger.info("✅ Bot services initialized successfully")
        logger.info("🤖 Starting Telegram bot...")
        bot.run()
    except Exception as e:
        logger.error(f"❌ Fatal error starting bot: {e}")
        logger.error(f"Python version: {sys.version}")
        logger.error(f"Python executable: {sys.executable}")
        traceback.print_exc()
        sys.exit(1)

